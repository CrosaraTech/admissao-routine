"""Resolução de função via planilha CBO.

Planilha esperada (funcoes_cbo.xlsx) com colunas:
  - usar         (obrigatório; "X" pra marcar cargos que o escritório usa)
  - nome_cargo   (string — nome do cargo cadastrado no eContador)
  - cbo          (string ou int — código CBO de 6 dígitos)
  - funcao_id    (string ou int — id da função no eContador)
  - codigo       (string — código interno usado pra resolução manual via UI)

Estratégia (regra do escritório — confirmada 28/05/2026):
  - SÓ procura match nos cargos marcados com X na coluna `usar`.
  - Se não achar match alto entre os X-marcados → pendência interna.
    Operador resolve manualmente via "Resolver pendência" → campo Função.
  - NÃO faz fallback automático pra planilha completa (9k+ cargos genéricos
    do eContador). Isso evita matches enganosos como "OPERADOR DE AGLUTINADOR"
    → "OPERADOR(A) DE CAIXA" só porque ambos começam com "OPERADOR".
"""

from __future__ import annotations

import json
import logging
import re
import unicodedata
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook


log = logging.getLogger("admissao.funcao")


CONFIANCA_ALTA = 0.85
CONFIANCA_AMBIGUA = 0.65


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.lower()).strip()


def _cbo_digitos(s: str | int | None) -> str:
    return re.sub(r"\D", "", str(s or ""))


def salvar_planilha(path: Path, items: list[dict]) -> None:
    """v2.16.47: regrava funcoes_cbo.xlsx a partir da lista em memoria.
    Mantem colunas: usar, nome_cargo, cbo, funcao_id, codigo.

    Uso: UI /cbo edita/adiciona/toggle X e salva de volta pra planilha
    substituir edicao manual no Excel.
    """
    from openpyxl import Workbook as _WB
    wb = _WB()
    ws = wb.active
    ws.title = "cargos"
    ws.append(["usar", "nome_cargo", "cbo", "funcao_id", "codigo"])
    for it in items:
        ws.append([
            "X" if it.get("usar") else "",
            str(it.get("nome_cargo") or "").strip(),
            str(it.get("cbo") or "").strip(),
            str(it.get("funcao_id") or "").strip(),
            str(it.get("codigo") or "").strip(),
        ])
    # Coluna A largura pequena, resto ampla
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    wb.save(path)


def carregar_planilha(path: Path) -> list[dict]:
    """Lê funcoes_cbo.xlsx e retorna lista de {nome_cargo, cbo, funcao_id, usar, codigo}."""
    if not path.exists():
        raise FileNotFoundError(
            f"Planilha CBO não encontrada em {path}. "
            f"Crie com colunas: nome_cargo, cbo, funcao_id"
        )
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Primeira linha = cabeçalho
    header = [_norm(str(c or "")) for c in rows[0]]

    def col_idx(*nomes_aceitos: str) -> int:
        for i, h in enumerate(header):
            if h in nomes_aceitos:
                return i
        return -1

    i_nome = col_idx("nome_cargo", "nome", "cargo")
    i_cbo = col_idx("cbo")
    i_id = col_idx("funcao_id", "id", "funcaoid")
    i_usar = col_idx("usar")  # opcional
    i_codigo = col_idx("codigo")  # opcional — usado pelo dialog Resolver Pendência

    if min(i_nome, i_cbo, i_id) < 0:
        raise ValueError(
            f"Planilha {path} sem colunas obrigatórias (nome_cargo, cbo, funcao_id). "
            f"Cabeçalho encontrado: {header}"
        )

    items: list[dict] = []
    for row in rows[1:]:
        if not row or not any(row):
            continue
        nome = str(row[i_nome] or "").strip()
        cbo = _cbo_digitos(row[i_cbo])
        fid = str(row[i_id] or "").strip()
        if not nome or not fid:
            continue
        usar = False
        if i_usar >= 0:
            usar = str(row[i_usar] or "").strip().upper() == "X"
        items.append({
            "nome_cargo": nome,
            "cbo": cbo,
            "funcao_id": fid,
            "usar": usar,
            "codigo": str(row[i_codigo] or "").strip() if i_codigo >= 0 else "",
        })
    return items


def _score_linha(cargo_norm: str, cbo: str, linha: dict) -> float:
    """Score combinado: bonus alto se CBO bate, senão similaridade textual."""
    nome_norm = _norm(linha["nome_cargo"])
    s = SequenceMatcher(None, cargo_norm, nome_norm).ratio() if cargo_norm else 0.0
    # Bonus tokens: todos os tokens do cargo aparecem no nome
    if cargo_norm and all(t in nome_norm for t in cargo_norm.split()):
        s = max(s, 0.9)
    # Bonus CBO: se o CBO bate, considera quase certo
    if cbo and linha.get("cbo") == cbo:
        s = max(s, 0.95)
    return s


def _resolver_em(
    candidatos: list[dict],
    cargo_norm: str,
    cbo: str,
    contexto: str,
) -> tuple[str | None, float, list[dict], str]:
    """Roda a lógica de match num subconjunto da planilha.

    Retorna o mesmo formato de `resolver_funcao`.
    `contexto`: string descritiva pro log ("X-marcado" ou "planilha completa").
    """
    if not candidatos:
        return None, 0.0, [], f"Nenhum candidato em {contexto}"

    scores: list[tuple[float, dict]] = [
        (_score_linha(cargo_norm, cbo, f), f) for f in candidatos
    ]
    scores.sort(key=lambda x: x[0], reverse=True)
    top_score, top = scores[0]

    if top_score >= CONFIANCA_ALTA:
        # Ambiguidade: top empata com o segundo
        if len(scores) >= 2 and (top_score - scores[1][0]) < 0.05:
            empatados = [f for s, f in scores if abs(s - top_score) < 0.05]

            # Caso especial: cargos eContador tem DUPLICATAS (mesmo nome
            # exato + mesmo CBO, IDs diferentes). Pra esses, qualquer ID
            # serve — pegamos o primeiro determinístico (menor id) sem
            # incomodar o cliente.
            nome_top = _norm(top["nome_cargo"])
            cbo_top = str(top.get("cbo") or "")
            duplicatas = [
                f for f in empatados
                if _norm(f["nome_cargo"]) == nome_top
                and str(f.get("cbo") or "") == cbo_top
            ]
            if len(duplicatas) == len(empatados):
                escolhido = min(duplicatas, key=lambda f: int(str(f["funcao_id"]) or "0"))
                log.info(
                    f"[{contexto}] '{escolhido['nome_cargo']}' "
                    f"(id={escolhido['funcao_id']}, conf={top_score:.0%}) "
                    f"— {len(duplicatas)} duplicata(s) do eContador, escolhi a 1ª"
                )
                return escolhido["funcao_id"], top_score, [], "ok"

            ambiguos = [f for s, f in scores[:5] if s >= CONFIANCA_AMBIGUA]
            return (
                None, top_score, ambiguos,
                f"Match alto mas ambíguo em {contexto} ({len(ambiguos)} cargos parecidos)",
            )
        log.info(
            f"[{contexto}] '{top['nome_cargo']}' (id={top['funcao_id']}, "
            f"conf={top_score:.0%})"
        )
        return top["funcao_id"], top_score, [], "ok"

    if top_score >= CONFIANCA_AMBIGUA:
        ambiguos = [f for s, f in scores[:5] if s >= CONFIANCA_AMBIGUA]
        return (
            None, top_score, ambiguos,
            f"Match com dúvida em {contexto} (melhor: {top['nome_cargo']} {top_score:.0%})",
        )

    return None, top_score, [], (
        f"Sem match em {contexto} (melhor: '{top['nome_cargo']}' {top_score:.0%})"
    )


def resolver_funcao(
    planilha: list[dict],
    cargo_extraido: str | None,
    cbo_extraido: str | int | None = None,
    eh_estagio: bool = False,
) -> tuple[str | None, float, list[dict], str]:
    """Tenta resolver a função SOMENTE entre cargos marcados com X.

    Mudança 28/05/2026: removido fallback automático pra planilha completa.
    Se não bater nos X-marcados, vira pendência e o operador resolve pela UI.

    v2.11.0 — estagiário:
      Quando `eh_estagio=True`, busca SÓ em alias de estágio. Se não tiver
      alias salvo, vira pendência interna IMEDIATAMENTE com motivo claro —
      não tenta resolver via X-marcados (CLT) pra evitar pegar função errada
      (cliente sempre tem função separada de estágio na planilha CBO).

    Retorna (funcao_id, confianca, candidatos_ambiguos, motivo).
      - funcao_id != None → resolveu (match alto entre X-marcados ou alias)
      - funcao_id None com ambíguos → operador escolhe manualmente
      - funcao_id None sem ambíguos → operador adiciona à planilha ou marca X
    """
    if not planilha:
        return None, 0.0, [], "Planilha CBO vazia"

    cbo = _cbo_digitos(cbo_extraido)
    cargo_norm = _norm(cargo_extraido or "")

    if not cargo_norm and not cbo:
        return None, 0.0, [], "Cargo e CBO não informados"

    # Step 0: alias global — operador já mapeou esse cargo antes.
    # Pra estágio, busca SÓ na fatia de estágio (sem fallback CLT).
    alias = consultar_funcao_alias(cargo_extraido or "", eh_estagio=eh_estagio)
    if alias:
        tipo = "ESTÁGIO" if eh_estagio else "CLT"
        log.info(
            f"[alias ✓ {tipo}] '{cargo_extraido}' → {alias['nome_cargo']} "
            f"(id={alias['funcao_id']})"
        )
        return alias["funcao_id"], 1.0, [], f"ok (alias {tipo.lower()})"

    # Pra estágio sem alias: pendência IMEDIATA (não tenta resolver via CLT).
    if eh_estagio:
        return None, 0.0, [], (
            f"Estágio detectado para cargo '{cargo_extraido}', mas não há alias "
            f"de estágio salvo. Cliente costuma cadastrar função separada "
            f"(ex: 'ESTAGIÁRIO DE LOJA') no eContador. Defina o código manualmente "
            f"e marque 'Salvar como alias permanente' pra próximos estagiários "
            f"do mesmo cargo."
        )

    marcados = [f for f in planilha if f.get("usar")]
    if not marcados:
        return None, 0.0, [], (
            "Nenhum cargo marcado com X na planilha CBO. "
            "Marque os cargos do escritório com X pra habilitar resolução automática."
        )

    log.info(f"🔎 Procurando match em {len(marcados)} cargo(s) marcado(s) com X")
    fid, conf, amb, msg = _resolver_em(marcados, cargo_norm, cbo, "X-marcado")
    if fid is not None:
        return fid, conf, [], "ok"

    # Sem match alto nos X-marcados — vira pendência. NÃO tenta planilha
    # completa (evita matches enganosos tipo "OPERADOR DE AGLUTINADOR" →
    # "OPERADOR DE CAIXA"). Operador resolve via UI digitando o código.
    return None, conf, amb, msg


# ─────────────────────────────────────────────────────────────────────────────
# Aliases GLOBAIS de cargo
#
# Diferente de funcao_overrides.json (que é por msg_id + nome_funcionario),
# aliases valem pra TODAS as futuras admissões com o mesmo cargo normalizado.
# Caso de uso: cliente sempre manda "AUXILIAR DE SERVIÇOS GERAIS" e o
# escritório quer mapear pro cargo "AUXILIAR DE LIMPEZA" — operador define
# UMA VEZ e nunca mais vira pendência.
#
# Formato do JSON:
# {
#   "auxiliar de servicos gerais": {
#     "funcao_id": "12345",
#     "nome_cargo": "AUXILIAR DE LIMPEZA",
#     "criado_em": "2026-06-01T14:30:00",
#     "observacoes": "Cliente X manda sempre assim"
#   },
#   ...
# }
# ─────────────────────────────────────────────────────────────────────────────

FUNCAO_ALIASES_FILE = Path(__file__).parent / "funcao_aliases.json"

# Prefixo na chave do alias quando é estágio (v2.11.0).
# Mantém compatibilidade com aliases CLT existentes — chaves antigas
# (sem prefixo) continuam valendo. Estágios novos viram chave dedicada
# tipo "_estagio_auxiliar de loja" → funcao_id da função "ESTAGIÁRIO DE LOJA".
PREFIXO_ESTAGIO = "_estagio_"


def _chave_alias(cargo: str, eh_estagio: bool = False) -> str:
    """Normaliza cargo + aplica prefixo de estágio se for o caso."""
    base = _norm(cargo or "")
    if not base:
        return ""
    return f"{PREFIXO_ESTAGIO}{base}" if eh_estagio else base


def carregar_funcao_aliases() -> dict:
    """Lê {cargo_normalizado: {funcao_id, nome_cargo, criado_em, observacoes}}.

    Estágios usam chaves com prefixo `_estagio_` no mesmo arquivo —
    operador vê tudo num lugar só, sem split de arquivo.

    Retorna {} se arquivo não existir ou estiver corrompido.
    """
    if not FUNCAO_ALIASES_FILE.exists():
        return {}
    try:
        with FUNCAO_ALIASES_FILE.open("r", encoding="utf-8") as fh:
            data = json.load(fh)
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, OSError) as e:
        log.warning(f"Falha lendo {FUNCAO_ALIASES_FILE}: {e}")
        return {}


def salvar_funcao_alias(
    cargo: str,
    funcao_id: str,
    nome_cargo: str,
    observacoes: str = "",
    eh_estagio: bool = False,
) -> None:
    """Salva (ou atualiza) um alias global de cargo.

    Args:
        eh_estagio: quando True, salva com prefixo `_estagio_` na chave.
            Estágios não compartilham alias com CLT — função pra "AUXILIAR
            DE LOJA" CLT é uma; pra "AUXILIAR DE LOJA" estagiário é outra
            (cliente costuma criar `ESTAGIÁRIO DE LOJA` na CBO).

    SANITY CHECK (v2.11.1):
        Se `nome_cargo` contém "ESTAGIO"/"ESTAGIARIO"/"ESTAGIARIA"
        (sem acento), FORÇA `eh_estagio=True` mesmo se foi chamado com
        False. Defesa em profundidade contra o bug do caso YURI (11/06/2026):
        operador digitou função 8860 "ESTAGIO EM SUPERMERCADO" mas a UI
        salvou como CLT porque a pendência veio de versão anterior à 2.11.0
        (sem a flag `eh_estagio` no resolucao). Risco real: próximo CLT
        com cargo "auxiliar de loja" mapearia pra função de estagiário.
    """
    # Sanity check pelo NOME DA FUNÇÃO que o operador digitou no eContador.
    # Se o nome contém marcador de estágio, é estágio — independente do que
    # o caller passou. Mais forte que confiar em flag persistida.
    nome_norm = _norm(nome_cargo or "")
    contem_marcador_estagio = bool(
        re.search(r"\bestagi[ao]\b|\bestagi[ao]ri[oa]\b", nome_norm)
    )
    if contem_marcador_estagio and not eh_estagio:
        log.warning(
            f"   ⚠ SANITY CHECK: nome_cargo '{nome_cargo}' contém marcador de "
            f"estágio mas chamada veio com eh_estagio=False — FORÇANDO True. "
            f"Provável bug do caller. Defesa em profundidade ativada."
        )
        eh_estagio = True

    chave = _chave_alias(cargo, eh_estagio)
    if not chave:
        raise ValueError("Cargo vazio — não dá pra salvar alias")
    if not str(funcao_id).strip():
        raise ValueError("funcao_id vazio — não dá pra salvar alias")

    aliases = carregar_funcao_aliases()
    aliases[chave] = {
        "funcao_id": str(funcao_id).strip(),
        "nome_cargo": nome_cargo.strip(),
        "criado_em": datetime.now().isoformat(timespec="seconds"),
        "observacoes": observacoes.strip(),
        "eh_estagio": bool(eh_estagio),
    }
    with FUNCAO_ALIASES_FILE.open("w", encoding="utf-8") as fh:
        json.dump(aliases, fh, ensure_ascii=False, indent=2)
    tipo = "ESTÁGIO" if eh_estagio else "CLT"
    log.info(
        f"[alias salvo {tipo}] '{cargo}' → {nome_cargo} (id={funcao_id}) "
        f"em {FUNCAO_ALIASES_FILE.name}"
    )


def consultar_funcao_alias(cargo: str, eh_estagio: bool = False) -> dict | None:
    """Procura alias global pra um cargo. Retorna entry ou None.

    Args:
        eh_estagio: quando True, busca SÓ na fatia de estágio (chave com
            prefixo `_estagio_`). NÃO faz fallback pra CLT — função CLT
            normal é INCORRETA pra estagiário (cargo diferente no eContador).
    """
    chave = _chave_alias(cargo, eh_estagio)
    if not chave:
        return None
    aliases = carregar_funcao_aliases()
    return aliases.get(chave)
