"""Gera/atualiza a planilha funcoes_cbo.xlsx puxando TODAS as funções do eContador.

Uso:
    python gerar_planilha_funcoes.py [--out funcoes_cbo.xlsx]

Lê ECONTADOR_TOKEN do .env e pagina /funcoes (cap=200/página, ~9k itens).
Gera planilha com colunas: usar, funcao_id, nome_cargo, cbo, codigo, externoid.

Coluna `usar`: marque com "X" os cargos que o escritório usa de verdade.
Quando o pipeline resolver uma função, ele:
  1. Faz match semântico por nome em TODA a planilha
  2. Entre os candidatos similares, prefere os que estão marcados com X

Re-rodar este script PRESERVA as marcas X existentes (indexadas por funcao_id).
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
import time
from pathlib import Path

from dotenv import load_dotenv
import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("gerar-planilha")


BASE_URL = "https://dp.pack.alterdata.com.br/api/v1"
PAGE_LIMIT = 200  # Cap real da API; > 200 retorna HTTP 500


def fetch_todas_funcoes(token: str) -> list[dict]:
    """Pagina todas as funções e retorna lista de attributes + id."""
    funcoes: list[dict] = []
    offset = 0
    with httpx.Client(
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/vnd.api+json",
        },
        timeout=60.0,
    ) as client:
        # Primeira chamada pra pegar total
        r = client.get(
            f"{BASE_URL}/funcoes",
            params={"page[limit]": PAGE_LIMIT, "page[offset]": 0},
        )
        r.raise_for_status()
        body = r.json()
        total = body.get("meta", {}).get("totalResourceCount")
        log.info(f"Total reportado pelo servidor: {total}")

        while True:
            if offset > 0:
                r = client.get(
                    f"{BASE_URL}/funcoes",
                    params={"page[limit]": PAGE_LIMIT, "page[offset]": offset},
                )
                r.raise_for_status()
                body = r.json()

            data = body.get("data", [])
            if not data:
                break

            for it in data:
                a = it.get("attributes") or {}
                funcoes.append({
                    "funcao_id": str(it.get("id", "")),
                    "nome_cargo": a.get("nome", ""),
                    "cbo": a.get("cbo", ""),
                    "codigo": a.get("codigo", ""),
                    "externoid": a.get("externoid", ""),
                })

            log.info(f"  offset={offset} → +{len(data)} (total acumulado: {len(funcoes)})")
            if len(data) < PAGE_LIMIT:
                break
            offset += PAGE_LIMIT
            # Pequena pausa pra não martelar o servidor
            time.sleep(0.1)

    return funcoes


def ler_x_existentes(path: Path) -> set[str]:
    """Lê marcas 'X' (coluna `usar`) de uma planilha existente, indexadas por funcao_id."""
    if not path.exists():
        return set()
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(c or "").strip().lower() for c in next(rows, [])]
        if "usar" not in header or "funcao_id" not in header:
            log.warning(f"Planilha existente em {path} sem colunas usar/funcao_id — não preservando marcas")
            return set()
        i_usar = header.index("usar")
        i_id = header.index("funcao_id")
        marcados: set[str] = set()
        for row in rows:
            usar_val = str(row[i_usar] or "").strip().upper()
            fid = str(row[i_id] or "").strip()
            if usar_val == "X" and fid:
                marcados.add(fid)
        log.info(f"📌 {len(marcados)} marcas X preservadas da planilha anterior")
        return marcados
    except Exception as e:
        log.warning(f"Falha lendo marcas X de {path}: {e}")
        return set()


def gravar_xlsx(funcoes: list[dict], path: Path, x_marcados: set[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "funcoes"

    # Ordem das colunas: 'usar' primeiro pra fácil edição manual
    headers = ["usar", "funcao_id", "nome_cargo", "cbo", "codigo", "externoid"]
    ws.append(headers)

    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    center = Alignment(horizontal="center")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Estilo das células 'usar' (centralizadas + fundo amarelo claro se marcadas)
    usar_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for f in funcoes:
        fid = f.get("funcao_id", "")
        usar = "X" if fid in x_marcados else ""
        ws.append([usar, fid, f.get("nome_cargo", ""), f.get("cbo", ""),
                   f.get("codigo", ""), f.get("externoid", "")])
        # Estiliza a célula 'usar' da linha recém-inserida
        cell = ws.cell(row=ws.max_row, column=1)
        cell.alignment = center
        if usar:
            cell.fill = usar_fill

    # Ajuste de largura das colunas (A=usar, B=id, C=nome, D=cbo, E=codigo, F=externoid)
    larguras = {"A": 6, "B": 12, "C": 60, "D": 10, "E": 12, "F": 38}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w

    # Freeze cabeçalho
    ws.freeze_panes = "A2"

    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Gera funcoes_cbo.xlsx do eContador")
    parser.add_argument(
        "--out",
        default=str(Path(__file__).parent / "funcoes_cbo.xlsx"),
        help="Caminho de saída (default: ./funcoes_cbo.xlsx)",
    )
    args = parser.parse_args()

    token = os.getenv("ECONTADOR_TOKEN")
    if not token:
        log.error("ECONTADOR_TOKEN não encontrado no ambiente (.env)")
        return 1

    out_path = Path(args.out)
    x_marcados = ler_x_existentes(out_path)

    log.info("Puxando todas as funções do eContador...")
    funcoes = fetch_todas_funcoes(token)
    log.info(f"✅ {len(funcoes)} funções coletadas")

    gravar_xlsx(funcoes, out_path, x_marcados)
    log.info(f"📊 Planilha salva em {out_path} ({out_path.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
