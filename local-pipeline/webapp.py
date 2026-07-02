"""webapp.py — interface web do AdmitER (v2.15.0, Flask + HTMX + Jinja).

Roda no MESMO PC do Tkinter (coexistem). Bind em 0.0.0.0:8080 → qualquer
máquina da LAN vê em http://<ip-do-pc>:8080. SEM autenticação — rede
confiável, mesma premissa do plano VNC/noVNC discutido antes.

Polling segue como hoje (Windows Task Scheduler chama `main.py` 2x/dia).
A web SÓ lê o estado + tem botão "Atualizar agora" que dispara uma
passada em thread separada — igual o botão da UI Tkinter.

Reusa:
  - dashboard_data.py — leitura unificada da planilha + payloads/
  - idempotencia, post_admissao — pra "Resolver pendência" e "POSTar"
  - billing — pro card "APIs do mês"

Start (manual):
    python webapp.py
    # → http://localhost:8080 no próprio PC
    # → http://<ip-do-pc-na-lan>:8080 de qualquer máquina da rede

Em produção: rode com `waitress` ou `pythonw` no startup do Windows
(ver README seção "Interface web").
"""
from __future__ import annotations

import json
import logging
import re
import socket
import sys
import threading
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from flask import (
    Flask, abort, jsonify, render_template, request, url_for,
)
from dotenv import load_dotenv

load_dotenv()

import dashboard_data as dd
import idempotencia
from ecotador_client import EContadorAPI
from post_admissao import postar_candidato_registrado


# ── Helpers de Gmail (usados em ações que devem fechar a thread) ──

def _abrir_gmail() -> "object | None":
    """Best-effort: tenta abrir GmailClient pra aplicar labels. Falha
    silenciosa se OAuth não estiver disponível (testes, .env incompleto)."""
    try:
        from gmail_client import GmailClient
        return GmailClient()
    except Exception as e:
        log.warning(f"[webapp] GmailClient indisponível: {type(e).__name__}: {e}")
        return None


def _fechar_thread_manual(gmail, msg_id: str, config) -> None:
    """Aplica label processado e remove pendente — usado quando o operador
    'marca como resolvido manualmente' (sem POST). Sem isso, a thread
    continua sendo reprocessada pelo Task Scheduler e queima tokens Claude
    (ciclo do RAIMUNDO documentado em idempotencia.py)."""
    if not gmail or not msg_id:
        return
    try:
        gmail.aplicar_label(msg_id, config.label_processado)
    except Exception as e:
        log.warning(f"[webapp] falha aplicando processado em {msg_id[:16]}: {e}")
    try:
        gmail.remover_label(msg_id, config.label_pendente)
    except Exception as e:
        log.warning(f"[webapp] falha removendo pendente de {msg_id[:16]}: {e}")

log = logging.getLogger("admissao.webapp")

# ── Paths ────────────────────────────────────────────────────────

ROOT = Path(__file__).parent
TEMPLATES_DIR = ROOT / "web" / "templates"
STATIC_DIR = ROOT / "web" / "static"

# Versão exibida no rodapé (lê de interface.py pra não duplicar)
# v2.15.0: APP_VERSION lida do interface.py via regex (não importa o módulo
# pra não puxar Tkinter). Regex aceita versão completa (suffixes -rc1, dev, etc.).
APP_VERSION = "desconhecida"
try:
    with (ROOT / "interface.py").open("r", encoding="utf-8") as _fh:
        _content = _fh.read(8192)
    _m = re.search(r'APP_VERSION\s*=\s*["\']([^"\']+)["\']', _content)
    if _m:
        APP_VERSION = _m.group(1)
    else:
        log.warning("[webapp] APP_VERSION não encontrada em interface.py")
except (OSError, AttributeError) as e:
    log.warning(f"[webapp] falha lendo APP_VERSION: {e}")


# ── Estado global thread-safe (passada em background) ────────────

class _PassadaState:
    """Status da passada em background ("Atualizar agora"). Thread-safe
    via lock simples — não tem leitura/escrita concorrente intensa."""
    def __init__(self):
        self._lock = threading.Lock()
        self.rodando: bool = False
        self.iniciada_em: str = ""
        self.terminada_em: str = ""
        self.erro: str | None = None
        # Resumo da última passada (n emails, n posts, custo USD)
        self.ultimo_resumo: dict = {}

    def snapshot(self) -> dict:
        with self._lock:
            return {
                "rodando": self.rodando,
                "iniciada_em": self.iniciada_em,
                "terminada_em": self.terminada_em,
                "erro": self.erro,
                "ultimo_resumo": dict(self.ultimo_resumo),
            }

    # Cooldown mínimo entre passadas (segundos) — anti-spam/anti-DoS local
    COOLDOWN_S = 30

    def marcar_iniciada(self) -> tuple[bool, str]:
        """Retorna (sucesso, motivo). Sucesso=False quando:
          - já tem uma passada rodando
          - última terminou há menos de COOLDOWN_S segundos
        """
        with self._lock:
            if self.rodando:
                return False, "Já tem uma passada rodando."
            if self.terminada_em:
                try:
                    ult = datetime.fromisoformat(self.terminada_em)
                    sobra = self.COOLDOWN_S - int((datetime.now() - ult).total_seconds())
                    if sobra > 0:
                        return False, f"Aguarde {sobra}s antes de rodar de novo."
                except ValueError:
                    pass
            self.rodando = True
            self.iniciada_em = datetime.now().isoformat(timespec="seconds")
            self.terminada_em = ""
            self.erro = None
            return True, "Passada iniciada — pode levar 1-2 min."

    def marcar_terminada(self, erro: str | None = None, resumo: dict | None = None) -> None:
        with self._lock:
            self.rodando = False
            self.terminada_em = datetime.now().isoformat(timespec="seconds")
            self.erro = erro
            if resumo:
                self.ultimo_resumo = dict(resumo)


PASSADA = _PassadaState()
IMPORTAR = _PassadaState()  # estado da importação manual (mesmo padrão)


# ── Captura de logs em memória pro "terminal de atividade" da web ─

class _LogBuffer(logging.Handler):
    """Buffer circular thread-safe que guarda as últimas N mensagens de log.
    Anexado aos loggers admissao.* na inicialização. Consumido pelo endpoint
    /htmx/atividade que renderiza no estilo terminal."""
    def __init__(self, capacidade: int = 500):
        super().__init__()
        from collections import deque
        self._lock = threading.Lock()
        self._buf: "deque[dict]" = deque(maxlen=capacidade)
        # Formato semelhante ao log do .bat (pra ficar familiar)
        self.setFormatter(logging.Formatter(
            "%(asctime)s [%(levelname)s] %(name)s: %(message)s",
            datefmt="%H:%M:%S",
        ))

    def emit(self, record: logging.LogRecord) -> None:
        try:
            msg = self.format(record)
            with self._lock:
                self._buf.append({
                    "ts": datetime.fromtimestamp(record.created).isoformat(timespec="seconds"),
                    "nivel": record.levelname,
                    "logger": record.name,
                    "msg": msg,
                })
        except Exception:
            pass  # log handler NUNCA pode levantar

    def snapshot(self, ultimas: int | None = None) -> list[dict]:
        with self._lock:
            itens = list(self._buf)
        if ultimas is not None and len(itens) > ultimas:
            return itens[-ultimas:]
        return itens

    def limpar(self) -> None:
        with self._lock:
            self._buf.clear()


LOG_BUFFER = _LogBuffer(capacidade=500)
LOG_BUFFER.setLevel(logging.INFO)
# Anexa SÓ ao root (pra evitar duplicação por propagação). Garante que o
# root aceita INFO — sem isso, INFOs do pipeline ficariam filtradas quando
# o webapp é importado via test_client (que não chama _setup_logging).
_root = logging.getLogger()
if _root.level == logging.NOTSET or _root.level > logging.INFO:
    _root.setLevel(logging.INFO)
_root.addHandler(LOG_BUFFER)


class _PollingLoopState:
    """Loop infinito de polling — mesmo que `python main.py --loop`. Mas
    rodando como thread dentro do Flask. Permite Iniciar/Parar pela web.

    Quando o polling está ativo, dispara `rodar_uma_passada` a cada
    `intervalo` segundos (lido do config.json). NÃO interfere com o
    Task Scheduler (eles são independentes; idempotência protege).
    """
    def __init__(self):
        self._lock = threading.Lock()
        self._stop_evt = threading.Event()
        self._thread: threading.Thread | None = None
        self.iniciado_em: str = ""
        self.parado_em: str = ""
        self.ultima_passada_em: str = ""
        self.n_passadas: int = 0
        self.erro: str | None = None

    def snapshot(self) -> dict:
        with self._lock:
            return {
                "rodando": self._thread is not None and self._thread.is_alive(),
                "iniciado_em": self.iniciado_em,
                "parado_em": self.parado_em,
                "ultima_passada_em": self.ultima_passada_em,
                "n_passadas": self.n_passadas,
                "erro": self.erro,
            }

    def iniciar(self) -> tuple[bool, str]:
        with self._lock:
            if self._thread is not None and self._thread.is_alive():
                return False, "Polling já está rodando."
            self._stop_evt.clear()
            self.iniciado_em = datetime.now().isoformat(timespec="seconds")
            self.parado_em = ""
            self.erro = None
            self.n_passadas = 0
            self._thread = threading.Thread(target=self._loop, daemon=True)
            self._thread.start()
            return True, "Polling iniciado."

    def parar(self) -> tuple[bool, str]:
        with self._lock:
            if self._thread is None or not self._thread.is_alive():
                return False, "Polling já estava parado."
            self._stop_evt.set()
            self.parado_em = datetime.now().isoformat(timespec="seconds")
            return True, "Pedido de parada enviado (termina após a passada atual)."

    def _loop(self):
        """Worker thread: roda passadas até receber stop_evt."""
        try:
            import os as _os
            sys.stdin = open(_os.devnull, "r")
        except Exception:
            pass
        while not self._stop_evt.is_set():
            try:
                log.info("[polling] Iniciando passada via loop interno…")
                from main import (
                    carregar_config, bootstrap_arquivos_locais, carregar_planilha,
                    ClaudeClient, rodar_uma_passada, PLANILHA_CBO,
                )
                config = carregar_config()
                bootstrap_arquivos_locais()
                planilha = carregar_planilha(PLANILHA_CBO)
                try:
                    config.confirmar_replies = False
                except Exception:
                    pass
                claude = ClaudeClient(
                    model=config.claude_model,
                    max_tokens=config.claude_max_tokens,
                    chamadas_verificacao=config.claude_chamadas_verificacao,
                )
                rodar_uma_passada(config, claude, planilha)
                with self._lock:
                    self.ultima_passada_em = datetime.now().isoformat(timespec="seconds")
                    self.n_passadas += 1
                intervalo = max(60, int(getattr(config, "intervalo", 300) or 300))
            except Exception as e:
                log.exception("[polling] Erro na passada")
                with self._lock:
                    self.erro = f"{type(e).__name__}: {e}"
                intervalo = 60  # backoff em caso de erro
            # Aguarda intervalo ou stop_evt
            self._stop_evt.wait(intervalo)
        log.info("[polling] Loop encerrado.")


POLLING = _PollingLoopState()


def _executar_passada_em_thread():
    """Roda uma passada do orquestrador. Importa main lazy pra acelerar
    o boot da web (main.py puxa Anthropic SDK etc)."""
    iniciada, _motivo = PASSADA.marcar_iniciada()
    if not iniciada:
        log.info(f"[webapp] Pedido de Atualizar agora ignorado: {_motivo}")
        return
    # Evita travar a thread se algum código chamar input() (defesa em profundidade)
    try:
        import os
        sys.stdin = open(os.devnull, "r")
    except Exception:
        pass
    try:
        log.info("[webapp] Iniciando passada via web…")
        from main import (
            carregar_config, bootstrap_arquivos_locais, carregar_planilha,
            ClaudeClient, rodar_uma_passada, PLANILHA_CBO,
        )
        config = carregar_config()
        bootstrap_arquivos_locais()
        planilha = carregar_planilha(PLANILHA_CBO)
        # Desliga confirmação interativa (sem TTY no Flask)
        try:
            config.confirmar_replies = False
        except Exception:
            pass
        claude = ClaudeClient(
            model=config.claude_model,
            max_tokens=config.claude_max_tokens,
            chamadas_verificacao=config.claude_chamadas_verificacao,
        )
        usage_antes = dict(claude.usage_total)
        rodar_uma_passada(config, claude, planilha)
        # Resumo simples (delta de tokens + custo)
        u = claude.usage_total
        resumo = {
            "n_calls": u["n_calls"] - usage_antes.get("n_calls", 0),
            "input_tokens": u["input_tokens"] - usage_antes.get("input_tokens", 0),
            "output_tokens": u["output_tokens"] - usage_antes.get("output_tokens", 0),
        }
        PASSADA.marcar_terminada(resumo=resumo)
        log.info(f"[webapp] Passada concluída: {resumo}")
    except Exception as e:
        log.exception("[webapp] Falha na passada")
        PASSADA.marcar_terminada(erro=f"{type(e).__name__}: {e}")


# ── Flask app ─────────────────────────────────────────────────────

app = Flask(
    __name__,
    template_folder=str(TEMPLATES_DIR),
    static_folder=str(STATIC_DIR),
    static_url_path="/static",
)


# v2.16.10: carrega cache de empresas em background no boot do webapp.
# Sem isso, dashboard_data.listar_entidades não consegue enriquecer
# `empresa="?"` via lookup, e pendências históricas continuam mostrando "?".
# Custo: 1 chamada paginada ao /empresas (~5-30s) em thread daemon — não
# bloqueia o boot do Flask. Tkinter já fazia isso em interface.py:1645.
def _carregar_cache_empresas_em_background():
    try:
        from main import recarregar_empresas_cache, carregar_config
        log.info("[webapp] Carregando cache de empresas em background...")
        config = carregar_config()
        api = EContadorAPI(config.base_url, config.token)
        try:
            cache = recarregar_empresas_cache(api)
            log.info(
                f"[webapp] ✓ Cache pronto: {len(cache)} CNPJs no whitelist"
            )
        finally:
            api.close()
    except Exception as e:
        log.warning(
            f"[webapp] ⚠ Falha carregando cache de empresas: "
            f"{type(e).__name__}: {e}. Pendências mostrarão '?' até "
            f"clicar em 'Atualizar cache' manualmente."
        )


threading.Thread(
    target=_carregar_cache_empresas_em_background,
    daemon=True, name="empresas-cache-boot",
).start()


@app.template_filter("cnpj_url")
def _filtro_cnpj_url(c: str) -> str:
    """Usado nos templates: {{ e.cnpj|cnpj_url }} → 'sem-cnpj' se vazio,
    senão dígitos. Evita gerar URLs com segmento vazio."""
    return c if (c and c.strip()) else _CNPJ_SENTINEL


@app.template_filter("data_iso")
def _filtro_data_iso(v) -> str:
    """v2.16.12: converte data BR (DD/MM/YYYY) pra ISO (YYYY-MM-DD) pro
    <input type="date"> renderizar valores extraídos pelo Claude. Aceita
    já-ISO, BR, ou retorna vazio. Não quebra com None/datetime."""
    if not v:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    # já ISO?
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    # BR DD/MM/YYYY
    if len(s) == 10 and s[2] == "/" and s[5] == "/":
        return f"{s[6:10]}-{s[3:5]}-{s[0:2]}"
    return ""


@app.template_filter("salario_num")
def _filtro_salario_num(v) -> str:
    """v2.16.12: salário pode chegar como string sentinela
    ('NÃO INFORMADO — pendente') no payload parcial. Pro <input>
    a gente só mostra se for número válido."""
    if v in (None, "", 0, "0"):
        return ""
    try:
        f = float(str(v).replace(",", "."))
        return f"{f:.2f}" if f > 0 else ""
    except (ValueError, TypeError):
        return ""


# Cache dos IPs (descobrir todos pode levar ~50-500ms; não muda durante uso)
_IPS_CACHE: list[str] = []


def _ips_acesso() -> list[str]:
    global _IPS_CACHE
    if not _IPS_CACHE:
        _IPS_CACHE = _ips_lan_todos()
    return _IPS_CACHE


@app.context_processor
def _inject_globals():
    """Variáveis disponíveis em todos os templates."""
    return {
        "APP_VERSION": APP_VERSION,
        "APP_NAME": "AdmitER",
        "now_iso": datetime.now().isoformat(timespec="seconds"),
        "contadores_sidebar": dd.resumo_contadores(),  # badge sidebar
        "polling": POLLING.snapshot(),  # estado do loop interno
        "config_atual": _ler_config_json(),  # pra preencher form de Configurações
        "ips_acesso": _ips_acesso(),  # lista de IPs pra outras máquinas
        "porta_acesso": 8080,  # porta do bind (TODO: ler de argparse)
    }


# ── Segurança: Origin check (CSRF-lite) + headers padrão ─────────

@app.before_request
def _csrf_origin_check():
    """v2.15.0 fix (review HIGH): rejeita POSTs cuja Origin/Referer não
    bate com o host atual. Premissa: rede LAN confiável, sem auth — mas
    impede que site malicioso visitado pelo operador force POSTs via
    <form auto-submit> ou <img src=POST_url>.

    Atualizar-agora, marcar-resolvido e postar são protegidos.
    GETs e endpoints HTMX de leitura passam livre.
    """
    if request.method != "POST":
        return None
    origem = request.headers.get("Origin") or request.headers.get("Referer", "")
    if not origem:
        # Sem Origin/Referer = req crafted (browser sempre manda em form POST)
        abort(403, description="CSRF: requisição sem Origin/Referer")
    parsed = urlparse(origem)
    # Bate exatamente com nosso host. request.host inclui porta.
    host_esperado = request.host
    if parsed.netloc != host_esperado:
        abort(403,
              description=f"CSRF: origem '{parsed.netloc}' não bate com '{host_esperado}'")
    return None


@app.after_request
def _security_headers(resp):
    """Headers básicos pra evitar clickjacking + sniffing.
    Sem CSP estrita ainda — HTMX inline trigger event handler precisa de
    unsafe-inline; refinar quando self-hostar htmx.js (TODO próximo)."""
    resp.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("Referrer-Policy", "same-origin")
    return resp


# ── Helpers de validação ─────────────────────────────────────────

_RE_NOME_VALIDO = re.compile(r"^[A-ZÀ-ÿa-zà-ÿ0-9 .'\-]{1,120}$")

# v2.15.9: sentinel pra CNPJ vazio na URL. Quando o Claude não conseguiu
# extrair CNPJ, a entidade fica com cnpj="" e a URL /pendencia/x/nome//
# dá 404 no Flask (rota não casa com segmento vazio). Solução: usa
# "sem-cnpj" como placeholder na URL e converte de volta no handler.
_CNPJ_SENTINEL = "sem-cnpj"


def _cnpj_da_url(cnpj: str) -> str:
    """Inverso do filtro `cnpj_url`: retorna "" quando recebe o sentinel."""
    return "" if cnpj == _CNPJ_SENTINEL else cnpj


def _validar_nome_path(nome: str) -> str:
    """Aborta 400 se o nome no path tem chars suspeitos. Evita injeção de
    quebras de linha (NDJSON corrupt) ou strings gigantes."""
    if not nome or not _RE_NOME_VALIDO.match(nome.strip()):
        abort(400, description=f"Nome inválido no path: {nome!r}")
    return nome.strip()


def _safe_redirect(url: str | None, fallback: str = "/") -> str:
    """Só permite URL relativa do próprio host. Evita open redirect e XSS
    via {{ redirect }} em template (review MEDIUM)."""
    if not url:
        return fallback
    p = urlparse(url)
    # URL precisa ser relativa (sem scheme/netloc) e começar com /
    if not p.netloc and not p.scheme and url.startswith("/"):
        return url
    return fallback


# ── Helpers internos ─────────────────────────────────────────────

def _eh_htmx() -> bool:
    """True quando o request veio do HTMX (header HX-Request)."""
    return request.headers.get("HX-Request") == "true"


def _entidade_por_chave(msg_id: str, nome: str, cnpj: str) -> dict | None:
    """Encontra a entidade exata na lista dedupada."""
    nome_u = (nome or "").upper().strip()
    for e in dd.listar_entidades(incluir_cargo_ia=True):
        if e["msg_id"] == msg_id and e["nome"].upper() == nome_u and e["cnpj"] == cnpj:
            return e
    return None


# ── Páginas ──────────────────────────────────────────────────────

@app.route("/")
def index():
    """Dashboard: cards + lista de pendentes recente."""
    contadores = dd.resumo_contadores()
    pendentes = [
        e for e in dd.listar_entidades(incluir_cargo_ia=True)
        if e["categoria"].startswith("pendente") or e["categoria"] == "falha_tecnica"
    ][:10]  # primeiras 10 mais recentes
    return render_template(
        "dashboard.html",
        contadores=contadores,
        pendentes=pendentes,
        passada=PASSADA.snapshot(),
        billing=_billing_resumido(),
    )


@app.route("/pendentes")
def pendentes():
    """Lista completa de entidades pendentes (cliente + interna + falha)."""
    todas = dd.listar_entidades(incluir_cargo_ia=True)
    pend = [
        e for e in todas
        if e["categoria"].startswith("pendente") or e["categoria"] == "falha_tecnica"
    ]
    por_dia = dd.agrupar_por_dia(pend)
    return render_template(
        "pendentes.html",
        por_dia=por_dia,
        total=len(pend),
        contadores=dd.resumo_contadores(),
    )


@app.route("/processadas")
def processadas():
    """Lista completa de admissões cadastradas."""
    todas = dd.listar_entidades()
    proc = [e for e in todas if e["categoria"] == "processada"]
    por_dia = dd.agrupar_por_dia(proc)
    return render_template(
        "processadas.html",
        por_dia=por_dia,
        total=len(proc),
    )


@app.route("/atividade")
def atividade():
    """Terminal de logs ao vivo — igual à 'Atividade recente' do Tkinter.
    Mostra os últimos eventos capturados pelo LOG_BUFFER, com auto-refresh."""
    return render_template("atividade.html", eventos=LOG_BUFFER.snapshot(200))


@app.route("/htmx/atividade")
def htmx_atividade():
    """Fragment com a lista atualizada (consumido por hx-trigger every 3s)."""
    return render_template(
        "_partials/atividade_log.html",
        eventos=LOG_BUFFER.snapshot(200),
    )


@app.route("/atividade/limpar", methods=["POST"])
def atividade_limpar():
    """Esvazia o buffer (não apaga o arquivo de log, só o que está em memória)."""
    LOG_BUFFER.limpar()
    return _resposta_ok_ou_htmx(
        "Histórico do terminal limpo.",
        redirect_url=url_for("atividade"),
    )


@app.route("/auditoria")
def auditoria():
    """Últimos 200 eventos do admissao_log.ndjson."""
    eventos = dd.auditoria_recente(n=200)
    return render_template("auditoria.html", eventos=eventos)


@app.route("/estatisticas")
def estatisticas():
    """Billing do mês + DirectData stats."""
    return render_template(
        "estatisticas.html",
        billing=_billing_resumido(),
        passada=PASSADA.snapshot(),
        contadores=dd.resumo_contadores(),
    )


@app.route("/pendencia/<msg_id>/<path:nome>/<cnpj>")
def pendencia_detalhe(msg_id: str, nome: str, cnpj: str):
    """Detalhe de uma pendência específica + payload salvo + duplicatas conhecidas."""
    nome = _validar_nome_path(nome)
    cnpj = _cnpj_da_url(cnpj)
    entidade = _entidade_por_chave(msg_id, nome, cnpj)
    if not entidade:
        abort(404, description="Entidade não encontrada")

    payload_path = dd.achar_payload_por_msg_e_nome(msg_id, nome)
    payload_doc = dd.carregar_payload_completo(payload_path) if payload_path else {}

    # Dados extraídos pelo Claude (pra mostrar resumo amigável)
    attrs = ((payload_doc.get("payload") or {}).get("data") or {}).get("attributes") or {}
    resolucao = payload_doc.get("resolucao") or {}
    cpf = attrs.get("cpf") or (resolucao.get("cpf") or "")
    duplicatas = []
    if cpf:
        try:
            duplicatas = idempotencia.consultar_duplicata(cpf, cnpj)
        except Exception:
            pass

    return render_template(
        "pendencia_detalhe.html",
        entidade=entidade,
        payload_path=payload_path.name if payload_path else None,
        payload_doc=payload_doc,
        attrs=attrs,
        resolucao=resolucao,
        duplicatas=duplicatas,
    )


# ── Ações (POST) ─────────────────────────────────────────────────

@app.route("/atualizar-agora", methods=["POST"])
def atualizar_agora():
    """Dispara uma passada em background. Não bloqueia o request.
    Cooldown de 30s entre passadas (anti-spam — review MEDIUM)."""
    snap = PASSADA.snapshot()
    if snap["rodando"]:
        return _renderizar_status_passada(mensagem="Já tem uma passada rodando.")
    if snap["terminada_em"]:
        try:
            ult = datetime.fromisoformat(snap["terminada_em"])
            sobra = _PassadaState.COOLDOWN_S - int((datetime.now() - ult).total_seconds())
            if sobra > 0:
                return _renderizar_status_passada(
                    mensagem=f"Aguarde {sobra}s antes de rodar de novo (anti-spam)."
                )
        except ValueError:
            pass
    t = threading.Thread(target=_executar_passada_em_thread, daemon=True)
    t.start()
    return _renderizar_status_passada(mensagem="Passada iniciada — pode levar 1-2 min.")


@app.route("/pendencia/<msg_id>/<path:nome>/<cnpj>/marcar-resolvido", methods=["POST"])
def marcar_resolvido_manual(msg_id: str, nome: str, cnpj: str):
    """Marca como resolvido manualmente (operador já cadastrou direto no
    eContador). Adiciona linha 'Cadastrado manualmente' na planilha + fecha
    a thread no Gmail (label processado + remove pendente).

    v2.15.0 fix (BLOCKER review): SEM aplicar o label, a próxima passada do
    Task Scheduler reprocessa o email do zero — gasta US$0.40 em Claude por
    cada pendência "resolvida manualmente". Em volume = sangramento real.
    """
    nome = _validar_nome_path(nome)
    cnpj = _cnpj_da_url(cnpj)
    from main import registrar_admissao_planilha, carregar_config

    # Recupera empresa real da entidade (review: planilha estava ficando com "?")
    entidade = _entidade_por_chave(msg_id, nome, cnpj)
    empresa = entidade["empresa"] if entidade else ""

    try:
        registrar_admissao_planilha(
            nome=nome, empresa=empresa, cnpj=cnpj,
            procedencia=(
                f"Cadastrado manualmente — via web "
                f"({datetime.now().strftime('%d/%m/%Y %H:%M')})"
            ),
            msg_id=msg_id,
        )
    except Exception as e:
        log.exception("Falha marcando resolvido")
        return _resposta_erro(f"{type(e).__name__}: {e}")

    # Fecha a thread no Gmail — best-effort, não falha o request se Gmail offline
    gmail = _abrir_gmail()
    if gmail is not None:
        try:
            config = carregar_config()
            _fechar_thread_manual(gmail, msg_id, config)
        except Exception as e:
            log.warning(f"[webapp] falha fechando thread {msg_id[:16]}: {e}")
        finally:
            try:
                gmail.close()
            except Exception:
                pass

    return _resposta_ok_ou_htmx(
        f"Marcado como resolvido manualmente: {nome}",
        redirect_url=url_for("pendentes"),
    )


# ── Reprocessar email específico (pendência sem payload) ─────────

def _reprocessar_msg_em_thread(msg_id: str):
    """Busca a mensagem original no Gmail por msg_id e roda processar_email
    do zero. Útil pra pendências antigas (pré-v2.15.3) que ficaram sem
    payload no disco — o reprocessamento agora SEMPRE salva payload parcial.
    Custo: ~US$0.15-0.40 em Claude (1 email, 1-2 chamadas)."""
    iniciada, motivo = IMPORTAR.marcar_iniciada()  # reusa IMPORTAR state pra UI
    if not iniciada:
        log.info(f"[webapp] Reprocessamento ignorado: {motivo}")
        return
    try:
        import os as _os
        sys.stdin = open(_os.devnull, "r")
    except Exception:
        pass
    try:
        log.info(f"[webapp] Reprocessando msg_id={msg_id[:16]}…")
        from main import (
            carregar_config, bootstrap_arquivos_locais, carregar_planilha,
            ClaudeClient, processar_email, PLANILHA_CBO,
        )
        from gmail_client import GmailClient
        config = carregar_config()
        bootstrap_arquivos_locais()
        planilha = carregar_planilha(PLANILHA_CBO)
        try:
            config.confirmar_replies = False
        except Exception:
            pass
        gmail = GmailClient()
        # Remove label pendente ANTES (senão reprocessar não muda nada)
        try:
            gmail.remover_label(msg_id, config.label_pendente)
        except Exception as e:
            log.debug(f"   nada pra remover: {e}")
        # Busca a mensagem
        msg = gmail.servico.users().messages().get(
            userId="me", id=msg_id, format="full"
        ).execute()
        claude = ClaudeClient(
            model=config.claude_model,
            max_tokens=config.claude_max_tokens,
            chamadas_verificacao=config.claude_chamadas_verificacao,
        )
        from ecotador_client import EContadorAPI as _API
        api = _API(config.base_url, config.token)
        try:
            processar_email(msg, gmail, claude, api, planilha, config)
        finally:
            api.close()
            gmail.close()
        IMPORTAR.marcar_terminada(resumo={"reprocessado": msg_id[:16]})
        log.info(f"[webapp] Reprocessamento concluído: {msg_id[:16]}")
    except Exception as e:
        log.exception("[webapp] Falha no reprocessamento")
        IMPORTAR.marcar_terminada(erro=f"{type(e).__name__}: {e}")


@app.route("/pendencia/<msg_id>/<path:nome>/<cnpj>/reprocessar", methods=["POST"])
def reprocessar_pendencia(msg_id: str, nome: str, cnpj: str):
    """Reprocessa o email do zero — gasta ~US$0.15 em Claude mas resolve
    pendências antigas sem payload no disco. Roda em thread bg."""
    nome = _validar_nome_path(nome)
    cnpj = _cnpj_da_url(cnpj)
    if IMPORTAR.snapshot()["rodando"]:
        return _resposta_erro("Já tem uma importação/reprocesso rodando — aguarde.")
    t = threading.Thread(
        target=_reprocessar_msg_em_thread, args=(msg_id,), daemon=True,
    )
    t.start()
    return _resposta_ok_ou_htmx(
        f"Reprocessamento de '{nome}' iniciado — pode levar 1-2 min. "
        f"Volte em alguns instantes pra ver os campos extraídos.",
        redirect_url=url_for("pendentes"),
    )


# ── Importação manual de arquivos ────────────────────────────────

# Limites pra upload (defesa contra DoS local)
_IMPORTAR_MAX_BYTES = 100 * 1024 * 1024  # 100 MB total
_IMPORTAR_EXTENSOES_OK = {".pdf", ".png", ".jpg", ".jpeg", ".gif", ".webp"}


def _executar_importacao_em_thread(arquivos: list[dict], corpo_texto: str):
    """Roda processar_arquivos_avulsos numa thread (libera o request).
    arquivos = [{filename, mime, data: bytes}, ...]"""
    iniciada, motivo = IMPORTAR.marcar_iniciada()
    if not iniciada:
        log.info(f"[webapp] Importação ignorada: {motivo}")
        return
    try:
        import os as _os
        sys.stdin = open(_os.devnull, "r")
    except Exception:
        pass
    try:
        log.info(f"[webapp] Iniciando importação de {len(arquivos)} arquivo(s)…")
        from main import (
            carregar_config, bootstrap_arquivos_locais, carregar_planilha,
            ClaudeClient, processar_arquivos_avulsos, PLANILHA_CBO,
        )
        from ecotador_client import EContadorAPI as _API
        config = carregar_config()
        bootstrap_arquivos_locais()
        planilha = carregar_planilha(PLANILHA_CBO)
        try:
            config.confirmar_replies = False
        except Exception:
            pass
        claude = ClaudeClient(
            model=config.claude_model,
            max_tokens=config.claude_max_tokens,
            chamadas_verificacao=config.claude_chamadas_verificacao,
        )
        api = _API(config.base_url, config.token)
        try:
            processar_arquivos_avulsos(
                arquivos, corpo_texto, claude, api, planilha, config,
            )
        finally:
            api.close()
        IMPORTAR.marcar_terminada(resumo={"n_arquivos": len(arquivos)})
        log.info(f"[webapp] Importação concluída: {len(arquivos)} arquivo(s)")
    except Exception as e:
        log.exception("[webapp] Falha na importação")
        IMPORTAR.marcar_terminada(erro=f"{type(e).__name__}: {e}")


@app.route("/importar", methods=["GET"])
def importar_form():
    """Página com form de upload pra processar arquivos sem precisar de email."""
    return render_template("importar.html", importar=IMPORTAR.snapshot())


@app.route("/importar", methods=["POST"])
def importar_post():
    """Recebe arquivos multipart + texto opcional. Dispara processamento em
    thread (igual /atualizar-agora) — não bloqueia o request."""
    if IMPORTAR.snapshot()["rodando"]:
        return _resposta_erro("Já tem uma importação rodando — aguarde terminar.")

    files = request.files.getlist("arquivos")
    if not files:
        return _resposta_erro("Nenhum arquivo enviado.")

    arquivos: list[dict] = []
    total_bytes = 0
    for f in files:
        if not f or not f.filename:
            continue
        ext = Path(f.filename).suffix.lower()
        if ext not in _IMPORTAR_EXTENSOES_OK:
            return _resposta_erro(
                f"Extensão '{ext}' não suportada. "
                f"Aceitos: {', '.join(sorted(_IMPORTAR_EXTENSOES_OK))}"
            )
        data = f.read()
        total_bytes += len(data)
        if total_bytes > _IMPORTAR_MAX_BYTES:
            return _resposta_erro(
                f"Tamanho total acima de {_IMPORTAR_MAX_BYTES // (1024*1024)}MB."
            )
        arquivos.append({
            "filename": f.filename,
            "mime": f.mimetype or "application/octet-stream",
            "data": data,
        })

    if not arquivos:
        return _resposta_erro("Nenhum arquivo válido pra processar.")

    corpo_texto = (request.form.get("corpo_texto") or "").strip()
    t = threading.Thread(
        target=_executar_importacao_em_thread,
        args=(arquivos, corpo_texto),
        daemon=True,
    )
    t.start()
    return _resposta_ok_ou_htmx(
        f"Importação iniciada com {len(arquivos)} arquivo(s) — "
        f"pode levar 1-2 min. Acompanhe em Pendentes/Processadas.",
        redirect_url=url_for("importar_form"),
    )


# ── Corrigir CNPJ do email (override) ────────────────────────────

# v2.16.3: API eContador aceita CPF (11 dígitos) ou CNPJ (14 dígitos) no
# campo cpfcnpj. Empresas pessoa-física são cadastradas com CPF do dono.
_RE_CPFCNPJ_DIGITS = re.compile(r"^\d{11}$|^\d{14}$")


@app.route("/pendencia/<msg_id>/<path:nome>/<cnpj>/corrigir-cnpj", methods=["POST"])
def corrigir_cnpj_email(msg_id: str, nome: str, cnpj: str):
    """Grava override de CNPJ (ou CPF) pro próximo reprocesso do email.
    O Claude vai receber o documento correto no prompt e não vai tentar
    adivinhar. v2.16.3: aceita CPF (11 dígitos) — empresas pessoa-física
    no eContador são cadastradas com o CPF do dono no mesmo campo cpfcnpj.

    Form field obrigatório: `cnpj_novo` (com ou sem pontuação)
    """
    nome = _validar_nome_path(nome)
    cnpj = _cnpj_da_url(cnpj)
    cnpj_novo = re.sub(r"\D", "", request.form.get("cnpj_novo", "") or "")
    if not _RE_CPFCNPJ_DIGITS.match(cnpj_novo):
        return _resposta_erro(
            "Documento inválido. Informe 11 dígitos (CPF) ou 14 dígitos "
            "(CNPJ), com ou sem pontuação."
        )

    try:
        from main import salvar_cnpj_override
        salvar_cnpj_override(msg_id, cnpj_novo)
    except Exception as e:
        log.exception("Falha gravando cnpj_override")
        return _resposta_erro(f"{type(e).__name__}: {e}")

    tipo = "CPF" if len(cnpj_novo) == 11 else "CNPJ"
    return _resposta_ok_ou_htmx(
        f"{tipo} {cnpj_novo} salvo como override pra '{nome}'. "
        f"No próximo reprocesso (botão Atualizar agora ou Task Scheduler) "
        f"o Claude vai usar esse documento direto.",
        redirect_url=url_for("pendentes"),
    )


# ── Ações de manutenção (Dashboard) ──────────────────────────────

@app.route("/maintenance/atualizar-cache-empresas", methods=["POST"])
def maintenance_atualizar_cache_empresas():
    """Recarrega o cache local de CNPJs do eContador (GET /empresas inteiro).
    Demora 5-10s — roda em thread pra não bloquear request."""
    def _worker():
        try:
            from main import recarregar_empresas_cache, carregar_config
            config = carregar_config()
            api = EContadorAPI(config.base_url, config.token)
            try:
                cache = recarregar_empresas_cache(api)
                log.info(f"[webapp] Cache empresas recarregado: {len(cache)} CNPJs")
            finally:
                api.close()
        except Exception as e:
            log.exception(f"[webapp] Falha recarregando cache: {e}")
    threading.Thread(target=_worker, daemon=True).start()
    return _resposta_ok_ou_htmx(
        "Recarregando cache de empresas em segundo plano (5-10s)…",
        redirect_url=url_for("index"),
    )


@app.route("/maintenance/recarregar-cbo", methods=["POST"])
def maintenance_recarregar_cbo():
    """Recarrega funcoes_cbo.xlsx. Útil quando você adicionou um cargo novo.
    A próxima passada vai usar a planilha atualizada."""
    try:
        from main import carregar_planilha, PLANILHA_CBO
        planilha = carregar_planilha(PLANILHA_CBO)
        n = len(planilha)
    except Exception as e:
        log.exception("Falha recarregando planilha CBO")
        return _resposta_erro(f"{type(e).__name__}: {e}")
    return _resposta_ok_ou_htmx(
        f"Planilha CBO recarregada: {n} cargos.",
        redirect_url=url_for("index"),
    )


@app.route("/controle/polling/iniciar", methods=["POST"])
def controle_polling_iniciar():
    ok, msg = POLLING.iniciar()
    if not ok:
        return _resposta_erro(msg)
    return _resposta_ok_ou_htmx(msg, redirect_url=url_for("index"))


@app.route("/controle/polling/parar", methods=["POST"])
def controle_polling_parar():
    ok, msg = POLLING.parar()
    if not ok:
        return _resposta_erro(msg)
    return _resposta_ok_ou_htmx(msg, redirect_url=url_for("index"))


@app.route("/maintenance/backup", methods=["POST"])
def maintenance_backup():
    """Copia admissoes.xlsx + payloads/ + billing.ndjson pra backups/<ts>/."""
    try:
        from main import fazer_backup_planilha_e_payloads
        dest = fazer_backup_planilha_e_payloads()
    except Exception as e:
        log.exception("Falha no backup")
        return _resposta_erro(f"{type(e).__name__}: {e}")
    if not dest:
        return _resposta_erro("Backup retornou vazio — veja logs.")
    return _resposta_ok_ou_htmx(
        f"Backup criado em {dest.name}.",
        redirect_url=url_for("index"),
    )


# ── Configurações (persistência em config.json) ──────────────────

# Allowlist do que pode ser editado via web (resto fica intocado no JSON)
_CONFIG_BOOLEANOS_PERMITIDOS = {
    "auto_email_pendencias",
    "sempre_mandar_sem_data_admissao",
    "sempre_mandar_sem_funcao",
    "reprocessar_pendentes_no_polling",
    "usar_data_atual_se_invalida",
    "dry_run",
    "postar_sem_departamento_quando_vazio",
    "directdata_pis_habilitado",
    "directdata_titulo_habilitado",
}
_CONFIG_INTEIROS_PERMITIDOS = {
    "polling_intervalo_segundos",
    "pausa_entre_emails_segundos",
}
# v2.16.19: strings enumeradas (radio buttons na UI)
_CONFIG_ENUMS_PERMITIDOS = {
    "auto_email_pendencias_modo": {"desligado", "rascunho", "direto"},
}


def _config_path() -> Path:
    """Resolve o path do config.json de forma robusta a mocks de testes:
    1) tenta `main.CONFIG_FILE` (permite tests patcharem)
    2) fallback pro ROOT/config.json local (path do próprio módulo)
    """
    try:
        import main as _m
        if hasattr(_m, "CONFIG_FILE") and _m.CONFIG_FILE:
            return Path(_m.CONFIG_FILE)
    except Exception:
        pass
    return ROOT / "config.json"


def _ler_config_json() -> dict:
    try:
        return json.loads(_config_path().read_text(encoding="utf-8"))
    except Exception as e:
        log.debug(f"[webapp] config.json: {type(e).__name__}: {e}")
        return {}


def _atualizar_config_json(updates: dict) -> None:
    """Escrita atômica: lê o JSON inteiro, atualiza só as chaves permitidas,
    regrava via tmp + os.replace. Resto fica intocado."""
    import os as _os
    import tempfile as _tempfile
    atual = _ler_config_json()
    for k, v in updates.items():
        if k in _CONFIG_BOOLEANOS_PERMITIDOS:
            atual[k] = bool(v)
        elif k in _CONFIG_INTEIROS_PERMITIDOS:
            try:
                atual[k] = int(v)
            except (ValueError, TypeError):
                raise ValueError(f"Inteiro inválido em '{k}': {v!r}")
        elif k in _CONFIG_ENUMS_PERMITIDOS:
            permitido = _CONFIG_ENUMS_PERMITIDOS[k]
            if str(v) not in permitido:
                raise ValueError(
                    f"Valor inválido em '{k}': {v!r} (esperado: {sorted(permitido)})"
                )
            atual[k] = str(v)
        else:
            raise ValueError(f"Chave de config não permitida: {k!r}")
    path = _config_path()
    fd, tmp = _tempfile.mkstemp(dir=str(path.parent), prefix=path.name, suffix=".tmp")
    try:
        with _os.fdopen(fd, "w", encoding="utf-8") as fh:
            json.dump(atual, fh, ensure_ascii=False, indent=2)
        _os.replace(tmp, path)
    except OSError:
        try:
            _os.unlink(tmp)
        except OSError:
            pass
        raise


@app.route("/configuracoes", methods=["GET"])
def configuracoes_pagina():
    """v2.16.9: página própria de configurações + manutenção (extraída do dashboard)."""
    return render_template("configuracoes.html")


@app.route("/configuracoes", methods=["POST"])
def configuracoes_salvar():
    """Form do card 'Configurações' do Dashboard. Aceita só chaves
    da allowlist; rejeita o resto.

    v2.15.4: dois modos de uso —
      A) form completo (botão "Salvar configurações" no fim) → salva tudo
      B) toggle individual (HTMX, request com `?campo=NOME`) → salva só esse
    """
    form = request.form
    campo_unico = request.args.get("campo")
    updates: dict = {}

    # Modo B: toggle individual via HTMX (auto-save)
    if campo_unico:
        if campo_unico in _CONFIG_BOOLEANOS_PERMITIDOS:
            updates[campo_unico] = form.get(campo_unico) in ("on", "1", "true", "yes")
        elif campo_unico in _CONFIG_ENUMS_PERMITIDOS:
            v = form.get(campo_unico, "")
            if v not in _CONFIG_ENUMS_PERMITIDOS[campo_unico]:
                return _resposta_erro(
                    f"{campo_unico}: valor '{v}' não permitido"
                )
            updates[campo_unico] = v
        elif campo_unico in _CONFIG_INTEIROS_PERMITIDOS:
            try:
                val = int(form[campo_unico])
            except (ValueError, KeyError):
                return _resposta_erro(f"{campo_unico} deve ser número inteiro.")
            # Validações de range
            if campo_unico == "polling_intervalo_segundos" and not (60 <= val <= 3600):
                return _resposta_erro("Intervalo: 60 a 3600 segundos.")
            if campo_unico == "pausa_entre_emails_segundos" and not (0 <= val <= 300):
                return _resposta_erro("Pausa entre emails: 0 a 300 segundos.")
            updates[campo_unico] = val
        else:
            return _resposta_erro(f"Campo não permitido: {campo_unico}")
    else:
        # Modo A: form completo
        for k in _CONFIG_BOOLEANOS_PERMITIDOS:
            updates[k] = form.get(k) in ("on", "1", "true", "yes")
        if "polling_intervalo_segundos" in form:
            try:
                val = int(form["polling_intervalo_segundos"])
            except ValueError:
                return _resposta_erro("Intervalo deve ser número inteiro.")
            if not (60 <= val <= 3600):
                return _resposta_erro("Intervalo deve estar entre 60 e 3600 segundos.")
            updates["polling_intervalo_segundos"] = val
        if "pausa_entre_emails_segundos" in form:
            try:
                val = int(form["pausa_entre_emails_segundos"])
            except ValueError:
                return _resposta_erro("Pausa entre emails deve ser número inteiro.")
            if not (0 <= val <= 300):
                return _resposta_erro("Pausa entre emails: 0 a 300 segundos.")
            updates["pausa_entre_emails_segundos"] = val

    try:
        _atualizar_config_json(updates)
    except (ValueError, OSError) as e:
        return _resposta_erro(f"Falha salvando configurações: {e}")

    # Para auto-save (HTMX), retorna fragment "Salvo" pra trocar inline
    if _eh_htmx() and campo_unico:
        from flask import make_response
        resp = make_response(render_template(
            "_partials/cfg_salvo.html",
            campo=campo_unico,
        ))
        return resp

    return _resposta_ok_ou_htmx(
        "Configurações salvas. Próxima passada vai usar os novos valores.",
        redirect_url=url_for("index"),
    )


@app.route("/maintenance/limpar-fingerprint", methods=["POST"])
def maintenance_limpar_fingerprint():
    """Apaga reprocesso_fp.json. O próximo reprocesso de qualquer pendência
    vai rodar mesmo sem mudança nas tabelas (útil quando você cadastrou
    algo direto no eContador e o sistema 'acha' que nada mudou)."""
    try:
        idempotencia.FP_FILE.unlink(missing_ok=True)
    except Exception as e:
        log.exception("Falha removendo fingerprint")
        return _resposta_erro(f"{type(e).__name__}: {e}")
    return _resposta_ok_ou_htmx(
        "Fingerprint de reprocesso limpo. Próximo Atualizar agora vai rodar "
        "tudo de novo (sem aviso 'nada mudou').",
        redirect_url=url_for("index"),
    )


@app.route("/pendencia/<msg_id>/excluir", methods=["POST"])
def excluir_pendencia(msg_id: str):
    """v2.16.37: remove TUDO relacionado a um msg_id — payloads, linhas
    da planilha, rascunhos, fingerprint, e remove labels do Gmail (o email
    volta a ficar sem nenhuma classificação, fora da fila).

    Útil pra pendências de falha técnica que não podem ser resolvidas
    pelo operador (ex: erro 529 do Anthropic), ou quando a entrada virou
    lixo e não vale a pena reprocessar.
    """
    msg_id = (msg_id or "").strip()
    if not msg_id or len(msg_id) < 12:
        return _resposta_erro("msg_id inválido.")

    apagados = {"payloads": 0, "planilha": 0, "rascunhos": 0}
    try:
        # 1) Payloads em disco
        from pathlib import Path as _Path
        for arq in _Path("payloads").glob(f"*{msg_id}*"):
            try:
                arq.unlink()
                apagados["payloads"] += 1
            except OSError as e:
                log.warning(f"[excluir] payload {arq.name}: {e}")

        # 2) Planilha admissoes.xlsx — todas as linhas com esse msg_id
        try:
            from openpyxl import load_workbook
            wb = load_workbook("admissoes.xlsx")
            ws = wb.active
            linhas_remover = []
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=2, values_only=True), start=2
            ):
                if not row:
                    continue
                cols = list(row) + [None] * (6 - len(row))
                mid = cols[5]
                if mid and str(mid).strip() == msg_id:
                    linhas_remover.append(row_idx)
            for idx in sorted(linhas_remover, reverse=True):
                ws.delete_rows(idx, 1)
            wb.save("admissoes.xlsx")
            apagados["planilha"] = len(linhas_remover)
        except FileNotFoundError:
            pass
        except Exception as e:
            log.warning(f"[excluir] planilha: {e}")

        # 3) Rascunhos com esse msg_id (lê do storage)
        try:
            import rascunhos_resposta as rr
            for rec in rr.listar(status=None, incluir_arquivados=True):
                if str(rec.get("msg_id", "")) == msg_id:
                    rec_id = rec.get("id")
                    if rec_id:
                        for arq in _Path("rascunhos").glob(f"{rec_id}_*.json"):
                            try:
                                arq.unlink()
                                apagados["rascunhos"] += 1
                            except OSError as e:
                                log.warning(f"[excluir] rasc {arq.name}: {e}")
        except Exception as e:
            log.warning(f"[excluir] rascunhos: {e}")

        # 4) Fingerprint de reprocesso
        try:
            fp = json.loads(idempotencia.FP_FILE.read_text(encoding="utf-8"))
            if msg_id in fp:
                fp.pop(msg_id, None)
                idempotencia.FP_FILE.write_text(
                    json.dumps(fp, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
        except (OSError, json.JSONDecodeError):
            pass

        # 5) CNPJ override (caso existisse)
        try:
            co_path = _Path("cnpj_overrides.json")
            if co_path.exists():
                co = json.loads(co_path.read_text(encoding="utf-8"))
                if msg_id in co:
                    co.pop(msg_id, None)
                    co_path.write_text(
                        json.dumps(co, ensure_ascii=False, indent=2),
                        encoding="utf-8",
                    )
        except (OSError, json.JSONDecodeError):
            pass

        # 6) Gmail labels — remove processado e pendente
        try:
            from gmail_client import GmailClient
            from main import carregar_config
            config = carregar_config()
            gmail = GmailClient()
            for label_attr in ("label_processado", "label_pendente"):
                try:
                    gmail.remover_label(msg_id, getattr(config, label_attr))
                except Exception:
                    pass
        except Exception as e:
            log.warning(f"[excluir] gmail: {e}")

        log.info(
            f"[excluir] msg_id={msg_id[:16]} removido: "
            f"{apagados['payloads']} payload(s), "
            f"{apagados['planilha']} linha(s) planilha, "
            f"{apagados['rascunhos']} rascunho(s)"
        )
    except Exception as e:
        log.exception(f"[excluir] Falha removendo {msg_id}")
        return _resposta_erro(f"{type(e).__name__}: {e}")

    return _resposta_ok_ou_htmx(
        f"Pendência {msg_id[:12]}… excluída. "
        f"({apagados['payloads']} payload, "
        f"{apagados['planilha']} linha planilha, "
        f"{apagados['rascunhos']} rascunho)",
        redirect_url=url_for("pendentes"),
    )


@app.route("/maintenance/reabrir-email", methods=["POST"])
def maintenance_reabrir_email():
    """v2.16.21: força reabertura de um email que já está marcado como
    processado no Gmail. Útil pra:
      - Re-rodar email com fix novo do pipeline (v2.16.20 detectou
        divergência de nome, mas email da Renata já estava fechado)
      - Corrigir admissão fechada errada que precisa de revisão
    Passos: remove label 'processado', adiciona 'pendente', limpa
    fingerprint daquele msg_id. Próximo polling reprocessa.
    """
    msg_id = (request.form.get("msg_id") or "").strip()
    if not msg_id or len(msg_id) < 12:
        return _resposta_erro(
            "msg_id inválido. Cole o ID hexadecimal do Gmail (ex: 19ee044535860deb)."
        )

    try:
        from gmail_client import GmailClient
        from main import carregar_config
        config = carregar_config()
        gmail = GmailClient()
        # Remove processado, adiciona pendente
        try:
            gmail.remover_label(msg_id, config.label_processado)
        except Exception as e:
            log.warning(f"[reabrir] remover processado falhou (talvez já não tinha): {e}")
        try:
            gmail.aplicar_label(msg_id, config.label_pendente)
        except Exception as e:
            log.warning(f"[reabrir] aplicar pendente falhou: {e}")
        # Limpa fingerprint deste msg_id pra forçar reprocesso
        try:
            fp = json.loads(idempotencia.FP_FILE.read_text(encoding="utf-8"))
            if msg_id in fp:
                fp.pop(msg_id, None)
                idempotencia.FP_FILE.write_text(
                    json.dumps(fp, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
        except (OSError, json.JSONDecodeError):
            pass  # fp não existia, ok
        log.info(f"[reabrir] msg_id={msg_id[:16]} reaberto pra reprocesso")
    except Exception as e:
        log.exception(f"[reabrir] Falha reabrindo {msg_id}")
        return _resposta_erro(f"{type(e).__name__}: {e}")

    return _resposta_ok_ou_htmx(
        f"Email {msg_id[:16]} reaberto. Clique em 'Atualizar agora' no dashboard "
        f"(ou aguarde próximo polling) pra ele ser reprocessado com as regras "
        f"mais recentes.",
        redirect_url=url_for("configuracoes_pagina"),
    )


# Allowlist explícita de campos que o operador PODE sobrescrever via form.
# Tudo fora dessa lista é rejeitado (proteção contra campos perigosos tipo
# `statusadmissao=2` que quebram o invariante "sempre 1" do CLAUDE.md).
_OVERRIDES_PERMITIDOS = {
    "salario", "admissao", "nascimento",
    "cep", "rua", "numero", "complemento", "bairro", "cidade",
    "cpf", "ctps", "diascontratoexperiencia", "primeiroemprego",
    "nomecargo", "nomedamae", "nomedopai",
    "email", "celular", "telefone",
}


def _parse_salario_br(v) -> float:
    """Aceita formatos comuns no Brasil: '1500', '1500,00', '1.500,00',
    'R$ 1.500,00', '1500.00'. Levanta ValueError se não der pra interpretar."""
    s = str(v).strip()
    # Remove R$, espaços, e qualquer letra
    import re as _re
    s = _re.sub(r"[^0-9,.\-]", "", s)
    if not s:
        raise ValueError("vazio")
    # Heurística: se tem ',' e '.', '.' é separador de milhar (formato BR)
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    return float(s)


@app.route("/pendencia/<msg_id>/<path:nome>/<cnpj>/postar", methods=["POST"])
def postar_pendencia(msg_id: str, nome: str, cnpj: str):
    """POSTa a admissão usando o payload salvo. Mesmo caminho do
    'Aplicar form e POSTar' da Tkinter, mas vindo da web.

    Form fields:
      - permitir_duplicata=on (quando há hit de idempotência)
      - overrides em _OVERRIDES_PERMITIDOS (salario, admissao, cep, …)
    """
    nome = _validar_nome_path(nome)
    cnpj = _cnpj_da_url(cnpj)
    payload_path = dd.achar_payload_por_msg_e_nome(msg_id, nome)
    if not payload_path:
        return _resposta_erro(
            "Não achei payload salvo pra essa pendência. "
            "Use 'Marcar como resolvido' ou rode uma passada nova."
        )
    doc = dd.carregar_payload_completo(payload_path)
    payload = doc.get("payload") or {}
    if "data" not in payload:
        return _resposta_erro("Payload sem 'data' — corrompido.")

    attrs = payload["data"].setdefault("attributes", {})

    # v2.16.29: flag pra forçar envio mesmo com campos faltando
    forcar_envio = request.form.get("forcar") in ("on", "1", "true", "yes")

    # Aplica overrides do form com ALLOWLIST (review §HIGH — proteção contra
    # campos perigosos tipo statusadmissao=2).
    sobrescreveu: list[str] = []
    for k, v in request.form.items():
        if k in ("permitir_duplicata", "csrf_token", "forcar") or not v:
            continue
        if k not in _OVERRIDES_PERMITIDOS:
            return _resposta_erro(
                f"Campo não permitido: '{k}'. "
                f"Permitidos: {', '.join(sorted(_OVERRIDES_PERMITIDOS))}"
            )
        try:
            if k == "salario":
                attrs[k] = _parse_salario_br(v)
            elif k in ("cpf", "numero", "diascontratoexperiencia", "ctps"):
                import re as _re
                d = _re.sub(r"\D", "", str(v))
                attrs[k] = int(d) if d else attrs.get(k)
            elif k == "primeiroemprego":
                attrs[k] = str(v).lower() in ("true", "1", "sim", "on", "yes")
            else:
                attrs[k] = str(v).strip()
            sobrescreveu.append(k)
        except Exception as e:
            return _resposta_erro(f"Campo '{k}' inválido: {v!r} ({e})")

    # Sanitiza ANTES de enviar (clamps 422, etc.) — falha aqui NÃO é silenciada
    # (review: era except: pass — bug invisível).
    try:
        from payload_builder import sanitizar_attributes
        payload["data"]["attributes"] = sanitizar_attributes(attrs)
    except Exception as e:
        log.exception("[webapp] sanitize falhou — abortando POST")
        return _resposta_erro(f"Sanitização do payload falhou: {e}")

    # v2.15.13: defaults do escritório — o payload parcial salvo quando o
    # Claude marca pendente raiz não tem os campos obrigatórios (primeiroemprego,
    # raca, vínculo, status, etc). Sem eles a API rejeita com HTTP 422.
    # Aplica TUDO que tem default documentado no briefing antes do POST.
    _attrs_p = payload["data"].setdefault("attributes", {})
    _rels_p = payload["data"].setdefault("relationships", {})

    # Attributes com default fixo
    _DEFAULTS_ATTRS = {
        "primeiroemprego": False,
        "possuideficiencia": False,
        "requersegurodesemprego": False,
        "diascontratoexperiencia": 30,
        "usuariocriacao": "PIPELINE-WEB",
    }
    for k, v in _DEFAULTS_ATTRS.items():
        if k not in _attrs_p:
            _attrs_p[k] = v

    # Relationships com default do escritório (ver briefing §4)
    _DEFAULTS_RELS = [
        ("statusadmissao",           "tipos-status-admissao",            "1"),  # Análise verde
        ("tipoadmissao",             "tipos-admissao",                   "1"),  # Admissão
        ("tipovinculotrabalhista",   "tipos-vinculos-trabalhista",       "60"), # CLT urbano determinado
        ("categoriawdp",             "tipos-categoria",                  "1"),  # Trabalhador
        ("formapagamento",           "tipos-forma-de-pagamento",         "4"),  # Mensal
        ("tipoidentidade",           "tipos-identidade",                 "1"),  # RG (workaround off-by-one)
        ("raca",                     "tipos-raca",                       "8"),  # Parda (v2.16.43: bug off-by-one FIXED por Alterdata)
        ("tipoDeDeficiencia",        "tipos-deficiencia",                "0"),  # Não possui
        ("statusatestadoocupacional","tipos-status-atestado-ocupacional","1"),  # Apto
        ("nacionalidade",            "paises",                           "105"),# Brasil
        ("paisnascimento",           "paises",                           "105"),
        ("pais",                     "paises",                           "105"),
        ("naturalidade",             "estados",                          "9"),  # Goiás
        ("ufidentidade",             "estados",                          "9"),
        ("ufctps",                   "estados",                          "9"),
        ("sexo",                     "tipos-sexo",                       "2"),  # F (mais comum no setor admin)
        ("estadocivil",              "tipos-estado-civil",               "1"),  # Solteiro
        ("escolaridade",             "tipos-escolaridade",               "7"),  # Médio completo
    ]
    for rel_nome, tipo, id_default in _DEFAULTS_RELS:
        if rel_nome not in _rels_p or not (_rels_p.get(rel_nome) or {}).get("data", {}).get("id"):
            _rels_p[rel_nome] = {"data": {"type": tipo, "id": id_default}}

    # CTPS derivada do CPF (regra do escritório) — só se não vier do Claude
    if "ctps" not in _attrs_p and _attrs_p.get("cpf"):
        try:
            cpf_str = str(int(_attrs_p["cpf"])).zfill(11)
            _attrs_p["ctps"] = int(cpf_str[:7])
            _attrs_p["seriectps"] = cpf_str[7:11]
        except (ValueError, TypeError):
            pass

    # v2.15.12: paridade com o orquestrador — tenta RESOLVER as relationships
    # faltantes ANTES de bloquear. A pendência salva geralmente vem só com
    # attrs (caminho raiz _pendente), sem empresa/função/estado resolvidos.
    # Em vez de exigir que o user faça tudo manual, a web faz o mesmo que
    # o orquestrador faria: bate na API pra resolver empresa, busca cargo
    # na planilha CBO, default Goiás pro estado.
    attrs_final = payload["data"]["attributes"]
    rels = payload["data"].setdefault("relationships", {})
    cnpj_resolucao = str((doc.get("resolucao") or {}).get("cnpj_empresa") or cnpj)
    empresa_id = (rels.get("empresa") or {}).get("data", {}).get("id")
    funcao_id = (rels.get("funcao") or {}).get("data", {}).get("id")
    estado_id = (rels.get("estado") or {}).get("data", {}).get("id")

    # Lê config uma vez só pra respeitar as flags do user
    cfg_atual = _ler_config_json()

    # ── Resolve EMPRESA pelo CNPJ ────────────────────────────────
    if not empresa_id and cnpj_resolucao:
        try:
            from main import carregar_config as _cf
            _cfg_obj = _cf()
            _api_resolve = EContadorAPI(_cfg_obj.base_url, _cfg_obj.token)
            try:
                eid, _attrs = _api_resolve.resolver_empresa(cnpj_resolucao)
                if eid:
                    rels["empresa"] = {"data": {"type": "empresas", "id": str(eid)}}
                    empresa_id = eid
                    log.info(f"[postar via web] empresa resolvida: CNPJ {cnpj_resolucao} → id={eid}")
            finally:
                _api_resolve.close()
        except Exception as e:
            log.warning(f"[postar via web] falha resolvendo empresa: {e}")

    # ── Resolve FUNÇÃO pela planilha CBO (e respeita flag) ───────
    if not funcao_id:
        cargo = attrs_final.get("nomecargo") or (doc.get("resolucao") or {}).get("cargo_extraido") or ""
        if cargo:
            try:
                from main import carregar_planilha, PLANILHA_CBO
                from funcao import resolver_funcao
                _planilha = carregar_planilha(PLANILHA_CBO)
                _fid, _conf, _amb, _msg = resolver_funcao(_planilha, cargo, None)
                if _fid:
                    rels["funcao"] = {"data": {"type": "funcoes", "id": str(_fid)}}
                    funcao_id = _fid
                    log.info(f"[postar via web] função resolvida: '{cargo}' → id={_fid} ({_conf:.0%})")
            except Exception as e:
                log.warning(f"[postar via web] falha resolvendo função: {e}")
    # v2.15.14: Flag "SEMPRE mandar SEM função" — antes mandava id=1
    # (que vira "ASSISTENTE DE CPD" no Desktop e confunde o DP). Agora
    # OMITE a relationship `funcao` do payload. DP escolhe no Desktop.
    if not funcao_id and cfg_atual.get("sempre_mandar_sem_funcao"):
        rels.pop("funcao", None)
        funcao_id = "OMITIDA"  # sentinela só pra passar checagem abaixo
        log.info("[postar via web] função omitida (flag 'sem função' ligada) — DP escolhe no Desktop")

    # ── Endereço padrão por CNPJ (v2.16.34) ──────────────────────
    # Aplica cadastro de endereço padrão da empresa antes da validação
    # final. Caso real (GABRIEL, 2026-06-22): MAMBORE cadastrada com
    # endereço da fazenda mas POST manual via web rejeitava com 'faltam
    # cep/rua/bairro/cidade' porque essa lógica era exclusiva do pipeline
    # orquestrador.
    try:
        import enderecos_padrao_empresa as _epe
        if cnpj_resolucao:
            preenchidos_end = _epe.aplicar_em_attrs(attrs_final, cnpj_resolucao)
            if preenchidos_end:
                log.info(
                    f"[postar via web] 🏠 Endereço padrão do CNPJ "
                    f"{cnpj_resolucao} aplicado: {', '.join(preenchidos_end[:3])}"
                )
    except Exception as e:
        log.warning(f"[postar via web] Falha aplicando endereço padrão: {e}")

    # ── ViaCEP: completa rua/bairro/cidade quando só temos o CEP (v2.16.40) ──
    # Caso real (24/06): operador colou o CEP na pendência e bateu o erro
    # "ainda faltam: endereço (rua, bairro, cidade)". Agora o webapp consulta
    # ViaCEP e preenche os campos derivados antes da validação final.
    try:
        cep_atual = attrs_final.get("cep")
        if cep_atual:
            faltam_end = [
                c for c in ("rua", "bairro", "cidade")
                if not (attrs_final.get(c) or "")
            ]
            if faltam_end:
                import enrichment as _enr
                dados_cep = _enr.enrich_from_cep(cep_atual)
                preench_vc = []
                for chave in ("rua", "bairro", "cidade"):
                    if not attrs_final.get(chave) and dados_cep.get(chave):
                        attrs_final[chave] = dados_cep[chave]
                        preench_vc.append(f"{chave}={dados_cep[chave]!r}")
                # Aproveita o estado_id do ViaCEP se ainda não temos
                if not estado_id and dados_cep.get("_estado_id"):
                    rels["estado"] = {
                        "data": {
                            "type": "estados",
                            "id": str(dados_cep["_estado_id"]),
                        }
                    }
                    estado_id = str(dados_cep["_estado_id"])
                    preench_vc.append(f"estado={estado_id} (UF {dados_cep.get('_uf')})")
                if preench_vc:
                    log.info(
                        f"[postar via web] 🌐 ViaCEP preencheu: "
                        f"{', '.join(preench_vc)}"
                    )
    except Exception as e:
        log.warning(f"[postar via web] Falha ViaCEP: {e}")

    # ── Resolve ESTADO: default Goiás (9) quando temos CEP ───────
    # Quase 100% das admissões da Crosara são GO. Caso ViaCEP não tenha
    # devolvido o _estado_id (sem internet, CEP novo etc), GO é o fallback.
    if not estado_id:
        if attrs_final.get("cep"):
            rels["estado"] = {"data": {"type": "estados", "id": "9"}}
            estado_id = "9"
            log.info("[postar via web] estado=9 (Goiás, default)")

    # ── Agora SIM checa o que ainda falta ────────────────────────
    # v2.16.36: separa em HARD requirement (empresa) vs soft (que forcar=on
    # pode pular). Lição do incidente GABRIEL/MAMBORE (2026-06-22):
    # forcar=on burlava até a validação de empresa, criando candidato órfão
    # no eContador (sem empresa vinculada) → não descia pro Desktop. Empresa
    # SEMPRE precisa estar cadastrada no eContador antes do POST.
    bloqueio_hard: list[str] = []
    soft_faltando: list[str] = []
    if not empresa_id:
        bloqueio_hard.append(
            f"EMPRESA — CNPJ {cnpj_resolucao or '(vazio)'} NÃO está cadastrada "
            f"no eContador. Sem isso, candidato fica ÓRFÃO e não desce pro "
            f"Alterdata Desktop. Cadastre a empresa primeiro no eContador "
            f"e depois reenvie. NEM 'Enviar mesmo assim' resolve isso."
        )
    if not funcao_id:
        soft_faltando.append(
            f"função — cargo '{attrs_final.get('nomecargo', '(vazio)')}' não bate "
            f"em nenhum cargo X-marcado na planilha CBO. Adicione o cargo na "
            f"planilha ou marque 'Mandar SEM função' nas Configurações."
        )
    end_faltando = [
        c for c in ("cep", "rua", "bairro", "cidade")
        if not (attrs_final.get(c) or "")
    ]
    if not estado_id and not attrs_final.get("cep"):
        end_faltando.append("estado")
    if end_faltando:
        soft_faltando.append("endereço (" + ", ".join(end_faltando) + ")")

    # Hard bloqueio NUNCA pula, nem com forcar=on
    if bloqueio_hard:
        return _resposta_erro(
            "🛑 BLOQUEIO HARD — não dá pra enviar de jeito nenhum:\n\n• "
            + "\n\n• ".join(bloqueio_hard)
            + ("\n\nAlém disso, faltam (mas isso forcar=on resolveria):\n\n• "
               + "\n\n• ".join(soft_faltando) if soft_faltando else "")
        )
    # Soft: bloqueia normal, libera com forcar=on
    if soft_faltando and not forcar_envio:
        return _resposta_erro(
            "Não dá pra enviar — ainda faltam:\n\n• "
            + "\n\n• ".join(soft_faltando)
        )
    if soft_faltando and forcar_envio:
        log.warning(
            f"[postar via web] ⚠ FORÇADO pelo operador — enviando com campos "
            f"faltando: {'; '.join(soft_faltando)[:200]}"
        )

    # CPF pra idempotência
    cpf_final = attrs_final.get("cpf")
    permitir_dup = request.form.get("permitir_duplicata") in ("on", "1", "true")

    # Wrapper único faz idempotência + POST + registro + label.
    # v2.15.0 fix: passar gmail+labels pra wrapper FECHAR a thread (sem isso
    # a próxima passada do Task Scheduler reprocessa o email → ciclo do
    # RAIMUNDO; cada reprocesso ~US$0.40).
    gmail = _abrir_gmail()
    try:
        from main import carregar_config
        config = carregar_config()
        api = EContadorAPI(config.base_url, config.token)
        try:
            res = postar_candidato_registrado(
                api, payload,
                cpf=cpf_final, cnpj=cnpj_resolucao, nome=nome,
                origem="web_resolver", msg_id=msg_id,
                permitir_duplicata=permitir_dup,
                payload_path=payload_path,
                gmail=gmail,
                label_processado=config.label_processado,
                label_pendente_remover=[msg_id] if msg_id else None,
            )
        finally:
            api.close()
    except Exception as e:
        log.exception("Falha postando pela web")
        return _resposta_erro(f"{type(e).__name__}: {e}")
    finally:
        if gmail is not None:
            try:
                gmail.close()
            except Exception:
                pass

    if res.ok:
        aviso_planilha = ""
        try:
            from main import registrar_admissao_planilha
            extra = " — POST pulado (idempotência)" if res.pulou else ""
            registrar_admissao_planilha(
                nome=nome,
                empresa=str((doc.get("resolucao") or {}).get("razao_social") or ""),
                cnpj=cnpj,
                procedencia=(
                    f"Cadastrado — candidato {res.candidato_id} "
                    f"(resolvido via web){extra}"
                ),
                msg_id=msg_id,
            )
        except Exception:
            log.exception("Falha registrando linha na planilha")
            aviso_planilha = " (atenção: planilha não atualizou — feche o Excel se estiver aberto)"

        msg = (
            f"Já estava cadastrado: candidato {res.candidato_id}.{aviso_planilha}"
            if res.pulou else
            f"Candidato {res.candidato_id} criado com sucesso.{aviso_planilha}"
        )
        return _resposta_ok_ou_htmx(msg, redirect_url=url_for("pendentes"))

    return _resposta_erro(
        f"Falha técnica no POST: {res.erro_ref}\n\n{(res.body_err or '')[:600]}"
    )


# ── Fragments HTMX (refresh parcial) ──────────────────────────────

@app.route("/htmx/contadores")
def htmx_contadores():
    """Re-renderiza só os cards do dashboard. Cliente pode hx-trigger every 30s."""
    return render_template(
        "_partials/contadores.html",
        contadores=dd.resumo_contadores(),
    )


@app.route("/htmx/status-passada")
def htmx_status_passada():
    """Cliente pode hx-trigger every 2s pra acompanhar o botão Atualizar."""
    return _renderizar_status_passada()


@app.route("/htmx/lista-pendentes")
def htmx_lista_pendentes():
    """Re-renderiza só a lista de pendentes (sem o layout)."""
    filtro = (request.args.get("q") or "").strip().lower()
    todas = dd.listar_entidades(incluir_cargo_ia=True)
    pend = [
        e for e in todas
        if e["categoria"].startswith("pendente") or e["categoria"] == "falha_tecnica"
    ]
    if filtro:
        def _bate(e):
            blob = " ".join((e["nome"], e["empresa"], e["cnpj"],
                              e["procedencia"], e.get("cargo_ia", "")))
            return filtro in blob.lower()
        pend = [e for e in pend if _bate(e)]
    return render_template(
        "_partials/lista_pendentes.html",
        por_dia=dd.agrupar_por_dia(pend),
        total=len(pend),
    )


# ── Perfis de remetente (v2.16.0) ────────────────────────────────

@app.route("/perfis")
def perfis_listar():
    """Tabela de remetentes ordenada por volume."""
    import perfis_remetente as pr
    return render_template("perfis.html", perfis=pr.listar_resumido())


@app.route("/perfis/<path:remetente>")
def perfis_detalhe(remetente):
    """Detalhe de um remetente. <path:> aceita 'rh@empresa.com' (com @ no path)."""
    import perfis_remetente as pr
    perfil = pr.perfil_de(remetente, consolidar_se_faltar=True)
    if not perfil:
        abort(404, description=f"Remetente '{remetente}' não tem histórico.")
    return render_template("perfil_detalhe.html",
                            remetente=remetente, perfil=perfil)


@app.route("/perfis/<path:remetente>/observacoes", methods=["POST"])
def perfis_salvar_observacoes(remetente):
    """Salva observações livres do operador pra esse remetente."""
    import perfis_remetente as pr
    obs = (request.form.get("observacoes") or "").strip()
    nome = (request.form.get("nome_apresentacao") or "").strip()
    if not pr.atualizar_observacoes(remetente, obs, nome):
        return _resposta_erro("Remetente inválido.")
    return _resposta_ok_ou_htmx(
        f"Observações de {remetente} salvas.",
        redirect_url=url_for("perfis_detalhe", remetente=remetente),
    )


@app.route("/perfis/<path:remetente>/salario-cargo", methods=["POST"])
def perfis_salvar_salario_cargo(remetente):
    """v2.16.39: cadastra (ou limpa) o salário manual de um cargo no perfil
    do remetente. Útil pra clientes que escrevem 'CONFIRMAR COM O DA MESMA
    FUNÇÃO' em vez de informar o valor.

    Form fields:
      - cargo (str, obrigatório)
      - valor (R$ — aceita formato BR; vazio/0 = remove cadastro)
    """
    import perfis_remetente as pr
    cargo = (request.form.get("cargo") or "").strip().upper()
    valor_raw = (request.form.get("valor") or "").strip()
    if not cargo:
        return _resposta_erro("Cargo é obrigatório.")
    valor = None
    if valor_raw:
        try:
            valor = _parse_salario_br(valor_raw)
        except ValueError:
            return _resposta_erro(
                f"Salário inválido: {valor_raw!r}. Use formato como 1500 ou 1.500,00."
            )
    if not pr.atualizar_salario_manual_cargo(remetente, cargo, valor):
        return _resposta_erro("Remetente ou cargo inválido.")
    msg = (f"Salário manual de {cargo!r} removido."
           if not valor else
           f"Salário manual de {cargo!r} = R$ {valor:.2f} salvo. "
           f"Próximas admissões usam.")
    return _resposta_ok_ou_htmx(
        msg, redirect_url=url_for("perfis_detalhe", remetente=remetente)
    )


@app.route("/perfis/<path:remetente>/defaults", methods=["POST"])
def perfis_salvar_defaults(remetente):
    """v2.16.4: salva os defaults_quando_ausente do perfil. Endereço da
    empresa (pra remetentes que não mandam comprovante) + flag omitir_aso."""
    import perfis_remetente as pr
    end = {
        "rua": (request.form.get("rua") or "").strip().upper(),
        "bairro": (request.form.get("bairro") or "").strip().upper(),
        "cidade": (request.form.get("cidade") or "").strip().upper(),
        "uf": (request.form.get("uf") or "").strip().upper(),
        "cep": re.sub(r"\D", "", request.form.get("cep") or ""),
    }
    # numero como int — aceita vazio (= 0)
    num_raw = (request.form.get("numero") or "").strip()
    end["numero"] = int(num_raw) if num_raw.isdigit() else 0
    # Se TODOS os campos estão vazios, salva como "sem endereço default"
    if not any(v for k, v in end.items() if k != "numero"):
        end = {}
    obs = (request.form.get("default_observacao") or "").strip()
    defaults = {
        "endereco": end,
        "omitir_aso": request.form.get("omitir_aso") == "on",
    }
    if obs:
        defaults["_observacao"] = obs
    if not pr.atualizar_defaults_quando_ausente(remetente, defaults):
        return _resposta_erro("Remetente inválido.")
    return _resposta_ok_ou_htmx(
        f"Defaults de {remetente} salvos. Próximas admissões usam.",
        redirect_url=url_for("perfis_detalhe", remetente=remetente),
    )


# ── Rascunhos de resposta (v2.16.19) ─────────────────────────────

@app.route("/respostas")
def respostas_listar():
    """Lista rascunhos pendentes de revisão (auto-email pendência modo=rascunho).
    v2.16.32: também mostra histórico de TODO email enviado (audit completo)."""
    import rascunhos_resposta as rr
    import auditoria_emails as ae
    pendentes = rr.listar(status=rr.STATUS_PENDENTE)
    arquivados = rr.listar(status=rr.STATUS_ENVIADO)[:20] + rr.listar(
        status=rr.STATUS_DESCARTADO
    )[:20]
    # v2.16.32: histórico de auditoria (TODO email enviado, manual OU automático)
    enviados_audit = ae.listar(desde_dias=60)
    return render_template(
        "respostas.html",
        pendentes=pendentes,
        arquivados=sorted(arquivados, key=lambda r: r.get("ts_acao", ""), reverse=True),
        enviados_audit=enviados_audit,
        n_pendentes=len(pendentes),
        n_enviados_24h=ae.contar_ultimas_24h(),
    )


@app.route("/respostas/<rid>")
def respostas_detalhe(rid):
    import rascunhos_resposta as rr
    rec = rr.carregar(rid)
    if not rec:
        abort(404, description="Rascunho não encontrado")
    return render_template("resposta_detalhe.html", rec=rec)


@app.route("/respostas/<rid>/salvar", methods=["POST"])
def respostas_salvar_edicao(rid):
    """Salva edição do corpo (auto-save HTMX)."""
    import rascunhos_resposta as rr
    corpo_editado = (request.form.get("corpo_editado") or "").rstrip() + "\n"
    rec = rr.atualizar(rid, corpo_editado=corpo_editado)
    if not rec:
        return _resposta_erro("Rascunho não encontrado")
    return _resposta_ok_ou_htmx(
        "Rascunho salvo.",
        redirect_url=url_for("respostas_detalhe", rid=rid),
    )


@app.route("/respostas/<rid>/aprovar", methods=["POST"])
def respostas_aprovar(rid):
    """Envia o rascunho via Gmail e marca como enviado."""
    import rascunhos_resposta as rr
    rec = rr.carregar(rid)
    if not rec:
        return _resposta_erro("Rascunho não encontrado")
    if rec["status"] != rr.STATUS_PENDENTE:
        return _resposta_erro(
            f"Rascunho já está como '{rec['status']}' — não dá pra reenviar"
        )
    # Aplica edição (se houver) e manda
    corpo = rr.corpo_final(rec)
    try:
        gmail = _abrir_gmail()
        if not gmail:
            return _resposta_erro(
                "Gmail indisponível no momento — tente de novo em alguns segundos."
            )
        # Reconstroi o "msg" minimal com threadId e From pra reply
        msg_pra_reply = {
            "id": rec["msg_id"],
            "threadId": rec.get("thread_id") or rec["msg_id"],
            "payload": {
                "headers": [
                    {"name": "From",
                     "value": (
                         f"{rec.get('remetente_nome') or ''} "
                         f"<{rec['remetente_email']}>"
                     ).strip()},
                    {"name": "Subject", "value": rec.get("assunto") or ""},
                ]
            },
        }
        from main import carregar_config
        config = carregar_config()
        gmail.responder_no_thread(
            msg_pra_reply, corpo=corpo, cc=config.email_dp or None
        )
        rr.marcar_enviado(rid, operador="web")
        # v2.16.32: AUDIT — registra envio manual aprovado via web
        try:
            import auditoria_emails as ae
            ae.registrar_envio(
                msg_id_original=rec["msg_id"],
                thread_id=rec.get("thread_id") or rec["msg_id"],
                destinatario_email=rec["remetente_email"],
                destinatario_nome=rec.get("remetente_nome", ""),
                assunto=rec.get("assunto", ""),
                corpo=corpo,
                cc=config.email_dp or None,
                origem="web_aprovado",
                operador="web",
                sucesso=True,
                contexto=rec.get("contexto") or {},
            )
        except Exception as e:
            log.warning(f"[audit] falha registrando aprovação manual: {e}")
    except Exception as e:
        log.exception(f"Falha enviando rascunho {rid}")
        # v2.16.32: registra falha pra ter trilha de auditoria
        try:
            import auditoria_emails as ae
            ae.registrar_envio(
                msg_id_original=rec["msg_id"],
                thread_id=rec.get("thread_id") or rec["msg_id"],
                destinatario_email=rec["remetente_email"],
                destinatario_nome=rec.get("remetente_nome", ""),
                assunto=rec.get("assunto", ""),
                corpo=corpo,
                origem="web_aprovado", operador="web",
                sucesso=False, erro=str(e)[:200],
            )
        except Exception:
            pass
        return _resposta_erro(f"Falha enviando: {type(e).__name__}: {e}")
    return _resposta_ok_ou_htmx(
        f"Email enviado pra {rec.get('remetente_email')}.",
        redirect_url=url_for("respostas_listar"),
    )


@app.route("/respostas/<rid>/descartar", methods=["POST"])
def respostas_descartar(rid):
    import rascunhos_resposta as rr
    rec = rr.carregar(rid)
    if not rec:
        return _resposta_erro("Rascunho não encontrado")
    motivo = (request.form.get("motivo") or "").strip()
    rr.marcar_descartado(rid, operador="web", motivo=motivo)
    return _resposta_ok_ou_htmx(
        "Rascunho descartado (não foi enviado).",
        redirect_url=url_for("respostas_listar"),
    )


@app.route("/perfis/recalcular", methods=["POST"])
def perfis_recalcular():
    """Força refresh de TODOS os perfis (lendo todos os payloads de novo).
    Demora alguns segundos em proporção ao volume."""
    import perfis_remetente as pr
    n = len(pr.consolidar_todos())
    return _resposta_ok_ou_htmx(
        f"{n} perfis recalculados.",
        redirect_url=url_for("perfis_listar"),
    )


# ── Debug: ler candidato direto da API eContador (v2.16.35) ─────

@app.route("/processadas/<msg_id>/<path:nome>/payload")
def processadas_ver_payload(msg_id: str, nome: str):
    """v2.16.35: mostra o payload enviado + (opcional) estado atual no
    eContador pra uma admissão processada. Útil pra auditoria e debug."""
    nome = _validar_nome_path(nome)
    payload_path = dd.achar_payload_por_msg_e_nome(msg_id, nome)
    if not payload_path:
        abort(404, description="Payload não encontrado pra essa admissão")
    doc = dd.carregar_payload_completo(payload_path)
    payload = doc.get("payload") or {}
    attrs = ((payload.get("data") or {}).get("attributes") or {})
    rels_raw = ((payload.get("data") or {}).get("relationships") or {})
    rels = {}
    for k, v in rels_raw.items():
        d = (v or {}).get("data")
        if isinstance(d, dict):
            rels[k] = {"id": d.get("id"), "type": d.get("type")}
    resol = doc.get("resolucao") or {}
    resultado = doc.get("resultado") or {}
    candidato_id = resultado.get("candidato_id")
    return render_template(
        "payload_view.html",
        msg_id=msg_id, nome=nome,
        attrs=attrs, rels=rels,
        resol=resol, resultado=resultado,
        candidato_id=candidato_id,
        payload_path=payload_path.name,
    )


@app.route("/api/candidato/<candidato_id>")
def api_candidato_detalhe(candidato_id):
    """Lê um candidato direto do eContador com relationships populadas.
    Útil pra verificar se um POST/PATCH foi persistido corretamente."""
    if not re.match(r"^\d{1,7}$", candidato_id):
        return jsonify({"erro": "candidato_id inválido"}), 400
    try:
        from main import carregar_config
        cfg = carregar_config()
        api = EContadorAPI(cfg.base_url, cfg.token)
        try:
            ok, body = api.get_candidato(candidato_id)
        finally:
            api.close()
        if not ok:
            return jsonify({"ok": False, **body}), 502
        # Resumo amigável + payload bruto
        data = body.get("data") or {}
        attrs = data.get("attributes") or {}
        rels = data.get("relationships") or {}
        rels_resumo = {}
        for k, v in rels.items():
            d = (v or {}).get("data")
            if isinstance(d, dict):
                rels_resumo[k] = d.get("id")
        resumo = {
            "id": data.get("id"),
            "nome": attrs.get("nome"),
            "cpf": attrs.get("cpf"),
            "admissao": attrs.get("admissao"),
            "salario": attrs.get("salario"),
            "endereco": {
                "cep": attrs.get("cep"), "rua": attrs.get("rua"),
                "bairro": attrs.get("bairro"), "cidade": attrs.get("cidade"),
            },
            "rels_resolvidas": rels_resumo,
        }
        return jsonify({"ok": True, "resumo": resumo, "raw": body})
    except Exception as e:
        log.exception("Falha lendo candidato")
        return jsonify({"ok": False, "erro": f"{type(e).__name__}: {e}"}), 500


# ── JSON API (pra ferramentas externas / curl / scripts) ──────────

@app.route("/api/status")
def api_status():
    """Snapshot do pipeline em JSON. Útil pra integrações externas."""
    return jsonify({
        "app_version": APP_VERSION,
        "status": dd.status_pipeline(),
        "passada": PASSADA.snapshot(),
    })


@app.route("/api/pendentes")
def api_pendentes():
    """Lista de entidades pendentes em JSON."""
    pend = [
        e for e in dd.listar_entidades(incluir_cargo_ia=False)
        if e["categoria"].startswith("pendente") or e["categoria"] == "falha_tecnica"
    ]
    return jsonify({"total": len(pend), "items": pend})


# ── Helpers de resposta ─────────────────────────────────────────

def _renderizar_status_passada(mensagem: str = "") -> str:
    """Retorna o HTML do "status box" da passada (usado em HTMX e em
    confirmação de ações)."""
    return render_template(
        "_partials/status_passada.html",
        passada=PASSADA.snapshot(),
        mensagem=mensagem,
    )


def _resposta_ok_ou_htmx(mensagem: str, redirect_url: str = "/") -> Any:
    """Em request HTMX, retorna fragment + HX-Redirect pra atualizar a página.
    Em request normal, retorna template flash + redirect manual.
    redirect_url é sempre sanitizado (review MEDIUM XSS)."""
    redirect_url = _safe_redirect(redirect_url, "/")
    if _eh_htmx():
        from flask import make_response
        resp = make_response(
            render_template("_partials/flash.html", tipo="ok", mensagem=mensagem)
        )
        resp.headers["HX-Trigger"] = "flash"
        resp.headers["HX-Redirect"] = redirect_url
        return resp
    return render_template("_partials/flash_full.html",
                            tipo="ok", mensagem=mensagem,
                            redirect=redirect_url)


def _resposta_erro(mensagem: str) -> Any:
    redirect_url = _safe_redirect(request.referrer, "/")
    if _eh_htmx():
        from flask import make_response
        resp = make_response(
            render_template("_partials/flash.html", tipo="erro", mensagem=mensagem)
        )
        return resp, 400
    return render_template(
        "_partials/flash_full.html",
        tipo="erro", mensagem=mensagem,
        redirect=redirect_url,
    ), 400


def _billing_resumido() -> dict:
    """{custo_usd, custo_brl, n_calls, ...} — soma do mês corrente."""
    try:
        from main import sum_billing_mes_atual
        from directdata_client import sum_directdata_mes_atual
        b = sum_billing_mes_atual()
        d = sum_directdata_mes_atual()
        return {
            "claude": b,
            "directdata": d,
        }
    except Exception as e:
        log.warning(f"Falha somando billing: {e}")
        return {"claude": {}, "directdata": {}}


# ── Boot ─────────────────────────────────────────────────────────

def _ip_lan() -> str:
    """Tenta descobrir o IP da máquina na LAN (pra mostrar no log de boot).
    Sem internet também funciona — só consulta a rota local.
    Timeout de 500ms pra não bloquear o boot em rede restritiva."""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(0.5)
        # Não envia pacote — só usa o roteamento local
        s.connect(("10.255.255.255", 1))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def _ips_lan_todos() -> list[str]:
    """Lista TODOS os IPv4 não-loopback da máquina (Ethernet, Wi-Fi, VPN, etc.).
    Usado pra mostrar "Acesso de outras máquinas" na sidebar — operador pode
    copiar qualquer um pra mandar pra colega da rede.

    Em alguns PCs com várias placas (Wi-Fi + Ethernet + VPN), pode ter 3+ IPs.
    Retorna ordenado, sem duplicatas.
    """
    ips: set[str] = set()
    # Estratégia 1: roteamento default (mais confiável — pega o IP da rota
    # padrão pra fora, geralmente o que outras máquinas da LAN alcançam)
    ip_default = _ip_lan()
    if ip_default and not ip_default.startswith("127."):
        ips.add(ip_default)
    # Estratégia 2: getaddrinfo no hostname — pega outros adaptadores
    try:
        for entry in socket.getaddrinfo(socket.gethostname(), None):
            ip = entry[4][0]
            if ":" in ip:  # ignora IPv6
                continue
            if ip.startswith("127.") or ip.startswith("169.254."):
                continue
            ips.add(ip)
    except Exception:
        pass
    return sorted(ips)


def _setup_logging():
    handler = logging.StreamHandler(sys.stdout)
    handler.setFormatter(logging.Formatter(
        "%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    ))
    root = logging.getLogger()
    if not root.handlers:
        root.addHandler(handler)
    root.setLevel(logging.INFO)
    # v2.15.0 fix (review LOW): silencia request log do werkzeug — paths
    # contêm PII (nomes/CNPJs/msg_ids) que ficariam visíveis na janela do .bat
    logging.getLogger("werkzeug").setLevel(logging.WARNING)
    # v2.15.4: anexa o buffer de logs aos loggers do pipeline pra alimentar
    # a aba Atividade. Evita duplicar (handler é singleton).
    LOG_BUFFER.setLevel(logging.INFO)
    if LOG_BUFFER not in root.handlers:
        root.addHandler(LOG_BUFFER)


def main(host: str = "0.0.0.0", port: int = 8080, debug: bool = False):
    _setup_logging()
    ip = _ip_lan()
    log.info("=" * 70)
    log.info(f"AdmitER Web v{APP_VERSION} — interface da rede")
    log.info("=" * 70)
    log.info(f"  Local:   http://localhost:{port}")
    log.info(f"  LAN:     http://{ip}:{port}")
    log.info("Pressione Ctrl+C pra parar.")
    log.info("=" * 70)

    if debug:
        # Dev server com reload em modo --debug
        app.run(host=host, port=port, debug=True, use_reloader=False, threaded=True)
        return

    # v2.15.0 fix (review HIGH): produção usa waitress (production-grade WSGI).
    # Fallback pro dev server se waitress não instalado (com aviso).
    try:
        from waitress import serve
        log.info("Servidor: waitress (production)")
        serve(app, host=host, port=port, threads=8)
    except ImportError:
        log.warning(
            "waitress não instalado — usando Flask dev server (Werkzeug). "
            "Pra produção: pip install waitress."
        )
        app.run(host=host, port=port, debug=False, use_reloader=False, threaded=True)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="AdmitER Web (Flask)")
    parser.add_argument("--host", default="0.0.0.0",
                        help="Bind host (default 0.0.0.0 = LAN inteira)")
    parser.add_argument("--port", type=int, default=8080)
    parser.add_argument("--debug", action="store_true",
                        help="Modo debug (auto-reload). NÃO use em produção.")
    args = parser.parse_args()
    main(host=args.host, port=args.port, debug=args.debug)
