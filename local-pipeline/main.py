"""Pipeline de admissão local — entry point.

Roda na máquina do escritório. Faz polling do Gmail a cada N segundos
(default 300 = 5 min), processa cada email pendente:

  1. Baixa corpo + anexos do email
  2. Manda tudo pro Claude (Vision) que retorna payload + cnpj/departamento
  3. Resolve empresa_id via GET /empresas
  4. Resolve departamento via 3 regras de negócio
  5. Resolve função via planilha CBO (com re-prompt ao Claude se ambíguo)
  6. POST /candidatos no eContador
  7. Aplica label no Gmail + envia notificação ao DP

Setup:
  pip install -r requirements.txt
  cp .env.example .env  # preencha as variáveis
  python main.py        # roda em loop

Para uma única passada (debug):
  python main.py --once
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import shutil
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv

load_dotenv()

from claude_client import ClaudeClient
from empresas_cache import (
    EmpresasCache,
    carregar_empresas_do_econtador,
    corrigir_cnpj_via_cache,
)
from enrichment import enrich_candidato
from departamento import resolver_departamento
from ecotador_client import EContadorAPI
from funcao import carregar_planilha, resolver_funcao
from gmail_client import GmailClient
from payload_builder import (
    CAMPOS_MANUAIS_DP,
    aplicar_regra_data_admissao,
    extrair_dados_consulta,
    finalizar_payload,
    normalizar_admissoes,
    validar_campos_obrigatorios,
)
# v2.14.1 — wrapper único de POST (idempotência + label + log).
# Antes da v2.14.0 cada caminho (UI, orquestrador, scripts) tinha o seu;
# JENIFFY virou candidato 7x, YURI 3x. Agora todo POST passa aqui.
from post_admissao import postar_candidato_registrado


# ============================================================
# Setup
# ============================================================

ROOT = Path(__file__).parent
CONFIG_FILE = ROOT / "config.json"
LOOKUPS_FILE = ROOT / "lookups.json"
DEPARTAMENTOS_FILE = ROOT / "departamentos.json"
PLANILHA_CBO = ROOT / "funcoes_cbo.xlsx"
LOG_FILE = ROOT / "admissao_log.ndjson"
PAYLOADS_DIR = ROOT / "payloads"
PLANILHA_ADMISSOES = ROOT / "admissoes.xlsx"
BILLING_FILE = ROOT / "billing.ndjson"
REGRAS_FILE = ROOT / "regras.json"
CNPJ_OVERRIDES_FILE = ROOT / "cnpj_overrides.json"
FUNCAO_OVERRIDES_FILE = ROOT / "funcao_overrides.json"
FUNCAO_ALIASES_FILE = ROOT / "funcao_aliases.json"

# Cache global de empresas (CNPJs cadastrados no eContador).
# Lazy load — populado na 1ª chamada de obter_empresas_cache().
# Usado pra auto-corrigir CNPJ com typo (gera variações e checa whitelist).
_EMPRESAS_CACHE: EmpresasCache | None = None


def obter_empresas_cache(api: "EContadorAPI | None" = None) -> EmpresasCache:
    """Retorna o cache global. Lazy load na 1ª chamada. Thread-safe pelo GIL.

    Se `api` for None E o cache ainda não foi carregado, retorna cache vazio
    (que faz o pipeline cair no fluxo atual sem auto-correção).
    """
    global _EMPRESAS_CACHE
    if _EMPRESAS_CACHE is None:
        if api is None:
            return EmpresasCache()  # vazio mas válido — não quebra pipeline
        _EMPRESAS_CACHE = carregar_empresas_do_econtador(api)
    return _EMPRESAS_CACHE


def recarregar_empresas_cache(api: "EContadorAPI") -> EmpresasCache:
    """Força recarregamento — usado pelo botão 'Atualizar cache' da UI."""
    global _EMPRESAS_CACHE
    _EMPRESAS_CACHE = carregar_empresas_do_econtador(api)
    return _EMPRESAS_CACHE


def carregar_funcao_overrides() -> dict[str, dict[str, str]]:
    """Le {msg_id: {nome_upper: funcao_id}} salvo pela UI quando o usuario
    precisa setar manualmente a funcao de uma admissao pendente."""
    if not FUNCAO_OVERRIDES_FILE.exists():
        return {}
    try:
        return json.loads(FUNCAO_OVERRIDES_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def salvar_funcao_override(msg_id: str, nome_funcionario: str, funcao_id: str) -> None:
    """Adiciona override pra (msg_id, funcionario) → funcao_id. Sobrescreve."""
    data = carregar_funcao_overrides()
    nome_key = nome_funcionario.upper().strip()
    data.setdefault(msg_id, {})[nome_key] = str(funcao_id)
    FUNCAO_OVERRIDES_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def obter_funcao_override(msg_id: str, nome_funcionario: str) -> str | None:
    """Retorna funcao_id se houver override pra esta admissao."""
    data = carregar_funcao_overrides()
    nome_key = (nome_funcionario or "").upper().strip()
    return (data.get(msg_id) or {}).get(nome_key)


def carregar_cnpj_overrides() -> dict[str, str]:
    """Le {msg_id: cnpj_correto} salvo pela UI quando o usuario corrige
    um CNPJ que o cliente escreveu errado nos documentos."""
    if not CNPJ_OVERRIDES_FILE.exists():
        return {}
    try:
        return json.loads(CNPJ_OVERRIDES_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def salvar_cnpj_override(msg_id: str, cnpj_correto: str) -> None:
    """Adiciona override pra este msg_id. Sobrescreve se ja existia."""
    data = carregar_cnpj_overrides()
    data[msg_id] = cnpj_correto
    CNPJ_OVERRIDES_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def aplicar_cnpj_override(resposta_claude: dict, msg_id: str) -> str | None:
    """Se houver override pra este msg_id, substitui o cnpj_empresa no
    response do Claude (top-level + cada bloco de admissoes). Retorna o
    CNPJ aplicado ou None se nao havia override.

    Rede de segurança: se o Claude ignorou o aviso pre-prompt e marcou
    `_pendente=true` POR CAUSA do CNPJ (motivo contém "cnpj"/"contratante"/
    "localizado"), tenta limpar o _pendente. Outras razões de pendência
    (data, dados pessoais) NÃO são afetadas.
    """
    overrides = carregar_cnpj_overrides()
    novo_cnpj = overrides.get(msg_id)
    if not novo_cnpj:
        return None
    resposta_claude["cnpj_empresa"] = novo_cnpj
    for bloco in resposta_claude.get("admissoes") or []:
        if isinstance(bloco, dict):
            bloco["cnpj_empresa"] = novo_cnpj

    # Rede de segurança: limpa _pendente se a única razão era CNPJ
    if resposta_claude.get("_pendente"):
        motivo = (resposta_claude.get("_motivo") or "").lower()
        eh_so_cnpj = any(termo in motivo for termo in (
            "cnpj", "contratante", "empresa não", "empresa nao",
            "empresa contratante", "localizado nos documentos",
        )) and not any(termo in motivo for termo in (
            # Se motivo cita também outros campos faltantes, mantém pendente
            "cpf", "rg", "ctps", "data de admissão", "data admissao",
            "nascimento", "nome da mãe", "nome da mae", "salário", "salario",
        ))
        if eh_so_cnpj:
            log.info("   🔧 Limpando _pendente — motivo era CNPJ e há override")
            resposta_claude.pop("_pendente", None)
            resposta_claude.pop("_motivo", None)
    return novo_cnpj


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("admissao")


@dataclass
class Config:
    base_url: str
    token: str
    label_entrada: str
    label_processado: str
    label_pendente: str
    email_dp: str
    intervalo: int
    dry_run: bool
    claude_model: str
    claude_max_tokens: int
    # Self-consistency: nº de chamadas independentes ao Claude por extração.
    # 1 = sem verificação (mais barato), 2 = double-check (recomendado),
    # 3+ = ensemble. Múltiplas chamadas mitigam casos em que o Claude
    # eventualmente perde campos visíveis (ex: RG em foto rotacionada).
    claude_chamadas_verificacao: int
    # v2.16.19: 3 modos de auto-email de pendência (substitui o bool antigo)
    #   "desligado": nada acontece (default — DP responde manualmente)
    #   "rascunho":  gera rascunho na fila /respostas, analista revisa e aprova
    #   "direto":    envia automaticamente sem revisão (legado, arriscado)
    # Pendências INTERNAS nunca disparam email (sempre manual via UI/eContador).
    auto_email_pendencias_modo: str
    # v2.14.0+ — quando True, empresa sem deptos no eContador NÃO vira
    # pendência interna: payload é montado sem a relationship `departamento`
    # e DP atribui no Desktop. Resolve 6 das 11 pendências abertas em 12/06.
    # OFF por default — validar com 1 cobaia antes (PATCHES.md §6.5).
    postar_sem_departamento_quando_vazio: bool
    # v2.14.1 (PATCHES.md §5) — pausa entre blocos/emails na mesma passada
    # pra não estourar 30k input-tokens/min do tier 1 da Anthropic. Default
    # 20s; em produção com tier maior, pode baixar pra 5 ou 0.
    pausa_entre_emails_segundos: int
    # Quando True, data de admissão ausente/no passado é substituída pela data
    # de hoje automaticamente em vez de virar pendência. Útil quando o cliente
    # esquece de informar e a admissão pode começar no mesmo dia. Default False
    # — operador resolve manualmente pela UI ("Resolver pendência").
    usar_data_atual_se_invalida: bool
    # Quando True, IGNORA toda lógica automática de data de admissão (regra do
    # ASO+1, deslocamento por horário do email, colapso de retroativa) e POSTA
    # a admissão SEM data — DP completa manualmente no Alterdata Desktop depois.
    # Override global do operador — útil pra evitar que admissões fiquem em
    # pendência por causa de data, deixando que o DP preencha no Desktop com
    # base no contrato físico. Default False.
    sempre_mandar_sem_data_admissao: bool
    # Quando True, IGNORA o resolver automático de função (X-marcados + alias)
    # e POSTA a admissão com função placeholder — DP completa manualmente no
    # Alterdata Desktop depois. Override global — útil quando o cargo da planilha
    # CBO ainda não está mapeado e o escritório quer subir mesmo assim. Default False.
    sempre_mandar_sem_funcao: bool
    # Quando True, a passada/polling REPROCESSA também emails que já têm a label
    # `pendente` no Gmail. Útil quando o cliente respondeu no thread pendente e
    # queremos tentar de novo sem o operador precisar clicar "Reprocessar email"
    # um por um. A label pendente é REMOVIDA do email antes do reprocessamento
    # (senão fica em loop infinito). Default False — comportamento histórico.
    reprocessar_pendentes_no_polling: bool
    # ─── TEMP-CONFIRMAR-REPLIES (REMOVER em produção) ───────────
    # Default LIGADO durante calibração: pergunta confirmação antes
    # de enviar cada email de pendência. Auto-desligado quando o
    # processo roda sem TTY (Task Scheduler, cron, etc.). Pode forçar
    # off com --no-ask.
    confirmar_replies: bool = True


def _extrair_from_subject(msg_gmail: dict) -> tuple[str, str]:
    """Extrai (From, Subject) dos headers da mensagem Gmail."""
    payload = msg_gmail.get("payload") or {}
    headers = payload.get("headers") or []
    from_ = ""
    subject = ""
    for h in headers:
        n = (h.get("name") or "").lower()
        v = h.get("value") or ""
        if n == "from":
            from_ = v
        elif n == "subject":
            subject = v
    return from_, subject


def _separar_email_nome(from_header: str) -> tuple[str, str]:
    """'Nome <email@dom.com>' → ('email@dom.com', 'Nome'). Se vier só email, nome=''."""
    if not from_header:
        return "", ""
    import re as _re
    m = _re.match(r"^\s*(.*?)\s*<([^>]+)>\s*$", from_header)
    if m:
        return m.group(2).strip(), m.group(1).strip().strip('"')
    return from_header.strip(), ""


def _normalizar_modo_email(raw: dict) -> str:
    """v2.16.19: aceita o novo `auto_email_pendencias_modo` (string enum)
    OU o legacy `auto_email_pendencias` (bool). Bool=True vira "rascunho"
    por padrão (mais seguro — analista revisa). Bool=False ou ausente → "desligado".
    """
    modo = (raw.get("auto_email_pendencias_modo") or "").strip().lower()
    if modo in ("desligado", "rascunho", "direto"):
        return modo
    # Legacy migration
    if raw.get("auto_email_pendencias") is True:
        return "rascunho"  # mais seguro que "direto" — sempre passa por humano
    return "desligado"


def carregar_config() -> Config:
    raw = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    api_cfg = raw.get("ecotador") or raw.get("econtador") or {}
    token = api_cfg.get("token") or os.getenv("ECONTADOR_TOKEN")
    if not token or token == "SEU_TOKEN_AQUI":
        token = os.getenv("ECONTADOR_TOKEN")
    if not token:
        raise ValueError("ECONTADOR_TOKEN ausente (defina no .env)")

    gmail = raw.get("gmail", {})
    anthropic_cfg = raw.get("anthropic", {})
    return Config(
        base_url=api_cfg.get("base_url", "https://dp.pack.alterdata.com.br/api/v1"),
        token=token,
        label_entrada=gmail.get("label_entrada", "ADMISSÃO"),
        label_processado=gmail.get("label_processado", "ADMISSÃO/processado"),
        label_pendente=gmail.get("label_pendente", "ADMISSÃO/pendente"),
        email_dp=raw.get("dp", {}).get("email_notificacao", ""),
        intervalo=int(raw.get("polling_intervalo_segundos", 300)),
        dry_run=bool(raw.get("dry_run", False)),
        claude_model=anthropic_cfg.get("model", "claude-sonnet-4-20250514"),
        claude_max_tokens=int(anthropic_cfg.get("max_tokens", 8192)),
        claude_chamadas_verificacao=int(anthropic_cfg.get("chamadas_verificacao", 2)),
        auto_email_pendencias_modo=_normalizar_modo_email(raw),
        usar_data_atual_se_invalida=bool(raw.get("usar_data_atual_se_invalida", False)),
        sempre_mandar_sem_data_admissao=bool(raw.get("sempre_mandar_sem_data_admissao", False)),
        sempre_mandar_sem_funcao=bool(raw.get("sempre_mandar_sem_funcao", False)),
        reprocessar_pendentes_no_polling=bool(raw.get("reprocessar_pendentes_no_polling", False)),
        postar_sem_departamento_quando_vazio=bool(
            raw.get("postar_sem_departamento_quando_vazio", False)
        ),
        pausa_entre_emails_segundos=int(raw.get("pausa_entre_emails_segundos", 20)),
    )


def bootstrap_arquivos_locais() -> None:
    """Copia lookups.json e departamentos.json da raiz se não existirem aqui."""
    pai = ROOT.parent
    pares = [
        (pai / "lookups.json", LOOKUPS_FILE),
        (pai / "departamentos.json", DEPARTAMENTOS_FILE),
    ]
    for origem, destino in pares:
        if destino.exists():
            continue
        if origem.exists():
            shutil.copy2(origem, destino)
            log.info(f"📋 Bootstrapped {destino.name} a partir de {origem}")
        else:
            log.warning(f"⚠ {origem} não existe — pulando bootstrap de {destino.name}")

    if not FUNCAO_ALIASES_FILE.exists():
        FUNCAO_ALIASES_FILE.write_text("{}", encoding="utf-8")


def carregar_regras() -> dict:
    """Carrega regras.json (exceções customizáveis pelo escritório).

    Arquivo é opcional. Chaves que começam com '_' são tratadas como
    documentação/exemplos e filtradas recursivamente — o escritório pode
    deixar exemplos como referência sem afetar o comportamento.

    Retorna {} se o arquivo não existir ou for JSON inválido.
    """
    if not REGRAS_FILE.exists():
        log.info("ℹ regras.json não encontrado — usando defaults")
        return {}
    try:
        raw = json.loads(REGRAS_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        log.error(f"regras.json inválido: {e} — ignorando, usando defaults")
        return {}

    def _filtrar_docs(obj):
        if isinstance(obj, dict):
            return {k: _filtrar_docs(v) for k, v in obj.items() if not k.startswith("_")}
        if isinstance(obj, list):
            return [_filtrar_docs(x) for x in obj]
        return obj

    regras = _filtrar_docs(raw)
    n_secoes = sum(1 for v in regras.values() if v)
    log.info(f"📋 regras.json carregado: {n_secoes} seção(ões) ativa(s)")
    return regras


def log_jsonl(entry: dict) -> None:
    """Append-only NDJSON. Sempre inclui `campos_faltantes` (manuais DP +
    bloqueios de validação, se houver)."""
    entry["timestamp"] = datetime.now().isoformat()
    entry.setdefault("campos_faltantes", {
        "manuais_dp": CAMPOS_MANUAIS_DP,
        "validacao_bloqueada": entry.pop("_validacao_bloqueada", []),
    })
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except OSError as e:
        log.warning(f"Falha escrevendo log: {e}")


# v2.16.11: labels que são responsabilidade DO ESCRITÓRIO (cadastro local),
# NÃO do cliente. Quando a única coisa faltando é uma dessas, a pendência
# deve ser categorizada como INTERNA — cliente não tem ação possível.
#   Empresa     → cadastro do CNPJ no eContador
#   Departamento → ditto (ou flag postar_sem_departamento)
#   Função      → planilha funcoes_cbo.xlsx ou alias cadastrado
_LABELS_INTERNOS = {"Empresa", "Departamento", "Função"}


def _procedencia_de(resultado: dict) -> str:
    """Constrói a frase curta de procedência pra coluna da planilha.

    Ordem de precedência (importante — v2.14.1):
      1. ok=True → "Cadastrado" (com nota se pulou por idempotência)
      2. falha_tecnica=True → "Falha técnica — HTTP <código>" (NUNCA "Pendente cliente"!)
      3. interno=True → "Pendente interno" (problema do escritório)
      4. campos_faltando todos internos → "Pendente interno — faltam: ..." (v2.16.11)
      5. campos_faltando ou motivo_cliente → "Pendente cliente" (falta info DELE)
      6. resto → "Falha técnica — <erro>" (catch-all defensivo)
    """
    if resultado.get("ok"):
        if resultado.get("dry_run"):
            return "Dry-run — não postado"
        cid = resultado.get("candidato_id")
        extra = resultado.get("procedencia_extra") or ""
        base = f"Cadastrado — candidato {cid}" if cid else "Cadastrado"
        return f"{base} ({extra})" if extra else base

    # v2.14.1 ITEM 8: HTTP 4xx/5xx ANTES de qualquer pendência cliente.
    # Bug real: 3 HTTP 422 saíram como "Pendente cliente" na auditoria do
    # 12/06. Cliente não tem culpa de erro nosso de payload.
    if resultado.get("falha_tecnica"):
        status = resultado.get("status_http")
        if status:
            return f"Falha técnica — HTTP {status}"
        erro = (resultado.get("erro_tecnico") or "?")[:120]
        return f"Falha técnica — {erro}"

    if resultado.get("interno"):
        erro = (resultado.get("erro_tecnico") or "?")[:120]
        return f"Pendente interno — {erro}"

    faltando = resultado.get("campos_faltando") or []
    if faltando:
        lista = ", ".join(faltando)[:120]
        # v2.16.11: separa responsabilidade — campos cadastrados pelo
        # escritório (Função/Empresa/Departamento) viram pendência INTERNA;
        # campos do cliente (Salário/CPF/Nome/...) ficam como CLIENTE.
        # Misto → cliente (cliente vê a pendência mesmo).
        if all(c in _LABELS_INTERNOS for c in faltando):
            return f"Pendente interno — faltam: {lista}"
        return f"Pendente cliente — faltam: {lista}"

    motivo = resultado.get("motivo_cliente")
    if motivo:
        return f"Pendente cliente — {motivo[:120]}"

    erro = (resultado.get("erro_tecnico") or "erro desconhecido")[:120]
    return f"Falha técnica — {erro}"


def registrar_admissao_planilha(
    nome: str | None,
    empresa: str | None,
    cnpj: str | None,
    procedencia: str,
    msg_id: str = "",
    ts: str | None = None,
) -> None:
    """Append uma linha em admissoes.xlsx. Cria com cabeçalho se não existir.

    6 colunas: Data/Hora | Nome | Empresa | CNPJ | Procedência | msg_id.
    A coluna msg_id é o ID da mensagem do Gmail (usada pra reprocessar
    pela UI e abrir o thread no Gmail). Pode ser vazia em casos de
    fallback ou linhas adicionadas manualmente.

    Append-only — cada execução pode adicionar várias linhas (1 por
    admissão processada, sucesso ou pendência).

    Migração: se a planilha existente tem só 4 colunas (versão antiga),
    novas linhas são escritas com 6 colunas e o openpyxl preserva as
    antigas como estão.
    """
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    headers = ["Data/Hora", "Nome do colaborador", "Empresa", "CNPJ", "Procedência", "msg_id"]
    timestamp = ts or datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        if PLANILHA_ADMISSOES.exists():
            wb = load_workbook(PLANILHA_ADMISSOES)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "admissoes"
            ws.append(headers)
            hfont = Font(bold=True, color="FFFFFF")
            hfill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=1, column=col)
                c.font = hfont
                c.fill = hfill
                c.alignment = Alignment(horizontal="center")
            # Larguras: data/hora, nome, empresa, cnpj, procedência, msg_id (estreito)
            larguras = {"A": 18, "B": 32, "C": 32, "D": 18, "E": 50, "F": 18}
            for col, w in larguras.items():
                ws.column_dimensions[col].width = w
            ws.freeze_panes = "A2"

        ws.append([
            timestamp,
            (nome or "?").strip(),
            (empresa or "?").strip(),
            (cnpj or "?").strip(),
            procedencia.strip(),
            (msg_id or "").strip(),
        ])
        wb.save(PLANILHA_ADMISSOES)
    except Exception as e:
        log.warning(f"Falha gravando admissoes.xlsx: {e}")


def sum_billing_mes_atual() -> dict:
    """Soma o billing do mês corrente lendo billing.ndjson.

    Retorna {n_passadas, n_calls, input_tokens, output_tokens, custo_usd}.
    Tudo zero se o arquivo não existe.
    """
    out = {"n_passadas": 0, "n_calls": 0, "input_tokens": 0, "output_tokens": 0, "custo_usd": 0.0}
    if not BILLING_FILE.exists():
        return out
    mes_atual = datetime.now().strftime("%Y-%m")
    try:
        with open(BILLING_FILE, "r", encoding="utf-8") as f:
            for linha in f:
                try:
                    e = json.loads(linha)
                except json.JSONDecodeError:
                    continue
                ts = e.get("timestamp", "")
                if not ts.startswith(mes_atual):
                    continue
                out["n_passadas"] += 1
                out["n_calls"] += int(e.get("n_calls") or 0)
                out["input_tokens"] += int(e.get("input_tokens") or 0)
                out["output_tokens"] += int(e.get("output_tokens") or 0)
                out["custo_usd"] += float(e.get("custo_usd") or 0)
    except OSError as e:
        log.warning(f"Erro lendo billing.ndjson: {e}")
    out["custo_usd"] = round(out["custo_usd"], 4)
    return out


def fazer_backup_planilha_e_payloads() -> Path | None:
    """Copia admissoes.xlsx + payloads/ pra pasta backups/<data>/.

    Retorna o path do backup criado, ou None se houver erro.
    """
    try:
        backup_root = ROOT / "backups"
        backup_root.mkdir(exist_ok=True)
        carimbo = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = backup_root / carimbo
        dest.mkdir()

        if PLANILHA_ADMISSOES.exists():
            shutil.copy2(PLANILHA_ADMISSOES, dest / PLANILHA_ADMISSOES.name)
        if BILLING_FILE.exists():
            shutil.copy2(BILLING_FILE, dest / BILLING_FILE.name)
        if PAYLOADS_DIR.exists():
            shutil.copytree(PAYLOADS_DIR, dest / "payloads", dirs_exist_ok=True)

        log.info(f"💾 Backup criado em {dest}")
        return dest
    except Exception as e:
        log.exception(f"Falha criando backup: {e}")
        return None


def salvar_payload(
    msg_id: str,
    metadados: dict,
    payload: dict,
    resolucao: dict | None = None,
    resultado: dict | None = None,
) -> Path:
    """Salva o payload completo + contexto em
    payloads/<timestamp>_<msg_id>[_i<indice>].json.

    Chamado em 2 momentos:
      1. Após montagem (antes do POST) — preserva o payload mesmo se crashar
      2. Após resposta do POST — atualiza com candidato_id ou erro

    Sobrescreve o arquivo do mesmo (msg_id, indice) — assim emails com N
    admissões geram N arquivos independentes (não 1 sobrescrito por todos).
    """
    PAYLOADS_DIR.mkdir(exist_ok=True)
    # Sufixo de admissão dentro do email — evita sobrescrita entre blocos
    indice = (resolucao or {}).get("indice")
    sufixo = f"_i{indice}" if indice else ""

    # Timestamp + msg_id curto pra evitar nomes gigantes
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arq = f"{ts}_{msg_id[:16]}{sufixo}.json"
    arq = PAYLOADS_DIR / nome_arq

    # Se já existe um arquivo dessa admissão específica (msg_id + indice),
    # sobrescreve (vem da 1ª chamada antes do POST — atualizamos com resultado)
    padrao = f"*_{msg_id[:16]}{sufixo}.json"
    existentes = sorted(PAYLOADS_DIR.glob(padrao))
    if existentes:
        arq = existentes[-1]

    doc = {
        "timestamp": datetime.now().isoformat(),
        "msg_id": msg_id,
        "remetente": metadados.get("remetente", ""),
        "assunto": metadados.get("assunto", ""),
        "data_email": metadados.get("data", ""),
        "resolucao": resolucao or {},
        "resultado": resultado or {"status": "preparado", "erro": None, "candidato_id": None},
        "payload": payload,
    }
    arq.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
    return arq


# ─── TEMP (REMOVER em produção): confirmação interativa de envio ──────
# Bloco coeso: 3 funções + 1 uso em _processar_seguro. Pesquise por
# "TEMP-CONFIRMAR-REPLIES" pra achar tudo de uma vez.

def _previa_reply(msg_orig: dict, corpo: str, cc: str | None) -> str:
    """TEMP-CONFIRMAR-REPLIES: monta preview legível pra exibir no terminal."""
    headers = {
        h["name"].lower(): h["value"]
        for h in msg_orig.get("payload", {}).get("headers", [])
    }
    de = headers.get("from", "?")
    assunto = headers.get("subject", "?")
    sep = "─" * 70
    corpo_trunc = corpo if len(corpo) <= 800 else corpo[:800] + "\n[...corpo truncado...]"
    return (
        f"\n{sep}\n"
        f"📨 PREVIEW DA RESPOSTA PRO CLIENTE\n"
        f"  Para:    {de}\n"
        f"  CC:      {cc or '(nenhum)'}\n"
        f"  Assunto: Re: {assunto[:80]}\n"
        f"\n{corpo_trunc}\n"
        f"{sep}"
    )


def _confirmar_envio(msg_orig: dict, corpo: str, cc: str | None) -> bool:
    """TEMP-CONFIRMAR-REPLIES: imprime preview e pergunta s/N. Default N.

    Se stdin não estiver disponível (fora do esperado, pq o caller já
    desliga via `confirmar_replies = False` sem TTY), retorna False
    (não envia) por precaução.
    """
    if not sys.stdin.isatty():
        log.warning("_confirmar_envio chamado sem TTY — abortando envio")
        return False
    print(_previa_reply(msg_orig, corpo, cc))
    try:
        resp = input("Enviar este email? [s/N]: ").strip().lower()
    except (EOFError, KeyboardInterrupt):
        print()  # quebra de linha após ^C
        return False
    return resp in ("s", "sim", "y", "yes")


# ─── fim do bloco TEMP-CONFIRMAR-REPLIES ──────────────────────────────


def _payload_parcial_de_dados(dp: dict) -> dict:
    """Converte `_dados_parciais` do Claude (chaves PT-BR variadas) num
    payload JSON:API parcial — pronto pra ser salvo em payloads/ e completado
    via UI. v2.15.3.

    Aceita as chaves comuns que o Claude usa em `_dados_parciais` quando
    desiste do email todo. Ignora valores vazios pra não poluir o payload.
    """
    if not isinstance(dp, dict):
        return {"data": {"type": "candidatos", "attributes": {}, "relationships": {}}}

    # Mapeamento das chaves PT-BR comuns do Claude → nomes da API JSON:API.
    # Pra chaves que já batem com o nome da API, manda direto.
    MAPA = {
        # Pessoais
        "nome": "nome",
        "nome_completo": "nome",
        "funcionario": "nome",
        "cpf": "cpf",
        "nascimento": "nascimento",
        "data_nascimento": "nascimento",
        "nomedamae": "nomedamae",
        "nome_mae": "nomedamae",
        "mae": "nomedamae",
        "nomedopai": "nomedopai",
        "nome_pai": "nomedopai",
        "pai": "nomedopai",
        "municipionascimento": "municipionascimento",
        "municipio_nascimento": "municipionascimento",
        "naturalidade_cidade": "municipionascimento",
        # Contratuais
        "admissao": "admissao",
        "data_admissao": "admissao",
        "salario": "salario",
        "salario_base": "salario",
        "nomecargo": "nomecargo",
        "cargo": "nomecargo",
        "primeiroemprego": "primeiroemprego",
        "primeiro_emprego": "primeiroemprego",
        # Documentos
        "identidade": "identidade",
        "rg": "identidade",
        "dataidentidade": "dataidentidade",
        "rg_data_emissao": "dataidentidade",
        "data_rg": "dataidentidade",
        "orgaoemissoridentidade": "orgaoemissoridentidade",
        "rg_orgao_emissor": "orgaoemissoridentidade",
        "ctps": "ctps",
        "seriectps": "seriectps",
        "ctps_serie": "seriectps",
        "datactps": "datactps",
        "ctps_data_emissao": "datactps",
        "pis": "pis",
        "datapis": "datapis",
        # Endereço
        "cep": "cep",
        "rua": "rua",
        "logradouro": "rua",
        "numero": "numero",
        "complemento": "complemento",
        "bairro": "bairro",
        "cidade": "cidade",
        # Contato
        "email": "email",
        "celular": "celular",
        "telefone": "telefone",
        # Título eleitor
        "tituloeleitor": "tituloeleitor",
        "titulo_eleitor": "tituloeleitor",
        "zonatituloeleitor": "zonatituloeleitor",
        "secaotituloeleitor": "secaotituloeleitor",
    }

    attrs: dict = {}
    for k, v in dp.items():
        if v in (None, "", [], {}):
            continue
        attr = MAPA.get(str(k).lower().strip())
        if not attr:
            continue
        # Não sobrescreve se já temos algo melhor
        if attr in attrs and attrs[attr]:
            continue
        attrs[attr] = v

    # v2.16.23: o Claude às vezes manda endereço como UMA string única
    # ('endereco' / 'endereco_completo' / 'enderecoResidencial') em vez de
    # cep/rua/bairro/cidade separados — geralmente quando marca o email
    # como pendente_claude_raiz e tá com pressa nos _dados_parciais.
    # Parseia a string e preenche os campos que ainda não vieram separados.
    # Caso real (JOSE ALBERTO TORRES LEAL, 2026-06-19):
    #   "RUA CAOLENITA NT, Q 007 L 1B, VILA OLIVEIRA, APARECIDA DE GOIANIA - GO, CEP 74956-140"
    end_raw = next(
        (dp.get(k) for k in ("endereco", "endereco_completo",
                              "enderecoresidencial", "logradouro_completo")
         if dp.get(k) and isinstance(dp.get(k), str)),
        None,
    )
    if end_raw and not all(attrs.get(k) for k in ("cep", "rua", "bairro", "cidade")):
        parsed = _parsear_endereco_string(end_raw)
        for k, v in parsed.items():
            if v and not attrs.get(k):
                attrs[k] = v

    return {
        "data": {
            "type": "candidatos",
            "attributes": attrs,
            "relationships": {},
        }
    }


def _parsear_endereco_string(endereco: str) -> dict:
    """v2.16.23: extrai cep/rua/numero/bairro/cidade de uma string de endereço
    formatada por extenso. Conservador — só preenche o que conseguir identificar
    com alta confiança. Casos cobertos:

      'RUA CAOLENITA NT, Q 007 L 1B, VILA OLIVEIRA, APARECIDA DE GOIANIA - GO, CEP 74956-140'
       → rua=RUA CAOLENITA NT, complemento=Q 007 L 1B, bairro=VILA OLIVEIRA,
         cidade=APARECIDA DE GOIANIA, cep=74956-140

      'Rua das Camélias, 123, Setor Bueno, Goiânia-GO, 74390-100'
       → rua, numero, bairro, cidade, cep
    """
    import re as _re
    if not endereco or not isinstance(endereco, str):
        return {}
    out: dict = {}
    # 1) CEP — sempre 8 dígitos (com ou sem hífen), opcionalmente "CEP" antes
    m_cep = _re.search(r"\bCEP\s*[:\-]?\s*(\d{2}\.?\d{3}-?\d{3})\b|(\d{5}-?\d{3})\b",
                       endereco, _re.IGNORECASE)
    if m_cep:
        digitos = _re.sub(r"\D", "", m_cep.group(0))
        if len(digitos) == 8:
            out["cep"] = f"{digitos[:5]}-{digitos[5:]}"
            endereco = endereco.replace(m_cep.group(0), "").strip(" ,;")
    # 2) Cidade-UF no fim — padrão "CIDADE - UF" ou "CIDADE/UF"
    m_uf = _re.search(r"([A-ZÀ-Ÿa-zà-ÿ][A-ZÀ-Ÿa-zà-ÿ\s]+?)\s*[\-/]\s*([A-Z]{2})\b",
                      endereco)
    if m_uf:
        out["cidade"] = m_uf.group(1).strip().upper()
        endereco = endereco.replace(m_uf.group(0), "").strip(" ,;-")
    # 3) Quebra o resto por vírgula
    partes = [p.strip() for p in endereco.split(",") if p.strip()]
    if not partes:
        return out
    # Primeira parte = rua (e talvez número junto)
    rua_e_num = partes[0]
    # Tenta separar número no fim (ex: "Rua das Camélias 123")
    m_num = _re.search(r"\b(\d{1,6})\s*$", rua_e_num)
    if m_num:
        out["numero"] = int(m_num.group(1))
        out["rua"] = rua_e_num[:m_num.start()].strip().upper()
    else:
        out["rua"] = rua_e_num.strip().upper()
    # Última parte (se não foi cidade) = bairro
    # Pesquisa de trás pra frente; ignora se parece complemento (Q X L Y)
    for p in reversed(partes[1:]):
        p_up = p.upper().strip()
        if _re.match(r"^Q\.?\s*\d+\s*L\.?\s*", p_up) or _re.match(
            r"^(N|N[ºO]\.?|NUM\.?)\s*\d+", p_up
        ):
            # Complemento (quadra/lote) ou número
            if "complemento" not in out:
                out["complemento"] = p_up
            continue
        if "bairro" not in out:
            out["bairro"] = p_up
            break
    return out


def _raise_pendencia(
    erro_tecnico: str,
    motivo_cliente: str,
    *,
    campos_faltando: list[str] | None = None,
    payload_parcial: dict | None = None,
) -> None:
    """Levanta ValueError anotada com o que o reply pro cliente precisa.

    - erro_tecnico: vai pro log/email DP (texto curto, técnico)
    - motivo_cliente: vai pro corpo da resposta amigável ao cliente
    - campos_faltando: opcional, lista estruturada (caminho do validador)
    - payload_parcial: opcional, dados que já foram extraídos
    """
    err = ValueError(erro_tecnico)
    err.motivo_cliente = motivo_cliente  # type: ignore[attr-defined]
    err.campos_faltando = campos_faltando or []  # type: ignore[attr-defined]
    err.payload_parcial = payload_parcial or {}  # type: ignore[attr-defined]
    raise err


def _lista_natural(itens: list[str]) -> str:
    """Junta itens em frase natural PT-BR: 'a', 'a e b', 'a, b e c'."""
    if not itens:
        return ""
    if len(itens) == 1:
        return itens[0]
    if len(itens) == 2:
        return f"{itens[0]} e {itens[1]}"
    return ", ".join(itens[:-1]) + " e " + itens[-1]


def email_resposta_cliente(
    payload: dict | None,
    faltantes: list[str] | None,
    motivo_livre: str | None = None,
) -> str:
    """Corpo conversacional pro cliente. Foco no que FALTA; não lista o que
    já recebemos (a info ficou em payloads/ pra auditoria — o cliente vê o
    próprio email original no thread se quiser conferir).

    3 caminhos:
      - faltantes (lista do validador) → "ainda preciso de A, B e C"
      - motivo_livre (Claude _pendente etc.) → repassa a frase do motivo
      - sem nada estruturado → pedido genérico pra reenviar docs
    """
    payload = payload or {}
    attrs = (payload.get("data") or {}).get("attributes") or {}
    dados_parciais = payload.get("_dados_parciais") or {}

    # Procura o nome em cascata pra personalizar (attrs → parciais → raiz)
    nome = None
    for fonte in (attrs, dados_parciais, payload):
        if not isinstance(fonte, dict):
            continue
        for k in ("nome", "nome_completo", "funcionario"):
            v = fonte.get(k)
            if isinstance(v, str) and v.strip():
                nome = v.strip()
                break
        if nome:
            break

    abertura_cadastro = (
        f"Pra fechar o cadastro de {nome} aqui no sistema"
        if nome
        else "Pra concluir esse cadastro aqui no sistema"
    )

    if faltantes:
        lista = _lista_natural(faltantes)
        miolo = (
            f"{abertura_cadastro}, ainda preciso de algumas informações "
            f"que não consegui identificar nos documentos: {lista}.\n\n"
            f"Pode me responder esse mesmo e-mail com esses dados? Não "
            f"precisa reenviar o que já mandou — só o que falta.\n\n"
        )
    elif motivo_livre:
        miolo = (
            f"{motivo_livre}\n\n"
            f"Pode me responder esse mesmo e-mail com o que falta? "
            f"Não precisa reenviar os documentos que já mandou.\n\n"
        )
    else:
        miolo = (
            f"Recebi sua mensagem mas não consegui ler os documentos da "
            f"admissão. Pode reenviar a ficha com RG, CPF, CTPS, comprovante "
            f"de endereço e ASO do(a) candidato(a)?\n\n"
        )

    return (
        f"Olá!\n\n"
        f"{miolo}"
        f"Qualquer dúvida, é só me chamar.\n\n"
        f"Atenciosamente,\n"
        f"DP — Crosara Contabilidade"
    )


# ============================================================
# Stub do Gmail pra importação manual de arquivos (v2.7.0)
# ============================================================

class GmailClientStub:
    """Substituto do GmailClient com métodos no-op pra processar admissões
    importadas manualmente pela UI (sem email real). Permite reusar
    `processar_admissao` sem ramificar o caminho do código.

    Importação manual NÃO tem:
      - msg_id real (usamos sintético: `manual_<timestamp>`)
      - thread pra responder (reply no thread vira no-op)
      - label pra aplicar/remover (vira no-op)
    """
    def __init__(self):
        self.label_ids = {}

    def aplicar_label(self, *a, **kw): pass
    def remover_label(self, *a, **kw): pass
    def responder_no_thread(self, *a, **kw): pass
    def _label_id(self, _label): return None
    def obter_thread(self, _tid): return {"messages": []}
    def extrair_corpo_thread(self, _t): return ""
    def baixar_anexos_thread(self, _t): return []
    def extrair_corpo(self, _m): return ""
    def baixar_anexos(self, _m): return []
    def extrair_metadados(self, _m): return {}
    def close(self): pass


def processar_arquivos_avulsos(
    arquivos: list[dict],
    corpo_texto: str,
    claude: ClaudeClient,
    api: EContadorAPI,
    planilha_cbo: list[dict],
    config: Config,
) -> None:
    """Processa admissão a partir de arquivos importados manualmente pela UI.

    Sintetiza um "email" com:
      - msg_id = "manual_<timestamp>"
      - corpo = texto que o operador digitou (opcional)
      - anexos = lista de dicts {filename, mime, data: bytes}
      - metadados = data atual + remetente "IMPORTACAO_MANUAL"

    Reusa `processar_admissao` via GmailClientStub. Resultados (sucesso/
    pendência) ficam em admissoes.xlsx e payloads/ como qualquer email.

    Args:
        arquivos: lista de dicts no mesmo formato de `gmail.baixar_anexos`:
            [{filename: str, mime: str, data: bytes}, ...]
        corpo_texto: texto opcional pra dar contexto ao Claude (ex: "Admissão
            da Maria, CNPJ 12345678000190, salário 2000")
    """
    import time
    msg_id = f"manual_{int(time.time())}"
    metadados = {
        "remetente": "IMPORTACAO_MANUAL",
        "assunto": "Importação manual de arquivos",
        "data": datetime.now().strftime("%a, %d %b %Y %H:%M:%S -0300"),
    }
    log.info(
        f"📥 Importação manual: {len(arquivos)} arquivo(s), "
        f"corpo: {len(corpo_texto)} chars, msg_id={msg_id}"
    )
    gmail = GmailClientStub()
    try:
        processar_admissao(
            msg_id=msg_id,
            msg_pra_resposta={},  # sem thread pra responder
            corpo=corpo_texto,
            anexos=arquivos,
            metadados=metadados,
            gmail=gmail, claude=claude, api=api,
            planilha_cbo=planilha_cbo, config=config,
            ids_label_pendente_remover=[],
        )
    except Exception as e:
        # Captura igual o _processar_seguro faz — mas sem aplicar label
        motivo_cliente = getattr(e, "motivo_cliente", None)
        log.exception(f"❌ Importação manual {msg_id}: {e}")
        # Registra fallback na planilha pra operador ver
        try:
            registrar_admissao_planilha(
                nome="? (importação manual)",
                empresa=None,
                cnpj=None,
                procedencia=(
                    f"Pendente cliente — {motivo_cliente}"
                    if motivo_cliente else
                    f"Falha técnica importação manual — {str(e)[:120]}"
                ),
                msg_id=msg_id,
            )
        except Exception:
            log.exception("Falha registrando linha de erro na planilha")


# ============================================================
# Processamento de 1 email
# ============================================================

def processar_email(
    msg: dict,
    gmail: GmailClient,
    claude: ClaudeClient,
    api: EContadorAPI,
    planilha_cbo: list[dict],
    config: Config,
) -> None:
    """Processa um email novo USANDO O THREAD INTEIRO.

    Mesmo emails "novos" (sem label processado/pendente) podem fazer parte
    de uma conversa onde o cliente já mandou docs em mensagens anteriores
    no mesmo thread (ex: 1º email só com a ficha, 2º email com o RG, 3º
    com o ASO). Sempre busca o thread completo pra dar ao Claude o
    contexto cumulativo — evita alucinação de campos não fornecidos no
    email isolado mas presentes em outras mensagens.

    Fallback: se houver falha pegando o thread, processa a msg solo.
    """
    thread_id = msg.get("threadId")
    if thread_id:
        try:
            thread = gmail.obter_thread(thread_id)
            msgs = thread.get("messages", []) or []
            if len(msgs) > 1:
                log.info(
                    f"   📚 Thread tem {len(msgs)} mensagens — agregando contexto completo"
                )
                corpo = gmail.extrair_corpo_thread(thread)
                anexos = gmail.baixar_anexos_thread(thread)
                metadados = gmail.extrair_metadados(msgs[0])
                processar_admissao(
                    msg_id=msg["id"],
                    msg_pra_resposta=msgs[-1],  # responde na msg mais recente
                    corpo=corpo,
                    anexos=anexos,
                    metadados=metadados,
                    gmail=gmail, claude=claude, api=api,
                    planilha_cbo=planilha_cbo, config=config,
                    ids_label_pendente_remover=[],
                )
                return
        except Exception:
            log.exception(
                f"   Falha pegando thread {thread_id} — caindo pra msg solo"
            )

    # Caminho msg isolada (thread só tem essa msg, ou falha de fetch)
    corpo = gmail.extrair_corpo(msg)
    anexos = gmail.baixar_anexos(msg)
    metadados = gmail.extrair_metadados(msg)
    processar_admissao(
        msg_id=msg["id"],
        msg_pra_resposta=msg,
        corpo=corpo,
        anexos=anexos,
        metadados=metadados,
        gmail=gmail, claude=claude, api=api,
        planilha_cbo=planilha_cbo, config=config,
        ids_label_pendente_remover=[],
    )


def processar_thread_resposta(
    thread: dict,
    gmail: GmailClient,
    claude: ClaudeClient,
    api: EContadorAPI,
    planilha_cbo: list[dict],
    config: Config,
) -> None:
    """Cliente respondeu num thread pendente — reprocessa com TUDO do thread."""
    msgs = thread.get("messages", [])
    if not msgs:
        raise ValueError("Thread vazio")

    # Identidade do log: usa msg original (mantém audit trail consistente)
    msg_original = msgs[0]
    msg_id = msg_original["id"]

    # Coleta agregada (corpo de todas + anexos de todas)
    corpo = gmail.extrair_corpo_thread(thread)
    anexos = gmail.baixar_anexos_thread(thread)
    metadados = gmail.extrair_metadados(msg_original)

    log.info(
        f"🔁 Reprocessando thread {thread.get('id', '?')[:16]} "
        f"({len(msgs)} mensagens) → msg_id_ref={msg_id}"
    )

    # Lista de msg ids com label pendente — vamos remover após reprocessar
    label_pendente_ids: list[str] = []
    pendente_id = gmail._label_id(config.label_pendente)
    if pendente_id:
        for m in msgs:
            if pendente_id in (m.get("labelIds") or []):
                label_pendente_ids.append(m["id"])

    processar_admissao(
        msg_id=msg_id,
        msg_pra_resposta=msgs[-1],  # responde na última msg (do cliente)
        corpo=corpo,
        anexos=anexos,
        metadados=metadados,
        gmail=gmail, claude=claude, api=api,
        planilha_cbo=planilha_cbo, config=config,
        ids_label_pendente_remover=label_pendente_ids,
    )


def processar_admissao(
    msg_id: str,
    msg_pra_resposta: dict,
    corpo: str,
    anexos: list[dict],
    metadados: dict,
    gmail: GmailClient,
    claude: ClaudeClient,
    api: EContadorAPI,
    planilha_cbo: list[dict],
    config: Config,
    ids_label_pendente_remover: list[str],
) -> None:
    """Orquestrador: Claude → quebra em N blocos de admissão → processa cada
    independentemente → agrega resultados → label + email no fim.

    Casos:
      - 1 admissão: igual ao comportamento antigo
      - N admissões no mesmo email: cada uma vira um candidato no eContador.
        Empresa/departamentos vêm de cache se compartilhados (mesmo CNPJ).
        Função pode variar por funcionário (cargo diferente cada).
        Se todas passarem → label processado, email DP com a lista.
        Se alguma falhar → label pendente, reply no thread com status por
        nome (quem foi cadastrado e quem ainda precisa de info).
    """
    # v2.16.20: delimitador visual claro de início de email no log
    log.info("┏" + "━" * 78)
    log.info(f"┃ 📧 INICIANDO email {msg_id[:16]} | {metadados.get('assunto', '')[:48]}")
    log.info(f"┃ De: {(metadados.get('remetente') or '?')[:60]}")
    log.info(f"┃ Corpo: {len(corpo)} chars | Anexos: {len(anexos)}")
    log.info("┗" + "━" * 78)

    if not corpo and not anexos:
        _raise_pendencia(
            "Email sem corpo nem anexos PDF/imagem",
            motivo_cliente=(
                "Não recebi nenhuma informação ou anexo nessa mensagem. "
                "Pode reenviar a ficha de admissão com os documentos "
                "(RG, CPF, CTPS, comprovante de endereço, ASO)?"
            ),
        )

    # 0.5. Override de CNPJ — INJETADO ANTES da chamada Claude.
    # Se o operador corrigiu o CNPJ via UI ("Corrigir CNPJ do email"),
    # avisamos o Claude no próprio prompt — assim ele NÃO marca _pendente=true
    # com motivo "CNPJ não localizado nos documentos". Sem isso, o pendente
    # do Claude desviaria o fluxo antes do override surtir efeito.
    cnpj_override_pre = carregar_cnpj_overrides().get(msg_id)
    corpo_efetivo = corpo
    if cnpj_override_pre:
        prefixo_override = (
            f"🔧 INSTRUÇÃO DO OPERADOR — CNPJ PRÉ-RESOLVIDO\n"
            f"O operador já confirmou via UI que o CNPJ da empresa contratante "
            f"deste email é EXATAMENTE: {cnpj_override_pre}\n"
            f"USE este valor em `cnpj_empresa` sem procurar nos documentos. "
            f"NÃO marque _pendente=true por causa de CNPJ ausente — está "
            f"resolvido externamente.\n\n"
            f"---\n\n"
        )
        corpo_efetivo = prefixo_override + corpo
        log.info(f"   🔧 CNPJ override injetado no prompt: {cnpj_override_pre}")

    # 1. Claude — 1 chamada que pode devolver N admissões
    resposta_claude = claude.gerar_payload(corpo_efetivo, metadados, anexos)

    # 1.5. Override de CNPJ — REDE DE SEGURANÇA (caso Claude ignore o aviso).
    # Substitui o cnpj_empresa direto na resposta.
    cnpj_override = aplicar_cnpj_override(resposta_claude, msg_id)
    if cnpj_override:
        log.info(f"   🔧 CNPJ override também aplicado pós-Claude: {cnpj_override}")

    # v2.16.14: respeitar flags config no caminho RAIZ-pendente.
    # Bug real: JOHN LENNON (18/06/2026) — Claude marcou _pendente=true no
    # topo com _motivo_codigo="DATA_ADMISSAO_AUSENTE", mas o caminho do
    # bloco individual (que tem AUTO_RESOLVIVEIS_CODIGOS) era ignorado
    # porque o Claude não chegou a dividir em blocos. Aqui aplicamos a
    # mesma lógica ANTES de gerar a pendência: se o único motivo bloqueante
    # é auto-resolvível pela flag, promove _dados_parciais e segue normal.
    if resposta_claude.get("_pendente"):
        cod_raiz = str(resposta_claude.get("_motivo_codigo") or "").upper().strip()
        cods_raiz_extras = resposta_claude.get("_motivos_codigos") or []
        if isinstance(cods_raiz_extras, list):
            cods_raiz = {cod_raiz, *(str(c).upper().strip() for c in cods_raiz_extras)} - {""}
        else:
            cods_raiz = {cod_raiz} - {""}
        AUTO_RAIZ = {"CTPS_AUSENTE", "ESTAGIO_INDEFINIDO"}
        if config.usar_data_atual_se_invalida or config.sempre_mandar_sem_data_admissao:
            AUTO_RAIZ = AUTO_RAIZ | {"DATA_ADMISSAO_AUSENTE"}
        if config.sempre_mandar_sem_funcao:
            AUTO_RAIZ = AUTO_RAIZ | {"CARGO_AUSENTE"}
        if cods_raiz and cods_raiz <= AUTO_RAIZ:
            log.info(
                f"   ℹ Raiz-pendente com motivos auto-resolvíveis "
                f"({', '.join(sorted(cods_raiz))}) — promovendo dados parciais "
                f"em vez de marcar pendente cliente"
            )
            dp_raiz = resposta_claude.get("_dados_parciais") or {}
            # Promove os dados parciais pra admissoes top-level (formato que
            # normalizar_admissoes consome). Limpa _pendente e segue fluxo
            # normal — _processar_um_bloco vai validar campos obrigatórios
            # e gerar pendência só se faltar algo de fato bloqueante.
            if dp_raiz and not resposta_claude.get("admissoes"):
                resposta_claude["admissoes"] = [{"data": {"attributes": _payload_parcial_de_dados(dp_raiz)["data"]["attributes"]}}]
            resposta_claude.pop("_pendente", None)
            resposta_claude.pop("_motivo", None)
            resposta_claude.pop("_motivo_codigo", None)
        # Após a promoção, pode ter saído do estado pendente — releia a flag
    if resposta_claude.get("_pendente"):
        motivo = resposta_claude.get("_motivo") or "Dados insuficientes"
        dp = resposta_claude.get("_dados_parciais") or {}

        # v2.16.50: extracao case-insensitive + fallback nos varios formatos que
        # Claude usa. Caso real EMERSON (2026-07-06): email tinha 'Nome: EMERSON
        # COSTA DOS SANTOS' explicito mas UI mostrou '(nome não extraído)'.
        # Provavelmente Claude retornou chave 'Nome' (case) ou 'candidato' ou
        # dentro de estrutura aninhada.
        def _get_ci(d: dict, *chaves):
            """Case-insensitive get; retorna primeiro valor nao-vazio."""
            if not isinstance(d, dict):
                return None
            keys_lower = {str(k).lower(): k for k in d.keys()}
            for c in chaves:
                real = keys_lower.get(c.lower())
                if real and d.get(real):
                    return str(d[real]).strip()
            return None

        nome_extraido = _get_ci(
            dp,
            "nome", "nome_completo", "nomecompleto", "funcionario",
            "candidato", "nome_candidato", "nome_funcionario",
            "colaborador", "nome_colaborador",
        )
        # v2.16.50: fallback — se Claude estruturou como admissoes[] em vez
        # de _dados_parciais mas ainda marcou _pendente, extrai do bloco.
        if not nome_extraido:
            blocos = resposta_claude.get("admissoes") or []
            if blocos and isinstance(blocos, list):
                _b0 = blocos[0] or {}
                _at = ((_b0.get("data") or {}).get("attributes") or {})
                if _at.get("nome"):
                    nome_extraido = str(_at["nome"]).strip()
        razao_extraida = _get_ci(
            dp,
            "razao_social_empresa", "razao_social", "razaosocial",
            "empresa", "nome_empresa", "nomeempresa",
        )
        cnpj_extraido = (
            _get_ci(dp, "cnpj_empresa", "cnpj", "cnpjempresa")
            or resposta_claude.get("cnpj_empresa")
            or ""
        )

        log.warning(
            f"   ⚠ Claude marcou como pendente: {motivo}\n"
            f"   📝 Registrando com dados parciais: nome={nome_extraido}, "
            f"empresa={razao_extraida}, cnpj={cnpj_extraido}"
        )

        # v2.15.3 (resposta direta a "o certo nao seria sempre ter o payload feito?"):
        # SEMPRE salvar payload parcial — mesmo no caminho raiz pendente.
        # Antes esse caminho não chamava salvar_payload e o operador via
        # "Nenhum payload encontrado pra esta pendência" na web/Tkinter.
        # Agora monta attrs a partir de _dados_parciais (chaves PT do Claude
        # mapeadas pros nomes da API JSON:API) e grava.
        try:
            payload_parcial = _payload_parcial_de_dados(dp)
            # v2.16.50: fallback — se nao pegou nome direto do dp, tenta do
            # payload normalizado (que mapeou chaves PT-BR pra API JSON:API).
            if not nome_extraido:
                _attrs = ((payload_parcial.get("data") or {}).get("attributes")
                          or {})
                if _attrs.get("nome"):
                    nome_extraido = str(_attrs["nome"]).strip()
            resolucao_parcial = {
                "indice": 1,
                "total": 1,
                "nome": nome_extraido or "(nome não extraído)",
                "cnpj_empresa": str(cnpj_extraido) if cnpj_extraido else "",
                "razao_social": razao_extraida,
                "claude_motivo": motivo,
                "claude_confianca": resposta_claude.get("_confianca"),
                "claude_dados_parciais": dp,
            }
            salvar_payload(
                msg_id, metadados, payload_parcial,
                resolucao=resolucao_parcial,
                resultado={
                    "status": "pendente_claude_raiz",
                    "candidato_id": None,
                    "erro": motivo,
                },
            )
            log.info(
                f"   💾 Payload parcial salvo ({len(payload_parcial['data']['attributes'])} "
                f"attrs) — operador pode completar e POSTar"
            )
        except Exception as e:
            log.warning(
                f"   ⚠ Falha salvando payload parcial (raiz pendente): "
                f"{type(e).__name__}: {e}"
            )

        resultado_pendente = {
            "indice": 1,
            # nome_extraido pode ter sido atualizado do payload_parcial acima
            "nome": nome_extraido or "(nome não extraído)",
            "ok": False,
            "interno": False,  # vai pro cliente — falta info do lado dele
            "candidato_id": None,
            "erro_tecnico": f"Claude marcou como pendente: {motivo}",
            "motivo_cliente": motivo,
            "campos_faltando": [],
            "payload_parcial": resposta_claude,
            "razao_social": razao_extraida,
            "cnpj_empresa": str(cnpj_extraido) if cnpj_extraido else "",
        }
        _finalizar_lote(
            [resultado_pendente], msg_id, msg_pra_resposta,
            ids_label_pendente_remover, gmail, config,
        )
        return

    # 3. Normaliza pra lista de blocos (1 ou N)
    blocos = normalizar_admissoes(resposta_claude)
    if not blocos:
        _raise_pendencia(
            "Resposta do Claude sem admissões processáveis",
            motivo_cliente=(
                "Não consegui identificar nenhuma admissão completa nos "
                "documentos. Pode reenviar a ficha com os dados do candidato?"
            ),
            payload_parcial=resposta_claude,
        )

    log.info(f"   📋 {len(blocos)} admissão(ões) detectada(s)")

    # 4. Caches compartilhados entre blocos do mesmo email (geralmente
    #    mesmo CNPJ → mesma empresa, mesmos departamentos)
    cache_empresa: dict[str, tuple[str | None, dict]] = {}
    cache_deptos: dict[str, list[dict]] = {}

    # 5. Processa cada bloco; pega exceptions de pendência individualmente
    resultados: list[dict] = []
    for i, bloco in enumerate(blocos, 1):
        attrs_pre = (bloco.get("data") or {}).get("attributes") or {}
        nome_pre = attrs_pre.get("nome") or f"funcionário #{i}"
        # Extrai CNPJ direto do bloco pra preservar mesmo em caso de
        # exception (antes da resolução de empresa)
        cnpj_pre = bloco.get("cnpj_empresa") or attrs_pre.get("cnpj_empresa") or ""
        try:
            r = _processar_um_bloco(
                bloco=bloco, indice=i, total=len(blocos),
                msg_id=msg_id, metadados=metadados,
                api=api, claude=claude, planilha_cbo=planilha_cbo,
                config=config, corpo=corpo, anexos=anexos,
                cache_empresa=cache_empresa, cache_deptos=cache_deptos,
            )
            resultados.append(r)
        except ValueError as e:
            log.exception(f"   ❌ [{i}/{len(blocos)}] {nome_pre}: {e}")
            # v2.16.10: tenta resolver razão social via cache em vez de gravar
            # None — assim a planilha e a UI mostram a empresa em vez de "?".
            # Cache lookup é O(1), grátis quando já carregado.
            razao_resolvida = None
            try:
                import re as _re_local
                cnpj_d = _re_local.sub(r"\D", "", str(cnpj_pre or ""))
                if cnpj_d and len(cnpj_d) in (11, 14):
                    info = obter_empresas_cache(None).info(cnpj_d) or {}
                    razao_resolvida = info.get("razao_social") or None
            except Exception:
                razao_resolvida = None
            resultados.append({
                "indice": i,
                "nome": nome_pre,
                "ok": False,
                "candidato_id": None,
                "erro_tecnico": str(e),
                "motivo_cliente": getattr(e, "motivo_cliente", None) or str(e),
                "campos_faltando": getattr(e, "campos_faltando", []) or [],
                "payload_parcial": getattr(e, "payload_parcial", None) or bloco,
                "razao_social": razao_resolvida,
                "cnpj_empresa": cnpj_pre,
                # v2.15.15: mesmo no erro carrega o msg_id da irmã (pode ter
                # mesclado e mesmo assim ter dado erro depois — fechar a antiga)
                "_pendencia_anterior_msg_id": bloco.get("_pendencia_anterior_msg_id"),
            })

    # 6. Agrega: label, reply ou email DP, log
    _finalizar_lote(
        resultados, msg_id, msg_pra_resposta,
        ids_label_pendente_remover, gmail, config,
    )


def _resultado_pendencia_interna(
    *, indice: int, nome: str, razao: str | None, cnpj: str | None,
    erro: str, diagnostico_dp: str, bloco: dict,
    # Parâmetros opcionais pra salvar payload parcial — UI consegue mostrar
    # botão "Aplicar form e POSTar" no dialog. Quando não passados, segue o
    # comportamento antigo (sem payload salvo).
    msg_id: str | None = None,
    metadados: dict | None = None,
    empresa_id: str | None = None,
    depto_id: str | None = None,
    funcao_id: str | None = None,
) -> dict:
    """Resultado pra pendência cujo problema é INTERNO (cargo/CBO/cadastro
    do escritório). NÃO comunica nada ao cliente — só registra na planilha.
    `interno: True` faz `_finalizar_lote` rotear pra esse caminho.

    Salva payload parcial em payloads/ sempre que tiver dados suficientes
    (msg_id + bloco com data.attributes). Isso permite que o dialog
    "Resolver pendência" mostre os campos editáveis + o botão "POSTar".
    """
    log.warning(f"      🔧 [{indice}] {nome}: pendência interna — {erro}")

    # Cargo extraído pelo Claude — usado pra exibir na tabela de pendências.
    # Pode vir do top-level (cargo_extraido) ou do attributes.nomecargo do bloco.
    cargo_ia = (
        bloco.get("cargo_extraido")
        or (bloco.get("data") or {}).get("attributes", {}).get("nomecargo")
        or ""
    )

    # Tenta salvar payload parcial com placeholders nos IDs ainda não resolvidos.
    # Permite ao operador completar via dialog (sobrescreve antes do POST).
    if msg_id:
        try:
            payload_parcial = finalizar_payload(
                bloco,
                empresa_id=empresa_id or "1",
                departamento_id=depto_id or "1",
                funcao_id=funcao_id or "1",
            )
            resolucao_parcial = {
                "indice": indice, "nome": nome,
                "cnpj_empresa": cnpj,
                "empresa_id": empresa_id, "razao_social": razao,
                "departamento_id": depto_id,
                "funcao_id": funcao_id,
                "erro": erro,
                "diagnostico_dp": diagnostico_dp,
                "cargo_extraido": cargo_ia,
                # v2.11.0: persiste flag de estágio pra UI usar ao salvar alias
                "eh_estagio": bool(bloco.get("_eh_estagio")),
            }
            salvar_payload(
                msg_id, metadados or {}, payload_parcial,
                resolucao=resolucao_parcial,
                resultado={
                    "status": "pendente_interna",
                    "candidato_id": None,
                    "erro": erro,
                },
            )
        except Exception as e:
            # Subir pra warning — quando o payload deveria ser salvo mas falhou,
            # o operador perde a capacidade de reprocessar via dialog e a coluna
            # "Cargo entendido pela IA" fica vazia. Sinal importante.
            log.warning(
                f"      ⚠ pendência interna sem payload salvo "
                f"[msg={msg_id[:16]} nome='{nome}' cnpj={cnpj}]: "
                f"{type(e).__name__}: {e}"
            )

    return {
        "indice": indice,
        "nome": nome,
        "ok": False,
        "interno": True,
        "candidato_id": None,
        "erro_tecnico": erro,
        "diagnostico_dp": diagnostico_dp,
        "motivo_cliente": None,
        "campos_faltando": [],
        "payload_parcial": bloco,
        "razao_social": razao,
        "cnpj_empresa": cnpj,
        # v2.15.15: carrega a msg_id da pendência irmã (se mesclou) pra
        # _finalizar_lote fechar a label pendente daquele email tbm.
        "_pendencia_anterior_msg_id": bloco.get("_pendencia_anterior_msg_id"),
    }


def _processar_um_bloco(
    *, bloco: dict, indice: int, total: int,
    msg_id: str, metadados: dict,
    api: EContadorAPI, claude: ClaudeClient, planilha_cbo: list[dict],
    config: Config, corpo: str, anexos: list[dict],
    cache_empresa: dict, cache_deptos: dict,
) -> dict:
    """Processa UM bloco de admissão. Retorna dict de resultado.
    Levanta ValueError (via _raise_pendencia) se pendência for específica deste bloco.
    """
    attrs = (bloco.get("data") or {}).get("attributes") or {}
    nome = attrs.get("nome") or f"funcionário #{indice}"
    log.info(f"   👤 [{indice}/{total}] {nome}")

    # v2.15.15: detecta pendência IRMÃ — mesmo candidato em email anterior
    # do mesmo cliente (cenário "cliente mandou email separado em vez de
    # responder no thread"). Mescla os attrs: novo vence empate (info mais
    # recente), antigos preenchem o que faltava. msg_id da irmã é guardado
    # no bloco pra _finalizar_lote fechar a label pendente do email antigo.
    bloco.pop("_pendencia_anterior_msg_id", None)  # garante reset
    try:
        import dashboard_data as _dd
        cnpj_bloco = bloco.get("cnpj_empresa") or attrs.get("cnpj_empresa") or ""
        irma = _dd.encontrar_pendencia_irma(
            cpf=attrs.get("cpf"),
            nome=nome,
            cnpj=cnpj_bloco,
            excluir_msg_id=msg_id,
        )
        if irma:
            attrs_antigos = irma.get("_attrs_antigos") or {}
            mesclados = dict(attrs_antigos)
            for k, v in attrs.items():
                if v not in (None, "", 0, [], {}):
                    mesclados[k] = v  # novo vence
            bloco.setdefault("data", {"type": "candidatos"})["attributes"] = mesclados
            attrs = mesclados
            bloco["_pendencia_anterior_msg_id"] = irma["msg_id"]
            log.info(
                f"      🔀 Mesclou pendência irmã msg_id={irma['msg_id'][:16]} "
                f"({len(attrs_antigos)} attrs antigos + novos = {len(mesclados)})"
            )
    except Exception as e:
        log.warning(f"      ⚠ Detecção de pendência irmã falhou: {type(e).__name__}: {e}")

    # ─── Detector de estágio (v2.11.0) ───────────────────────────────
    # Marca o bloco como estagiário se o email indicar (palavras-chave em
    # assunto/corpo OU filename dos anexos). Pipeline usa essa flag pra:
    #   - Buscar SÓ aliases de função tipo "_estagio_<cargo>"
    #   - Forçar pendência interna se não há alias salvo (cliente tem função
    #     separada de estágio no eContador)
    # Não toca em outros campos do payload por enquanto — vou cobaiar
    # `tipovinculotrabalhista` etc. depois pra descobrir quais mais mudam.
    try:
        from estagio import detectar as _detectar_estagio
        assunto_meta = (metadados or {}).get("assunto", "") or ""
        anexos_nomes = [a.get("filename", "") for a in (anexos or [])]
        deteccao = _detectar_estagio(
            assunto=assunto_meta,
            corpo=corpo or "",
            anexos_filenames=anexos_nomes,
        )
        if deteccao.eh_estagio:
            bloco["_eh_estagio"] = True
            log.info(
                f"      🎓 ESTÁGIO detectado (confiança {deteccao.confianca:.0%}): "
                f"{', '.join(deteccao.evidencias[:3])}"
                f"{'...' if len(deteccao.evidencias) > 3 else ''}"
            )
    except Exception as e:
        log.warning(f"      ⚠ Falha no detector de estágio: {type(e).__name__}: {e}")

    # ─── Pre-enrichment (v2.4.1) ─────────────────────────────────────
    # Roda enrichment (ViaCEP + cadastro escritório + DirectData) JÁ AQUI,
    # antes de qualquer detecção de pendência. Por quê:
    #   - Pendência cliente do Claude: às vezes ele desiste por achar que
    #     faltam dados, mas tem CPF/CEP que dariam pra completar via APIs.
    #     Caso real: Pedro Henrique — Claude marcou _pendente, mas tinha CPF
    #     (DirectData preencheria nome/nasc/mãe/pai) e CEP (ViaCEP, endereço).
    #   - Pendência interna (CNPJ/depto/função): payload parcial salvo fica
    #     mais completo pro DP resolver no dialog.
    # Custo: DirectData só dispara se tem CPF. Sem CPF = R$ 0.
    # Skip por economia (no enrichment) evita gasto duplo.
    try:
        bloco = enrich_candidato(bloco)
        enrich_meta_early = bloco.pop("_enrich_meta", {})
        n_cep = len(enrich_meta_early.get("fields_filled_by_cep", []))
        n_cpf = len(enrich_meta_early.get("fields_filled_by_cpf", []))
        if n_cep or n_cpf:
            log.info(
                f"      ✨ Pre-enrichment: +{n_cep} via CEP, +{n_cpf} via CPF "
                f"(antes de detectar pendência)"
            )
        if enrich_meta_early.get("cep_suspeito"):
            cs = enrich_meta_early["cep_suspeito"]
            log.warning(
                f"      🔍 SANITY CHECK CEP suspeito: Claude leu cidade="
                f"'{cs['cidade_claude']}' mas CEP {cs['cep']} é de "
                f"'{cs['cidade_viacep']}'. Endereço NÃO sobrescrito."
            )
        # Marca que enrichment já rodou — evita rodar de novo na linha ~1334
        # (idempotente mas poluiria log). Cache RAM cobre DirectData/ViaCEP.
        bloco["_pre_enriched"] = True
        # Re-extrai attrs do bloco enriquecido pra usar daqui pra frente
        attrs = (bloco.get("data") or {}).get("attributes") or {}
    except Exception as e:
        log.warning(f"      ⚠ Falha no pre-enrichment: {type(e).__name__}: {e}")

    # ─── Salário padrão por cliente (v2.12.0) ────────────────────────
    # Se o cliente é recorrente e o escritório já tem cadastrado o salário
    # padrão pra esse CNPJ + cargo, preenche automaticamente. Resolve o
    # caso real: cliente manda "salário base" sem valor → antes virava
    # pendência cliente; agora se já temos o cadastro, sobe direto.
    # v2.16.15: agora detecta strings sentinela ("NÃO INFORMADO — pendente",
    # "SALARIO BASE", "A COMBINAR") que o Claude às vezes devolve em vez de
    # vazio — antes a checagem `not ... or in (0,...)` deixava passar.
    def _eh_salario_vazio(v) -> bool:
        if v is None or v == "":
            return True
        if isinstance(v, bool):  # bool é subclasse de int — proteção
            return True
        if isinstance(v, (int, float)):
            return v <= 0
        # string: tenta converter pra número; se não der ou der <=0 → vazio
        try:
            s = str(v).replace("R$", "").replace(".", "").replace(",", ".").strip()
            return float(s) <= 0
        except (ValueError, TypeError):
            return True
    try:
        if _eh_salario_vazio(attrs.get("salario")):
            from salarios_padrao import consultar as _consultar_salario_padrao
            cnpj_bloco = (
                bloco.get("cnpj_empresa")
                or attrs.get("cnpj_empresa")
                or ""
            )
            cargo_bloco = (
                bloco.get("cargo_extraido")
                or attrs.get("nomecargo")
                or ""
            )
            if cnpj_bloco and cargo_bloco:
                entry = _consultar_salario_padrao(cnpj_bloco, cargo_bloco)
                if entry and entry.get("salario"):
                    valor = float(entry["salario"])
                    attrs["salario"] = valor
                    bloco.setdefault("data", {})["attributes"] = attrs
                    log.info(
                        f"      💰 Salário R$ {valor:.2f} aplicado do cadastro padrão "
                        f"(CNPJ {cnpj_bloco}, cargo '{cargo_bloco}', "
                        f"fonte={entry.get('fonte', '?')})"
                    )
                else:
                    log.info(
                        f"      💡 Salário ausente e sem cadastro padrão pra "
                        f"CNPJ {cnpj_bloco} + cargo '{cargo_bloco}' — cadastre "
                        f"em salarios_padrao.json pra próxima admissão descer sozinha"
                    )
    except Exception as e:
        log.warning(f"      ⚠ Falha consultando salário padrão: {type(e).__name__}: {e}")

    # ─── Defaults do perfil do remetente (v2.16.0) ───────────────────
    # Fase 3 da memória: aplica padrões aprendidos do REMETENTE quando o
    # bloco ainda está faltando campos. O salário padrão acima é por CNPJ;
    # esse aqui é por pessoa (rh@cliente.com). Útil quando o mesmo email
    # gerencia múltiplos CNPJs e o padrão do humano é mais constante que
    # o da empresa.
    # Só preenche o que está cadastrado no perfil e foi detectado como
    # omissão habitual. Nunca chuta. Idempotente — não sobrescreve.
    try:
        import perfis_remetente as _pr
        rem = (metadados or {}).get("remetente", "")
        if rem:
            cnpj_perf = bloco.get("cnpj_empresa") or attrs.get("cnpj_empresa") or ""
            preenchidos = _pr.aplicar_defaults_do_perfil(bloco, rem, cnpj_perf)
            if preenchidos:
                attrs = (bloco.get("data") or {}).get("attributes") or {}
                log.info(
                    f"      🧠 Defaults do perfil de {rem}: "
                    f"{', '.join(preenchidos)}"
                )
    except Exception as e:
        log.warning(
            f"      ⚠ Falha aplicando defaults do perfil: {type(e).__name__}: {e}"
        )

    # ─── Endereço padrão por CNPJ (v2.16.27) ──────────────────────────
    # Para fazendas/sítios onde os peões moram no alojamento. Cadastro
    # estruturado por CNPJ em `enderecos_padrao_empresa.json`. Idempotente:
    # NUNCA sobrescreve dado que veio dos documentos.
    # v2.16.30: fallback por REMETENTE quando CNPJ vazio — usa
    # remetente_aliases.json pra inferir o CNPJ e tentar o cadastro.
    try:
        import enderecos_padrao_empresa as _epe
        cnpj_end = bloco.get("cnpj_empresa") or attrs.get("cnpj_empresa") or ""
        # Fallback: se CNPJ vazio, tenta inferir pelo remetente
        if not cnpj_end:
            try:
                from remetente_empresa import consultar_alias_exato as _consultar_alias
                rem = (metadados or {}).get("remetente", "")
                if rem:
                    alias = _consultar_alias(rem)
                    if alias and alias.get("cnpj"):
                        cnpj_end = alias["cnpj"]
                        log.info(
                            f"      🏠 CNPJ inferido pelo remetente {rem}: "
                            f"{cnpj_end} (pra tentar endereço padrão)"
                        )
            except Exception:
                pass
        if cnpj_end:
            preenchidos_end = _epe.aplicar_em_attrs(attrs, cnpj_end)
            if preenchidos_end:
                bloco.setdefault("data", {})["attributes"] = attrs
                log.info(
                    f"      🏠 Endereço padrão do CNPJ {cnpj_end} aplicado: "
                    f"{', '.join(preenchidos_end[:3])}"
                    + (f" (+{len(preenchidos_end)-3} mais)" if len(preenchidos_end) > 3 else "")
                )
    except Exception as e:
        log.warning(
            f"      ⚠ Falha aplicando endereço padrão: {type(e).__name__}: {e}"
        )

    # Caso especial: Claude marcou _pendente=true neste bloco individual.
    # Acontece quando ele extraiu dados de 1+ funcionário num email com N
    # admissões, mas desistiu de alguns por falta de campos críticos. Em vez
    # de descartar silenciosamente, vira pendência cliente com motivo claro.
    if bloco.get("_pendente"):
        motivo_cliente = bloco.get("_motivo") or "Dados insuficientes pra processar"
        # v2.14.1 (PATCHES.md §3.3): matcher por CÓDIGO (briefing §3.8 enum).
        # Decisão programática em substring de texto livre era frágil — o
        # Claude muda a redação a cada release e o matcher quebrava silencioso.
        # Agora: enum fechado decide; matching textual abaixo é fallback pra
        # respostas antigas/cacheadas SEM `_motivo_codigo`.
        AUTO_RESOLVIVEIS_CODIGOS = {
            "CTPS_AUSENTE",
            "ESTAGIO_INDEFINIDO",
        }
        # DATA_ADMISSAO_AUSENTE só é auto-resolvível com flag config
        if config.usar_data_atual_se_invalida or config.sempre_mandar_sem_data_admissao:
            AUTO_RESOLVIVEIS_CODIGOS = AUTO_RESOLVIVEIS_CODIGOS | {"DATA_ADMISSAO_AUSENTE"}
        BLOQUEANTES_CODIGOS = {
            "SALARIO_AUSENTE",
            "CPF_AUSENTE_OU_INVALIDO",
            "RG_AUSENTE",
            "ENDERECO_INCOMPLETO",
            "NOME_MAE_AUSENTE",
            "NASCIMENTO_AUSENTE",
            "CARGO_AUSENTE",
            "CARGOS_DIVERGENTES_SEM_ASO",
            "DOC_ILEGIVEL",
            "DOCS_PESSOA_AUSENTES",
            "CNPJ_NAO_LOCALIZADO",
        }
        cod_principal = str(bloco.get("_motivo_codigo") or "").upper().strip()
        cods_extras = bloco.get("_motivos_codigos") or []
        if isinstance(cods_extras, list):
            cods = {cod_principal, *(str(c).upper().strip() for c in cods_extras)} - {""}
        else:
            cods = {cod_principal} - {""}

        # Caminho A: temos códigos → decisão limpa
        if cods:
            tem_bloqueante = bool(cods & BLOQUEANTES_CODIGOS)
            todos_auto = cods <= AUTO_RESOLVIVEIS_CODIGOS  # subset (qualquer não-AUTO bloqueia)
            if not tem_bloqueante and todos_auto:
                log.info(
                    f"      ℹ [{indice}] {nome}: ignorando _pendente — "
                    f"motivo_codigo auto-resolvível ({', '.join(sorted(cods))})"
                )
                # Promove _dados_parciais → attrs e segue (mesmo path do
                # matching textual abaixo, mas SEM passar pelo if textual)
                dp_attrs = bloco.get("_dados_parciais") or {}
                attrs_merged = dict(dp_attrs)
                attrs_merged.update(attrs)  # attrs já preenchidos têm prioridade
                bloco["data"] = {
                    "type": "candidatos",
                    "attributes": attrs_merged,
                    "relationships": (bloco.get("data") or {}).get("relationships") or {},
                }
                attrs = attrs_merged
                bloco.pop("_pendente", None)
                bloco.pop("_motivo", None)
                # CONTINUA pro resto do fluxo (resolver empresa/depto/função)
                # — pula completamente o matching textual abaixo.
                # Variável-sentinela pra sinalizar pro código abaixo: já tratamos
                bloco["_motivo_codigo_tratou"] = True
            elif tem_bloqueante:
                log.warning(
                    f"      ⚠ [{indice}] {nome}: _motivo_codigo bloqueante "
                    f"({', '.join(sorted(cods & BLOQUEANTES_CODIGOS))}) — "
                    f"pendência cliente"
                )
                # Cai pro matching textual abaixo já que ele faz o save do
                # payload parcial + return — não duplicar lógica aqui.

        # Fallback: matching textual (PATCHES.md §3.3 obs — fallback legacy
        # pra respostas sem `_motivo_codigo`). Se o caminho A já tratou,
        # `_motivo_codigo_tratou` está True e o resto do bloco "if _pendente:"
        # é ignorado via early-skip do `if bloco.get("_pendente")` original
        # (a flag foi removida no caminho A).
        ja_tratado_por_codigo = bool(bloco.pop("_motivo_codigo_tratou", False))

        # Caminho A já promoveu _dados_parciais → segue fluxo normal sem
        # cair no matching textual. O motivo_lower abaixo é só pra log/save.
        if not ja_tratado_por_codigo:
            # FALSOS POSITIVOS: Claude marca pendente por campos que o backend
            # resolve automaticamente. Verifica se o motivo pode ser tratado:
            #   - CTPS/seriectps: derivadas do CPF (regra do escritório)
            #   - Data de admissão: se config.usar_data_atual_se_invalida=True,
            #     usamos hoje em vez de pedir
            # Quando todos os termos do motivo podem ser auto-resolvidos, ignora
            # _pendente e segue o fluxo normal (enrichment + aplicar_regra_data).
            motivo_lower = motivo_cliente.lower()
            dp_attrs = bloco.get("_dados_parciais") or {}
            tem_cpf = bool(dp_attrs.get("cpf") or attrs.get("cpf"))

            # Lista de termos que aparecem em motivos auto-resolvíveis
            termos_auto_resolviveis: list[str] = []
            if tem_cpf:
                termos_auto_resolviveis += ["ctps", "seriectps"]
            if config.usar_data_atual_se_invalida:
                termos_auto_resolviveis += ["data de admissão", "data de admissao",
                                             "data admiss", "data admissao",
                                             "informada no email"]
            # ASO admissional NÃO é obrigatório no payload — DP completa manual
            # no Desktop quando faltar. Não bloqueia admissão. Termos típicos que
            # Claude usa quando desiste por causa disso:
            termos_auto_resolviveis += [
                "dataatestadoocupacional",
                "data do exame admissional",
                "data do exame ocupacional",
                "data do atestado",
                "exame admissional",
                "aso admissional",
                "atestado ocupacional",
                "atestado admissional",
            ]
            # Campos com DEFAULT do escritório — Claude não devia marcar pendente
            # por causa deles, mas às vezes marca. Estado civil = Solteiro (id=1)
            # e escolaridade = Médio (id=7) são aplicados em apply_fixed_defaults.
            termos_auto_resolviveis += [
                "estado civil",
                "estadocivil",
                "escolaridade",
                "datactps",
                "data da ctps",
                "data de emissão da ctps",
                "série da ctps",
                "serie da ctps",
            ]
            # Estagiário: Claude às vezes "desiste" alegando que estágio não é CLT
            # nem gera S-2200 (juridicamente correto, mas operacionalmente errado —
            # o pipeline da Crosara trata estagiário como admissão normal). v2.11.3
            # Caso real (JESSYKA 11/06/2026): pendência cliente com motivo "Contrato
            # de estágio (Lei 11.788/2008) NÃO é vínculo CLT — não gera payload de
            # admissão eSocial S-2200". Pipeline detecta o estágio depois e mapeia
            # função correta — Claude só precisa extrair os dados.
            termos_auto_resolviveis += [
                "lei 11.788",
                "lei 11788",
                "nao e vinculo clt",
                "não é vínculo clt",
                "não é vinculo clt",
                "nao é vinculo clt",
                "não gera payload de admissão esocial",
                "nao gera payload de admissao esocial",
                "não gera s-2200",
                "nao gera s-2200",
                "não gera s2200",
                "nao gera s2200",
                "esocial s-2200",
                "esocial s2200",
                "termo de compromisso de estagio",
                "termo de compromisso de estágio",
            ]
            # Claude às vezes pede "confirmação cruzada" de campos que ELE MESMO já
            # extraiu (caso real GABRIELY: "CPF documento mostra X mas precisa de
            # confirmação cruzada"). Se o motivo cita "confirmar"/"confirmação"
            # E os campos críticos estão preenchidos nos attrs, ignora.
            attrs_tem_cpf = bool(attrs.get("cpf") or dp_attrs.get("cpf"))
            attrs_tem_admissao = bool(attrs.get("admissao") or dp_attrs.get("admissao"))
            attrs_tem_nome = bool(attrs.get("nome") or dp_attrs.get("nome"))
            attrs_tem_salario = bool(attrs.get("salario") or dp_attrs.get("salario"))
            if attrs_tem_cpf and "cpf" in motivo_lower and (
                "confirma" in motivo_lower or "cruzada" in motivo_lower
                or "precisa" in motivo_lower
            ):
                termos_auto_resolviveis += ["cpf"]  # despendentar, CPF já tá lá
            if attrs_tem_admissao and (
                "confirma" in motivo_lower or "inferid" in motivo_lower
            ):
                termos_auto_resolviveis += ["data de admissão", "data admiss"]
            # v2.12.0: se salário foi preenchido pelo cadastro padrão (CNPJ+cargo)
            # antes da detecção de pendente, despendentar motivos sobre salário.
            if attrs_tem_salario and (
                "sal" in motivo_lower and (
                    "base" in motivo_lower
                    or "informad" in motivo_lower
                    or "valor" in motivo_lower
                    or "num[ée]rico" in motivo_lower
                    or "numerico" in motivo_lower
                )
            ):
                termos_auto_resolviveis += ["salário", "salario"]

            # Termos que REALMENTE são problema do cliente (mantém pendência)
            # NOTA: removido "escolaridade" (tem default), trocado por checagem
            # contextual acima pra "cpf" (só bloqueia se CPF ausente de verdade).
            termos_bloqueantes = (
                "cep", "rua", "endereço", "endereco", "rg",
                "identidade", "salário", "salario", "nome", "cpf",
                "nascimento", "mãe", "mae", "função", "funcao",
            )

            tem_termo_auto = any(t in motivo_lower for t in termos_auto_resolviveis)
            # Bloqueante real: termo bloqueante presente E não foi adicionado
            # ao auto-resolvíveis pela checagem contextual de "confirmação".
            tem_termo_bloqueante = any(
                t in motivo_lower for t in termos_bloqueantes
                if t not in termos_auto_resolviveis
            )

            if tem_termo_auto and not tem_termo_bloqueante:
                quais = [t for t in termos_auto_resolviveis if t in motivo_lower]
                log.info(
                    f"      ℹ [{indice}] {nome}: ignorando _pendente — motivo "
                    f"auto-resolvido por config/regra ({', '.join(quais)})"
                )
                # Promove _dados_parciais pra data.attributes pra seguir o fluxo
                attrs_merged = dict(dp_attrs)
                attrs_merged.update(attrs)  # attrs já preenchidos têm prioridade
                bloco["data"] = {
                    "type": "candidatos",
                    "attributes": attrs_merged,
                    "relationships": (bloco.get("data") or {}).get("relationships") or {},
                }
                attrs = attrs_merged
                # Limpa flags pra não confundir downstream
                bloco.pop("_pendente", None)
                bloco.pop("_motivo", None)
                # Segue o fluxo normal
            else:
                log.warning(f"      ⚠ [{indice}] {nome}: Claude marcou _pendente — {motivo_cliente[:80]}")

                # Salva payload parcial pra UI poder mostrar os dados extraídos +
                # permitir operador preencher os campos faltantes via dialog.
                payload_parcial = {
                    "data": {
                        "type": "candidatos",
                        "attributes": dict(attrs),
                        "relationships": {},
                    }
                }
                resolucao_parcial = {
                    "indice": indice, "total": total, "nome": nome,
                    "cnpj_empresa": bloco.get("cnpj_empresa") or attrs.get("cnpj_empresa") or "",
                    "claude_motivo": motivo_cliente,
                    "claude_confianca": bloco.get("_confianca"),
                    "claude_dados_parciais": bloco.get("_dados_parciais") or {},
                }
                try:
                    salvar_payload(
                        msg_id, metadados, payload_parcial,
                        resolucao=resolucao_parcial,
                        resultado={
                            "status": "pendente_claude", "candidato_id": None,
                            "erro": motivo_cliente,
                        },
                    )
                except Exception as e:
                    log.warning(
                        f"      ⚠ Falha salvando payload parcial "
                        f"[msg={msg_id[:16]} nome='{nome}']: "
                        f"{type(e).__name__}: {e}"
                    )

                return {
                    "_pendencia_anterior_msg_id": bloco.get("_pendencia_anterior_msg_id"),
                    "indice": indice,
                    "nome": nome,
                    "ok": False,
                    "interno": False,  # vai pro cliente — falta info do lado dele
                    "candidato_id": None,
                    "erro_tecnico": f"Claude marcou bloco como pendente: {motivo_cliente}",
                    "motivo_cliente": motivo_cliente,
                    "campos_faltando": [],
                    "payload_parcial": bloco,
                    "razao_social": None,
                    "cnpj_empresa": bloco.get("cnpj_empresa") or attrs.get("cnpj_empresa") or "",
                }

    dados = extrair_dados_consulta(bloco)
    cnpj = dados["cnpj_empresa"]
    if not cnpj:
        # CNPJ não foi extraído dos documentos. Antes de mandar pra pendência
        # interna, tentamos inferir a empresa a partir do remetente do email:
        #   1. Alias exato (cache `remetente_aliases.json`)
        #   2. Domínio do email (`@modelofarma.com.br` → empresa MODELOFARMA)
        #   3. Razão social fuzzy match no assunto/corpo do email
        # v2.10.0: estratégias do módulo remetente_empresa.
        try:
            from remetente_empresa import resolver as inferir_empresa_por_remetente
            remetente_email = (metadados or {}).get("remetente", "") or ""
            assunto = (metadados or {}).get("assunto", "") or ""
            texto_busca = f"{assunto}\n\n{corpo or ''}"
            cache_global = obter_empresas_cache(api)
            inferencia = inferir_empresa_por_remetente(
                remetente=remetente_email,
                texto_email=texto_busca,
                cache=cache_global,
            )
        except Exception as e:
            log.warning(f"      ⚠ Falha na inferência de empresa por remetente: {type(e).__name__}: {e}")
            inferencia = None

        if inferencia and inferencia.get("cnpj"):
            cnpj = inferencia["cnpj"]
            log.info(
                f"      🎯 CNPJ INFERIDO ({inferencia['estrategia']}): {cnpj} "
                f"({inferencia.get('razao_social', '?')}) — substituindo CNPJ vazio"
            )
            # Atualiza dados pra propagar adiante (departamento, função usam)
            dados["cnpj_empresa"] = cnpj
        else:
            # CNPJ não extraído E inferência também falhou → problema NOSSO
            # (DP olha o remetente/assinatura do email pra deduzir qual
            # cliente é). NÃO vai pro cliente.
            return _resultado_pendencia_interna(
                indice=indice, nome=nome, razao=None, cnpj=None,
                erro="CNPJ da empresa contratante não foi extraído do email/anexos",
                diagnostico_dp=(
                    f"Claude não identificou CNPJ pra {nome} e as estratégias de "
                    f"inferência (alias de remetente, domínio, razão social) "
                    f"também não bateram. DP pode olhar o remetente do email "
                    f"({(metadados or {}).get('remetente', '?')}), assinatura, "
                    f"ou logo nos documentos pra descobrir a empresa contratante. "
                    f"Use 'Corrigir CNPJ' + marque 'Salvar como alias permanente' "
                    f"pra esse remetente nunca mais cair em pendência."
                ),
                bloco=bloco,
                msg_id=msg_id, metadados=metadados,
            )

    # Empresa (cache por email + GET /empresas)
    if cnpj in cache_empresa:
        empresa_id, empresa_attrs = cache_empresa[cnpj]
    else:
        empresa_id, empresa_attrs = api.resolver_empresa(cnpj)
        cache_empresa[cnpj] = (empresa_id, empresa_attrs)

    # Auto-correção via cache global de CNPJs: se o GET /empresas falhou,
    # tenta gerar variações do CNPJ (1-2 dígitos) e checar contra a whitelist
    # local. Resolve caso real: PDF com typo (Ekoplastic 09.401 → 09.491).
    cnpj_corrigido = None
    if not empresa_id:
        cache_global = obter_empresas_cache(api)
        candidato, motivo = corrigir_cnpj_via_cache(cnpj, cache_global)
        if candidato and motivo in ("corrigido_1d", "corrigido_2d"):
            log.info(
                f"      🎯 CNPJ corrigido automaticamente via cache: "
                f"{cnpj} → {candidato} ({motivo})"
            )
            cnpj_corrigido = candidato
            # Atualiza cnpj usado daqui pra frente
            cnpj = candidato
            # Tenta de novo o GET com o CNPJ correto
            if candidato in cache_empresa:
                empresa_id, empresa_attrs = cache_empresa[candidato]
            else:
                empresa_id, empresa_attrs = api.resolver_empresa(candidato)
                cache_empresa[candidato] = (empresa_id, empresa_attrs)

    if not empresa_id:
        # CNPJ não cadastrado no eContador (mesmo após auto-correção).
        # NÃO vai pro cliente — DP resolve internamente.
        return _resultado_pendencia_interna(
            indice=indice, nome=nome, razao=None, cnpj=cnpj,
            erro=f"CNPJ {cnpj} não está cadastrado no eContador",
            diagnostico_dp=(
                f"O CNPJ {cnpj} (extraído pra {nome}) não retornou empresa "
                f"em GET /empresas?filter[cpfcnpj]={cnpj}, e a busca por "
                f"variações no cache também falhou. Possibilidades: "
                f"(a) Crosara ainda não fez onboarding dessa empresa no "
                f"eContador — cadastrar primeiro; (b) CNPJ tem typo grande "
                f"demais (>2 dígitos) — operador pode usar 'Corrigir CNPJ'."
            ),
            bloco=bloco,
            msg_id=msg_id, metadados=metadados,
        )
    razao = empresa_attrs.get("nome", "?")
    log.info(f"      🏢 Empresa {empresa_id}: {razao}")

    # Departamento (cache de listagem por empresa)
    if empresa_id in cache_deptos:
        deptos_api = cache_deptos[empresa_id]
    else:
        deptos_api = api.listar_departamentos(empresa_id)
        cache_deptos[empresa_id] = deptos_api
    depto_id, depto_msg = resolver_departamento(
        empresa_id=empresa_id, cnpj_empresa=cnpj, razao_social=razao,
        deptos_api=deptos_api, departamento_sugerido=dados["departamento_sugerido"],
        departamentos_json_paths=[DEPARTAMENTOS_FILE, ROOT.parent / "departamentos.json"],
        # v2.14.1: liga REGRA 0 quando empresa tem 0 deptos no eContador.
        # Flag fica false até validarmos com cobaia (PATCHES.md §6.5).
        permitir_sem_departamento=getattr(
            config, "postar_sem_departamento_quando_vazio", False
        ),
    )

    # Fallback inteligente: se o resolver determinístico não bateu, pede pro
    # Claude escolher o departamento com base no nome do cargo. Geralmente o
    # cargo já diz o setor (Motorista→Transporte, Vendedor→Vendas, etc.).
    if depto_msg != "ok" and deptos_api and len(deptos_api) > 1 and dados.get("cargo"):
        try:
            id_escolhido, motivo_ia = claude.escolher_departamento_por_cargo(
                cargo=dados["cargo"],
                deptos=deptos_api,
                cbo=dados.get("cbo"),
            )
            if id_escolhido:
                depto_id = id_escolhido
                depto_msg = "ok-ia"
                log.info(
                    f"      🗂 Depto (IA): {depto_id} — {motivo_ia}"
                )
        except Exception:
            log.exception("      Falha no fallback IA pra depto — seguindo")

    # v2.14.1: depto_msg pode ser "ok", "ok-ia" ou "ok (sem departamento ...)"
    # (REGRA 0 quando empresa tem 0 deptos e flag postar_sem_departamento ON).
    # Tratamos qualquer string que COMEÇA com "ok" como sucesso de resolução.
    depto_resolvido = depto_msg == "ok-ia" or str(depto_msg).startswith("ok")
    if not depto_resolvido:
        # Sem resolução determinística NEM fallback IA — pendência interna.
        # Cliente não vê (problema do nosso cadastro), DP recebe a lista.
        deptos_disponiveis = (
            "; ".join(f"{d['nome']} (id={d['id']})" for d in deptos_api[:15])
            if deptos_api else "(empresa sem departamentos retornados pela API)"
        )
        return _resultado_pendencia_interna(
            indice=indice, nome=nome, razao=razao, cnpj=cnpj,
            erro=f"Departamento não resolvido: {depto_msg}",
            diagnostico_dp=(
                f"Empresa {razao} (CNPJ {cnpj}) — {depto_msg}\n"
                f"Cargo extraído: {dados.get('cargo') or '(nenhum)'}\n"
                f"Departamentos disponíveis no eContador: {deptos_disponiveis}\n"
                f"O fallback de IA também não conseguiu escolher.\n"
                f"Resolver: (a) configurar `departamento_default_id` em "
                f"departamentos.json pra essa empresa, ou (b) adicionar "
                f"`nome_variantes` que cubra o nome usado no email, ou "
                f"(c) cadastrar manualmente escolhendo um departamento."
            ),
            bloco=bloco,
            msg_id=msg_id, metadados=metadados,
            empresa_id=empresa_id,
            # depto_id deixa None → placeholder "1"
        )
    if depto_msg == "ok":
        log.info(f"      🗂 Depto: {depto_id}")

    # Função — SÓ resolve automaticamente se bater num X-marcado da planilha.
    # Sem fallback Claude (custoso) e sem fallback planilha completa (engana).
    # Quando não bate, vira pendência e o operador resolve via UI digitando
    # o código no campo "Função (código)" do dialog "Resolver pendência".
    #
    # Override manual via UI tem prioridade sobre o resolver automático.
    # v2.15.5: `sempre_mandar_sem_funcao` mudou de comportamento — agora é
    # CONDICIONAL: tenta resolver primeiro; se não conseguir, usa placeholder
    # em vez de virar pendência. Antes era absoluto (pulava o resolver).
    override_funcao_id = obter_funcao_override(msg_id, nome)
    if override_funcao_id:
        funcao_id, conf, ambiguos, fmsg = override_funcao_id, 1.0, [], "ok (override manual via UI)"
        log.info(f"      💼 Função override aplicado: id={funcao_id} (correção manual)")
    else:
        # v2.11.0: passar eh_estagio pra busca de alias direcionada
        # (estágio tem aliases separados de CLT — função no eContador difere)
        funcao_id, conf, ambiguos, fmsg = resolver_funcao(
            planilha_cbo, dados["cargo"], dados["cbo"],
            eh_estagio=bool(bloco.get("_eh_estagio")),
        )

    # v2.15.14: se não conseguiu resolver E a flag está LIGADA, omite a
    # relationship `funcao` do payload (em vez de mandar id=1 que vira
    # uma função real qualquer — "ASSISTENTE DE CPD" no caso real).
    # DP escolhe a função correta no Alterdata Desktop. Sentinela _SEM_FUNCAO
    # pra distinguir "não pôde resolver" de "intencionalmente omitir".
    if funcao_id is None and config.sempre_mandar_sem_funcao:
        funcao_id = ""  # string vazia → payload_builder omite a relationship
        conf, fmsg = 1.0, "ok (omitida — DP escolhe no Desktop)"
        log.info(
            f"      💼 Função não resolvida — omitindo relationship 'funcao' "
            f"(flag 'sem função se não identificada' ligada). DP escolhe no Desktop."
        )

    if funcao_id is None:
        # Pendência interna — operador resolve no dialog preenchendo o código
        candidatos_str = (
            "Candidatos parecidos (mas sem confiança alta): "
            + "; ".join(
                f"{f['nome_cargo']} (id={f['funcao_id']}, cbo={f.get('cbo') or '?'})"
                for f in ambiguos[:5]
            )
            if ambiguos else
            "Sem cargos parecidos entre os X-marcados."
        )
        # Diagnóstico específico pra estágio (v2.11.0): cliente costuma ter
        # função separada (ex: ESTAGIÁRIO DE LOJA com CBO próprio); pedir
        # código manualmente e oferecer salvar como alias permanente.
        if bloco.get("_eh_estagio"):
            diag = (
                f"🎓 ESTÁGIO detectado pra {nome} (cargo extraído: "
                f"'{dados.get('cargo')}').\n"
                f"O cliente provavelmente cadastrou uma função separada de estágio "
                f"no eContador (ex: 'ESTAGIÁRIO DE LOJA', CBO específico).\n\n"
                f"Resolver: dê 2× clique nesta linha → preencha 'Função (código)' "
                f"com o código da função de ESTÁGIO correspondente, e MARQUE "
                f"'Salvar como alias permanente'. Próximos estagiários do mesmo "
                f"cargo deste cliente sobem automáticos.\n\n"
                f"⚠️ NÃO use a função CLT normal — o eContador trata estagiário "
                f"como vínculo diferente."
            )
            erro_msg = f"ESTÁGIO sem alias salvo pro cargo '{dados.get('cargo')}'"
        else:
            diag = (
                f"Cargo extraído: '{dados.get('cargo')}' / "
                f"CBO '{dados.get('cbo') or '?'}'.\n"
                f"{candidatos_str}\n"
                f"Resolver: dê 2× clique nesta linha e preencha 'Função (código)' "
                f"com o código da coluna `codigo` da planilha CBO. "
                f"Caso o cargo não esteja na planilha, adicione uma linha com "
                f"X em `usar` e clique '🔄 Recarregar planilha CBO'."
            )
            erro_msg = f"Função não encontrada nos X-marcados da planilha CBO ({fmsg})"

        return _resultado_pendencia_interna(
            indice=indice, nome=nome, razao=razao, cnpj=cnpj,
            erro=erro_msg,
            diagnostico_dp=diag,
            bloco=bloco,
            msg_id=msg_id, metadados=metadados,
            empresa_id=empresa_id, depto_id=depto_id,
            # funcao_id deixa None → placeholder "1"
        )
    if funcao_id:
        log.info(f"      💼 Função: {funcao_id} ({conf:.0%})")

    # Payload final + sanitização
    payload = finalizar_payload(bloco, empresa_id, depto_id, funcao_id)

    # Enriquecimento via APIs externas (ViaCEP, CPF lookup, defaults fixos).
    # Roda APÓS o Claude porque os inputs (cep, cpf) saem dos documentos.
    # Não quebra o fluxo se as APIs falharem — degrada graciosamente.
    # Skip se pre-enrichment já rodou (v2.4.1) — evita log duplicado.
    if not bloco.get("_pre_enriched"):
        payload = enrich_candidato(payload)
        enrich_meta = payload.pop("_enrich_meta", {})
        if enrich_meta.get("fields_filled_by_cep") or enrich_meta.get("fields_filled_by_cpf"):
            log.info(
                f"      ✨ Enrichment: +{len(enrich_meta.get('fields_filled_by_cep', []))} "
                f"via CEP, +{len(enrich_meta.get('fields_filled_by_cpf', []))} via CPF "
                f"({len(enrich_meta.get('fields_still_missing', []))} dependem de OCR)"
            )
    else:
        enrich_meta = {}

    # ─── Sanity checks pós-Claude (v2.4.0) ───────────────────────────
    # Pega erros de troca de dígito (0/6, 3/8 etc.) que o Claude não viu.
    # Não bloqueia o POST — só log warning visível pro operador investigar.
    _attrs = (payload.get("data") or {}).get("attributes") or {}
    if enrich_meta.get("cep_suspeito"):
        cs = enrich_meta["cep_suspeito"]
        log.warning(
            f"      🔍 SANITY CHECK falhou — CEP '{cs['cep']}': Claude leu "
            f"cidade='{cs['cidade_claude']}' mas ViaCEP aponta '{cs['cidade_viacep']}'. "
            f"Provável troca de dígito. Endereço NÃO foi sobrescrito — DP confere."
        )
    try:
        nasc = _attrs.get("nascimento")
        adm = _attrs.get("admissao")
        if nasc and adm:
            d_nasc = datetime.strptime(nasc, "%Y-%m-%d")
            d_adm = datetime.strptime(adm, "%Y-%m-%d")
            idade = (d_adm - d_nasc).days // 365
            if idade < 14 or idade > 75:
                log.warning(
                    f"      🔍 SANITY CHECK: idade na admissão = {idade} anos "
                    f"(nascimento {nasc}, admissão {adm}). Faixa esperada: 14-75. "
                    f"Provável erro de leitura em uma das datas."
                )
        # Datas de documentos devem ser posteriores ao nascimento
        if nasc:
            d_nasc = datetime.strptime(nasc, "%Y-%m-%d")
            for campo_doc in ("datactps", "dataidentidade", "datapis"):
                v = _attrs.get(campo_doc)
                if not v:
                    continue
                try:
                    d_doc = datetime.strptime(v, "%Y-%m-%d")
                    if d_doc < d_nasc:
                        log.warning(
                            f"      🔍 SANITY CHECK: {campo_doc}={v} é ANTES do "
                            f"nascimento={nasc}. Impossível — erro de leitura."
                        )
                except ValueError:
                    pass
    except (ValueError, TypeError):
        pass  # datas mal-formadas não devem quebrar o pipeline

    # Regra de negócio da data de admissão (default = ASO + 1 dia,
    # rejeita se < hoje). Pendência aqui é DO CLIENTE — ele que tem
    # que confirmar/atualizar a data.
    #
    # Hora do email vira parâmetro pra regra: se admissão == hoje, desloca
    # +1 dia (email até 15h) ou +2 dias (após 15h). Cai em FDS → terça.
    data_email_dt = None
    try:
        from email.utils import parsedate_to_datetime
        raw_data = metadados.get("data") or metadados.get("date") or ""
        if raw_data:
            data_email_dt = parsedate_to_datetime(raw_data)
    except Exception:
        data_email_dt = None

    # v2.15.5: `sempre_mandar_sem_data_admissao` mudou de comportamento.
    # Antes era ABSOLUTO (removia mesmo se Claude identificou); agora é
    # CONDICIONAL: tenta aplicar a regra normal (ASO+1, deslocamento, etc.);
    # SE não conseguiu (ex: Claude não extraiu nem ASO), omite a data
    # ao invés de virar pendência cliente. DP completa no Desktop.
    payload, erro_data = aplicar_regra_data_admissao(
        payload,
        usar_atual_se_invalida=config.usar_data_atual_se_invalida,
        data_email_dt=data_email_dt,
    )
    if erro_data and config.sempre_mandar_sem_data_admissao:
        attrs_payload = (payload.get("data") or {}).get("attributes") or {}
        attrs_payload.pop("admissao", None)
        attrs_payload.pop("dataterminocontrato", None)
        log.info(
            f"      📅 Sem data identificada — mandando vazio "
            f"(flag 'sem data se não identificada' ligada). DP completa no Desktop."
        )
        erro_data = None
    if erro_data:
        log.warning(f"      📅 Data de admissão inválida ({nome}): {erro_data}")
        _raise_pendencia(
            f"Data de admissão inválida pra {nome}: {erro_data}",
            motivo_cliente=erro_data,
            payload_parcial=payload,
        )

    # Snapshot pra auditoria
    resolucao = {
        "indice": indice, "total": total, "nome": nome,
        "cnpj_empresa": cnpj, "empresa_id": empresa_id,
        "razao_social": razao,
        "departamento_id": depto_id, "departamento_motivo": depto_msg,
        "departamento_sugerido": dados.get("departamento_sugerido"),
        "funcao_id": funcao_id, "funcao_confianca": round(conf, 4),
        "cargo_extraido": dados.get("cargo"),
        "cbo_extraido": dados.get("cbo"),
        "enrich_meta": enrich_meta,
    }

    # v2.16.18: rede de segurança ANTES da validação — infere relationships
    # ausentes que dá pra derivar deterministicamente de outros campos.
    # Caso real (JOHN LENNON): Claude extraiu municipionascimento="SITIO
    # NOVO DO TOCANTINS" mas esqueceu a rel `naturalidade` → infere TO=27.
    try:
        from inferir_lookups import inferir_relationships_ausentes
        novas_rels = inferir_relationships_ausentes(payload)
        if novas_rels:
            payload.setdefault("data", {}).setdefault("relationships", {}).update(novas_rels)
    except Exception as e:
        log.warning(
            f"      ⚠ Falha inferindo relationships: {type(e).__name__}: {e}"
        )

    # Validação determinística
    # v2.16.5: quando REGRA 0 atua (empresa sem deptos + flag ON), o
    # finalizar_payload omite intencionalmente a relationship `departamento`.
    # O validador precisa saber disso pra não marcar como "faltando".
    # v2.16.16: extende a mesma lógica pra `funcao` e `admissao` — as flags
    # `sempre_mandar_sem_funcao` e `sempre_mandar_sem_data_admissao` já
    # passam por AUTO_RESOLVIVEIS no _pendente do Claude, mas se o campo
    # cai como vazio no payload final (sem motivo_codigo), o validador
    # marcava como pendente cliente. Agora respeita as flags aqui também.
    from departamento import SEM_DEPARTAMENTO
    ignorar_rels = set()
    ignorar_attrs = set()
    if depto_id == SEM_DEPARTAMENTO:
        ignorar_rels.add("departamento")
    # Função vazia + flag ON → omite validação (finalizar_payload já omitiu rel)
    if (not funcao_id or not str(funcao_id).strip()) and getattr(
        config, "sempre_mandar_sem_funcao", False
    ):
        ignorar_rels.add("funcao")
    # Data admissão ausente + flag ON → omite validação do attribute
    _attrs_pay = (payload.get("data") or {}).get("attributes") or {}
    if (not _attrs_pay.get("admissao")) and (
        getattr(config, "sempre_mandar_sem_data_admissao", False)
        or getattr(config, "usar_data_atual_se_invalida", False)
    ):
        # Quando data não vem, dataterminocontrato também não — DP digita ambas
        ignorar_attrs.add("admissao")
        # dataterminocontrato não é obrigatório no validador (não está em
        # ATTRS_OBRIGATORIOS), mas se virar, ignora aqui também.
    faltando = validar_campos_obrigatorios(
        payload,
        ignorar_rels=ignorar_rels or None,
        ignorar_attrs=ignorar_attrs or None,
    )
    if faltando:
        log.error(f"      ⛔ Campos faltando ({nome}): {faltando}")
        salvar_payload(
            msg_id, metadados, payload, resolucao=resolucao,
            resultado={
                "status": "pendente_validacao", "candidato_id": None,
                "erro": "Campos obrigatórios faltando",
                "campos_faltando": faltando,
            },
        )
        _raise_pendencia(
            f"Campos faltando para {nome}: {', '.join(faltando)}",
            motivo_cliente="",  # ignorado quando campos_faltando está populado
            campos_faltando=faltando,
            payload_parcial=payload,
        )

    arq_payload = salvar_payload(msg_id, metadados, payload, resolucao=resolucao)

    if config.dry_run:
        log.info(f"      DRY-RUN — pulando POST de {nome}")
        salvar_payload(
            msg_id, metadados, payload, resolucao=resolucao,
            resultado={"status": "dry_run", "candidato_id": None, "erro": None},
        )
        return {
            "_pendencia_anterior_msg_id": bloco.get("_pendencia_anterior_msg_id"),
            "indice": indice, "nome": nome, "ok": True,
            "candidato_id": None, "dry_run": True,
            "razao_social": razao, "cnpj_empresa": cnpj,
            "empresa_id": empresa_id, "departamento_id": depto_id, "funcao_id": funcao_id,
            "payload_path": str(arq_payload.name),
        }

    # v2.14.1: POST passa pelo wrapper único — idempotência (cobra duplicata
    # ANTES), registra_post DEPOIS, grava resultado no JSON, log NDJSON.
    # Sem isso, reprocesso de email multi-pessoa criava candidato duplicado
    # mesmo com a v2.14.0 (UI tinha guarda, orquestrador não).
    attrs_para_post = (payload.get("data") or {}).get("attributes") or {}
    cpf_para_post = attrs_para_post.get("cpf")
    res = postar_candidato_registrado(
        api, payload,
        cpf=cpf_para_post, cnpj=cnpj, nome=nome,
        origem="orquestrador", msg_id=msg_id,
        payload_path=arq_payload,
        # gmail/labels não passados aqui: _finalizar_lote decide o label do
        # email inteiro (pode ter outras pendências). Wrapper só registra.
    )
    # v2.16.20: divergência de nome na duplicata (caso JOYCE/Thaynara) — o
    # post_admissao bloqueou o PULO pra evitar marcar admissão errada como
    # sucesso. Routeamos como pendência cliente pra ele revisar/confirmar.
    if not res.ok and getattr(res, "motivo_cliente", None):
        log.warning(
            f"      ⚠ {nome}: divergência de nome com candidato existente — "
            f"gerando pendência cliente pra revisão"
        )
        salvar_payload(
            msg_id, metadados, payload, resolucao=resolucao,
            resultado={
                "status": "divergencia_nome", "candidato_id": None,
                "erro": res.erro_tecnico or "Divergência de nome",
            },
        )
        return {
            "_pendencia_anterior_msg_id": bloco.get("_pendencia_anterior_msg_id"),
            "indice": indice, "nome": nome, "ok": False,
            "candidato_id": None,
            "motivo_cliente": res.motivo_cliente,
            "erro_tecnico": res.erro_tecnico,
            "campos_faltando": [],
            "razao_social": razao, "cnpj_empresa": cnpj,
            "empresa_id": empresa_id, "departamento_id": depto_id,
            "funcao_id": funcao_id,
            "payload_path": str(arq_payload.name),
        }

    if res.pulou:
        log.info(
            f"      ⏭ {nome}: já cadastrado (candidato {res.candidato_id}) — POST pulado"
        )
        return {
            "_pendencia_anterior_msg_id": bloco.get("_pendencia_anterior_msg_id"),
            "indice": indice, "nome": nome, "ok": True,
            "candidato_id": res.candidato_id,
            "razao_social": razao, "cnpj_empresa": cnpj,
            "empresa_id": empresa_id, "departamento_id": depto_id,
            "funcao_id": funcao_id,
            "payload_path": str(arq_payload.name),
            "procedencia_extra": "já cadastrado — POST pulado (idempotência)",
        }
    if res.ok:
        candidato_id = res.candidato_id
        log.info(f"      ✅ Candidato {candidato_id} criado pra {nome}")
        # Auto-aprendizado de remetente → CNPJ (v2.10.0).
        # Em admissão bem-sucedida, salva o mapeamento `remetente → cnpj` SE ainda
        # não existir. Próximos emails do mesmo remetente caem na estratégia 1
        # (alias exato) automaticamente. Não sobrescreve aliases manuais existentes.
        try:
            from remetente_empresa import (
                consultar_alias_exato as _consultar_alias,
                salvar_alias as _salvar_alias,
            )
            remetente_email = (metadados or {}).get("remetente", "") or ""
            if remetente_email and cnpj and not _consultar_alias(remetente_email):
                _salvar_alias(
                    remetente=remetente_email,
                    cnpj=cnpj,
                    razao_social=razao or "",
                    fonte="auto",
                )
        except Exception as e:
            log.debug(f"   auto-aprendizado de remetente pulado: {type(e).__name__}: {e}")
        # Auto-aprendizado de salário padrão por CNPJ + cargo (v2.12.0).
        # Em admissão bem-sucedida com salário explícito (não veio do próprio
        # cadastro padrão), salva pra essa combinação CNPJ + cargo. Próximas
        # admissões com "salário base" sem valor do mesmo cliente sobem auto.
        try:
            from salarios_padrao import (
                salvar as _salvar_sal_padrao,
                consultar as _consultar_sal_padrao,
            )
            attrs_final = (payload.get("data") or {}).get("attributes") or {}
            salario_final = attrs_final.get("salario")
            cargo_final = attrs_final.get("nomecargo") or dados.get("cargo") or ""
            if (
                salario_final
                and float(salario_final) > 0
                and cnpj
                and cargo_final
            ):
                # Não sobrescrever cadastros manuais com valor diferente
                # (operador pode ter setado exceção). Só auto-aprende se não
                # existe OU se o valor é igual.
                existente = _consultar_sal_padrao(cnpj, cargo_final)
                if existente is None or existente.get("fonte") == "auto":
                    _salvar_sal_padrao(
                        cnpj=cnpj,
                        cargo=cargo_final,
                        salario=float(salario_final),
                        razao_social=razao or "",
                        fonte="auto",
                    )
        except Exception as e:
            log.debug(f"   auto-aprendizado de salário padrão pulado: {type(e).__name__}: {e}")

        # v2.16.46: descarta rascunhos de resposta ainda pendentes deste msg_id.
        # Cobre caso em que pipeline reprocessa email e resolve pendencia sem
        # operador tocar — rascunho antigo (perguntando dado ao cliente) fica
        # sem sentido depois do POST OK.
        try:
            import rascunhos_resposta as _rr
            _n = _rr.descartar_por_msg_id(
                msg_id, operador="auto-orquestrador",
                motivo=f"admissao resolvida (candidato {candidato_id})",
            )
            if _n:
                log.info(f"   descartou {_n} rascunho(s) pendente(s) do msg_id {msg_id[:16]}")
        except Exception as e:
            log.debug(f"   auto-descarte rascunhos pulado: {type(e).__name__}: {e}")

        return {
            "_pendencia_anterior_msg_id": bloco.get("_pendencia_anterior_msg_id"),
            "indice": indice, "nome": nome, "ok": True,
            "candidato_id": candidato_id,
            "razao_social": razao, "cnpj_empresa": cnpj,
            "empresa_id": empresa_id, "departamento_id": depto_id, "funcao_id": funcao_id,
            "payload_path": str(arq_payload.name),
        }

    # POST falhou (HTTP 4xx/5xx ou exception) — v2.14.1 (ITEM 8):
    # ISSO É FALHA TÉCNICA DO ESCRITÓRIO, NÃO PENDÊNCIA DE CLIENTE.
    # Bug real: 3 HTTP 422 saíram como "Pendente cliente — Falha técnica..."
    # na auditoria 12/06; se auto_email_pendencias ligar, cliente recebia
    # email cobrando informação por erro nosso.
    # Marcamos `falha_tecnica=True` (não `interno` — pendência interna é
    # categoria diferente: problema de cadastro NOSSO no eContador, não
    # erro HTTP). `_procedencia_de` renderiza como "Falha técnica — ...".
    # SEM motivo_cliente → reply do thread NÃO cita o erro pro cliente.
    log.error(
        f"      ❌ POST falhou pra {nome}: {res.erro_ref} "
        f"(HTTP {res.status_http or '?'})\n{res.body_err[:200]}"
    )
    # salvar_payload com resultado: o wrapper já gravou via _atualizar_payload_em_disco,
    # mas chamamos de novo aqui pra anexar o `resolucao` completo (o wrapper só
    # gravou o campo `resultado`).
    salvar_payload(
        msg_id, metadados, payload, resolucao=resolucao,
        resultado={"status": "falha_post", "candidato_id": None,
                   "erro": res.erro_ref, "body": res.body_err[:2000],
                   "origem": "orquestrador"},
    )
    return {
        "_pendencia_anterior_msg_id": bloco.get("_pendencia_anterior_msg_id"),
        "indice": indice, "nome": nome, "ok": False,
        "candidato_id": None,
        "falha_tecnica": True,  # ← v2.14.1: NÃO é pendência cliente nem interna
        "status_http": res.status_http,
        "erro_tecnico": f"{res.erro_ref}: {res.body_err[:200]}",
        # motivo_cliente intencionalmente OMITIDO — _finalizar_lote não cita
        # falha técnica pro cliente (problema é nosso).
        "campos_faltando": [],
        "razao_social": razao, "cnpj_empresa": cnpj,
        "payload_path": str(arq_payload.name),
    }


def _localizar_bloco_correspondente(blocos: list[dict], attrs_alvo: dict) -> dict | None:
    """Acha o bloco em `blocos` cujo CPF ou nome bate com `attrs_alvo`."""
    cpf_alvo = str(attrs_alvo.get("cpf", "")).strip() or None
    nome_alvo = (attrs_alvo.get("nome") or "").strip().upper() or None
    for b in blocos:
        a = (b.get("data") or {}).get("attributes") or {}
        if cpf_alvo and str(a.get("cpf", "")).strip() == cpf_alvo:
            return b
        if nome_alvo and (a.get("nome") or "").strip().upper() == nome_alvo:
            return b
    # Fallback: se só tem 1 bloco, retorna ele
    return blocos[0] if len(blocos) == 1 else None


def _finalizar_lote(
    resultados: list[dict],
    msg_id: str,
    msg_pra_resposta: dict,
    ids_label_pendente_remover: list[str],
    gmail: GmailClient,
    config: Config,
) -> None:
    """Após processar N admissões: decide label, manda reply consolidado pro
    cliente (se aplicável), registra TODOS na planilha de admissões, loga.

    Não envia email separado pro DP — todas as informações de status (sucesso,
    pendente cliente, pendente interno, falha técnica) ficam consolidadas em
    admissoes.xlsx.
    """
    if not resultados:
        log.warning("   _finalizar_lote chamado sem resultados")
        return

    # v2.15.15: fecha pendências irmãs detectadas pelo merge no _processar_um_bloco.
    # Pra cada bloco que mesclou com um email anterior do mesmo candidato:
    #   - remove a label `pendente` do email anterior no Gmail
    #   - aplica a label `processado` (esse email "morreu", agora segue no atual)
    #   - registra na planilha que foi mesclado
    irmas_pra_fechar = [
        r.get("_pendencia_anterior_msg_id") for r in resultados
        if r.get("_pendencia_anterior_msg_id")
    ]
    for mid_irma in set(irmas_pra_fechar):
        try:
            gmail.remover_label(mid_irma, config.label_pendente)
        except Exception as e:
            log.debug(f"   nada de pendente em {mid_irma[:16]}: {e}")
        try:
            gmail.aplicar_label(mid_irma, config.label_processado)
        except Exception as e:
            log.warning(f"   Falha aplicando processado em irmã {mid_irma}: {e}")
        try:
            # Acha qual resultado mesclou com essa irmã e usa nome/empresa dele
            r_assoc = next((r for r in resultados
                            if r.get("_pendencia_anterior_msg_id") == mid_irma), None)
            nome_assoc = (r_assoc or {}).get("nome", "?")
            registrar_admissao_planilha(
                nome=nome_assoc, empresa=(r_assoc or {}).get("razao_social", ""),
                cnpj=(r_assoc or {}).get("cnpj_empresa", ""),
                procedencia=(
                    f"Mesclada com email mais recente ({msg_id[:16]}) — "
                    f"informações somadas e enviadas via lá"
                ),
                msg_id=mid_irma,
            )
            log.info(f"   🔀 Fechou pendência irmã {mid_irma[:16]} (mesclada com {msg_id[:16]})")
        except Exception as e:
            log.warning(f"   Falha registrando mesclagem na planilha: {e}")

    n_ok = sum(1 for r in resultados if r["ok"])
    n_total = len(resultados)
    sucessos = [r for r in resultados if r["ok"]]
    # v2.14.1: 3 categorias distintas de falha
    falhas_tecnicas = [r for r in resultados if not r["ok"] and r.get("falha_tecnica")]
    falhas_internas = [r for r in resultados if not r["ok"] and r.get("interno") and not r.get("falha_tecnica")]
    falhas_cliente = [
        r for r in resultados
        if not r["ok"] and not r.get("interno") and not r.get("falha_tecnica")
    ]
    log.info(
        f"   📊 Lote: {n_ok}/{n_total} OK | "
        f"{len(falhas_cliente)} cliente | "
        f"{len(falhas_internas)} interno | "
        f"{len(falhas_tecnicas)} técnica"
    )

    todos_ok = (n_ok == n_total)

    if todos_ok:
        # v2.15.7: SEMPRE remove label pendente do msg_id atual quando dá
        # sucesso — não só dos ids_label_pendente_remover (que cobrem só
        # threads de retentativa). Sem isso, um email que foi pendente
        # NUMA passada e depois processou COM SUCESSO em outra passada
        # ficava com AS DUAS labels (caso TATIANE 15/06/2026): "pendente"
        # da 1ª passada + "processado" da 2ª. O Gmail não substitui labels
        # ao aplicar, então precisa remover explicitamente.
        try:
            gmail.remover_label(msg_id, config.label_pendente)
        except Exception as e:
            log.debug(f"   nada de pendente em {msg_id[:16]}: {e}")
        # Limpa pendente de msgs antigas também
        for mid in ids_label_pendente_remover:
            try:
                gmail.remover_label(mid, config.label_pendente)
            except Exception as e:
                log.warning(f"   Falha removendo pendente de {mid}: {e}")
        gmail.aplicar_label(msg_id, config.label_processado)
    else:
        gmail.aplicar_label(msg_id, config.label_pendente)

        # Reply no thread — só se há algo a comunicar ao cliente. Inclui
        # sucessos, falhas_cliente e menciona pendências internas suavemente.
        # v2.16.19: 3 modos (auto_email_pendencias_modo):
        #   "desligado" → nada acontece (DP responde manualmente)
        #   "rascunho"  → grava na fila /respostas pra analista revisar
        #   "direto"    → envia automaticamente (comportamento antigo)
        modo_email = getattr(config, "auto_email_pendencias_modo", "desligado")
        if (
            msg_pra_resposta
            and modo_email in ("rascunho", "direto")
            and (sucessos or falhas_cliente or falhas_internas)
        ):
            try:
                # v2.16.26: extrai From cedo pra passar nome do destinatário
                # pro template (saudação personalizada "Oi Polyana, tudo bem?")
                meta = msg_pra_resposta if isinstance(msg_pra_resposta, dict) else {}
                from_, subject = _extrair_from_subject(meta)
                rem_email, rem_nome = _separar_email_nome(from_)
                # v2.16.28: flags ativas → filtros do template (não cobrar
                # do cliente coisa que pipeline já resolve sozinho)
                flags_auto = set()
                if getattr(config, "sempre_mandar_sem_data_admissao", False) \
                        or getattr(config, "usar_data_atual_se_invalida", False):
                    flags_auto.add("sempre_mandar_sem_data_admissao")
                if getattr(config, "sempre_mandar_sem_funcao", False):
                    flags_auto.add("sempre_mandar_sem_funcao")
                corpo = _corpo_reply_lote(
                    resultados,
                    destinatario_nome=rem_nome,
                    flags_auto_resolvido=flags_auto,
                )
                if modo_email == "rascunho":
                    # Grava rascunho pra revisão humana — NÃO envia
                    try:
                        import rascunhos_resposta as _rasc
                        contexto = {
                            "n_sucessos": len(sucessos),
                            "n_falhas_cliente": len(falhas_cliente),
                            "n_falhas_internas": len(falhas_internas),
                            "nomes_sucesso": [r["nome"] for r in sucessos],
                            "nomes_pendente": [r["nome"] for r in falhas_cliente],
                            "razao_social": next(
                                (r.get("razao_social") for r in resultados if r.get("razao_social")),
                                "",
                            ),
                        }
                        _rasc.criar_rascunho(
                            msg_id=msg_id,
                            thread_id=meta.get("threadId") or msg_id,
                            remetente_email=rem_email,
                            remetente_nome=rem_nome,
                            assunto=subject,
                            corpo_proposto=corpo,
                            contexto=contexto,
                        )
                        log.info(
                            f"   📝 Rascunho gravado pra revisão na fila /respostas "
                            f"({len(sucessos)} OK + {len(falhas_cliente)} cliente "
                            f"+ {len(falhas_internas)} interna)"
                        )
                    except Exception as e:
                        log.warning(
                            f"   ⚠ Falha gravando rascunho: {type(e).__name__}: {e}"
                        )
                elif config.confirmar_replies and not _confirmar_envio(
                    msg_pra_resposta, corpo, config.email_dp or None
                ):
                    log.warning("   ✋ Envio cancelado pelo usuário (--ask)")
                else:
                    # v2.16.32: HARD FREIO — modo "direto" exige arquivo sentinela
                    # explícito no disco. Incidente Gabriel/Esther (2026-06-22):
                    # sistema mandou email automático com texto ruim sem o DP
                    # saber. Agora: pra enviar direto, precisa criar arquivo
                    # 'PERMITIR_ENVIO_DIRETO.flag' manualmente. Sem ele, cai
                    # silenciosamente pro modo rascunho.
                    sentinela_path = Path(__file__).parent / "PERMITIR_ENVIO_DIRETO.flag"
                    if not sentinela_path.exists():
                        log.warning(
                            f"   🛑 HARD FREIO: modo 'direto' configurado mas "
                            f"arquivo PERMITIR_ENVIO_DIRETO.flag NÃO EXISTE. "
                            f"NÃO vou enviar — gravando como rascunho pra "
                            f"revisão em /respostas."
                        )
                        try:
                            import rascunhos_resposta as _rasc
                            _rasc.criar_rascunho(
                                msg_id=msg_id,
                                thread_id=meta.get("threadId") or msg_id,
                                remetente_email=rem_email,
                                remetente_nome=rem_nome,
                                assunto=subject,
                                corpo_proposto=corpo,
                                contexto={
                                    "n_sucessos": len(sucessos),
                                    "n_falhas_cliente": len(falhas_cliente),
                                    "n_falhas_internas": len(falhas_internas),
                                    "nomes_pendente": [r["nome"] for r in falhas_cliente],
                                    "razao_social": next(
                                        (r.get("razao_social") for r in resultados
                                         if r.get("razao_social")), ""),
                                    "fallback_de_modo_direto": True,
                                },
                            )
                        except Exception as e:
                            log.warning(f"   ⚠ Falha gravando rascunho de fallback: {e}")
                    else:
                        try:
                            gmail.responder_no_thread(
                                msg_pra_resposta, corpo=corpo,
                                cc=config.email_dp or None,
                            )
                            log.info(
                                f"   📨 Reply DIRETO enviado "
                                f"({len(sucessos)} OK + {len(falhas_cliente)} cliente "
                                f"+ {len(falhas_internas)} interna)"
                            )
                            # v2.16.32: AUDIT obrigatório
                            try:
                                import auditoria_emails as _ae
                                _ae.registrar_envio(
                                    msg_id_original=msg_id,
                                    thread_id=meta.get("threadId") or msg_id,
                                    destinatario_email=rem_email,
                                    destinatario_nome=rem_nome,
                                    assunto=subject,
                                    corpo=corpo,
                                    cc=config.email_dp or None,
                                    origem="pipeline_direto",
                                    operador="<automático>",
                                    sucesso=True,
                                    contexto={
                                        "nomes_pendente": [r["nome"] for r in falhas_cliente],
                                    },
                                )
                            except Exception as e:
                                log.warning(f"   ⚠ Audit falhou: {e}")
                        except Exception as e:
                            log.exception("Reply DIRETO falhou — registrando como erro")
                            try:
                                import auditoria_emails as _ae
                                _ae.registrar_envio(
                                    msg_id_original=msg_id,
                                    thread_id=meta.get("threadId") or msg_id,
                                    destinatario_email=rem_email,
                                    destinatario_nome=rem_nome,
                                    assunto=subject, corpo=corpo,
                                    origem="pipeline_direto",
                                    sucesso=False, erro=str(e)[:200],
                                )
                            except Exception:
                                pass
                            raise
            except Exception:
                log.exception("Falha enviando reply consolidado")

    # Planilha de admissões — 1 linha por funcionário, sucesso ou pendência.
    # msg_id permite reprocessar pela UI e abrir o thread no Gmail.
    for r in resultados:
        registrar_admissao_planilha(
            nome=r.get("nome"),
            empresa=r.get("razao_social"),
            cnpj=r.get("cnpj_empresa"),
            procedencia=_procedencia_de(r),
            msg_id=msg_id,
        )
    log.info(f"   📊 Planilha admissoes.xlsx atualizada (+{len(resultados)} linha(s))")

    # v2.16.0: re-consolida perfis de remetente após o lote.
    # Garante que o perfil reflete o evento que acabou de acontecer (sucesso
    # ou pendência), pra próxima admissão do mesmo cliente já se beneficiar
    # do contexto atualizado no prompt do Claude. Custo: O(N) em payloads
    # locais, sub-segundo até centenas de milhares.
    try:
        import perfis_remetente as _pr
        _pr.consolidar_todos()
    except Exception as e:
        log.warning(f"   ⚠ Falha re-consolidando perfis: {type(e).__name__}: {e}")

    # v2.16.1: salva fingerprint da tentativa pra guarda do polling
    # automático. Em sucesso, a msg recebe label `processado` e nunca mais
    # é retentada — mas salvar o fp não atrapalha. Em pendência, isso é
    # o que impede o reprocesso cego em cada passada.
    # Pega o thread_msgs_count do gmail pra detectar quando o cliente
    # responder (gatilho pra reprocessar automático).
    try:
        from idempotencia import salvar_fingerprint_reprocesso
        n_msgs_thread = None
        try:
            threadId = (
                msg_pra_resposta.get("threadId") if msg_pra_resposta else None
            ) or msg_id
            _t = gmail.obter_thread(threadId)
            n_msgs_thread = len(_t.get("messages") or [])
        except Exception:
            pass
        salvar_fingerprint_reprocesso(msg_id, n_msgs_thread)
    except Exception as e:
        log.warning(f"   ⚠ Falha salvando fingerprint: {type(e).__name__}: {e}")

    # v2.16.20: delimitador visual claro de fim de email no log
    n_ok = sum(1 for r in resultados if r.get("ok"))
    n_pend = len(resultados) - n_ok
    icone_final = "✅" if n_pend == 0 else "⚠" if n_ok else "❌"
    log.info("┏" + "━" * 78)
    log.info(
        f"┃ {icone_final} FIM email {msg_id[:16]} — "
        f"{n_ok} processada(s) · {n_pend} pendente(s)"
    )
    log.info("┗" + "━" * 78)

    # NDJSON técnico (separado da planilha, pra audit/debug interno)
    for r in resultados:
        log_jsonl({
            "msg_id": msg_id,
            "status": (
                "sucesso" if r["ok"]
                else (
                    "pendente_interno" if r.get("interno")
                    else ("pendente_validacao" if r.get("campos_faltando") else "erro")
                )
            ),
            "indice": r.get("indice"),
            "nome": r.get("nome"),
            "candidato_id": r.get("candidato_id"),
            "erro": r.get("erro_tecnico"),
            "_validacao_bloqueada": r.get("campos_faltando", []),
        })


def _primeiro_nome_destinatario(nome_completo: str) -> str:
    """v2.16.26: pega só o primeiro nome de uma string como 'Polyana Sousa'.
    Filtra prefixos comuns que não são nome (Sra., Sr., Dr., RH, DP, etc.)."""
    if not nome_completo:
        return ""
    # Versão sem ponto pra match com .strip(".,;:") aplicado no token
    PREFIXOS_IGNORAR = {"sr", "sra", "dr", "dra", "rh", "dp",
                         "depto", "departamento"}
    partes = nome_completo.strip().split()
    for p in partes:
        p_low = p.lower().strip(".,;:")
        if p_low and p_low not in PREFIXOS_IGNORAR and len(p_low) >= 2:
            return p.strip(".,;:").capitalize()
    return ""


def _humanizar_nome(nome: str) -> str:
    """JOAO BATISTA → João Batista (capitaliza, mantém preposições minúsculas)."""
    if not nome:
        return ""
    PREP_MIN = {"DE", "DA", "DO", "DAS", "DOS", "E"}
    out = []
    for w in nome.split():
        if w.upper() in PREP_MIN:
            out.append(w.lower())
        else:
            out.append(w.capitalize())
    return " ".join(out)


def _parsear_lista_faltam(texto: str) -> list[str]:
    """v2.16.26: o Claude às vezes manda motivo_cliente como
    'Termo aditivo de estágio com dados incompletos. Faltam: (1) DATA DE
    NASCIMENTO — não consta; (2) NOME DA MÃE — não consta; ...'
    Esta função extrai os itens (1), (2), (3)... e devolve só os nomes
    de campo limpos pra virar bullets. Retorna [] se não casar o padrão.
    """
    import re
    if not texto or "faltam" not in texto.lower():
        return []
    # Pega tudo depois de "Faltam:" (case-insensitive)
    m = re.search(r"falta(?:m|ndo|riam)?\s*:?\s*(.+?)(?:\.|\bObs?\.|$)",
                  texto, re.IGNORECASE | re.DOTALL)
    if not m:
        return []
    resto = m.group(1)
    # Quebra por marcadores: (1) (2) (3) ou 1) 2) ou ; ou • etc.
    itens = re.split(r"\s*[\(\[]?\d{1,2}[\)\]]\s*|\s*;\s*|\s*•\s*", resto)
    out = []
    for it in itens:
        if not it or not it.strip():
            continue
        # Pega só até primeiro ' — ' ou ' - ' ou ',' (corta a descrição)
        it = re.split(r"\s+[—–-]\s+|\bn[ãa]o\s+consta", it, maxsplit=1)[0]
        it = it.strip(" ,.;:()").strip()
        # Limpa "do estagiário", "do funcionário" etc no fim
        it = re.sub(r"\b(do|da|de)\s+(estagi[áa]rio|funcion[áa]rio|colaborador|candidato)$",
                    "", it, flags=re.IGNORECASE).strip()
        if it and len(it) > 2 and len(it) < 80:
            # Sentence case quando tudo MAIÚSCULO ou se primeira palavra é caps.
            # Preserva acrônimos curtos (CEP, CPF, RG, PIS, CNPJ, CNH, CTPS, UF).
            ACRONIMOS = {"CEP", "CPF", "RG", "PIS", "CNPJ", "CNH", "CTPS",
                          "UF", "PIS/PASEP", "DV", "ASO"}
            palavras = it.split()
            if palavras and palavras[0].isupper() and palavras[0] not in ACRONIMOS:
                # SEXO e ESTADO CIVIL ausentes → Sexo e estado civil ausentes
                novas = []
                for i, w in enumerate(palavras):
                    wu = w.upper().strip(".,;:")
                    if wu in ACRONIMOS:
                        novas.append(w.upper())
                    elif i == 0:
                        novas.append(w.capitalize())
                    elif w.isupper():
                        novas.append(w.lower())
                    else:
                        novas.append(w)
                it = " ".join(novas)
            out.append(it)
    return out


def _filtrar_motivos_auto_resolvidos(
    texto: str,
    flags_ativas: set[str],
) -> str:
    """v2.16.28: remove menções a campos que o pipeline já resolve sozinho
    (não precisam virar cobrança ao cliente).

    Quando `sempre_mandar_sem_data_admissao` ON, o pipeline envia admissão
    sem data → DP digita no Desktop. Cobrar do cliente nesse caso é ruído.
    Idem pra `sempre_mandar_sem_funcao`.

    Funciona em texto livre (motivo_cliente do Claude) e em bullets curtos.
    Remove a frase/bullet inteiro que cita o campo, conservando o resto.
    """
    if not texto or not flags_ativas:
        return texto or ""
    import re as _re

    PADROES = {
        "data_admissao": [
            r"\bdata\s+(de\s+)?admiss(ão|ao)\b",
            r"\bd(t|a)\.?\s*adm(iss(ão|ao))?\b",
        ],
        "funcao": [
            r"\bfun(ç|c)(ão|ao)\s+do?\s+(funcion[áa]rio|estagi[áa]rio|cargo)\b",
            r"\bcargo\s+do?\s+(funcion[áa]rio|estagi[áa]rio|colaborador)\b",
        ],
    }
    rx_lista = []
    if "sempre_mandar_sem_data_admissao" in flags_ativas:
        rx_lista.extend(PADROES["data_admissao"])
    if "sempre_mandar_sem_funcao" in flags_ativas:
        rx_lista.extend(PADROES["funcao"])
    if not rx_lista:
        return texto

    rx_combinado = "|".join(f"(?:{p})" for p in rx_lista)

    # 1) Remove BULLETS inteiros que mencionem o padrão
    linhas = texto.split("\n")
    linhas_ok = []
    for ln in linhas:
        ln_strip = ln.lstrip(" •-*").strip()
        if ln_strip and ln.startswith((" •", "•", " -", "-", "*", "  •")):
            if _re.search(rx_combinado, ln_strip, _re.IGNORECASE):
                continue
        linhas_ok.append(ln)
    texto2 = "\n".join(linhas_ok)

    # 2) Remove ITENS numerados "(N) DATA DE ADMISSÃO: ..." dentro de motivo
    #    livre. Estratégia: encontra todos os itens "N) ..." (até próximo N)
    #    ou fim), descarta os que matcham o padrão, junta o resto.
    #    Suporta "(N)" e "N)" como separadores.
    # Regex divide o texto antes de cada marcador "(N)" ou "N)"
    partes = _re.split(r"(?<!\d)(\(?\d{1,2}\))\s+", texto2)
    # split com grupo captura: [pre, '(1)', conteudo1, '(2)', conteudo2, ...]
    if len(partes) >= 3:
        pre = partes[0]
        items = list(zip(partes[1::2], partes[2::2]))  # [(marker, content), ...]
        kept = [pre.strip()] if pre.strip() else []
        i = 1
        for _marker, content in items:
            if _re.search(rx_combinado, content, _re.IGNORECASE):
                continue  # descarta esse item inteiro
            # Re-numera sequencialmente
            kept.append(f"{i}) {content.strip()}")
            i += 1
        texto2 = " ".join(kept)

    # 3) Limpa pontuação / espaços duplos sobrando
    texto2 = _re.sub(r"\s{2,}", " ", texto2)
    texto2 = _re.sub(r"\s*;\s*;\s*", "; ", texto2)
    texto2 = _re.sub(r"\s*\n\s*\n\s*\n+", "\n\n", texto2)
    texto2 = texto2.strip()

    return texto2


def _corpo_reply_lote(resultados: list[dict],
                      destinatario_nome: str = "",
                      flags_auto_resolvido: set[str] | None = None) -> str:
    """Reply consolidado no thread.

    v2.16.26: tom humano, saudação personalizada + bullets de campos faltantes.
    v2.16.28: filtra campos que o pipeline resolve sozinho (flags) — não cita
    pro cliente coisas que ele não precisa providenciar.

    3 categorias de funcionário no mesmo email:
      - sucessos: "Já cadastrei X."
      - falhas_cliente: lista bullets de campos que faltam.
      - falhas_internas: "Pra W, estou organizando alguns detalhes" (suave).
    """
    flags_auto_resolvido = flags_auto_resolvido or set()
    # Pra bullets estruturados — termos a remover
    BULLETS_AUTO = set()
    if "sempre_mandar_sem_data_admissao" in flags_auto_resolvido:
        BULLETS_AUTO.update({"data de admissão", "data de admissao",
                              "data de admissao", "data admissao"})
    if "sempre_mandar_sem_funcao" in flags_auto_resolvido:
        BULLETS_AUTO.update({"função", "funcao", "cargo"})
    sucessos = [r for r in resultados if r["ok"]]
    # v2.14.1: NUNCA citar falha técnica pro cliente (erro nosso)
    falhas_cliente = [
        r for r in resultados
        if not r["ok"] and not r.get("interno") and not r.get("falha_tecnica")
    ]
    falhas_internas = [
        r for r in resultados
        if not r["ok"] and (r.get("interno") or r.get("falha_tecnica"))
    ]

    blocos: list[str] = []

    if sucessos:
        nomes = [_humanizar_nome(r["nome"]) for r in sucessos]
        if len(nomes) == 1:
            blocos.append(f"Já cadastrei {nomes[0]}.")
        else:
            blocos.append(f"Já cadastrei: {_lista_natural(nomes)}.")

    for r in falhas_cliente:
        nome = _humanizar_nome(r["nome"])
        # Prioridade: campos_faltando (estruturado) > parser do motivo > texto cru
        bullets: list[str] = []
        if r.get("campos_faltando"):
            bullets = list(r["campos_faltando"])
        elif r.get("motivo_cliente"):
            bullets = _parsear_lista_faltam(r["motivo_cliente"])
        # v2.16.28: filtra bullets de campos que o pipeline já resolve sozinho
        # (flags sempre_mandar_sem_data_admissao / sempre_mandar_sem_funcao)
        if BULLETS_AUTO:
            def _eh_auto(b: str) -> bool:
                b_low = b.lower().strip().rstrip(":").rstrip(".")
                # bate se b começa com OU contém termo auto-resolvido
                return any(termo in b_low for termo in BULLETS_AUTO)
            bullets = [b for b in bullets if not _eh_auto(b)]
        if bullets:
            # Bullet list humana
            linha_abertura = (
                f"Pra cadastrar a admissão{' do' if nome else ''} "
                f"{nome}, ainda preciso de:"
            ).rstrip()
            bullets_fmt = "\n".join(f"  • {b}" for b in bullets)
            blocos.append(f"{linha_abertura}\n{bullets_fmt}")
        elif r.get("motivo_cliente"):
            # Fallback: texto livre do Claude — SUAVIZADO + filtrado
            motivo = r["motivo_cliente"].strip()
            # v2.16.28: remove menções aos campos que o pipeline resolve sozinho
            motivo = _filtrar_motivos_auto_resolvidos(motivo, flags_auto_resolvido)
            # Remove jargões técnicos comuns
            for jargao, substituto in [
                ("dados incompletos para cadastro", "alguns dados que faltam"),
                ("Termo aditivo de estágio", "no aditivo de estágio"),
                ("documento", "doc"),
            ]:
                motivo = motivo.replace(jargao, substituto)
            blocos.append(f"Pra {nome}: {motivo}")
        else:
            blocos.append(
                f"Pra {nome}: não consegui processar — "
                f"pode revisar os documentos?"
            )

    if falhas_internas:
        nomes = [_humanizar_nome(r["nome"]) for r in falhas_internas]
        nomes_str = _lista_natural(nomes)
        if len(nomes) == 1:
            blocos.append(
                f"Pra {nomes_str}, estou organizando alguns detalhes "
                f"aqui do nosso lado — te aviso assim que estiver pronto."
            )
        else:
            blocos.append(
                f"Pra {nomes_str}, estou organizando alguns detalhes "
                f"aqui do nosso lado — te aviso assim que estiverem prontos."
            )

    miolo = "\n\n".join(blocos)

    primeiro_nome = _primeiro_nome_destinatario(destinatario_nome)
    saudacao = f"Oi {primeiro_nome}, tudo bem?" if primeiro_nome else "Olá!"

    pedido_resposta = (
        "Pode me responder esse mesmo e-mail com essas info? Não precisa "
        "reenviar os documentos que já mandou.\n\n"
        if falhas_cliente else ""
    )

    return (
        f"{saudacao}\n\n"
        f"{miolo}\n\n"
        f"{pedido_resposta}"
        f"Qualquer dúvida, é só me chamar.\n\n"
        f"Abraço,\n"
        f"DP — Crosara Contabilidade"
    )


# ============================================================
# Loop principal
# ============================================================

def rodar_uma_passada(config: Config, claude: ClaudeClient, planilha: list[dict]) -> None:
    gmail = GmailClient()
    api = EContadorAPI(config.base_url, config.token)
    # Snapshot do billing antes da passada — usado pra calcular delta no fim
    usage_antes = dict(claude.usage_total)
    n_emails_processados = 0
    try:
        # ---- 1. Emails NOVOS (e PENDENTES se config.reprocessar_pendentes_no_polling)
        # Quando reprocessar_pendentes=True (v2.9.0), também pega emails que já
        # estão com label `pendente` e tenta de novo. A label é removida ANTES
        # de processar pra evitar loop infinito — se a admissão continuar
        # pendente, o fluxo de _processar_seguro reaplica a label no final.
        emails = gmail.buscar_emails_pendentes(
            config.label_entrada, config.label_processado, config.label_pendente,
            incluir_pendentes=config.reprocessar_pendentes_no_polling,
        )
        novos = []
        retentativas = []
        for msg in emails:
            if gmail._msg_tem_label(msg, config.label_pendente):
                retentativas.append(msg)
            else:
                novos.append(msg)
        if novos:
            log.info(f"📥 {len(novos)} email(s) novo(s)")
        if retentativas:
            log.info(f"🔁 {len(retentativas)} email(s) pendente(s) pra retentar")
        n_emails_processados += len(emails)

        # v2.14.1: pausa entre emails (PATCHES.md §5) — espaça as chamadas
        # Claude pra não estourar 30k input-tokens/min do tier 1. Org reportou
        # 3 erros 429 com 2 emails seguidos. Pulamos a pausa antes do PRIMEIRO
        # email (sem mensagem anterior pra esperar).
        pausa_s = max(0, int(getattr(config, "pausa_entre_emails_segundos", 20) or 0))
        for idx_email, msg in enumerate(emails):
            msg_id = msg["id"]
            if pausa_s and idx_email > 0:
                log.info(f"   ⏳ Pausando {pausa_s}s entre emails (anti-429)")
                time.sleep(pausa_s)
            # Retentativa: remove label pendente ANTES de processar (caso contrário,
            # se a admissão der pendente de novo, fica tentando reaplicar label
            # já existente; também evita confusão visual na UI).
            if gmail._msg_tem_label(msg, config.label_pendente):
                # v2.16.1: pula reprocesso quando NADA mudou desde a última
                # tentativa. Evita queimar tokens Claude reprocessando email
                # idêntico em cada passada do polling. Critério de mudança:
                # cliente respondeu OU operador cadastrou alias/salário/etc.
                # Conta as msgs da thread Gmail pra detectar resposta nova.
                try:
                    from idempotencia import deve_pular_reprocesso_auto
                    thread = gmail.obter_thread(msg.get("threadId") or msg_id)
                    n_msgs = len(thread.get("messages") or [])
                    pular, motivo = deve_pular_reprocesso_auto(msg_id, n_msgs)
                    if pular:
                        log.info(
                            f"   ⏭ [{msg_id[:16]}] pulando reprocesso — {motivo}. "
                            f"(Cadastre alias/salário ou aguarde resposta do "
                            f"cliente pra disparar nova tentativa.)"
                        )
                        continue
                    if motivo:
                        log.info(f"   🔁 [{msg_id[:16]}] reprocessando — {motivo}")
                except Exception as e:
                    log.warning(
                        f"   ⚠ Falha checando fingerprint de reprocesso "
                        f"({type(e).__name__}): {e}. Vou reprocessar mesmo assim."
                    )
                try:
                    gmail.remover_label(msg_id, config.label_pendente)
                    log.info(f"   🔁 [{msg_id[:16]}] removendo label pendente pra retentar")
                except Exception:
                    log.exception(f"   ⚠ Falha removendo label pendente de {msg_id[:16]}")
            _processar_seguro(
                lambda m=msg: processar_email(m, gmail, claude, api, planilha, config),
                msg_id=msg_id,
                msg_pra_label=msg_id,
                msg_pra_resposta=msg,
                gmail=gmail, config=config,
            )

        # ---- 2. Threads PENDENTES com resposta do cliente ------------
        threads = gmail.buscar_threads_aguardando_cliente(config.label_pendente)
        if threads:
            log.info(f"🔁 {len(threads)} thread(s) com resposta do cliente")
        n_emails_processados += len(threads)
        for idx_thread, thread in enumerate(threads):
            tid = thread.get("id", "?")
            msgs = thread.get("messages", []) or []
            ref_id = msgs[0]["id"] if msgs else tid
            ultima_msg = msgs[-1] if msgs else None
            # v2.14.1: pausa entre threads (PATCHES.md §5)
            if pausa_s and (idx_thread > 0 or emails):
                log.info(f"   ⏳ Pausando {pausa_s}s antes da próxima thread (anti-429)")
                time.sleep(pausa_s)
            _processar_seguro(
                lambda t=thread: processar_thread_resposta(
                    t, gmail, claude, api, planilha, config
                ),
                msg_id=ref_id,
                msg_pra_label=msgs[-1]["id"] if msgs else ref_id,
                msg_pra_resposta=ultima_msg,
                gmail=gmail, config=config,
            )
    finally:
        api.close()
        _registrar_billing_passada(claude, usage_antes, n_emails_processados)


def _registrar_billing_passada(
    claude: ClaudeClient,
    usage_antes: dict,
    n_emails_processados: int,
) -> None:
    """Calcula delta de billing da passada (tokens + custo) e registra em
    billing.ndjson + log do terminal. Útil pra acompanhar custo da API
    Claude ao longo do tempo (mensal etc.)."""
    u_agora = claude.usage_total
    delta = {
        "n_calls": u_agora["n_calls"] - usage_antes.get("n_calls", 0),
        "input_tokens": u_agora["input_tokens"] - usage_antes.get("input_tokens", 0),
        "output_tokens": u_agora["output_tokens"] - usage_antes.get("output_tokens", 0),
        "cache_creation_input_tokens": (
            u_agora["cache_creation_input_tokens"]
            - usage_antes.get("cache_creation_input_tokens", 0)
        ),
        "cache_read_input_tokens": (
            u_agora["cache_read_input_tokens"]
            - usage_antes.get("cache_read_input_tokens", 0)
        ),
    }

    if delta["n_calls"] == 0:
        log.info("💰 Nenhuma chamada ao Claude nessa passada — sem custo")
        return

    custo = claude.estimar_custo_usd(
        delta["input_tokens"], delta["output_tokens"],
        delta["cache_creation_input_tokens"], delta["cache_read_input_tokens"],
    )

    log.info(
        f"💰 Passada: {delta['n_calls']} chamada(s) Claude, "
        f"{delta['input_tokens']:,} input + {delta['output_tokens']:,} output tokens "
        f"em {n_emails_processados} email(s) → US$ {custo:.4f}"
    )

    try:
        entry = {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "model": claude.model,
            "n_emails_processados": n_emails_processados,
            **delta,
            "custo_usd": round(custo, 6),
        }
        with open(BILLING_FILE, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except OSError as e:
        log.warning(f"Falha gravando billing.ndjson: {e}")


def _processar_seguro(
    fn,
    msg_id: str,
    msg_pra_label: str,
    msg_pra_resposta: dict | None,
    gmail: GmailClient,
    config: Config,
) -> None:
    """Executa fn() capturando exceptions:
      - aplica label pendente
      - envia reply amigável NO THREAD original (CC pro DP) — único canal de
        comunicação em casos de pendência, evitando email seco duplicado
      - loga em admissao_log.ndjson
    """
    try:
        fn()
    except Exception as e:
        log.exception(f"❌ {msg_id}: {e}")

        motivo_cliente = getattr(e, "motivo_cliente", None)
        campos_faltando = getattr(e, "campos_faltando", None) or []
        payload_parcial = getattr(e, "payload_parcial", None)
        e_e_pendencia = bool(motivo_cliente or campos_faltando)

        # 1. Label pendente
        try:
            gmail.aplicar_label(msg_pra_label, config.label_pendente)
        except Exception:
            log.exception("Falha aplicando label pendente")

        # 2. Resposta no thread (preferido) OU email seco pro DP (fallback)
        reply_enviado = False
        envio_pulado_pelo_usuario = False  # TEMP-CONFIRMAR-REPLIES
        if msg_pra_resposta and e_e_pendencia:
            try:
                corpo = email_resposta_cliente(
                    payload_parcial, campos_faltando, motivo_livre=motivo_cliente
                )
                # ─── TEMP-CONFIRMAR-REPLIES (remover este if em produção) ─
                if config.confirmar_replies and not _confirmar_envio(
                    msg_pra_resposta, corpo, config.email_dp or None
                ):
                    log.warning("   ✋ Envio cancelado pelo usuário (--ask)")
                    envio_pulado_pelo_usuario = True
                # ─── fim TEMP-CONFIRMAR-REPLIES ───────────────────────────
                else:
                    gmail.responder_no_thread(
                        msg_pra_resposta, corpo=corpo, cc=config.email_dp or None
                    )
                    reply_enviado = True
                    log.info(
                        f"   📨 Reply enviado no thread "
                        f"(cliente + DP em CC)"
                    )
            except Exception:
                log.exception("Falha enviando reply no thread")

        # Sem email seco pro DP — informações de pendência ficam consolidadas
        # em admissoes.xlsx. Aqui (catch de erro orquestrador-level), o
        # pipeline pode não ter dado tempo de chegar em _finalizar_lote pra
        # popular a planilha; registramos uma linha fallback com o que temos.
        try:
            registrar_admissao_planilha(
                nome="?",
                empresa=None,
                cnpj=None,
                procedencia=f"Falha técnica orquestrador — {str(e)[:120]}",
                msg_id=msg_id,
            )
        except Exception:
            log.exception("Falha registrando linha de erro na planilha")

        log_jsonl({
            "msg_id": msg_id,
            "status": "pendente_validacao" if campos_faltando else "erro",
            "erro": str(e),
            "_validacao_bloqueada": campos_faltando,
        })


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Pipeline de admissão local da Crosara. "
                    "Padrão: roda UMA passada (emails novos + threads aguardando "
                    "cliente) e sai — feito pra Windows Task Scheduler."
    )
    parser.add_argument(
        "--loop",
        action="store_true",
        help="Em vez de uma passada, fica em loop infinito fazendo polling "
             "(intervalo do config.json). Útil pra rodar manualmente em terminal.",
    )
    parser.add_argument(
        "--once",
        action="store_true",
        help="(Compatibilidade) idêntico ao padrão.",
    )
    # ─── TEMP-CONFIRMAR-REPLIES (remover em produção) ────────────────
    # Default: confirma antes de cada email (modo calibração). Auto-desliga
    # quando não tem TTY (Task Scheduler). Flag --no-ask força off sempre.
    parser.add_argument(
        "--no-ask",
        action="store_true",
        help="[TEMP] Desliga a confirmação interativa antes de cada email "
             "de pendência. Default é PERGUNTAR; só passe esta flag se quiser "
             "que o pipeline envie sem perguntar (ex: rodando em script).",
    )
    parser.add_argument(
        "--ask",
        action="store_true",
        help="[TEMP/DEPRECATED] No-op — confirmação já é default. Mantido "
             "pra compatibilidade com atalhos existentes.",
    )
    # ─── fim TEMP-CONFIRMAR-REPLIES ──────────────────────────────────
    args = parser.parse_args()

    log.info("=" * 70)
    log.info("Pipeline de Admissão Local — Crosara Contabilidade")
    log.info("=" * 70)

    try:
        config = carregar_config()
    except (FileNotFoundError, ValueError, json.JSONDecodeError) as e:
        log.error(f"Erro de configuração: {e}")
        return 1

    # ─── TEMP-CONFIRMAR-REPLIES (remover em produção) ────────────────
    # Resolução do modo confirmação:
    #   1. --no-ask explícito → SEMPRE desliga
    #   2. Sem TTY (Task Scheduler/redirect) → desliga automaticamente
    #   3. Caso contrário → default LIGADO (configurado na Config)
    if args.no_ask:
        config.confirmar_replies = False
        log.info("ℹ Confirmação interativa DESLIGADA por --no-ask")
    elif not sys.stdin.isatty():
        config.confirmar_replies = False
        log.info(
            "ℹ Sem TTY (provavelmente Task Scheduler) — "
            "confirmação interativa DESLIGADA automaticamente"
        )
    else:
        log.warning(
            "⚠ MODO CONFIRMAR ATIVO (default): cada email de pendência "
            "exigirá confirmação no terminal antes do envio. "
            "Use --no-ask pra desligar."
        )
    # ─── fim TEMP-CONFIRMAR-REPLIES ──────────────────────────────────

    bootstrap_arquivos_locais()

    # Regras customizáveis pelo escritório (exceções, defaults, observações).
    # Arquivo opcional — pipeline funciona normal mesmo se vazio/inexistente.
    # Conforme o escritório for definindo regras concretas, vão sendo aplicadas
    # nos pontos apropriados do pipeline (cargos_forcados, cnpjs_excecoes, etc.).
    _regras = carregar_regras()  # noqa: F841 — wiring concreto vem em commits futuros

    try:
        planilha = carregar_planilha(PLANILHA_CBO)
        log.info(f"📊 Planilha CBO: {len(planilha)} cargos")
    except (FileNotFoundError, ValueError) as e:
        log.error(f"Planilha CBO inválida: {e}")
        return 1

    try:
        claude = ClaudeClient(
            model=config.claude_model,
            max_tokens=config.claude_max_tokens,
            chamadas_verificacao=config.claude_chamadas_verificacao,
        )
    except Exception as e:
        log.error(f"Falha inicializando Claude: {e}")
        return 2

    if config.dry_run:
        log.warning("⚠ DRY-RUN ATIVO — não envia POSTs nem emails")

    if not args.loop:
        # Padrão: uma única passada e encerra (ideal pro Task Scheduler)
        log.info("▶ Executando passada única...")
        try:
            rodar_uma_passada(config, claude, planilha)
        except Exception as e:
            log.exception(f"Erro na passada: {e}")
            return 3
        log.info("✅ Passada concluída. Encerrando.")
        return 0

    # --loop: polling contínuo (uso manual)
    log.info(f"⏱  Modo loop ativo — polling a cada {config.intervalo}s. Ctrl+C pra parar.")
    while True:
        try:
            rodar_uma_passada(config, claude, planilha)
        except Exception as e:
            log.exception(f"Erro na passada: {e}")
        try:
            time.sleep(config.intervalo)
        except KeyboardInterrupt:
            log.info("Encerrado pelo usuário.")
            return 0


if __name__ == "__main__":
    sys.exit(main())
