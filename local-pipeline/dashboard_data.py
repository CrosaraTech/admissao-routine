"""dashboard_data.py — leitura compartilhada entre Tkinter e web (v2.15.0).

Toda a lógica de "ler o estado atual do pipeline pra mostrar" mora aqui.
Antes ficava espalhada no interface.py (Tkinter); agora a interface web
(webapp.py) reusa as mesmas funções — uma única fonte de verdade.

Nada aqui depende de Tkinter, Flask ou qualquer framework de UI: só
arquivos no disco (planilha XLSX + payloads/ + ndjson) + módulos do
pipeline (idempotencia, billing).
"""
from __future__ import annotations

import json
import logging
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

log = logging.getLogger("admissao.dashboard_data")

_DIR = Path(__file__).parent
PLANILHA_ADMISSOES = _DIR / "admissoes.xlsx"
PAYLOADS_DIR = _DIR / "payloads"
ADMISSAO_LOG = _DIR / "admissao_log.ndjson"

DIAS_PENDENCIA_VELHA = 3  # entidade pendente há ≥ 3 dias = "velha" (destaque)


# ── Tipos ────────────────────────────────────────────────────────

# Cada linha da planilha vira um dict desse formato após dedup:
#   {
#     "ts": "2026-06-13 10:00:00",
#     "nome": "INGRIDE DA SILVA JACINTO",
#     "empresa": "MODELOFARMA LTDA",
#     "cnpj": "10560396000185",
#     "procedencia": "Pendente cliente — faltam: ...",
#     "msg_id": "abc123def456",
#     "categoria": "processada" | "pendente_cliente" | "pendente_interna" | "falha_tecnica",
#     "velha": bool,
#     "dia": "2026-06-13",
#     "hora": "10:00:00",
#     "cargo_ia": "..." (só pra pendências, opcional),
#   }


# ── Leitura crua da planilha ─────────────────────────────────────

def _ler_linhas_brutas() -> list[dict]:
    """Lê admissoes.xlsx e retorna lista de dicts (uma por linha, sem dedup)."""
    if not PLANILHA_ADMISSOES.exists():
        return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(PLANILHA_ADMISSOES, read_only=True)
        ws = wb.active
    except Exception as e:
        log.warning(f"Falha lendo {PLANILHA_ADMISSOES}: {e}")
        return []

    linhas: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        cols = list(row) + [""] * (6 - len(row))
        primeira = str(cols[0] or "")
        if ("-" in primeira and ":" in primeira) or primeira == "":
            ts, nome, empresa, cnpj, procedencia, msg_id = cols[:6]
        else:
            # planilha legacy (4 cols sem ts/msg_id)
            nome, empresa, cnpj, procedencia = cols[:4]
            ts, msg_id = "", ""
        linhas.append({
            "ts": str(ts or ""),
            "nome": str(nome or "").strip(),
            "empresa": str(empresa or "").strip(),
            "cnpj": str(cnpj or "").strip(),
            "procedencia": str(procedencia or ""),
            "msg_id": str(msg_id or "").strip(),
        })
    return linhas


def _categorizar(procedencia: str) -> str:
    """Mapeia a string de procedência pra uma categoria fechada.

    v2.16.11: detecta "Pendente cliente — faltam: Função/Empresa/Departamento"
    como INTERNA — recategoriza histórico mal-classificado antes do fix sem
    precisar reprocessar. Quando há campo cliente misturado (ex.: Salário +
    Função), continua cliente (mais pertinente pra exibir ao operador).
    """
    p = (procedencia or "").lower()
    if p.startswith("cadastrado") or p.startswith("dry-run"):
        return "processada"
    if p.startswith("pendente interno"):
        return "pendente_interna"
    if p.startswith("falha técnica") or p.startswith("falha tecnica"):
        return "falha_tecnica"
    if p.startswith("pendente"):
        # Recategorização: se "faltam:" só tem campos internos, é interna
        if "faltam:" in p:
            partes = p.split("faltam:", 1)[1].strip().rstrip(".")
            campos = [c.strip() for c in partes.split(",") if c.strip()]
            internos = {"função", "funcao", "empresa", "departamento"}
            if campos and all(c in internos for c in campos):
                return "pendente_interna"
        return "pendente_cliente"
    return "outro"


def _eh_velha(ts: str, agora: datetime | None = None) -> bool:
    if not ts:
        return False
    agora = agora or datetime.now()
    try:
        dt = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
        return (agora - dt).days >= DIAS_PENDENCIA_VELHA
    except (ValueError, TypeError):
        return False


# ── Dedup por entidade ───────────────────────────────────────────

def _nome_canonico(n: str) -> str:
    """v2.16.6: chave canônica pra dedup robusta. Tolera variações de extração
    do Claude no mesmo email (reprocessos onde o nome saiu levemente diferente).

    Aplica: remove acentos, uppercase, remove TODOS os espaços e pontuação.
    Casos cobertos pelo bug real:
      "LINDOMAR IRACI DE SANTANA"  → "LINDOMARIRACIDESANTANA"
      "LINDOMAR IRACI DE SANT ANA" → "LINDOMARIRACIDESANTANA"  (mesma chave)
    Multi-pessoa no mesmo email continua separado porque os nomes começam
    diferente (LINDOMAR vs ALEXANDRE → chaves distintas).
    """
    import re as _re
    import unicodedata as _ud
    n = n or ""
    n = _ud.normalize("NFD", n).encode("ASCII", "ignore").decode("ASCII")
    return _re.sub(r"[^A-Z0-9]", "", n.upper())


def listar_entidades(*, incluir_cargo_ia: bool = False) -> list[dict]:
    """Retorna a lista de admissões DEDUPADA por entidade.

    Cada admissão aparece UMA vez, com o estado do ÚLTIMO evento.
    Bug histórico (v2.14.0): planilha é append-only, então a Tkinter mostrava
    8x o RAIMUNDO, 7x a JENIFFY. Agora cada um é 1 linha só.

    v2.16.6: chave de dedup mudou de (msg_id, nome_upper, cnpj) para
    (msg_id, nome_canonico). Razão:
      - CNPJ pode ser preenchido depois (override) — antes mudava chave
      - Claude pode extrair nome levemente diferente entre reprocessos
        ("SANTANA" vs "SANT ANA") — antes virava entidade nova
      - Multi-pessoa no mesmo lote continua separado (nomes diferentes)

    Args:
        incluir_cargo_ia: se True, popula `cargo_ia` lendo payloads/ (mais lento).
    """
    linhas = _ler_linhas_brutas()
    if not linhas:
        return []

    # v2.15.6: normaliza CNPJ pra dígitos puros antes da dedup. Sem isso,
    # uma mesma admissão registrada 2× (uma com "12.222.431/0001-81" formatado
    # e outra com "12222431000181" puro) aparecia como entidades diferentes
    # E a URL gerada com CNPJ formatado quebrava o roteamento Flask (a barra
    # do CNPJ era interpretada como separador de path → "Bad Request").
    import re as _re
    def _norm_cnpj(c: str) -> str:
        return _re.sub(r"\D", "", c or "")

    # 1ª passada: pega o ÚLTIMO evento por (msg_id, nome_canonico).
    # Pra cada chave, mantém os campos do evento mais recente — incluindo
    # CNPJ atualizado se veio depois (ex.: via override do operador).
    entidade_ultimo: dict[tuple, dict] = {}
    for r in linhas:
        cnpj_digits = _norm_cnpj(r["cnpj"])
        chave = (r["msg_id"], _nome_canonico(r["nome"]))
        ant = entidade_ultimo.get(chave)
        if ant is None or r["ts"] > ant["ts"]:
            entry = dict(r)
            entry["cnpj"] = cnpj_digits  # sempre dígitos puros pra adiante
            entidade_ultimo[chave] = entry

    # 2ª passada: enriquece (categoria, dia/hora, velha, cargo_ia opcional)
    agora = datetime.now()
    saida: list[dict] = []
    # v2.16.10: empresa "?" + CNPJ válido → consulta cache de empresas pra
    # exibir razão social na UI. Bug histórico: pendências geradas via
    # exception em _processar_um_bloco gravavam razao_social=None mesmo
    # tendo CNPJ resolvido (main.py:1237). Cache lookup é O(1) em RAM.
    _cache_empresas = None
    def _empresa_via_cache(cnpj_d: str) -> str:
        nonlocal _cache_empresas
        if not cnpj_d or len(cnpj_d) not in (11, 14):
            return ""
        if _cache_empresas is None:
            try:
                from main import obter_empresas_cache
                _cache_empresas = obter_empresas_cache(None)
            except Exception:
                _cache_empresas = False  # sinaliza "falhou ao carregar"
                return ""
        if _cache_empresas is False:
            return ""
        info = _cache_empresas.info(cnpj_d) or {}
        return info.get("razao_social", "") or ""

    for e in entidade_ultimo.values():
        cat = _categorizar(e["procedencia"])
        ts = e["ts"]
        dia, hora = ("(sem data)", "")
        if ts and " " in ts:
            dia, hora = ts.split(" ", 1)
        e["categoria"] = cat
        e["dia"] = dia
        e["hora"] = hora
        e["velha"] = _eh_velha(ts, agora) if cat.startswith("pendente") else False
        if incluir_cargo_ia and cat.startswith("pendente"):
            e["cargo_ia"] = cargo_ia_de(e["msg_id"])
        else:
            e["cargo_ia"] = ""
        # v2.16.10: enriquecer empresa vazia/"?" quando CNPJ é válido
        if (not e.get("empresa") or e["empresa"] in ("?", "—", "")) and e.get("cnpj"):
            razao = _empresa_via_cache(e["cnpj"])
            if razao:
                e["empresa"] = razao
        saida.append(e)

    saida.sort(key=lambda e: e["ts"], reverse=True)
    return saida


def resumo_contadores() -> dict[str, int]:
    """Conta entidades por categoria. Útil pros cards do dashboard."""
    out = {
        "processadas": 0,
        "pendentes_cliente": 0,
        "pendentes_internas": 0,
        "falhas_tecnicas": 0,
        "pendentes_velhas": 0,
        "total_pendentes": 0,
    }
    for e in listar_entidades():
        if e["categoria"] == "processada":
            out["processadas"] += 1
        elif e["categoria"] == "pendente_cliente":
            out["pendentes_cliente"] += 1
            out["total_pendentes"] += 1
        elif e["categoria"] == "pendente_interna":
            out["pendentes_internas"] += 1
            out["total_pendentes"] += 1
        elif e["categoria"] == "falha_tecnica":
            out["falhas_tecnicas"] += 1
        if e.get("velha"):
            out["pendentes_velhas"] += 1
    # v2.16.19: rascunhos de resposta pendentes de aprovação
    try:
        import rascunhos_resposta as _rr
        out["rascunhos_pendentes"] = _rr.contar_pendentes()
    except Exception:
        out["rascunhos_pendentes"] = 0
    return out


def agrupar_por_dia(entidades: list[dict]) -> dict[str, list[dict]]:
    """{ 'YYYY-MM-DD': [entidades] } — pros templates renderizarem por dia.
    Dias mais recentes primeiro."""
    por_dia: dict[str, list[dict]] = {}
    for e in entidades:
        por_dia.setdefault(e["dia"], []).append(e)
    return dict(sorted(por_dia.items(), reverse=True))


# ── Payloads ─────────────────────────────────────────────────────

def cargo_ia_de(msg_id: str) -> str:
    """Cargo que o Claude extraiu (lê do JSON mais recente da pendência).
    Sem cache — caller decide se quer cachear."""
    if not msg_id or not PAYLOADS_DIR.exists():
        return ""
    try:
        arqs = sorted(PAYLOADS_DIR.glob(f"*_{msg_id[:16]}*.json"), reverse=True)
        for arq in arqs:
            try:
                doc = json.loads(arq.read_text(encoding="utf-8"))
            except (json.JSONDecodeError, OSError):
                continue
            payload = doc.get("payload") or {}
            attrs = (payload.get("data") or {}).get("attributes") or {}
            cargo = (
                (doc.get("resolucao") or {}).get("cargo_extraido")
                or attrs.get("nomecargo")
                or payload.get("cargo_extraido")
            )
            if cargo:
                return str(cargo)
    except Exception:
        pass
    return ""


def achar_payload_por_msg_e_nome(msg_id: str, nome: str) -> Path | None:
    """Path do JSON da admissão em payloads/. msg_id é primário; nome
    é tiebreaker quando há N pessoas no mesmo email."""
    if not PAYLOADS_DIR.exists():
        return None
    if not msg_id and not nome:
        return None
    nome_upper = (nome or "").upper().strip()
    if msg_id:
        arqs = sorted(PAYLOADS_DIR.glob(f"*_{msg_id[:16]}*.json"), reverse=True)
    else:
        arqs = sorted(PAYLOADS_DIR.glob("*.json"), reverse=True)
    # Tiebreaker por nome
    for arq in arqs:
        try:
            doc = json.loads(arq.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not nome_upper:
            return arq
        # Confere se o nome bate
        attrs = ((doc.get("payload") or {}).get("data") or {}).get("attributes") or {}
        nome_payload = str(attrs.get("nome") or "").upper().strip()
        nome_resol = str((doc.get("resolucao") or {}).get("nome") or "").upper().strip()
        if nome_upper in (nome_payload, nome_resol):
            return arq
    # Fallback: primeiro arquivo do msg_id
    return arqs[0] if arqs else None


def carregar_payload_completo(path: Path | None) -> dict:
    """Lê o JSON inteiro do payload. {} em erro."""
    if not path or not Path(path).exists():
        return {}
    try:
        return json.loads(Path(path).read_text(encoding="utf-8"))
    except Exception as e:
        log.warning(f"Falha lendo {path}: {e}")
        return {}


# ── Auditoria ───────────────────────────────────────────────────

def encontrar_pendencia_irma(
    cpf: str | int | None,
    nome: str,
    cnpj: str,
    excluir_msg_id: str = "",
) -> dict | None:
    """v2.15.15: Procura uma pendência ABERTA do mesmo candidato (provavelmente
    veio dum email anterior do mesmo cliente).

    Estratégia de match (em ordem de confiança):
      1. CPF normalizado igual (match forte) — preciso
      2. Nome upper + CNPJ digits iguais (match fraco) — quando CPF faltava
         num dos lados

    Retorna a entidade pendente irmã (mesmo dict que listar_entidades) ou
    None se nenhuma bate. NÃO retorna a própria msg_id (excluir_msg_id).
    Só considera pendentes — processadas/falhas técnicas não entram.
    """
    cpf_d = re.sub(r"\D", "", str(cpf or ""))
    nome_u = (nome or "").upper().strip()
    cnpj_d = re.sub(r"\D", "", str(cnpj or ""))

    for ent in listar_entidades():
        if not ent["categoria"].startswith("pendente"):
            continue
        if ent["msg_id"] == excluir_msg_id:
            continue
        # Match forte por CPF — precisa ler o payload antigo pra pegar o CPF
        if cpf_d:
            path = achar_payload_por_msg_e_nome(ent["msg_id"], ent["nome"])
            doc = carregar_payload_completo(path) if path else {}
            attrs_old = ((doc.get("payload") or {}).get("data") or {}).get("attributes") or {}
            cpf_old = re.sub(r"\D", "", str(attrs_old.get("cpf") or ""))
            if cpf_old and cpf_old == cpf_d:
                ent["_payload_path"] = str(path) if path else ""
                ent["_attrs_antigos"] = attrs_old
                return ent
        # Match fraco por nome+CNPJ (quando algum CPF faltava)
        if nome_u and cnpj_d:
            if ent["nome"].upper() == nome_u and ent["cnpj"] == cnpj_d:
                path = achar_payload_por_msg_e_nome(ent["msg_id"], ent["nome"])
                doc = carregar_payload_completo(path) if path else {}
                attrs_old = ((doc.get("payload") or {}).get("data") or {}).get("attributes") or {}
                ent["_payload_path"] = str(path) if path else ""
                ent["_attrs_antigos"] = attrs_old
                return ent
    return None


def auditoria_recente(n: int = 200) -> list[dict]:
    """Últimas N entradas do admissao_log.ndjson (ordem do mais recente
    pra mais antigo). Cada linha já é um dict — só parseia.

    v2.15.0 fix (review MEDIUM): usa collections.deque(maxlen=n) pra evitar
    carregar arquivo inteiro na memória. Antes lia tudo e cortava o tail —
    com NDJSON de 100MB+, 5 hits paralelos = 500MB RAM no processo Flask.
    """
    if not ADMISSAO_LOG.exists():
        return []
    from collections import deque
    eventos: list[dict] = []
    try:
        with open(ADMISSAO_LOG, "r", encoding="utf-8") as f:
            ultimas = deque(f, maxlen=n)
    except OSError:
        return []
    for linha in ultimas:
        s = linha.strip()
        if not s:
            continue
        try:
            eventos.append(json.loads(s))
        except json.JSONDecodeError:
            continue
    eventos.reverse()
    return eventos


# ── Status/Health ────────────────────────────────────────────────

def status_pipeline() -> dict:
    """Snapshot do estado: última passada, total processado, fila."""
    contadores = resumo_contadores()
    ultima = ""
    if ADMISSAO_LOG.exists():
        try:
            with open(ADMISSAO_LOG, "rb") as f:
                # Pega só a última linha (eficiente em arquivos grandes)
                f.seek(0, 2)
                tamanho = f.tell()
                bloco = min(4096, tamanho)
                f.seek(tamanho - bloco)
                ultima_linha = f.read().decode("utf-8", errors="ignore").strip().split("\n")[-1]
            if ultima_linha:
                try:
                    e = json.loads(ultima_linha)
                    ultima = str(e.get("timestamp", ""))
                except json.JSONDecodeError:
                    pass
        except OSError:
            pass
    return {
        "ultima_passada": ultima,
        **contadores,
    }
