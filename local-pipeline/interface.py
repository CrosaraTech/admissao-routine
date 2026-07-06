"""interface.py — GUI desktop pro pipeline de admissão (Tkinter).

Visual com identidade da Crosara Contabilidade:
  • Sidebar navy (#344E5C) com logo + nav
  • Conteúdo em bege (#F4DDC8) com cards brancos
  • Accent laranja (#D95C32) pra primary + active state
  • Log "terminal" em navy escuro com timestamps em laranja

6 abas (sidebar): Principal · Processadas · Pendentes · Auditoria ·
Estatísticas · Regras. Polling em thread separada, fila thread-safe
pra atualizar a UI. Reaproveita toda a lógica do main.py.

Uso:
    python interface.py
"""

from __future__ import annotations

import json
import logging
import queue
import threading
import time
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import messagebox, scrolledtext, ttk

from dotenv import load_dotenv
load_dotenv()

# ============================================================
# Paleta de cores Crosara (extraída do Logotipo Crosara - CMYK-04.jpg)
# ============================================================

COR_NAVY = "#344E5C"          # sidebar, header dark, log terminal
COR_NAVY_DARK = "#2A4050"     # log slightly darker
COR_NAVY_HOVER = "#3F5C6E"    # nav button hover
COR_NAVY_ACTIVE = "#2A4050"   # nav button when active (não usado, vai laranja)
COR_BEIGE = "#F4DDC8"         # main background
COR_BEIGE_DARK = "#E8C9AE"    # subtle differentiation
COR_WHITE = "#FFFFFF"         # cards
COR_ORANGE = "#D95C32"        # accent / primary / active nav
COR_ORANGE_DARK = "#B84A24"   # primary button hover
COR_TEXT_DARK = "#2A4050"     # text on light bg
COR_TEXT_LIGHT = "#F4DDC8"    # text on dark bg (matches beige)
COR_TEXT_MUTED = "#7A8896"    # muted/secondary text
COR_BORDER = "#D8C5AE"        # subtle border on cards
COR_SUCCESS = "#2E7D32"       # contador processadas (verde)
COR_WARNING = "#D95C32"       # contador pendentes (laranja Crosara)
COR_DANGER = "#C62828"        # contador >3d (vermelho)
COR_INFO = "#1565C0"          # custo claude (azul)

# Pillow opcional (logo na sidebar)
try:
    from PIL import Image, ImageTk
    _HAS_PIL = True
except ImportError:
    _HAS_PIL = False

LOGO_PATH = Path(__file__).parent / "Logotipo Crosara - CMYK-04.jpg"

# Logo do app — AdmitER. Usuário salva admitir-logo.png na pasta;
# o .ico é auto-gerado a partir do .png (via Pillow) se faltar.
ADMITER_LOGO_PNG = Path(__file__).parent / "admitir-logo.png"
ADMITER_LOGO_ICO = Path(__file__).parent / "admitir-logo.ico"
APP_NAME = "AdmitER"
APP_VERSION = "2.16.50"
APP_TAGLINE = "Pipeline de Admissão Automatizada"
APP_VENDOR = "CrosaraTech"


# ============================================================
# Paleta dark mode (slots espelham os do tema light)
# ============================================================

# Mapeamento por "slot" — cada cor da paleta light tem uma equivalente dark.
# O toggle dark mode varre o widget tree e troca bg/fg por essa correspondência.
_PALETA_LIGHT = {
    "main_bg":     COR_BEIGE,        # #F4DDC8
    "card_bg":     COR_WHITE,        # #FFFFFF
    "sidebar_bg":  COR_NAVY,         # #344E5C
    "log_bg":      COR_NAVY_DARK,    # #2A4050
    "sub_bg":      COR_BEIGE_DARK,   # #E8C9AE
    "text_dark":   COR_TEXT_DARK,    # #2A4050
    "text_light":  COR_TEXT_LIGHT,   # #F4DDC8
    "text_muted":  COR_TEXT_MUTED,   # #7A8896
    "border":      COR_BORDER,       # #D8C5AE
}

_PALETA_DARK = {
    "main_bg":     "#1E2128",
    "card_bg":     "#2A2D34",
    "sidebar_bg":  "#161922",
    "log_bg":      "#0F1117",
    "sub_bg":      "#252830",
    "text_dark":   "#E8E8E8",
    "text_light":  "#F4DDC8",       # accent claro mantém pra brilho
    "text_muted":  "#9098A4",
    "border":      "#3D404A",
}

from claude_client import ClaudeClient
from ecotador_client import AUDIT_FILE as ECONTADOR_AUDIT_FILE, EContadorAPI
from gmail_client import GmailClient
from cnpj_utils import formatar_cnpj, so_digitos, validar_cnpj
from ecotador_client import EContadorAPI
from directdata_client import sum_directdata_mes_atual
from main import (
    PAYLOADS_DIR, PLANILHA_ADMISSOES, PLANILHA_CBO, REGRAS_FILE,
    bootstrap_arquivos_locais, carregar_config, carregar_planilha,
    carregar_regras, fazer_backup_planilha_e_payloads,
    recarregar_empresas_cache,
    registrar_admissao_planilha, rodar_uma_passada,
    salvar_cnpj_override, salvar_funcao_override, sum_billing_mes_atual,
)
from payload_builder import LABELS_AMIGAVEIS
from funcao import salvar_funcao_alias
import idempotencia
# v2.14.1: wrapper único de POST — substitui chamada direta api.post_candidato
# nos dois caminhos da UI (Enviar mesmo assim + Resolver pendência). Os messagebox
# de confirmação ficam aqui na UI; idempotência/log/label/payload no wrapper.
from post_admissao import postar_candidato_registrado

# Toast Windows (opcional — fallback pra popup Tk se plyer não instalado)
try:
    from plyer import notification as _plyer_notif  # noqa
    _HAS_PLYER = True
except ImportError:
    _HAS_PLYER = False


# Dias após os quais uma pendência é considerada "velha" → highlight visual
DIAS_PENDENCIA_VELHA = 3

# Map reverso: label amigável → chave do attribute no payload
# (usado pelo form estruturado de resolver pendência)
LABEL_PRA_ATTR = {v: k for k, v in LABELS_AMIGAVEIS.items()}

# Attrs que são data (aceitam dd/mm/aaaa, dd-mm-aaaa ou YYYY-MM-DD)
DATE_ATTRS = {
    "admissao", "nascimento", "dataatestadoocupacional",
    "dataterminocontrato", "datapis",
    "emissaocnh", "validadecnh", "primeiraemissaocnh",
}

# Hints de preenchimento por tipo de campo
HINTS_FORM = {
    "admissao": "dd/mm/aaaa (ex: 28/05/2026)",
    "nascimento": "dd/mm/aaaa (ex: 12/03/1990)",
    "dataatestadoocupacional": "dd/mm/aaaa (ex: 10/05/2026)",
    "dataidentidade": "dd/mm/aaaa (data emissão RG)",
    "salario": "Número decimal (ex: 1722.25)",
    "cpf": "Apenas dígitos (ex: 12345678901)",
    "diascontratoexperiencia": "Inteiro (default 30)",
    "primeiroemprego": "true ou false",
    "cep": "00000-000 (com hífen)",
    "rua": "UPPERCASE, sem vírgula",
    "bairro": "UPPERCASE, sem vírgula",
    "cidade": "UPPERCASE, sem UF",
    "numero": "Inteiro (0 = sem número)",
    "identidade": "Só dígitos (ex: 5932277)",
    "orgaoemissoridentidade": "Sigla curta (ex: SSP/PE) — máx 9 chars",
    "ctps": "Inteiro (número da CTPS)",
    "seriectps": "String (ex: 0030)",
    "pis": "11 dígitos (preserva zeros)",
    "_codigo_funcao": "Código da planilha CBO (ex: 009332 ou 9332)",
}

# Pseudo-labels que NÃO mapeiam pra um attribute direto do payload — exigem
# tratamento especial em _aplicar_form_e_postar (ex: função vira relationship).
PSEUDO_LABELS = {
    "Função (código)": "_codigo_funcao",
}


# ============================================================
# Logging handler — ecoa TODOS os logs do logging module pra UI
# ============================================================

class _TkQueueLogHandler(logging.Handler):
    """Handler de logging que joga registros formatados na fila da UI.

    Thread-safe via queue.Queue. Roda no thread que originou o log;
    o consumidor (_consumir_fila) puxa do main thread Tk.

    Filtra ruido de DEBUG do googleapiclient/urllib3/httpx no nivel INFO —
    sem isso a UI vira soup de HTTP requests.
    """

    _SILENCIAR_PREFIXOS = (
        "googleapiclient.discovery",
        "urllib3.connectionpool",
        "google.auth.transport",
    )

    def __init__(self, ui_queue: queue.Queue):
        super().__init__(level=logging.INFO)
        self.ui_queue = ui_queue
        self.setFormatter(logging.Formatter(
            "%(asctime)s [%(levelname)s] %(name)s: %(message)s",
            datefmt="%H:%M:%S",
        ))

    def emit(self, record: logging.LogRecord) -> None:
        # Filtra ruido de bibliotecas pra UI nao virar log de HTTP request
        for prefixo in self._SILENCIAR_PREFIXOS:
            if record.name.startswith(prefixo):
                return
        try:
            msg = self.format(record)
            self.ui_queue.put(("log_raw", msg))
        except Exception:
            self.handleError(record)


# ============================================================
# Janela principal
# ============================================================

class PipelineGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_NAME} — {APP_TAGLINE}")
        self.geometry("1300x800")
        self.minsize(1000, 600)

        self._init_state()
        self._aplicar_icone_janela()
        self._build_ui()
        self._consumir_fila()  # loop periódico thread→UI
        self._refresh_tabelas()

        # Fechar limpo
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    # ---- Estado / inicialização --------------------------------

    def _init_state(self):
        bootstrap_arquivos_locais()
        try:
            self.config = carregar_config()
            # GUI controla via toggle, não via prompt interativo
            self.config.confirmar_replies = False
        except Exception as e:
            messagebox.showerror("Erro carregando config", str(e))
            raise

        try:
            self.planilha = carregar_planilha(PLANILHA_CBO)
        except Exception as e:
            messagebox.showerror("Erro carregando planilha CBO", str(e))
            raise

        carregar_regras()  # log info; wiring vem em commits futuros

        self.claude = ClaudeClient(
            model=self.config.claude_model,
            max_tokens=self.config.claude_max_tokens,
            chamadas_verificacao=self.config.claude_chamadas_verificacao,
        )

        self.polling = False
        self.polling_thread: threading.Thread | None = None
        self.gui_q: queue.Queue = queue.Queue()

        # Espelha logs do logging.* na UI (mesmo nivel/formato do terminal)
        root_logger = logging.getLogger()
        self._ui_log_handler = _TkQueueLogHandler(self.gui_q)
        root_logger.addHandler(self._ui_log_handler)
        if root_logger.level == logging.NOTSET or root_logger.level > logging.INFO:
            root_logger.setLevel(logging.INFO)

        # Carrega cache de CNPJs em RAM no background — usado pra auto-correção
        # de typos (caso real Ekoplastic). 30s na 1ª paginação, sem bloquear UI.
        # Roda 1× por sessão. Não persiste em disco (portável: roda igual em
        # qualquer máquina sem state local).
        # IMPORTANTE: arrancar SÓ depois de self.gui_q existir (thread loga via gui_q).
        threading.Thread(
            target=self._carregar_empresas_cache_background,
            daemon=True,
        ).start()

        # Tk vars
        self.status_var = tk.StringVar(value="Parado")
        # Quando status muda, atualiza cor do dot da pill (verde/laranja/cinza)
        self.status_var.trace_add("write", lambda *a: self._atualizar_status_dot())
        self.ultima_var = tk.StringVar(value="—")
        self.proxima_var = tk.StringVar(value="—")
        self.auto_email_var = tk.BooleanVar(value=self.config.auto_email_pendencias)
        # Overrides globais (v2.6.0)
        self.sempre_sem_data_var = tk.BooleanVar(value=self.config.sempre_mandar_sem_data_admissao)
        self.sempre_sem_funcao_var = tk.BooleanVar(value=self.config.sempre_mandar_sem_funcao)
        # Reprocessar pendentes (v2.9.0)
        self.reprocessar_pendentes_var = tk.BooleanVar(
            value=self.config.reprocessar_pendentes_no_polling
        )
        self.intervalo_var = tk.IntVar(value=self.config.intervalo)
        self.contador_proc_var = tk.StringVar(value="0")
        self.contador_pend_var = tk.StringVar(value="0")
        self.contador_velha_var = tk.StringVar(value="0")
        self.billing_var = tk.StringVar(value="US$ 0.00")
        # Custo agregado de TODAS as APIs externas no mês (Claude USD → BRL +
        # DirectData BRL). Cotação fixa simples (5.00 BRL/USD) — não vamos
        # consultar API de câmbio só pra display de billing.
        self.billing_total_brl_var = tk.StringVar(value="R$ 0,00")
        self.cotacao_usd_brl = 5.00
        # Limite mensal de billing pra alerta visual (default 50 USD)
        self.billing_limite_usd = 50.0
        # Filtros de busca por tab
        self.filtro_proc_var = tk.StringVar()
        self.filtro_pend_var = tk.StringVar()
        self.filtro_audit_var = tk.StringVar()
        self.filtro_proc_var.trace_add("write", lambda *a: self._refresh_tabelas())
        self.filtro_pend_var.trace_add("write", lambda *a: self._refresh_tabelas())
        self.filtro_audit_var.trace_add("write", lambda *a: self._refresh_auditoria())

        # Tracking pra disparar toast quando chega pendência nova
        self._msg_ids_pendentes_conhecidos: set[str] = set()
        self._primeira_carga = True

        # Dark mode (afeta log + treeviews + algumas frames)
        self.dark_mode_var = tk.BooleanVar(value=False)

    # ---- Construção das abas -----------------------------------

    def _build_ui(self):
        self._aplicar_tema_crosara()
        self.configure(background=COR_BEIGE)

        # Layout: sidebar | (header + content)
        container = tk.Frame(self, bg=COR_BEIGE)
        container.pack(fill="both", expand=True)

        self._build_sidebar(container)

        # Área direita (header + conteúdo)
        right = tk.Frame(container, bg=COR_BEIGE)
        right.pack(side="left", fill="both", expand=True)
        self._build_header(right)

        # Holder do conteúdo (swap entre tabs)
        self.content_holder = tk.Frame(right, bg=COR_BEIGE)
        self.content_holder.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        # Frames de cada "aba" (sem Notebook — sidebar nav controla)
        self.tab_main = tk.Frame(self.content_holder, bg=COR_BEIGE)
        self.tab_proc = tk.Frame(self.content_holder, bg=COR_BEIGE)
        self.tab_pend = tk.Frame(self.content_holder, bg=COR_BEIGE)
        self.tab_audit = tk.Frame(self.content_holder, bg=COR_BEIGE)
        self.tab_api = tk.Frame(self.content_holder, bg=COR_BEIGE)
        self.tab_stats = tk.Frame(self.content_holder, bg=COR_BEIGE)
        self.tab_regras = tk.Frame(self.content_holder, bg=COR_BEIGE)

        self._tab_frames = {
            "principal":   self.tab_main,
            "processadas": self.tab_proc,
            "pendentes":   self.tab_pend,
            "auditoria":   self.tab_audit,
            "api":         self.tab_api,
            "estatisticas": self.tab_stats,
            "regras":      self.tab_regras,
        }

        self._build_main()
        self._build_processadas()
        self._build_pendentes()
        self._build_auditoria()
        self._build_api_econtador()
        self._build_estatisticas()
        self._build_regras()

        # Mostra Principal por padrão
        self._switch_tab("principal")

    def _aplicar_tema_crosara(self):
        """Configura ttk.Style com a paleta da marca."""
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        # Frame, Label
        style.configure("TFrame", background=COR_BEIGE)
        style.configure("TLabel", background=COR_BEIGE, foreground=COR_TEXT_DARK)
        style.configure("TLabelframe", background=COR_BEIGE, foreground=COR_TEXT_DARK,
                        bordercolor=COR_BORDER, relief="solid")
        style.configure("TLabelframe.Label", background=COR_BEIGE, foreground=COR_TEXT_DARK)

        # Cards (frames brancos)
        style.configure("Card.TFrame", background=COR_WHITE, relief="solid", borderwidth=1)
        style.configure("Card.TLabel", background=COR_WHITE, foreground=COR_TEXT_DARK)

        # Buttons
        style.configure("TButton", background=COR_WHITE, foreground=COR_TEXT_DARK,
                        padding=8, borderwidth=1, focusthickness=0)
        style.map("TButton",
                  background=[("active", COR_BEIGE_DARK)],
                  foreground=[("active", COR_TEXT_DARK)])

        # Primary button (laranja)
        style.configure("Primary.TButton", background=COR_ORANGE, foreground=COR_WHITE,
                        padding=10, borderwidth=0, font=("Segoe UI", 10, "bold"))
        style.map("Primary.TButton",
                  background=[("active", COR_ORANGE_DARK), ("disabled", "#C9C9C9")],
                  foreground=[("disabled", "#888")])

        # Treeview
        style.configure("Treeview",
                        background=COR_WHITE, fieldbackground=COR_WHITE,
                        foreground=COR_TEXT_DARK, rowheight=26, bordercolor=COR_BORDER)
        style.configure("Treeview.Heading",
                        background=COR_NAVY, foreground=COR_WHITE,
                        font=("Segoe UI", 9, "bold"), padding=5, relief="flat")
        style.map("Treeview",
                  background=[("selected", COR_ORANGE)],
                  foreground=[("selected", COR_WHITE)])

        # Entry, Combobox, Spinbox
        for w in ("TEntry", "TCombobox", "TSpinbox"):
            style.configure(w, fieldbackground=COR_WHITE, foreground=COR_TEXT_DARK,
                            bordercolor=COR_BORDER, relief="solid")

        # Notebook (usado só no ResolverPendenciaDialog)
        style.configure("TNotebook", background=COR_BEIGE, borderwidth=0)
        style.configure("TNotebook.Tab", background=COR_BEIGE_DARK, foreground=COR_TEXT_DARK,
                        padding=(12, 6))
        style.map("TNotebook.Tab",
                  background=[("selected", COR_WHITE)],
                  foreground=[("selected", COR_TEXT_DARK)])

        # Checkbutton, Scrollbar
        style.configure("TCheckbutton", background=COR_BEIGE, foreground=COR_TEXT_DARK)
        style.configure("TScrollbar", background=COR_BEIGE, troughcolor=COR_BEIGE_DARK,
                        bordercolor=COR_BEIGE_DARK, arrowcolor=COR_NAVY)

    def _build_sidebar(self, parent: tk.Frame):
        side = tk.Frame(parent, bg=COR_NAVY, width=230)
        side.pack(side="left", fill="y")
        side.pack_propagate(False)

        # Logo header — AdmitER (logo do app) + texto + tagline
        header = tk.Frame(side, bg=COR_NAVY, height=110)
        header.pack(fill="x", pady=(20, 25), padx=15)
        header.pack_propagate(False)

        admiter_logo = self._carregar_logo_admiter_sidebar()
        if admiter_logo:
            self._admiter_logo_ref = admiter_logo
            tk.Label(header, image=admiter_logo, bg=COR_NAVY).pack(side="left", padx=(0, 10))
            txt_frame = tk.Frame(header, bg=COR_NAVY)
            txt_frame.pack(side="left", anchor="center")
            tk.Label(txt_frame, text=APP_NAME, bg=COR_NAVY, fg=COR_TEXT_LIGHT,
                     font=("Segoe UI", 18, "bold")).pack(anchor="w")
            tk.Label(txt_frame, text=f"by {APP_VENDOR}", bg=COR_NAVY,
                     fg=COR_ORANGE, font=("Segoe UI", 8, "bold")).pack(anchor="w")
        else:
            # Fallback: só texto (caso admitir-logo.png não exista)
            tk.Label(header, text=APP_NAME, bg=COR_NAVY, fg=COR_ORANGE,
                     font=("Segoe UI", 24, "bold")).pack(anchor="w")
            tk.Label(header, text=f"by {APP_VENDOR}", bg=COR_NAVY,
                     fg=COR_TEXT_LIGHT, font=("Segoe UI", 9)).pack(anchor="w")

        # Nav buttons
        self._nav_buttons: dict[str, tk.Button] = {}
        nav_items = [
            ("principal",   "Principal"),
            ("processadas", "Processadas"),
            ("pendentes",   "Pendentes"),
            ("auditoria",   "Auditoria"),
            ("api",         "API eContador"),
            ("estatisticas", "Estatísticas"),
            ("regras",      "Regras"),
        ]
        for key, label in nav_items:
            btn = tk.Button(
                side, text=f"   {label}", anchor="w",
                bg=COR_NAVY, fg=COR_TEXT_LIGHT,
                activebackground=COR_NAVY_HOVER, activeforeground=COR_WHITE,
                bd=0, relief="flat", padx=20, pady=10,
                font=("Segoe UI", 11),
                cursor="hand2",
                command=lambda k=key: self._switch_tab(k),
            )
            btn.pack(fill="x", padx=10, pady=2)
            self._nav_buttons[key] = btn

        # Rodapé do sidebar — botão Sobre + assinatura
        rodape = tk.Frame(side, bg=COR_NAVY)
        rodape.pack(side="bottom", fill="x", padx=10, pady=15)
        tk.Button(
            rodape, text="ⓘ  Sobre o AdmitER", anchor="w",
            bg=COR_NAVY, fg=COR_TEXT_LIGHT,
            activebackground=COR_NAVY_HOVER, activeforeground=COR_WHITE,
            bd=0, relief="flat", padx=20, pady=8,
            font=("Segoe UI", 9),
            cursor="hand2",
            command=self._abrir_sobre,
        ).pack(fill="x", pady=2)
        tk.Label(rodape, text=f"v{APP_VERSION} · by {APP_VENDOR}",
                 bg=COR_NAVY, fg=COR_TEXT_MUTED,
                 font=("Segoe UI", 7)).pack(pady=(8, 0))

    def _carregar_logo_admiter_sidebar(self):
        """Logo do AdmitER pra topo do sidebar. Tamanho compacto (~60px)."""
        if not _HAS_PIL or not ADMITER_LOGO_PNG.exists():
            return None
        try:
            img = Image.open(ADMITER_LOGO_PNG)
            if img.mode != "RGBA":
                img = img.convert("RGBA")
            w, h = img.size
            target_h = 60
            new_w = int(w * target_h / h)
            img = img.resize((new_w, target_h), Image.LANCZOS)
            return ImageTk.PhotoImage(img)
        except Exception:
            return None

    def _carregar_logo_sidebar(self):
        if not _HAS_PIL or not LOGO_PATH.exists():
            return None
        try:
            img = Image.open(LOGO_PATH)
            w, h = img.size
            target_h = 50
            new_w = int(w * target_h / h)
            img = img.resize((new_w, target_h), Image.LANCZOS)
            return ImageTk.PhotoImage(img)
        except Exception:
            return None

    def _build_header(self, parent: tk.Frame):
        header = tk.Frame(parent, bg=COR_BEIGE, height=70)
        header.pack(fill="x", padx=20, pady=(20, 10))
        header.pack_propagate(False)

        self._titulo_var = tk.StringVar(value="Pipeline de Admissão")
        tk.Label(header, textvariable=self._titulo_var, bg=COR_BEIGE,
                 fg=COR_TEXT_DARK, font=("Segoe UI", 20, "bold")).pack(side="left")

        # Status pill à direita
        self._status_pill = tk.Frame(header, bg=COR_WHITE, padx=15, pady=6, bd=1, relief="solid")
        self._status_pill.pack(side="right", pady=10)
        self._status_dot = tk.Label(self._status_pill, text="●", bg=COR_WHITE,
                                    fg="#888", font=("Segoe UI", 14))
        self._status_dot.pack(side="left", padx=(0, 6))
        tk.Label(self._status_pill, textvariable=self.status_var, bg=COR_WHITE,
                 fg=COR_TEXT_DARK, font=("Segoe UI", 10, "bold")).pack(side="left")

    def _switch_tab(self, name: str):
        """Mostra o frame da tab `name`, esconde os outros, destaca o botão."""
        for n, f in self._tab_frames.items():
            if n == name:
                f.pack(fill="both", expand=True)
            else:
                f.pack_forget()

        # Update sidebar — botão ativo em laranja
        for k, btn in self._nav_buttons.items():
            if k == name:
                btn.configure(bg=COR_ORANGE, fg=COR_WHITE,
                              activebackground=COR_ORANGE_DARK,
                              activeforeground=COR_WHITE)
            else:
                btn.configure(bg=COR_NAVY, fg=COR_TEXT_LIGHT,
                              activebackground=COR_NAVY_HOVER,
                              activeforeground=COR_WHITE)

        # Update title
        titulos = {
            "principal": "Pipeline de Admissão",
            "processadas": "Admissões Processadas",
            "pendentes": "Admissões Pendentes",
            "auditoria": "Auditoria",
            "api": "API eContador — Chamadas HTTP",
            "estatisticas": "Estatísticas",
            "regras": "Regras de Negócio",
        }
        self._titulo_var.set(titulos.get(name, "Pipeline"))

    def _build_main(self):
        # Stats: 4 cards lado a lado
        stats_row = tk.Frame(self.tab_main, bg=COR_BEIGE)
        stats_row.pack(fill="x", pady=(10, 15))

        self._card_stat(stats_row, "Processadas", self.contador_proc_var, COR_SUCCESS).pack(
            side="left", fill="both", expand=True, padx=(0, 10))
        self._card_stat(stats_row, "Pendentes", self.contador_pend_var, COR_WARNING).pack(
            side="left", fill="both", expand=True, padx=10)
        self._card_stat(stats_row, f"Pendência >{DIAS_PENDENCIA_VELHA}d",
                        self.contador_velha_var, COR_DANGER).pack(
            side="left", fill="both", expand=True, padx=10)
        billing_card = self._card_stat(stats_row, "APIs externas/mês",
                                       self.billing_total_brl_var, COR_INFO,
                                       valor_font_size=18)
        billing_card.pack(side="left", fill="both", expand=True, padx=(10, 0))
        # Referência pro label de valor pra mudar cor (alerta de limite)
        self.billing_label = billing_card._valor_label  # set dentro de _card_stat

        # Controle (card branco)
        f_ctrl_card = tk.Frame(self.tab_main, bg=COR_WHITE, bd=1, relief="solid",
                                highlightbackground=COR_BORDER, highlightthickness=0)
        f_ctrl_card.pack(fill="x", pady=10)
        tk.Label(f_ctrl_card, text="Controle", bg=COR_WHITE, fg=COR_TEXT_DARK,
                 font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=15, pady=(12, 8))

        row1 = tk.Frame(f_ctrl_card, bg=COR_WHITE)
        row1.pack(fill="x", padx=15, pady=(0, 6))
        self.btn_iniciar = ttk.Button(row1, text="Iniciar polling",
                                       style="Primary.TButton",
                                       command=self._iniciar_polling)
        self.btn_iniciar.pack(side="left", padx=(0, 8))
        self.btn_parar = ttk.Button(row1, text="Parar",
                                     command=self._parar_polling, state="disabled")
        self.btn_parar.pack(side="left", padx=8)
        ttk.Button(row1, text="Rodar 1 passada", command=self._rodar_unica).pack(side="left", padx=8)

        row2 = tk.Frame(f_ctrl_card, bg=COR_WHITE)
        row2.pack(fill="x", padx=15, pady=(0, 12))
        ttk.Button(row2, text="📥  Importar arquivos",
                   style="Primary.TButton",
                   command=self._abrir_importar_arquivos).pack(side="left", padx=(0, 8))
        ttk.Button(row2, text="Backup agora", command=self._backup_agora).pack(side="left", padx=8)
        ttk.Button(row2, text="Atualizar tabelas", command=self._refresh_tabelas).pack(side="left", padx=8)

        # Configurações (card)
        f_set = tk.Frame(self.tab_main, bg=COR_WHITE, bd=1, relief="solid")
        f_set.pack(fill="x", pady=10)
        tk.Label(f_set, text="Configurações", bg=COR_WHITE, fg=COR_TEXT_DARK,
                 font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=15, pady=(12, 8))

        ttk.Checkbutton(
            f_set,
            text="Enviar email de pendência automaticamente pro cliente "
                 "(APENAS pra campos externos — internas sempre manuais)",
            variable=self.auto_email_var,
            command=self._on_auto_email_change,
        ).pack(anchor="w", padx=15, pady=3)

        # Overrides globais (v2.6.0) — postar admissão SEM esses campos pra DP
        # completar manualmente no Alterdata Desktop. Evita pendências quando
        # o escritório prefere subir tudo e ajustar depois.
        ttk.Checkbutton(
            f_set,
            text="SEMPRE mandar a admissão SEM data de admissão "
                 "(DP completa manualmente no Desktop)",
            variable=self.sempre_sem_data_var,
            command=self._on_sempre_sem_data_change,
        ).pack(anchor="w", padx=15, pady=3)

        ttk.Checkbutton(
            f_set,
            text="SEMPRE mandar a admissão SEM função "
                 "(DP completa manualmente no Desktop)",
            variable=self.sempre_sem_funcao_var,
            command=self._on_sempre_sem_funcao_change,
        ).pack(anchor="w", padx=15, pady=3)

        # Reprocessar pendentes no polling (v2.9.0)
        ttk.Checkbutton(
            f_set,
            text="REPROCESSAR emails com label \"pendente\" na passada/polling "
                 "(remove a label e tenta de novo)",
            variable=self.reprocessar_pendentes_var,
            command=self._on_reprocessar_pendentes_change,
        ).pack(anchor="w", padx=15, pady=3)

        intv = tk.Frame(f_set, bg=COR_WHITE)
        intv.pack(anchor="w", padx=15, pady=(8, 0))
        tk.Label(intv, text="Intervalo de polling (segundos):",
                 bg=COR_WHITE, fg=COR_TEXT_DARK).pack(side="left")
        ttk.Spinbox(intv, from_=60, to=3600, increment=30,
                    textvariable=self.intervalo_var, width=8).pack(side="left", padx=8)
        # Conversão dinâmica segundos → minutos (atualiza em tempo real)
        self.intervalo_min_var = tk.StringVar()
        tk.Label(intv, textvariable=self.intervalo_min_var,
                 bg=COR_WHITE, fg=COR_TEXT_MUTED, font=("Segoe UI", 8)).pack(side="left")
        self.intervalo_var.trace_add("write", lambda *a: self._atualizar_label_intervalo())
        self._atualizar_label_intervalo()  # estado inicial

        ttk.Checkbutton(
            f_set, text="Modo escuro",
            variable=self.dark_mode_var,
            command=self._toggle_dark_mode,
        ).pack(anchor="w", padx=15, pady=(8, 12))

        # Atividade recente (log) — terminal style com brand colors
        f_log_card = tk.Frame(self.tab_main, bg=COR_NAVY_DARK, bd=0)
        f_log_card.pack(fill="both", expand=True, pady=(10, 0))
        tk.Label(f_log_card, text="Atividade recente",
                 bg=COR_NAVY_DARK, fg=COR_TEXT_MUTED,
                 font=("Segoe UI", 10)).pack(anchor="w", padx=15, pady=(10, 5))

        self.log_text = tk.Text(
            f_log_card, height=14, font=("Consolas", 10),
            bg=COR_NAVY_DARK, fg=COR_BEIGE, bd=0, padx=15, pady=10,
            insertbackground=COR_BEIGE, wrap="word",
            state="disabled",
        )
        self.log_text.pack(fill="both", expand=True, padx=0, pady=(0, 10))
        # Tag pra colorir timestamps em laranja
        self.log_text.tag_configure("ts", foreground=COR_ORANGE)

    def _card_stat(self, parent, titulo, var_valor, cor_valor, valor_font_size=28):
        """Cria um card branco com título cinza e valor grande colorido."""
        card = tk.Frame(parent, bg=COR_WHITE, bd=1, relief="solid",
                        highlightbackground=COR_BORDER, highlightthickness=0)
        tk.Label(card, text=titulo, bg=COR_WHITE, fg=COR_TEXT_MUTED,
                 font=("Segoe UI", 10)).pack(anchor="w", padx=18, pady=(15, 4))
        lbl = tk.Label(card, textvariable=var_valor, bg=COR_WHITE, fg=cor_valor,
                       font=("Segoe UI", valor_font_size, "bold"))
        lbl.pack(anchor="w", padx=18, pady=(0, 15))
        card._valor_label = lbl  # pra alterar cor depois (alerta billing)
        return card

    def _build_processadas(self):
        f = ttk.Frame(self.tab_proc)
        f.pack(fill="both", expand=True, padx=15, pady=15)

        # Toolbar
        bar = ttk.Frame(f)
        bar.pack(fill="x", pady=(0, 8))
        ttk.Button(bar, text="🔄  Atualizar", command=self._refresh_tabelas).pack(side="left", padx=5)
        ttk.Button(bar, text="📂  Abrir payloads/", command=self._abrir_pasta_payloads).pack(side="left", padx=5)
        ttk.Button(bar, text="📧  Abrir email no Gmail",
                   command=lambda: self._abrir_email_gmail(self.tree_proc)).pack(side="left", padx=5)

        # Filtro de busca
        ttk.Label(bar, text="  Buscar:").pack(side="left", padx=(20, 4))
        ttk.Entry(bar, textvariable=self.filtro_proc_var, width=30).pack(side="left")

        # Tabela — agrupada por dia (v2.8.0)
        cols = ("ts", "nome", "empresa", "cnpj", "procedencia")
        self.tree_proc = ttk.Treeview(f, columns=cols, show="tree headings", height=22)
        # Coluna #0 (tree) = agrupador "📅 dd/mm/aaaa — N admissões"
        self.tree_proc.heading("#0", text="Dia")
        self.tree_proc.column("#0", width=200, anchor="w", stretch=False)
        headers = {
            "ts": ("Hora", 80),
            "nome": ("Nome do colaborador", 240),
            "empresa": ("Empresa", 220),
            "cnpj": ("CNPJ", 140),
            "procedencia": ("Procedência", 420),
        }
        for c, (txt, w) in headers.items():
            self.tree_proc.heading(c, text=txt)
            self.tree_proc.column(c, width=w, anchor="w")

        sb = ttk.Scrollbar(f, orient="vertical", command=self.tree_proc.yview)
        self.tree_proc.configure(yscrollcommand=sb.set)
        self.tree_proc.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        # Tag pro nó agrupador (dia) — cor distintiva
        self.tree_proc.tag_configure(
            "dia_pai",
            background="#37474f", foreground="#fff",
            font=("Segoe UI", 9, "bold"),
        )

    def _build_pendentes(self):
        f = ttk.Frame(self.tab_pend)
        f.pack(fill="both", expand=True, padx=15, pady=15)

        # Toolbar — 2 linhas agrupadas por finalidade
        # Linha 1: ações sobre a pendência SELECIONADA
        bar_acoes = ttk.Frame(f)
        bar_acoes.pack(fill="x", pady=(0, 4))
        tk.Label(bar_acoes, text="Ação da pendência selecionada:",
                 bg=COR_BEIGE, fg=COR_TEXT_MUTED,
                 font=("Segoe UI", 8)).pack(side="left", padx=(0, 6))
        ttk.Button(bar_acoes, text="🔧  Resolver",
                   command=self._resolver_selecionada).pack(side="left", padx=3)
        ttk.Button(bar_acoes, text="🚀  Enviar mesmo assim",
                   command=self._enviar_mesmo_assim).pack(side="left", padx=3)
        ttk.Button(bar_acoes, text="🏢  Corrigir CNPJ",
                   command=self._corrigir_cnpj).pack(side="left", padx=3)
        ttk.Button(bar_acoes, text="🔢  Definir função",
                   command=self._definir_funcao_por_codigo).pack(side="left", padx=3)
        ttk.Button(bar_acoes, text="🔁  Reprocessar email",
                   command=self._reprocessar_selecionada).pack(side="left", padx=3)
        ttk.Button(bar_acoes, text="📧  Abrir no Gmail",
                   command=lambda: self._abrir_email_gmail(self.tree_pend)).pack(side="left", padx=3)

        # Linha 2: utilitários globais (recarregar caches, atualizar lista)
        bar_util = ttk.Frame(f)
        bar_util.pack(fill="x", pady=(0, 8))
        tk.Label(bar_util, text="Utilitários:",
                 bg=COR_BEIGE, fg=COR_TEXT_MUTED,
                 font=("Segoe UI", 8)).pack(side="left", padx=(0, 6))
        ttk.Button(bar_util, text="🔄  Recarregar planilha CBO",
                   command=self._recarregar_planilha_cbo).pack(side="left", padx=3)
        ttk.Button(bar_util, text="🏢  Atualizar cache empresas",
                   command=self._atualizar_empresas_cache).pack(side="left", padx=3)
        ttk.Button(bar_util, text="🔃  Atualizar lista",
                   command=self._refresh_tabelas).pack(side="left", padx=3)

        # Filtro de busca (na linha de utilitários, à direita)
        ttk.Label(bar_util, text="  Buscar:").pack(side="left", padx=(20, 4))
        ttk.Entry(bar_util, textvariable=self.filtro_pend_var, width=30).pack(side="left")

        ttk.Label(f, text=(
            "(agrupado por dia — clique no triângulo ▶ pra expandir/colapsar | "
            "clique 2× na linha pra resolver | amarelo/laranja = pendência >3 dias)"
        ), foreground="#666").pack(anchor="w", pady=(0, 5))

        # Tabela — agrupada por dia (v2.8.0)
        cols = ("ts", "nome", "empresa", "cnpj", "tipo", "cargo_ia", "procedencia")
        self.tree_pend = ttk.Treeview(f, columns=cols, show="tree headings", height=22)
        # Coluna #0 (tree) = agrupador "📅 dd/mm/aaaa — N pendências"
        self.tree_pend.heading("#0", text="Dia")
        self.tree_pend.column("#0", width=200, anchor="w", stretch=False)
        headers = {
            "ts": ("Hora", 80),
            "nome": ("Nome do colaborador", 200),
            "empresa": ("Empresa", 200),
            "cnpj": ("CNPJ", 130),
            "tipo": ("Tipo", 80),
            "cargo_ia": ("Cargo entendido pela IA", 180),
            "procedencia": ("Procedência / motivo", 320),
        }
        for c, (txt, w) in headers.items():
            self.tree_pend.heading(c, text=txt)
            self.tree_pend.column(c, width=w, anchor="w")

        sb = ttk.Scrollbar(f, orient="vertical", command=self.tree_pend.yview)
        self.tree_pend.configure(yscrollcommand=sb.set)
        self.tree_pend.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        self.tree_pend.bind("<Double-1>", lambda e: self._resolver_selecionada())

        # Cores por tipo + idade + agrupador
        self.tree_pend.tag_configure("interno", background="#fff3e0")
        self.tree_pend.tag_configure("cliente", background="#e3f2fd")
        self.tree_pend.tag_configure("velha_interno", background="#ffab40", foreground="#000")
        self.tree_pend.tag_configure("velha_cliente", background="#ffd180", foreground="#000")
        self.tree_pend.tag_configure(
            "dia_pai",
            background="#37474f", foreground="#fff",
            font=("Segoe UI", 9, "bold"),
        )

    # ---- Consumidor da fila thread→UI --------------------------

    def _consumir_fila(self):
        """Chamado a cada 200ms pelo Tkinter main loop. Lê eventos
        postados pela thread de polling e atualiza UI."""
        try:
            while True:
                kind, data = self.gui_q.get_nowait()
                if kind == "status":
                    self.status_var.set(data)
                elif kind == "ultima":
                    self.ultima_var.set(data)
                elif kind == "proxima":
                    self.proxima_var.set(data)
                elif kind == "log":
                    self._append_log(data)
                elif kind == "log_raw":
                    self._append_log_raw(data)
                elif kind == "refresh":
                    self._refresh_tabelas()
        except queue.Empty:
            pass
        self.after(200, self._consumir_fila)

    def _append_log(self, msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        # Timestamp em laranja, resto em bege
        self.log_text.insert("end", f"[{ts}] ", ("ts",))
        self.log_text.insert("end", f"{msg}\n")
        self.log_text.see("end")
        n_lines = int(self.log_text.index("end-1c").split(".")[0])
        if n_lines > 5000:
            self.log_text.delete("1.0", "2500.0")
        self.log_text.configure(state="disabled")

    def _append_log_raw(self, msg: str):
        """Mensagem ja formatada por _TkQueueLogHandler (com asctime/level/name).
        Mostra crua, sem prepender timestamp adicional."""
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"{msg}\n")
        self.log_text.see("end")
        n_lines = int(self.log_text.index("end-1c").split(".")[0])
        if n_lines > 5000:
            self.log_text.delete("1.0", "2500.0")
        self.log_text.configure(state="disabled")

    def _atualizar_label_intervalo(self):
        """Conversão dinâmica segundos → minutos pra exibir ao lado do spinbox.

        Exemplos:
            60  → '= 1 min'
            90  → '= 1 min 30 s'
            300 → '= 5 min'
            3600 → '= 1 h'
        """
        try:
            segundos = int(self.intervalo_var.get())
        except (tk.TclError, ValueError):
            self.intervalo_min_var.set("(valor inválido)")
            return
        if segundos <= 0:
            self.intervalo_min_var.set("(zero)")
            return
        if segundos >= 3600:
            horas = segundos / 3600
            self.intervalo_min_var.set(
                f"= {horas:.0f} h" if horas == int(horas) else f"= {horas:.2f} h"
            )
            return
        minutos = segundos // 60
        resto = segundos % 60
        if minutos == 0:
            self.intervalo_min_var.set(f"= {resto} s")
        elif resto == 0:
            unidade = "min" if minutos > 1 else "min"
            self.intervalo_min_var.set(f"= {minutos} {unidade}")
        else:
            self.intervalo_min_var.set(f"= {minutos} min {resto} s")

    # ---- Ações de controle -------------------------------------

    def _abrir_importar_arquivos(self):
        """Abre dialog pra importar arquivos manualmente (sem email Gmail).
        Útil quando docs chegam por WhatsApp/Drive/pendrive."""
        ImportarArquivosDialog(
            parent=self,
            config=self.config,
            claude=self.claude,
            planilha_cbo=self.planilha,
            gui_q=self.gui_q,
        )

    def _iniciar_polling(self):
        if self.polling:
            return
        self.config.intervalo = max(60, int(self.intervalo_var.get()))
        self.polling = True
        self.polling_thread = threading.Thread(target=self._loop_polling, daemon=True)
        self.polling_thread.start()
        self.btn_iniciar.configure(state="disabled")
        self.btn_parar.configure(state="normal")
        self.gui_q.put(("log", f"▶ Polling iniciado (intervalo {self.config.intervalo}s)"))

    def _parar_polling(self):
        if not self.polling:
            return
        self.polling = False
        self.btn_parar.configure(state="disabled")
        self.btn_iniciar.configure(state="normal")
        self.gui_q.put(("status", "Parando..."))
        self.gui_q.put(("log", "Pedido de parada — vai parar após a passada atual"))

    def _rodar_unica(self):
        if self.polling:
            messagebox.showinfo("Polling ativo",
                                "O polling já está rodando passadas automaticamente.\nPara rodar uma extra, primeiro pare o polling.")
            return
        threading.Thread(target=self._executar_passada, daemon=True).start()

    def _executar_passada(self):
        self.gui_q.put(("status", "Executando..."))
        self.gui_q.put(("log", "Iniciando passada única..."))
        try:
            rodar_uma_passada(self.config, self.claude, self.planilha)
            self.gui_q.put(("log", "Passada única concluída"))
        except Exception as e:
            self.gui_q.put(("log", f"Erro na passada: {e}"))
        self.gui_q.put(("status", "Parado"))
        self.gui_q.put(("ultima", datetime.now().strftime("%d/%m/%Y %H:%M:%S")))
        self.gui_q.put(("refresh", None))

    def _loop_polling(self):
        while self.polling:
            self.gui_q.put(("status", "Executando passada..."))
            self.gui_q.put(("log", "Iniciando passada..."))
            try:
                rodar_uma_passada(self.config, self.claude, self.planilha)
                self.gui_q.put(("log", "Passada concluída"))
            except Exception as e:
                self.gui_q.put(("log", f"Erro na passada: {e}"))

            self.gui_q.put(("ultima", datetime.now().strftime("%d/%m/%Y %H:%M:%S")))
            self.gui_q.put(("refresh", None))

            t_end = time.time() + self.config.intervalo
            while time.time() < t_end and self.polling:
                t_left = int(t_end - time.time())
                self.gui_q.put(("status", f"Aguardando próxima ({t_left}s)"))
                self.gui_q.put(("proxima", datetime.fromtimestamp(t_end).strftime("%H:%M:%S")))
                time.sleep(1)

        self.gui_q.put(("status", "Parado"))
        self.gui_q.put(("proxima", "—"))
        self.gui_q.put(("log", "Polling parado"))

    def _on_auto_email_change(self):
        self.config.auto_email_pendencias = bool(self.auto_email_var.get())
        estado = "LIGADO ✉" if self.config.auto_email_pendencias else "DESLIGADO 🔇"
        self.gui_q.put(("log", f"⚙ Auto-email de pendências (apenas externas): {estado}"))

    def _persistir_flag_config(self, chave: str, valor: bool) -> None:
        """Reescreve config.json com uma flag boolean atualizada. Preserva
        todas as outras chaves intactas. Falha silenciosa não trava UI."""
        try:
            from main import CONFIG_FILE
            raw = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            raw[chave] = valor
            CONFIG_FILE.write_text(
                json.dumps(raw, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        except Exception as e:
            self.gui_q.put((
                "log",
                f"⚠ Falha persistindo {chave}={valor} em config.json: "
                f"{type(e).__name__}: {e}"
            ))

    def _on_sempre_sem_data_change(self):
        valor = bool(self.sempre_sem_data_var.get())
        self.config.sempre_mandar_sem_data_admissao = valor
        self._persistir_flag_config("sempre_mandar_sem_data_admissao", valor)
        estado = "LIGADO 📅⛔" if valor else "DESLIGADO ✅"
        self.gui_q.put((
            "log",
            f"⚙ Sempre mandar admissão SEM data: {estado} "
            f"(DP completa no Desktop quando ligado)"
        ))

    def _on_sempre_sem_funcao_change(self):
        valor = bool(self.sempre_sem_funcao_var.get())
        self.config.sempre_mandar_sem_funcao = valor
        self._persistir_flag_config("sempre_mandar_sem_funcao", valor)
        estado = "LIGADO 💼⛔" if valor else "DESLIGADO ✅"
        self.gui_q.put((
            "log",
            f"⚙ Sempre mandar admissão SEM função: {estado} "
            f"(DP completa no Desktop quando ligado)"
        ))

    def _on_reprocessar_pendentes_change(self):
        valor = bool(self.reprocessar_pendentes_var.get())
        self.config.reprocessar_pendentes_no_polling = valor
        self._persistir_flag_config("reprocessar_pendentes_no_polling", valor)
        estado = "LIGADO 🔁" if valor else "DESLIGADO ✅"
        self.gui_q.put((
            "log",
            f"⚙ Reprocessar pendentes no polling: {estado} "
            f"(emails com label pendente são retentados a cada passada)"
        ))

    def _abrir_pasta_payloads(self):
        if PAYLOADS_DIR.exists():
            import os
            os.startfile(PAYLOADS_DIR)  # Windows
        else:
            messagebox.showinfo("Sem payloads", "Pasta payloads/ ainda não existe.")

    # ---- Atualização das tabelas -------------------------------

    def _ler_cargo_ia_do_payload(self, msg_id: str) -> str:
        """Procura em payloads/ o JSON mais recente da pendência (por msg_id)
        e retorna o cargo que o Claude extraiu. Cache em RAM por msg_id.

        Usado pra exibir a coluna 'Cargo entendido pela IA' na tabela
        de pendências — ajuda o operador a ver de cara se a extração foi
        ruim (ex: planilha CBO sem o cargo, ou Claude errou o nome)."""
        if not msg_id:
            return ""
        cache = getattr(self, "_cache_cargo_ia", None)
        if cache is None:
            cache = {}
            self._cache_cargo_ia = cache
        if msg_id in cache:
            return cache[msg_id]
        try:
            if not PAYLOADS_DIR.exists():
                cache[msg_id] = ""
                return ""
            arqs = sorted(PAYLOADS_DIR.glob(f"*_{msg_id[:16]}*.json"), reverse=True)
            for arq in arqs:
                try:
                    doc = json.loads(arq.read_text(encoding="utf-8"))
                except (json.JSONDecodeError, OSError):
                    continue
                # Fontes em ordem de preferência:
                # 1. resolucao.cargo_extraido (formato novo desde v2.2.0)
                # 2. payload.data.attributes.nomecargo (extraído pelo Claude)
                # 3. payload.cargo_extraido (formato top-level de bloco antigo)
                payload = doc.get("payload") or {}
                attrs = (payload.get("data") or {}).get("attributes") or {}
                cargo = (
                    (doc.get("resolucao") or {}).get("cargo_extraido")
                    or attrs.get("nomecargo")
                    or payload.get("cargo_extraido")
                )
                if cargo:
                    cache[msg_id] = str(cargo)
                    return cache[msg_id]
        except Exception:
            pass
        cache[msg_id] = ""
        return ""

    def _refresh_tabelas(self):
        """Lê admissoes.xlsx, popula as 2 tabelas, aplica filtros,
        destaca pendências velhas, atualiza contadores e billing."""
        # Invalida cache de cargo IA — payloads novos podem ter aparecido
        self._cache_cargo_ia = {}
        for tv in (self.tree_proc, self.tree_pend):
            for item in tv.get_children():
                tv.delete(item)

        # Atualiza billing/limite
        self._atualizar_billing()

        if not PLANILHA_ADMISSOES.exists():
            self.contador_proc_var.set("0")
            self.contador_pend_var.set("0")
            self.contador_velha_var.set("0")
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(PLANILHA_ADMISSOES, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) <= 1:
                self.contador_proc_var.set("0")
                self.contador_pend_var.set("0")
                self.contador_velha_var.set("0")
                return

            filtro_p = (self.filtro_proc_var.get() or "").strip().lower()
            filtro_e = (self.filtro_pend_var.get() or "").strip().lower()
            hoje = datetime.now()

            # Agrupador por dia (v2.8.0) — coleta primeiro, agrupa por data e
            # insere depois com nó pai expansível. Mais legível pro operador
            # quando há volume de admissões acumulado.
            proc_por_dia: dict[str, list[tuple]] = {}
            pend_por_dia: dict[str, list[tuple]] = {}

            n_proc = 0
            n_pend = 0
            n_velha = 0

            # v2.14.1 (ITEM 6): dedup de pendências por ENTIDADE (msg_id+nome+cnpj).
            # admissoes.xlsx é append-only — cada tentativa de POST grava 1 linha.
            # Bug real 12/06: aba mostrava "68 pendentes" mas a fila era 11.
            # Solução: 1ª passada agrega o ÚLTIMO estado por (msg_id+nome+cnpj);
            # se o último estado é "Cadastrado", a entidade saiu da fila.
            # A aba Auditoria continua mostrando TODOS os eventos.
            entidade_ultimo: dict[tuple, dict] = {}
            for row in rows[1:]:
                if not row or not any(row):
                    continue
                cols = list(row) + [""] * (6 - len(row))
                primeira = str(cols[0] or "")
                if ("-" in primeira and ":" in primeira) or primeira == "":
                    ts, nome, empresa, cnpj, procedencia, msg_id = cols[:6]
                else:
                    nome, empresa, cnpj, procedencia = cols[:4]
                    ts = ""
                    msg_id = ""
                ts_s = str(ts or "")
                nome_s = str(nome or "").strip()
                empresa_s = str(empresa or "").strip()
                cnpj_s = str(cnpj or "").strip()
                proc_s = str(procedencia or "")
                msg_id_s = str(msg_id or "").strip()
                chave = (msg_id_s, nome_s.upper(), cnpj_s)
                ant = entidade_ultimo.get(chave)
                # Mantém o mais RECENTE (ts maior). Empate em ts vazio → sobrescreve.
                if ant is None or ts_s > ant.get("ts", ""):
                    entidade_ultimo[chave] = {
                        "ts": ts_s, "nome": nome_s, "empresa": empresa_s,
                        "cnpj": cnpj_s, "procedencia": proc_s,
                        "msg_id": msg_id_s,
                    }
            # Pra preservar comportamento atual (mais recente em cima nas listas),
            # processo as entidades em ordem reversa de ts.
            linhas_dedup = sorted(
                entidade_ultimo.values(), key=lambda e: e["ts"], reverse=True,
            )
            for entidade in linhas_dedup:
                row = (
                    entidade["ts"], entidade["nome"], entidade["empresa"],
                    entidade["cnpj"], entidade["procedencia"], entidade["msg_id"],
                )
                if not row or not any(row):
                    continue
                # 6 colunas (novas) ou 4 (legacy): detecta
                cols = list(row) + [""] * (6 - len(row))
                # Heurística: se primeira coluna parece data (tem '-' ou ':'),
                # é planilha nova. Senão é legacy (4 cols: nome,empresa,cnpj,proc).
                primeira = str(cols[0] or "")
                if ("-" in primeira and ":" in primeira) or primeira == "":
                    ts, nome, empresa, cnpj, procedencia, msg_id = cols[:6]
                else:
                    nome, empresa, cnpj, procedencia = cols[:4]
                    ts = ""
                    msg_id = ""

                ts = str(ts or "")
                nome = str(nome or "")
                empresa = str(empresa or "")
                cnpj = str(cnpj or "")
                procedencia = str(procedencia or "")
                msg_id = str(msg_id or "")
                proc_low = procedencia.lower()

                # Extrai dia (YYYY-MM-DD) e hora (HH:MM:SS) do timestamp
                if ts and " " in ts:
                    dia_iso, hora_str = ts.split(" ", 1)
                else:
                    dia_iso, hora_str = "(sem data)", ts

                # Concatena tudo pra match de filtro
                hay_texto = " ".join((ts, nome, empresa, cnpj, procedencia)).lower()

                if proc_low.startswith("cadastrado") or proc_low.startswith("dry-run"):
                    if filtro_p and filtro_p not in hay_texto:
                        continue
                    proc_por_dia.setdefault(dia_iso, []).append((
                        hora_str, nome, empresa, cnpj, procedencia, msg_id
                    ))
                    n_proc += 1
                elif proc_low.startswith("pendente") or proc_low.startswith("falha"):
                    cargo_ia = self._ler_cargo_ia_do_payload(msg_id)
                    # Inclui cargo na busca pra operador filtrar por cargo extraído
                    hay_texto_pend = hay_texto + " " + cargo_ia.lower()
                    if filtro_e and filtro_e not in hay_texto_pend:
                        continue
                    tipo = "interno" if "interno" in proc_low else "cliente"
                    # Detecta pendência velha
                    velha = False
                    try:
                        dt_row = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
                        if (hoje - dt_row).days >= DIAS_PENDENCIA_VELHA:
                            velha = True
                            n_velha += 1
                    except (ValueError, TypeError):
                        pass
                    pend_por_dia.setdefault(dia_iso, []).append((
                        hora_str, nome, empresa, cnpj, tipo, cargo_ia, procedencia,
                        msg_id, velha
                    ))
                    n_pend += 1

            # Insere os agrupadores de dia + filhos. Dias mais recentes primeiro.
            # Dia de hoje fica EXPANDIDO; os outros colapsados (operador clica).
            hoje_iso = hoje.strftime("%Y-%m-%d")

            def _fmt_dia(iso: str) -> str:
                """ISO YYYY-MM-DD → 'dd/mm/aaaa' (ou '(sem data)')."""
                if iso == "(sem data)":
                    return iso
                try:
                    d = datetime.strptime(iso, "%Y-%m-%d")
                    return d.strftime("%d/%m/%Y")
                except ValueError:
                    return iso

            for dia in sorted(proc_por_dia.keys(), reverse=True):
                filhos = proc_por_dia[dia]
                rotulo = f"📅 {_fmt_dia(dia)} — {len(filhos)} admissão{'s' if len(filhos) != 1 else ''}"
                if dia == hoje_iso:
                    rotulo += "  • HOJE"
                pai = self.tree_proc.insert(
                    "", "end", text=rotulo, tags=("dia_pai",),
                    open=(dia == hoje_iso),
                )
                for (hora, nome, empresa, cnpj, procedencia, msg_id) in filhos:
                    self.tree_proc.insert(
                        pai, "end",
                        values=(hora, nome, empresa, cnpj, procedencia),
                        tags=(msg_id,),
                    )

            for dia in sorted(pend_por_dia.keys(), reverse=True):
                filhos = pend_por_dia[dia]
                # Contadores por tipo no rótulo
                n_int = sum(1 for f in filhos if f[4] == "interno")
                n_cli = sum(1 for f in filhos if f[4] == "cliente")
                resumo = []
                if n_int:
                    resumo.append(f"{n_int} interno{'s' if n_int != 1 else ''}")
                if n_cli:
                    resumo.append(f"{n_cli} cliente{'s' if n_cli != 1 else ''}")
                rotulo = (
                    f"📅 {_fmt_dia(dia)} — {len(filhos)} "
                    f"pendência{'s' if len(filhos) != 1 else ''}"
                )
                if resumo:
                    rotulo += f" ({', '.join(resumo)})"
                if dia == hoje_iso:
                    rotulo += "  • HOJE"
                pai = self.tree_pend.insert(
                    "", "end", text=rotulo, tags=("dia_pai",),
                    open=(dia == hoje_iso),
                )
                for (hora, nome, empresa, cnpj, tipo, cargo_ia,
                     procedencia, msg_id, velha) in filhos:
                    tags = [tipo, msg_id]
                    if velha:
                        tags.append(f"velha_{tipo}")
                    self.tree_pend.insert(
                        pai, "end",
                        values=(hora, nome, empresa, cnpj, tipo, cargo_ia, procedencia),
                        tags=tags,
                    )

            self.contador_proc_var.set(str(n_proc))
            self.contador_pend_var.set(str(n_pend))
            self.contador_velha_var.set(str(n_velha))

            # Badge no botão "Pendentes" do sidebar — mostra contagem entre parênteses
            btn_pend = self._nav_buttons.get("pendentes") if hasattr(self, "_nav_buttons") else None
            if btn_pend is not None:
                if n_pend > 0:
                    btn_pend.configure(text=f"   Pendentes  ({n_pend})")
                else:
                    btn_pend.configure(text="   Pendentes")

            # Detecta pendências NOVAS (não vistas antes nessa sessão) e dispara toast
            try:
                ids_atuais = self._coletar_msg_ids_pendentes()
                if self._primeira_carga:
                    self._msg_ids_pendentes_conhecidos = ids_atuais
                    self._primeira_carga = False
                else:
                    novas = ids_atuais - self._msg_ids_pendentes_conhecidos
                    if novas:
                        self._notify_toast(
                            "⚠ Nova pendência",
                            f"{len(novas)} nova(s) admissão(ões) pendente(s) na fila.\n"
                            f"Veja na aba Pendentes.",
                        )
                    self._msg_ids_pendentes_conhecidos = ids_atuais
            except Exception:
                pass

            # Alerta de billing — toast quando passar do limite
            if hasattr(self, "_billing_alertou_dessa_sessao"):
                pass
            else:
                self._billing_alertou_dessa_sessao = False
            b = sum_billing_mes_atual()
            if (
                b["custo_usd"] >= self.billing_limite_usd
                and not self._billing_alertou_dessa_sessao
            ):
                self._notify_toast(
                    "💰 Limite de billing atingido",
                    f"O custo da API Claude no mês corrente passou de US$ "
                    f"{self.billing_limite_usd:.2f} (atual: US$ {b['custo_usd']:.4f}).",
                )
                self._billing_alertou_dessa_sessao = True

            # Atualiza outras tabs também
            if hasattr(self, "tree_audit"):
                self._refresh_auditoria()
            if hasattr(self, "tree_api"):
                self._refresh_api_econtador()
            if hasattr(self, "stats_body"):
                self._refresh_estatisticas()
        except Exception as e:
            self.gui_q.put(("log", f"⚠ Erro lendo planilha: {e}"))

    def _coletar_msg_ids_pendentes(self) -> set[str]:
        """Conjunto de msg_ids que estão como pendência na planilha agora.

        v2.14.1 (ITEM 6): dedup por (msg_id+nome+cnpj). Antes contava cada
        TENTATIVA — então 8 reprocessos do RAIMUNDO faziam 8 'novas pendências'
        no toast. Agora só conta a entidade UMA vez, pelo último estado dela.
        """
        ids = set()
        if not PLANILHA_ADMISSOES.exists():
            return ids
        try:
            from openpyxl import load_workbook
            wb = load_workbook(PLANILHA_ADMISSOES, read_only=True)
            ws = wb.active
            # 1ª passada: pega o ÚLTIMO evento por entidade
            entidade_ultimo: dict[tuple, dict] = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                cols = list(row) + [""] * (6 - len(row))
                primeira = str(cols[0] or "")
                if ("-" in primeira and ":" in primeira) or primeira == "":
                    ts, nome, _emp, cnpj, proc, mid = cols[:6]
                else:
                    nome, _emp, cnpj, proc = cols[:4]
                    ts, mid = "", ""
                chave = (str(mid or ""), str(nome or "").upper(), str(cnpj or ""))
                ant = entidade_ultimo.get(chave)
                if ant is None or str(ts or "") > ant.get("ts", ""):
                    entidade_ultimo[chave] = {"ts": str(ts or ""), "proc": str(proc or ""), "mid": str(mid or "")}
            # 2ª passada: só entidades cujo último estado é pendência/falha
            for e in entidade_ultimo.values():
                pl = e["proc"].lower()
                if (pl.startswith("pendente") or pl.startswith("falha")) and e["mid"]:
                    ids.add(e["mid"])
        except Exception:
            pass
        return ids

    def _atualizar_billing(self):
        """Lê billing.ndjson (Claude USD) + directdata_audit.ndjson (BRL) e
        soma TUDO em BRL pro card 'APIs externas/mês'. Atualiza cor com
        gradiente do limite (verde < 50%, laranja 50-100%, vermelho >100%)."""
        try:
            # Claude (USD)
            resumo_cl = sum_billing_mes_atual()
            custo_usd = resumo_cl["custo_usd"]
            self.billing_var.set(f"US$ {custo_usd:.2f}")  # mantido pra Estatísticas

            # Direct Data (BRL)
            resumo_dd = sum_directdata_mes_atual()
            custo_brl_dd = resumo_dd["custo_brl"]

            # Total agregado em BRL (Claude USD × cotação + DirectData BRL)
            total_brl = custo_usd * self.cotacao_usd_brl + custo_brl_dd
            self.billing_total_brl_var.set(f"R$ {total_brl:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

            # Cor baseada em limite (USD pra manter retro-compat com config)
            limite_total_brl = self.billing_limite_usd * self.cotacao_usd_brl
            if total_brl >= limite_total_brl:
                self.billing_label.configure(foreground="#c62828")
            elif total_brl >= limite_total_brl * 0.5:
                self.billing_label.configure(foreground="#e65100")
            else:
                self.billing_label.configure(foreground="#1565c0")
        except Exception:
            pass

    def _msg_id_from_selection(self, tree: ttk.Treeview) -> str | None:
        sel = tree.selection()
        if not sel:
            return None
        iid = sel[0]
        # v2.8.0: se selecionou nó pai (dia), não há msg_id — ignora silencioso
        if not tree.parent(iid):
            return None
        tags = tree.item(iid)["tags"] or []
        # tags pode ter "interno"/"cliente"/"velha_*"/"dia_pai" + msg_id
        for t in tags:
            t = str(t)
            if t and t not in ("interno", "cliente", "velha_interno",
                                "velha_cliente", "dia_pai"):
                return t
        return None

    def _abrir_email_gmail(self, tree: ttk.Treeview):
        msg_id = self._msg_id_from_selection(tree)
        if not msg_id:
            messagebox.showinfo("Selecione",
                                "Selecione uma linha primeiro.\n\n"
                                "(Linhas antigas — antes do timestamp/msg_id — não conseguem abrir.)")
            return
        import webbrowser
        url = f"https://mail.google.com/mail/u/0/#all/{msg_id}"
        webbrowser.open(url)

    def _reprocessar_selecionada(self):
        msg_id = self._msg_id_from_selection(self.tree_pend)
        if not msg_id:
            messagebox.showinfo("Selecione",
                                "Selecione uma pendência com msg_id (planilha nova).\n\n"
                                "Linhas legacy precisam ser resolvidas com os outros botões.")
            return
        # v2.14.0 — guarda anti-churn: avisa se NADA mudou desde a última
        # tentativa deste msg_id (caso RAIMUNDO: 8 cliques idênticos/dia).
        texto = (
            "Vou:\n"
            "  1. Remover labels processado/pendente da mensagem no Gmail\n"
            "  2. Rodar uma passada agora pra que o pipeline pegue de novo\n\n"
        )
        aviso = None
        try:
            aviso = idempotencia.aviso_reprocesso(msg_id)
        except Exception:
            pass
        if aviso:
            texto += aviso + "\n\n"
        texto += "Confirma?"
        if not messagebox.askyesno("Confirmar reprocessar", texto,
                                   default=messagebox.NO if aviso else messagebox.YES):
            return
        threading.Thread(
            target=self._reprocessar_worker, args=(msg_id,), daemon=True
        ).start()

    def _reprocessar_worker(self, msg_id: str):
        self.gui_q.put(("log", f"🔁 Reprocessando msg {msg_id[:16]}..."))
        try:
            gmail = GmailClient()
            for lbl in (self.config.label_processado, self.config.label_pendente):
                try:
                    gmail.remover_label(msg_id, lbl)
                except Exception:
                    pass
            self.gui_q.put(("log", "   ✓ Labels removidas — rodando passada..."))
            rodar_uma_passada(self.config, self.claude, self.planilha)
            idempotencia.salvar_fingerprint_reprocesso(msg_id)
            self.gui_q.put(("log", "✓ Reprocessamento concluído"))
        except Exception as e:
            self.gui_q.put(("log", f"❌ Erro reprocessando: {e}"))
        self.gui_q.put(("refresh", None))

    def _corrigir_cnpj(self):
        """Quando o cliente errou o CNPJ nos documentos: usuário informa o
        correto, salvamos um override pro msg_id, removemos as labels do
        Gmail e re-processamos. O override é aplicado em processar_admissao
        antes de qualquer chamada ao eContador — não pede nada pro cliente."""
        from tkinter import simpledialog

        msg_id = self._msg_id_from_selection(self.tree_pend)
        if not msg_id:
            messagebox.showinfo(
                "Selecione",
                "Selecione uma pendência com msg_id (planilha nova) antes "
                "de corrigir o CNPJ."
            )
            return

        # Lista todas as linhas pendentes do mesmo email
        # v2.8.0: itera filhos (admissões) pulando os agrupadores de dia
        afetados = []
        for iid in self._iter_filhos_admissao(self.tree_pend):
            tags = [str(t) for t in (self.tree_pend.item(iid)["tags"] or [])]
            if msg_id in tags:
                valores = self.tree_pend.item(iid)["values"]
                # cols (v2.8.0): hora, nome, empresa, cnpj, tipo, cargo_ia, procedencia
                nome = str(valores[1]) if len(valores) > 1 else "?"
                cnpj_atual = str(valores[3]) if len(valores) > 3 else "?"
                afetados.append((nome, cnpj_atual))

        if not afetados:
            messagebox.showinfo("Nada a corrigir", "Não achei linhas pendentes desse email.")
            return

        # Pega o CNPJ atual (que veio errado da extração) pra mostrar
        cnpj_extraido = afetados[0][1]
        lista_nomes = "\n".join(f"  • {n}" for n, _ in afetados)

        novo = simpledialog.askstring(
            "Corrigir CNPJ do email",
            f"O CNPJ atualmente extraído é: {cnpj_extraido}\n\n"
            f"Esta correção será aplicada a {len(afetados)} pendência(s) "
            f"deste email:\n{lista_nomes}\n\n"
            "Digite o CNPJ correto (com ou sem formatação):",
            parent=self,
        )
        if not novo:
            return

        cnpj_limpo = so_digitos(novo)
        if not validar_cnpj(cnpj_limpo):
            messagebox.showerror(
                "CNPJ inválido",
                f"O CNPJ '{novo}' não passa na validação de dígito verificador.\n\n"
                "Confira se digitou corretamente."
            )
            return

        if not messagebox.askyesno(
            "Confirmar correção",
            f"CNPJ correto: {formatar_cnpj(cnpj_limpo)}\n\n"
            f"Vai aplicar esta correção e reprocessar {len(afetados)} pendência(s).\n"
            "O cliente NÃO recebe email — correção feita internamente.\n\n"
            "Continuar?"
        ):
            return

        # Salva override e reprocessa
        salvar_cnpj_override(msg_id, cnpj_limpo)
        self.gui_q.put(("log",
            f"🏢 CNPJ override salvo pra msg {msg_id[:16]}: "
            f"{cnpj_extraido} → {formatar_cnpj(cnpj_limpo)}"))

        # Oferecer salvar como alias permanente do remetente (v2.10.0).
        # Tenta achar o remetente no payload salvo desse msg_id; se achar,
        # pergunta. Se não achar, pula silenciosamente.
        try:
            remetente = self._buscar_remetente_no_payload(msg_id)
            if remetente:
                if messagebox.askyesno(
                    "💾 Salvar como alias permanente?",
                    f"Quer que TODO email vindo de\n\n"
                    f"   {remetente}\n\n"
                    f"use automaticamente o CNPJ {formatar_cnpj(cnpj_limpo)}?\n\n"
                    f"(Recomendado se este remetente sempre representa a mesma "
                    f"empresa — você nunca mais precisa corrigir CNPJ pra ele.)"
                ):
                    from remetente_empresa import salvar_alias as _salvar_alias
                    _salvar_alias(
                        remetente=remetente,
                        cnpj=cnpj_limpo,
                        razao_social=afetados[0][0] if afetados else "",
                        fonte="manual",
                    )
                    self.gui_q.put((
                        "log",
                        f"💾 Alias permanente salvo: {remetente} → {formatar_cnpj(cnpj_limpo)}"
                    ))
        except Exception as e:
            self.gui_q.put((
                "log",
                f"⚠ Falha oferecendo alias permanente: {type(e).__name__}: {e}"
            ))

        threading.Thread(
            target=self._reprocessar_worker, args=(msg_id,), daemon=True
        ).start()

    def _buscar_remetente_no_payload(self, msg_id: str) -> str:
        """Procura em payloads/ o JSON mais recente desse msg_id e extrai
        o campo `remetente` dos metadados salvos. Retorna '' se não achar."""
        if not msg_id or not PAYLOADS_DIR.exists():
            return ""
        try:
            for arq in sorted(
                PAYLOADS_DIR.glob(f"*_{msg_id[:16]}*.json"), reverse=True
            ):
                try:
                    doc = json.loads(arq.read_text(encoding="utf-8"))
                except (OSError, json.JSONDecodeError):
                    continue
                remetente = doc.get("remetente") or ""
                if remetente:
                    return str(remetente)
        except Exception:
            pass
        return ""

    def _carregar_empresas_cache_background(self):
        """Roda no startup em thread separada — carrega o cache de CNPJs
        do eContador em RAM pra auto-correção de typos.

        Falha silenciosa: se não conseguir carregar (eContador offline, token
        ruim), o pipeline continua normal — só a auto-correção fica desligada.
        """
        self.gui_q.put(("log", "📥 Carregando cache de empresas (1x por sessão)..."))
        api = EContadorAPI(self.config.base_url, self.config.token)
        try:
            cache = recarregar_empresas_cache(api)
            if cache.carregado:
                self.gui_q.put(("log",
                    f"✓ Cache de empresas pronto: {len(cache)} CNPJs no whitelist "
                    f"(auto-correção de typos habilitada)"))
            else:
                self.gui_q.put(("log",
                    "⚠ Cache de empresas vazio — auto-correção desabilitada"))
        except Exception as e:
            self.gui_q.put(("log", f"⚠ Falha carregando cache de empresas: {e}"))
        finally:
            api.close()

    def _atualizar_empresas_cache(self):
        """Força recarregamento do cache (botão na UI). Útil quando o operador
        cadastra empresa nova no eContador e quer aplicar AGORA sem reabrir."""
        threading.Thread(
            target=self._carregar_empresas_cache_background,
            daemon=True,
        ).start()
        messagebox.showinfo(
            "Atualizando cache",
            "Cache de empresas sendo recarregado em background.\n"
            "Acompanhe no log de atividade (canto inferior)."
        )

    def _recarregar_planilha_cbo(self):
        """Releita funcoes_cbo.xlsx do disco. Útil quando você editou a
        planilha externamente (adicionou função nova, marcou X) e quer que
        a próxima passada use já o conteúdo novo, sem fechar/reabrir o app."""
        try:
            self.planilha = carregar_planilha(PLANILHA_CBO)
            marcadas = sum(1 for f in self.planilha if str(f.get("usar", "")).strip().upper() == "X")
            msg = (f"✓ Planilha CBO recarregada: {len(self.planilha)} função(ões) "
                   f"total, {marcadas} marcada(s) com X")
            self.gui_q.put(("log", msg))
            messagebox.showinfo("Recarregada", msg)
        except Exception as e:
            self.gui_q.put(("log", f"❌ Falha recarregando planilha CBO: {e}"))
            messagebox.showerror("Erro", f"Falha recarregando planilha:\n{e}")

    def _definir_funcao_por_codigo(self):
        """Quando uma admissão pendou por função ambígua/não cadastrada:
        usuário informa o código (coluna 'codigo' da planilha CBO, ex 009332).
        Busca o funcao_id correspondente, salva override pra (msg_id, nome) e
        reprocessa. A função fica fixada pra essa admissão específica."""
        from tkinter import simpledialog

        msg_id = self._msg_id_from_selection(self.tree_pend)
        if not msg_id:
            messagebox.showinfo("Selecione",
                                "Selecione uma pendência com msg_id antes de definir função.")
            return

        # Pega nome do funcionário da linha selecionada
        sel = self.tree_pend.selection()
        valores = self.tree_pend.item(sel[0])["values"]
        nome = str(valores[1]) if len(valores) > 1 else ""
        if not nome or nome.startswith("(nome"):
            messagebox.showerror("Sem nome",
                                 "Não consigo identificar o nome do funcionário nessa linha.")
            return

        codigo = simpledialog.askstring(
            "Definir função por código",
            f"Funcionário: {nome}\n\n"
            "Digite o código da função (coluna 'codigo' da planilha CBO).\n"
            "Exemplo: 009332 (OPERADOR(A) DE AGLUTINADOR)\n\n"
            "Código:",
            parent=self,
        )
        if not codigo:
            return

        codigo_limpo = codigo.strip()
        # Busca na planilha em memória (forma robusta: aceita "9332" ou "009332")
        achadas = []
        for f in self.planilha:
            cod_planilha = str(f.get("codigo") or "").strip()
            # Normaliza zeros à esquerda pra comparar
            if cod_planilha == codigo_limpo or cod_planilha.lstrip("0") == codigo_limpo.lstrip("0"):
                achadas.append(f)

        if not achadas:
            messagebox.showerror(
                "Código não encontrado",
                f"O código '{codigo_limpo}' não está na planilha CBO em memória.\n\n"
                "Verifique:\n"
                "  1. O código está mesmo na planilha (abra funcoes_cbo.xlsx)\n"
                "  2. Se você acabou de adicionar, clique 🔄 Recarregar planilha CBO"
            )
            return

        if len(achadas) > 1:
            msgs = "\n".join(
                f"  • id={f['funcao_id']}, cbo={f.get('cbo')}, nome={f.get('nome_cargo')}"
                for f in achadas[:5]
            )
            messagebox.showerror(
                "Múltiplas funções com mesmo código",
                f"Achei {len(achadas)} funções com código '{codigo_limpo}':\n{msgs}\n\n"
                "Edite a planilha e deixe um código único."
            )
            return

        funcao = achadas[0]
        funcao_id = str(funcao["funcao_id"])
        nome_cargo = funcao.get("nome_cargo", "?")
        cbo = funcao.get("cbo", "?")

        if not messagebox.askyesno(
            "Confirmar função",
            f"Funcionário: {nome}\n\n"
            f"Função encontrada:\n"
            f"  • Código:    {codigo_limpo}\n"
            f"  • Nome:      {nome_cargo}\n"
            f"  • CBO:       {cbo}\n"
            f"  • funcao_id: {funcao_id}\n\n"
            "Vou salvar este override e reprocessar o email. Continuar?"
        ):
            return

        salvar_funcao_override(msg_id, nome, funcao_id)
        self.gui_q.put(("log",
            f"🔢 Função override salva: {nome} → {nome_cargo} (id {funcao_id})"))
        threading.Thread(
            target=self._reprocessar_worker, args=(msg_id,), daemon=True
        ).start()

    def _backup_agora(self):
        threading.Thread(target=self._backup_worker, daemon=True).start()

    def _backup_worker(self):
        self.gui_q.put(("log", "💾 Criando backup..."))
        try:
            dest = fazer_backup_planilha_e_payloads()
            if dest:
                self.gui_q.put(("log", f"✓ Backup criado em {dest}"))
            else:
                self.gui_q.put(("log", "❌ Falha criando backup"))
        except Exception as e:
            self.gui_q.put(("log", f"❌ Erro no backup: {e}"))

    # ---- Aba Auditoria (todas as admissões, todas as tentativas) ---

    def _build_auditoria(self):
        f = ttk.Frame(self.tab_audit)
        f.pack(fill="both", expand=True, padx=15, pady=15)

        bar = ttk.Frame(f)
        bar.pack(fill="x", pady=(0, 8))
        ttk.Label(bar, text="Histórico completo (todas as tentativas, ordem cronológica reversa):").pack(side="left")
        ttk.Label(bar, text="  Buscar:").pack(side="left", padx=(20, 4))
        ttk.Entry(bar, textvariable=self.filtro_audit_var, width=30).pack(side="left")
        ttk.Button(bar, text="🔄 Atualizar", command=self._refresh_auditoria).pack(side="left", padx=15)

        cols = ("ts", "nome", "empresa", "cnpj", "procedencia", "msg_id")
        self.tree_audit = ttk.Treeview(f, columns=cols, show="headings", height=22)
        headers = {
            "ts": ("Data/Hora", 140),
            "nome": ("Nome", 220),
            "empresa": ("Empresa", 220),
            "cnpj": ("CNPJ", 130),
            "procedencia": ("Procedência", 380),
            "msg_id": ("msg_id", 140),
        }
        for c, (txt, w) in headers.items():
            self.tree_audit.heading(c, text=txt)
            self.tree_audit.column(c, width=w, anchor="w")

        sb = ttk.Scrollbar(f, orient="vertical", command=self.tree_audit.yview)
        self.tree_audit.configure(yscrollcommand=sb.set)
        self.tree_audit.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        # Cores por categoria
        self.tree_audit.tag_configure("ok", background="#e8f5e9")
        self.tree_audit.tag_configure("pend_cli", background="#e3f2fd")
        self.tree_audit.tag_configure("pend_int", background="#fff3e0")
        self.tree_audit.tag_configure("falha", background="#ffebee")

    def _refresh_auditoria(self):
        if not hasattr(self, "tree_audit"):
            return
        for item in self.tree_audit.get_children():
            self.tree_audit.delete(item)
        if not PLANILHA_ADMISSOES.exists():
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(PLANILHA_ADMISSOES, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) <= 1:
                return
            filtro = (self.filtro_audit_var.get() or "").strip().lower()
            for row in reversed(rows[1:]):
                if not row or not any(row):
                    continue
                cols = list(row) + [""] * (6 - len(row))
                primeira = str(cols[0] or "")
                if ("-" in primeira and ":" in primeira) or primeira == "":
                    ts, nome, empresa, cnpj, procedencia, msg_id = cols[:6]
                else:
                    nome, empresa, cnpj, procedencia = cols[:4]
                    ts, msg_id = "", ""
                ts = str(ts or ""); nome = str(nome or ""); empresa = str(empresa or "")
                cnpj = str(cnpj or ""); procedencia = str(procedencia or "")
                msg_id = str(msg_id or "")
                hay = " ".join((ts, nome, empresa, cnpj, procedencia)).lower()
                if filtro and filtro not in hay:
                    continue
                proc_low = procedencia.lower()
                if proc_low.startswith("cadastrado") or proc_low.startswith("dry-run"):
                    tag = "ok"
                elif "interno" in proc_low:
                    tag = "pend_int"
                elif proc_low.startswith("pendente"):
                    tag = "pend_cli"
                else:
                    tag = "falha"
                self.tree_audit.insert(
                    "", "end",
                    values=(ts, nome, empresa, cnpj, procedencia, msg_id[:16] if msg_id else ""),
                    tags=(tag,),
                )
        except Exception as e:
            self.gui_q.put(("log", f"⚠ Erro lendo auditoria: {e}"))

    # ---- Aba API eContador (auditoria das chamadas HTTP) ----------

    def _build_api_econtador(self):
        f = self.tab_api

        # Toolbar
        bar = tk.Frame(f, bg=COR_BEIGE)
        bar.pack(fill="x", pady=(10, 8))
        ttk.Button(bar, text="🔄  Atualizar", command=self._refresh_api_econtador).pack(side="left", padx=(0, 8))
        ttk.Button(bar, text="👁  Ver JSON completo (seleção)",
                   command=self._abrir_detalhe_api).pack(side="left", padx=8)
        ttk.Button(bar, text="📂  Abrir arquivo audit",
                   command=self._abrir_arquivo_audit).pack(side="left", padx=8)

        # Filtro
        self.filtro_api_var = tk.StringVar()
        self.filtro_api_var.trace_add("write", lambda *a: self._refresh_api_econtador())
        ttk.Label(bar, text="  Buscar:").pack(side="left", padx=(20, 4))
        ttk.Entry(bar, textvariable=self.filtro_api_var, width=30).pack(side="left")

        self.apenas_falhas_var = tk.BooleanVar(value=False)
        self.apenas_falhas_var.trace_add("write", lambda *a: self._refresh_api_econtador())
        ttk.Checkbutton(bar, text="Apenas falhas",
                        variable=self.apenas_falhas_var).pack(side="left", padx=15)

        # Contador resumido
        self.api_resumo_var = tk.StringVar(value="—")
        tk.Label(bar, textvariable=self.api_resumo_var,
                 bg=COR_BEIGE, fg=COR_TEXT_MUTED,
                 font=("Segoe UI", 9)).pack(side="right", padx=(0, 8))

        tk.Label(f, text="(Clique 2× na linha pra ver o JSON completo BEFORE+AFTER)",
                 bg=COR_BEIGE, fg=COR_TEXT_MUTED,
                 font=("Segoe UI", 9)).pack(anchor="w", pady=(0, 5))

        # Tabela
        cols = ("ts", "corr_id", "operation", "status", "duration", "resumo")
        self.tree_api = ttk.Treeview(f, columns=cols, show="headings", height=22)
        headers = {
            "ts":        ("Quando", 150),
            "corr_id":   ("Corr ID", 80),
            "operation": ("Operação", 150),
            "status":    ("Status", 80),
            "duration":  ("Duração", 80),
            "resumo":    ("Resumo", 480),
        }
        for c, (txt, w) in headers.items():
            self.tree_api.heading(c, text=txt)
            self.tree_api.column(c, width=w, anchor="w")

        sb = ttk.Scrollbar(f, orient="vertical", command=self.tree_api.yview)
        self.tree_api.configure(yscrollcommand=sb.set)
        self.tree_api.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        self.tree_api.bind("<Double-1>", lambda e: self._abrir_detalhe_api())

        # Tags por status
        self.tree_api.tag_configure("ok",       background="#E8F5E9", foreground=COR_TEXT_DARK)
        self.tree_api.tag_configure("fail",     background="#FFEBEE", foreground=COR_TEXT_DARK)
        self.tree_api.tag_configure("slow",     background="#FFF8E1")
        self.tree_api.tag_configure("orphan",   background="#F5F5F5")  # before sem after

        # Cache pareado: corr_id → {"before": entry, "after": entry}
        self._api_entries_cache: dict[str, dict] = {}

    def _refresh_api_econtador(self):
        if not hasattr(self, "tree_api"):
            return
        for item in self.tree_api.get_children():
            self.tree_api.delete(item)

        if not ECONTADOR_AUDIT_FILE.exists():
            self.api_resumo_var.set("(econtador_audit.ndjson ainda não existe)")
            self._api_entries_cache = {}
            return

        # Lê todas as linhas e agrupa por corr_id
        pares: dict[str, dict] = {}
        try:
            with open(ECONTADOR_AUDIT_FILE, "r", encoding="utf-8") as fh:
                for linha in fh:
                    linha = linha.strip()
                    if not linha:
                        continue
                    try:
                        e = json.loads(linha)
                    except json.JSONDecodeError:
                        continue
                    cid = e.get("corr_id") or ""
                    phase = e.get("phase") or ""
                    if not cid or phase not in ("before", "after"):
                        continue
                    pares.setdefault(cid, {})[phase] = e
        except OSError as exc:
            self.api_resumo_var.set(f"Erro lendo: {exc}")
            return

        self._api_entries_cache = pares

        # Conta totais
        total = len(pares)
        falhas = sum(
            1 for p in pares.values()
            if "after" in p and not p["after"].get("success", True)
        )
        slow = sum(
            1 for p in pares.values()
            if "after" in p and (p["after"].get("duration_ms") or 0) >= 5000
        )

        # Filtros
        filtro = (self.filtro_api_var.get() or "").strip().lower()
        apenas_falhas = bool(self.apenas_falhas_var.get())

        # Ordena por timestamp (do after se existir, senão before)
        def _ts(par):
            return (par.get("after") or par.get("before") or {}).get("timestamp", "")

        ordenados = sorted(pares.items(), key=lambda kv: _ts(kv[1]), reverse=True)

        mostrados = 0
        for cid, par in ordenados:
            before = par.get("before") or {}
            after = par.get("after")
            op = (before.get("operation")
                  or (after.get("operation") if after else "?"))
            ts_completo = _ts(par)
            ts_display = ts_completo[:19].replace("T", " ") if ts_completo else "?"

            if after is None:
                # Orphan — só BEFORE
                status_txt = "..."
                tag = "orphan"
                duration = "—"
                resumo = "BEFORE sem AFTER (processo crashou?)"
            else:
                ok = bool(after.get("success"))
                code = after.get("status_code", "?")
                if ok:
                    status_txt = f"✓ {code}"
                    tag = "ok"
                else:
                    status_txt = f"✗ {code}"
                    tag = "fail"
                dur_ms = after.get("duration_ms")
                duration = f"{dur_ms} ms" if dur_ms is not None else "—"
                if dur_ms and dur_ms >= 5000 and tag == "ok":
                    tag = "slow"
                resumo = self._resumo_after(op, after, before)

            if apenas_falhas and tag != "fail":
                continue
            hay = " ".join((ts_display, cid, op, status_txt, duration, resumo)).lower()
            if filtro and filtro not in hay:
                continue

            self.tree_api.insert(
                "", "end",
                iid=cid,
                values=(ts_display, cid, op, status_txt, duration, resumo[:200]),
                tags=(tag,),
            )
            mostrados += 1

        self.api_resumo_var.set(
            f"{mostrados} chamada(s) visível(eis) | "
            f"total {total} · falhas {falhas} · lentas (>5s) {slow}"
        )

    @staticmethod
    def _resumo_after(op: str, after: dict, before: dict) -> str:
        """Constrói o resumo da linha baseado na operação."""
        if "EXCEPTION" in (after.get("exception") or "") or after.get("exception"):
            return f"EXCEPTION: {after.get('exception')[:150]}"
        if op == "POST /candidatos":
            if after.get("success"):
                snap = (after.get("input_snapshot") or {})
                cid = after.get("candidato_id", "?")
                return f"candidato_id={cid} | nome={snap.get('nome', '?')}"
            body = after.get("body_preview") or ""
            snap = (after.get("input_snapshot") or {})
            return f"nome={snap.get('nome', '?')} | erro: {body[:120]}"
        if op == "GET /empresas":
            if after.get("success"):
                if after.get("empresa_id"):
                    return f"empresa_id={after['empresa_id']} | razão: {after.get('razao_social', '?')}"
                return "0 resultados"
            inp = before.get("input") or {}
            return f"cnpj={inp.get('cnpj', '?')} | erro {after.get('status_code')}"
        if op == "GET /departamentos":
            if after.get("success"):
                return f"{after.get('n_resultados', 0)} departamento(s)"
            inp = before.get("input") or {}
            return f"empresa_id={inp.get('empresa_id', '?')} | erro {after.get('status_code')}"
        # Default
        return str(after)[:120]

    def _abrir_detalhe_api(self):
        sel = self.tree_api.selection()
        if not sel:
            messagebox.showinfo("Selecione",
                                "Selecione uma linha na tabela primeiro.")
            return
        cid = sel[0]
        par = self._api_entries_cache.get(cid)
        if not par:
            messagebox.showinfo("Não encontrado", f"Sem dados pro corr_id {cid}")
            return

        # Modal com 2 painéis: BEFORE | AFTER
        win = tk.Toplevel(self)
        win.title(f"Auditoria eContador — corr_id {cid}")
        win.geometry("1100x650")
        win.transient(self)

        before = par.get("before") or {}
        after = par.get("after")

        head = tk.Frame(win, bg=COR_WHITE, padx=15, pady=12)
        head.pack(fill="x")
        tk.Label(head, text=f"Operação: {before.get('operation') or (after or {}).get('operation', '?')}",
                 font=("Segoe UI", 12, "bold"), bg=COR_WHITE).pack(anchor="w")
        tk.Label(head, text=f"corr_id: {cid}",
                 fg=COR_TEXT_MUTED, bg=COR_WHITE).pack(anchor="w")
        if after:
            ok = after.get("success")
            cor = COR_SUCCESS if ok else COR_DANGER
            tk.Label(head,
                     text=f"Resultado: {'✓ SUCESSO' if ok else '✗ FALHA'} "
                          f"(status {after.get('status_code', '?')}, "
                          f"{after.get('duration_ms', '?')} ms)",
                     fg=cor, bg=COR_WHITE,
                     font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(4, 0))
        else:
            tk.Label(head, text="(sem AFTER — processo crashou antes de receber resposta?)",
                     fg=COR_DANGER, bg=COR_WHITE).pack(anchor="w", pady=(4, 0))

        body = tk.Frame(win, bg=COR_BEIGE)
        body.pack(fill="both", expand=True, padx=10, pady=10)
        body.grid_columnconfigure(0, weight=1)
        body.grid_columnconfigure(1, weight=1)
        body.grid_rowconfigure(0, weight=1)

        # BEFORE panel
        f_b = tk.Frame(body, bg=COR_WHITE, bd=1, relief="solid")
        f_b.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        tk.Label(f_b, text="BEFORE", font=("Segoe UI", 11, "bold"),
                 bg=COR_WHITE, fg=COR_ORANGE).pack(anchor="w", padx=10, pady=(10, 5))
        txt_b = scrolledtext.ScrolledText(f_b, font=("Consolas", 9), wrap="none")
        txt_b.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        txt_b.insert("1.0", json.dumps(before, ensure_ascii=False, indent=2))
        txt_b.configure(state="disabled")

        # AFTER panel
        f_a = tk.Frame(body, bg=COR_WHITE, bd=1, relief="solid")
        f_a.grid(row=0, column=1, sticky="nsew", padx=(5, 0))
        tk.Label(f_a, text="AFTER", font=("Segoe UI", 11, "bold"),
                 bg=COR_WHITE, fg=COR_SUCCESS if (after and after.get("success")) else COR_DANGER
                 ).pack(anchor="w", padx=10, pady=(10, 5))
        txt_a = scrolledtext.ScrolledText(f_a, font=("Consolas", 9), wrap="none")
        txt_a.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        if after:
            txt_a.insert("1.0", json.dumps(after, ensure_ascii=False, indent=2))
        else:
            txt_a.insert("1.0", "(sem registro de AFTER)")
        txt_a.configure(state="disabled")

        ttk.Button(win, text="Fechar", command=win.destroy).pack(pady=10)

    def _abrir_arquivo_audit(self):
        if ECONTADOR_AUDIT_FILE.exists():
            import os
            os.startfile(ECONTADOR_AUDIT_FILE)
        else:
            messagebox.showinfo("Sem arquivo", "econtador_audit.ndjson ainda não foi criado (nenhuma chamada eContador feita ainda).")

    # ---- Aba Estatísticas (redesign com cards + barras) -----------

    # Cores adicionais usadas nas barras de status (paleta complementar
    # ao tema Crosara, similar à proposta enviada pelo usuário)
    _COR_TRILHA = "#F1EFE8"     # fundo cinza claro das barras
    _COR_VERDE_OK = "#4E8A6B"   # cadastrado
    _COR_AMBAR = "#E0A030"      # pendente interno
    _COR_BADGE_TXT = "#B25529"  # texto do badge de motivo

    def _build_estatisticas(self):
        f = self.tab_stats

        # Toolbar superior (botão Recalcular à direita)
        toolbar = tk.Frame(f, bg=COR_BEIGE)
        toolbar.pack(fill="x", pady=(10, 12))
        ttk.Button(toolbar, text="↻  Recalcular",
                   command=self._refresh_estatisticas).pack(side="right")

        # Container scrollável (conteúdo pode ser longo)
        wrap = tk.Frame(f, bg=COR_BEIGE)
        wrap.pack(fill="both", expand=True)

        canvas = tk.Canvas(wrap, bg=COR_BEIGE, highlightthickness=0)
        sb = ttk.Scrollbar(wrap, orient="vertical", command=canvas.yview)
        self.stats_body = tk.Frame(canvas, bg=COR_BEIGE)

        self.stats_body.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")),
        )
        _win_id = canvas.create_window((0, 0), window=self.stats_body, anchor="nw")
        canvas.bind(
            "<Configure>",
            lambda e: canvas.itemconfigure(_win_id, width=e.width),
        )
        canvas.configure(yscrollcommand=sb.set)

        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        # Scroll com mousewheel
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

    def _refresh_estatisticas(self):
        if not hasattr(self, "stats_body"):
            return

        # Limpa conteúdo anterior
        for w in self.stats_body.winfo_children():
            w.destroy()

        body = self.stats_body
        b = sum_billing_mes_atual()
        mes = datetime.now().strftime("%B / %Y").capitalize()

        # ===== BILLING SECTION =====
        tk.Label(body, text=f"BILLING — {mes.upper()}",
                 bg=COR_BEIGE, fg=COR_TEXT_MUTED,
                 font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(4, 8))

        # 4 cards de métrica
        metricas = tk.Frame(body, bg=COR_BEIGE)
        metricas.pack(fill="x", pady=(0, 10))
        for i in range(4):
            metricas.grid_columnconfigure(i, weight=1)

        def _metric(col, label, valor, cor_valor):
            c = tk.Frame(metricas, bg=COR_WHITE, bd=1, relief="solid",
                         highlightbackground=COR_BORDER, highlightthickness=0)
            c.grid(row=0, column=col, sticky="nsew",
                   padx=(0 if col == 0 else 6, 6 if col < 3 else 0))
            tk.Label(c, text=label, bg=COR_WHITE, fg=COR_TEXT_MUTED,
                     font=("Segoe UI", 10)).pack(anchor="w", padx=14, pady=(12, 2))
            tk.Label(c, text=valor, bg=COR_WHITE, fg=cor_valor,
                     font=("Segoe UI", 19, "bold")).pack(anchor="w", padx=14, pady=(0, 12))

        custo_total = b["custo_usd"]
        custo_med = (custo_total / b["n_passadas"]) if b["n_passadas"] else 0.0
        _metric(0, "Custo total", f"US$ {custo_total:.2f}", COR_ORANGE)
        _metric(1, "Chamadas API", str(b["n_calls"]), COR_TEXT_DARK)
        _metric(2, "Passadas", str(b["n_passadas"]), COR_TEXT_DARK)
        _metric(3, "Custo médio/adm.", f"US$ {custo_med:.3f}", COR_TEXT_DARK)

        # 3 kv-cards (tokens + limite)
        kv = tk.Frame(body, bg=COR_BEIGE)
        kv.pack(fill="x", pady=(0, 16))
        for i in range(3):
            kv.grid_columnconfigure(i, weight=1)

        def _kv(col, label, valor, valor_cor=COR_TEXT_DARK):
            c = tk.Frame(kv, bg=COR_WHITE, bd=1, relief="solid")
            c.grid(row=0, column=col, sticky="nsew",
                   padx=(0 if col == 0 else 6, 6 if col < 2 else 0))
            row = tk.Frame(c, bg=COR_WHITE)
            row.pack(fill="x", padx=14, pady=12)
            tk.Label(row, text=label, bg=COR_WHITE, fg=COR_TEXT_MUTED,
                     font=("Segoe UI", 10)).pack(side="left")
            tk.Label(row, text=valor, bg=COR_WHITE, fg=valor_cor,
                     font=("Segoe UI", 12, "bold")).pack(side="right")

        _kv(0, "Input tokens", f"{b['input_tokens']:,}".replace(",", "."))
        _kv(1, "Output tokens", f"{b['output_tokens']:,}".replace(",", "."))
        limite_cor = COR_DANGER if custo_total >= self.billing_limite_usd else COR_TEXT_DARK
        _kv(2, "Limite mensal", f"US$ {self.billing_limite_usd:.2f}", limite_cor)

        # ===== DIRECT DATA SECTION =====
        try:
            dd = sum_directdata_mes_atual()
        except Exception:
            dd = None

        if dd and (dd["n_chamadas"] > 0 or dd["n_cache_hits"] > 0):
            tk.Label(body, text="DIRECT DATA (CPF) — EFETIVIDADE DAS APIS",
                     bg=COR_BEIGE, fg=COR_TEXT_MUTED,
                     font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(16, 8))

            dd_metricas = tk.Frame(body, bg=COR_BEIGE)
            dd_metricas.pack(fill="x", pady=(0, 10))
            for i in range(4):
                dd_metricas.grid_columnconfigure(i, weight=1)

            def _dd_metric(col, label, valor, cor_valor):
                c = tk.Frame(dd_metricas, bg=COR_WHITE, bd=1, relief="solid",
                             highlightbackground=COR_BORDER, highlightthickness=0)
                c.grid(row=0, column=col, sticky="nsew",
                       padx=(0 if col == 0 else 6, 6 if col < 3 else 0))
                tk.Label(c, text=label, bg=COR_WHITE, fg=COR_TEXT_MUTED,
                         font=("Segoe UI", 10)).pack(anchor="w", padx=14, pady=(12, 2))
                tk.Label(c, text=valor, bg=COR_WHITE, fg=cor_valor,
                         font=("Segoe UI", 19, "bold")).pack(anchor="w", padx=14, pady=(0, 12))

            cache_pct = int(dd["taxa_cache_hit"] * 100)
            sucesso_pct = int(dd["taxa_sucesso"] * 100)
            cor_sucesso = COR_SUCCESS if sucesso_pct >= 80 else (
                COR_WARNING if sucesso_pct >= 50 else COR_DANGER
            )
            _dd_metric(0, "Custo total",
                       f"R$ {dd['custo_brl']:.2f}".replace(".", ","), COR_ORANGE)
            _dd_metric(1, "Chamadas pagas", str(dd["n_chamadas"]), COR_TEXT_DARK)
            _dd_metric(2, "Cache hits (R$ 0)", f"{dd['n_cache_hits']} ({cache_pct}%)",
                       COR_SUCCESS)
            _dd_metric(3, "Taxa de sucesso", f"{sucesso_pct}%", cor_sucesso)

            # Breakdown por API
            tk.Label(body, text="Quebra por API:",
                     bg=COR_BEIGE, fg=COR_TEXT_MUTED,
                     font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(6, 4))

            tabela_api = tk.Frame(body, bg=COR_WHITE, bd=1, relief="solid",
                                   highlightbackground=COR_BORDER, highlightthickness=0)
            tabela_api.pack(fill="x", pady=(0, 16))

            # Header
            cabecalho = tk.Frame(tabela_api, bg=COR_BEIGE)
            cabecalho.pack(fill="x")
            for col, txt, w in [
                (0, "API", 240), (1, "Chamadas", 100), (2, "Sucesso", 100),
                (3, "Latência média", 130), (4, "Custo", 100),
            ]:
                tk.Label(cabecalho, text=txt, bg=COR_BEIGE, fg=COR_TEXT_DARK,
                         font=("Segoe UI", 9, "bold"), width=w//8,
                         anchor="w").pack(side="left", padx=8, pady=6)

            apis_labels = {
                "cadastro_basico": "Cadastro Básico (R$ 0,16)",
                "pis":             "PIS Ministério Trabalho (R$ 0,36)",
                "titulo":          "TSE Título Eleitor (R$ 0,72)",
            }
            for api_key, label_amigavel in apis_labels.items():
                info = dd["por_api"].get(api_key, {})
                if info.get("n", 0) == 0:
                    continue
                n = info["n"]
                suc = info.get("sucesso", 0)
                pct = f"{int(suc/n*100)}%" if n > 0 else "—"
                dur = dd["duracao_media_ms"].get(api_key)
                dur_str = f"{dur} ms" if dur is not None else "—"
                custo = info.get("custo_brl", 0.0)

                row = tk.Frame(tabela_api, bg=COR_WHITE)
                row.pack(fill="x")
                tk.Label(row, text=label_amigavel, bg=COR_WHITE,
                         fg=COR_TEXT_DARK, font=("Segoe UI", 9),
                         width=30, anchor="w").pack(side="left", padx=8, pady=6)
                tk.Label(row, text=str(n), bg=COR_WHITE, fg=COR_TEXT_DARK,
                         font=("Segoe UI", 9), width=12,
                         anchor="w").pack(side="left", padx=8, pady=6)
                tk.Label(row, text=pct, bg=COR_WHITE,
                         fg=COR_SUCCESS if int(suc/n*100 if n else 0) >= 80 else COR_WARNING,
                         font=("Segoe UI", 9, "bold"), width=12,
                         anchor="w").pack(side="left", padx=8, pady=6)
                tk.Label(row, text=dur_str, bg=COR_WHITE, fg=COR_TEXT_DARK,
                         font=("Segoe UI", 9), width=16,
                         anchor="w").pack(side="left", padx=8, pady=6)
                tk.Label(row, text=f"R$ {custo:.2f}".replace(".", ","),
                         bg=COR_WHITE, fg=COR_TEXT_DARK,
                         font=("Segoe UI", 9), width=12,
                         anchor="w").pack(side="left", padx=8, pady=6)

        # ===== REFERÊNCIA DE CUSTOS POR FERRAMENTA =====
        tk.Label(body, text="REFERÊNCIA DE CUSTOS POR FERRAMENTA",
                 bg=COR_BEIGE, fg=COR_TEXT_MUTED,
                 font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(16, 8))

        # Cotação informativa
        tk.Label(
            body,
            text=f"(Cotação USD→BRL usada pra estimativas: R$ {self.cotacao_usd_brl:.2f})",
            bg=COR_BEIGE, fg=COR_TEXT_MUTED, font=("Segoe UI", 8, "italic"),
        ).pack(anchor="w", pady=(0, 6))

        ref_card = tk.Frame(body, bg=COR_WHITE, bd=1, relief="solid",
                            highlightbackground=COR_BORDER, highlightthickness=0)
        ref_card.pack(fill="x", pady=(0, 6))

        # Header
        h = tk.Frame(ref_card, bg=COR_BEIGE)
        h.pack(fill="x")
        for txt, w in [
            ("Ferramenta", 280), ("Tipo", 100),
            ("Custo médio/admissão", 200), ("Observação", 380),
        ]:
            tk.Label(h, text=txt, bg=COR_BEIGE, fg=COR_TEXT_DARK,
                     font=("Segoe UI", 9, "bold"), width=w // 8,
                     anchor="w").pack(side="left", padx=8, pady=6)

        # Linhas — ferramenta, tipo, custo médio, observação, cor
        ferramentas = [
            ("Claude API (Anthropic Sonnet 4.6)", "Variável",
             "R$ 0,30 – R$ 1,45",
             "$3/MTok input + $15/MTok output. Cresce com nº de admissões/anexos",
             COR_ORANGE),
            ("Direct Data — Cadastro Básico", "Fixo",
             "R$ 0,16",
             "Nome, mãe, nascimento, sexo, endereço completo",
             COR_TEXT_DARK),
            ("Direct Data — PIS Ministério", "Fixo",
             "R$ 0,36",
             "Número PIS (preserva zeros à esquerda)",
             COR_TEXT_DARK),
            ("Direct Data — TSE Título Eleitor", "Fixo",
             "R$ 0,72",
             "Título, zona, seção (latência ~40s)",
             COR_TEXT_DARK),
            ("eContador (Alterdata E-plugin)", "Assinatura",
             "R$ 0,00",
             "Mensalidade já paga; API ilimitada",
             COR_SUCCESS),
            ("Gmail API (Google)", "Gratuita",
             "R$ 0,00",
             "Quota generosa, nunca atingimos",
             COR_SUCCESS),
            ("ViaCEP", "Gratuita",
             "R$ 0,00",
             "Sem rate limit oficial + cache em RAM",
             COR_SUCCESS),
            ("PyMuPDF (compressão de PDF)", "Open source",
             "R$ 0,00",
             "Lib local — comprime PDFs antes do Claude",
             COR_SUCCESS),
            ("UnRAR / rarfile", "Gratuitos",
             "R$ 0,00",
             "Descompacta .rar dos clientes",
             COR_SUCCESS),
        ]

        for nome, tipo, custo, obs, cor_custo in ferramentas:
            row = tk.Frame(ref_card, bg=COR_WHITE)
            row.pack(fill="x")
            tk.Label(row, text=nome, bg=COR_WHITE, fg=COR_TEXT_DARK,
                     font=("Segoe UI", 9), width=35,
                     anchor="w").pack(side="left", padx=8, pady=5)
            tk.Label(row, text=tipo, bg=COR_WHITE, fg=COR_TEXT_MUTED,
                     font=("Segoe UI", 9), width=12,
                     anchor="w").pack(side="left", padx=8, pady=5)
            tk.Label(row, text=custo, bg=COR_WHITE, fg=cor_custo,
                     font=("Segoe UI", 9, "bold"), width=25,
                     anchor="w").pack(side="left", padx=8, pady=5)
            tk.Label(row, text=obs, bg=COR_WHITE, fg=COR_TEXT_MUTED,
                     font=("Segoe UI", 9), width=47,
                     anchor="w").pack(side="left", padx=8, pady=5)

        # Resumo de cenários
        tk.Label(body,
                 text="Cenários típicos (custo total por admissão):",
                 bg=COR_BEIGE, fg=COR_TEXT_MUTED,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(12, 4))

        cenarios_card = tk.Frame(body, bg=COR_WHITE, bd=1, relief="solid",
                                  highlightbackground=COR_BORDER, highlightthickness=0)
        cenarios_card.pack(fill="x", pady=(0, 16))

        h2 = tk.Frame(cenarios_card, bg=COR_BEIGE)
        h2.pack(fill="x")
        for txt, w in [
            ("Cenário", 320), ("Claude", 140), ("Direct Data", 140), ("Total", 140),
        ]:
            tk.Label(h2, text=txt, bg=COR_BEIGE, fg=COR_TEXT_DARK,
                     font=("Segoe UI", 9, "bold"), width=w // 8,
                     anchor="w").pack(side="left", padx=8, pady=6)

        cenarios = [
            ("Simples (1 admissão, docs limpos)",
             "R$ 0,30", "R$ 1,24", "R$ 1,54"),
            ("Médio (1 admissão, docs ruins)",
             "R$ 0,75", "R$ 1,24", "R$ 1,99"),
            ("Complexo (3 admissões + RAR + reprompt)",
             "R$ 1,45", "R$ 3,72", "R$ 5,17"),
            ("Reprocessamento (cache hit + override)",
             "R$ 0,20", "R$ 0,00", "R$ 0,20"),
        ]
        for desc, cl, dd, tot in cenarios:
            row = tk.Frame(cenarios_card, bg=COR_WHITE)
            row.pack(fill="x")
            tk.Label(row, text=desc, bg=COR_WHITE, fg=COR_TEXT_DARK,
                     font=("Segoe UI", 9), width=40,
                     anchor="w").pack(side="left", padx=8, pady=5)
            tk.Label(row, text=cl, bg=COR_WHITE, fg=COR_TEXT_MUTED,
                     font=("Segoe UI", 9), width=17,
                     anchor="w").pack(side="left", padx=8, pady=5)
            tk.Label(row, text=dd, bg=COR_WHITE, fg=COR_TEXT_MUTED,
                     font=("Segoe UI", 9), width=17,
                     anchor="w").pack(side="left", padx=8, pady=5)
            tk.Label(row, text=tot, bg=COR_WHITE, fg=COR_ORANGE,
                     font=("Segoe UI", 9, "bold"), width=17,
                     anchor="w").pack(side="left", padx=8, pady=5)

        # ===== STATS DA PLANILHA =====
        if not PLANILHA_ADMISSOES.exists():
            tk.Label(body, text="(Planilha admissoes.xlsx ainda não existe.)",
                     bg=COR_BEIGE, fg=COR_TEXT_MUTED).pack(pady=20)
            return

        try:
            from collections import Counter
            from openpyxl import load_workbook
            wb = load_workbook(PLANILHA_ADMISSOES, read_only=True)
            ws = wb.active
            rows = [r for r in list(ws.iter_rows(values_only=True))[1:] if r and any(r)]

            cont_status = Counter()
            cont_empresa = Counter()
            cont_motivo = Counter()
            for row in rows:
                cols = list(row) + [""] * (6 - len(row))
                primeira = str(cols[0] or "")
                if ("-" in primeira and ":" in primeira) or primeira == "":
                    _ts, _nome, empresa, _cnpj, procedencia, _mid = cols[:6]
                else:
                    _nome, empresa, _cnpj, procedencia = cols[:4]
                empresa = str(empresa or "(sem identificação)").strip() or "(sem identificação)"
                procedencia = str(procedencia or "")
                pl = procedencia.lower()
                if pl.startswith("cadastrado"):
                    cont_status["Cadastrado"] += 1
                elif pl.startswith("dry-run"):
                    cont_status["Dry-run"] += 1
                elif "interno" in pl:
                    cont_status["Pendente interno"] += 1
                elif pl.startswith("pendente"):
                    cont_status["Pendente cliente"] += 1
                else:
                    cont_status["Falha"] += 1
                cont_empresa[empresa] += 1
                if "—" in procedencia and (pl.startswith("pendente") or pl.startswith("falha")):
                    motivo = procedencia.split("—", 1)[1].strip()[:100]
                    cont_motivo[motivo] += 1

            total = sum(cont_status.values())
            ok_count = cont_status.get("Cadastrado", 0) + cont_status.get("Dry-run", 0)
            taxa_sucesso = (ok_count / total * 100) if total else 0

            # ===== STATUS + EMPRESAS (lado a lado) =====
            meio = tk.Frame(body, bg=COR_BEIGE)
            meio.pack(fill="x", pady=(0, 14))
            meio.grid_columnconfigure(0, weight=11)
            meio.grid_columnconfigure(1, weight=9)

            # --- Card Status (com barras) ---
            card_st = tk.Frame(meio, bg=COR_WHITE, bd=1, relief="solid")
            card_st.grid(row=0, column=0, sticky="nsew", padx=(0, 6))

            top_st = tk.Frame(card_st, bg=COR_WHITE)
            top_st.pack(fill="x", padx=18, pady=(16, 12))
            tk.Label(top_st, text="Admissões por status",
                     bg=COR_WHITE, fg=COR_TEXT_DARK,
                     font=("Segoe UI", 12, "bold")).pack(side="left")
            sufx = f"{total} no total · {taxa_sucesso:.1f}% sucesso".replace(".", ",")
            tk.Label(top_st, text=sufx,
                     bg=COR_WHITE, fg=COR_TEXT_MUTED,
                     font=("Segoe UI", 9)).pack(side="right")

            cores_status = {
                "Cadastrado": self._COR_VERDE_OK,
                "Pendente cliente": COR_ORANGE,
                "Pendente interno": self._COR_AMBAR,
                "Falha": COR_DANGER,
                "Dry-run": COR_INFO,
            }
            # Ordem fixa pra exibição consistente
            ordem = ["Cadastrado", "Pendente cliente", "Pendente interno", "Falha", "Dry-run"]
            for label in ordem:
                qtd = cont_status.get(label, 0)
                if qtd == 0 and label not in ("Cadastrado", "Pendente cliente", "Pendente interno", "Falha"):
                    continue
                pct = (qtd / total * 100) if total else 0
                self._stat_barra(card_st, label, qtd, pct, cores_status[label])
            tk.Frame(card_st, bg=COR_WHITE, height=6).pack()

            # --- Card Empresas ---
            card_emp = tk.Frame(meio, bg=COR_WHITE, bd=1, relief="solid")
            card_emp.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
            tk.Label(card_emp, text="Top empresas",
                     bg=COR_WHITE, fg=COR_TEXT_DARK,
                     font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=18, pady=(16, 8))
            top_emp = cont_empresa.most_common(8)
            for i, (nome, qtd) in enumerate(top_emp):
                row = tk.Frame(card_emp, bg=COR_WHITE)
                row.pack(fill="x", padx=18, pady=3)
                cor_nome = COR_TEXT_MUTED if nome.startswith("(") else COR_TEXT_DARK
                cor_qtd = COR_ORANGE if i == 0 else COR_TEXT_DARK
                tk.Label(row, text=nome, bg=COR_WHITE, fg=cor_nome,
                         font=("Segoe UI", 10), anchor="w").pack(side="left", fill="x", expand=True)
                tk.Label(row, text=str(qtd), bg=COR_WHITE, fg=cor_qtd,
                         font=("Segoe UI", 11, "bold")).pack(side="right")
            tk.Frame(card_emp, bg=COR_WHITE, height=10).pack()

            # ===== MOTIVOS =====
            card_mot = tk.Frame(body, bg=COR_WHITE, bd=1, relief="solid")
            card_mot.pack(fill="x")
            tk.Label(card_mot, text="Top motivos de pendência / falha",
                     bg=COR_WHITE, fg=COR_TEXT_DARK,
                     font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=18, pady=(16, 10))
            top_mot = cont_motivo.most_common(10)
            if not top_mot:
                tk.Label(card_mot, text="(nenhuma pendência registrada)",
                         bg=COR_WHITE, fg=COR_TEXT_MUTED,
                         font=("Segoe UI", 10)).pack(anchor="w", padx=18, pady=(0, 16))
            else:
                for qtd, motivo in [(q, m) for m, q in top_mot]:
                    row = tk.Frame(card_mot, bg=COR_WHITE)
                    row.pack(fill="x", padx=18, pady=4)
                    # Badge laranja claro com a quantidade
                    badge = tk.Label(
                        row, text=str(qtd),
                        bg=COR_BEIGE, fg=self._COR_BADGE_TXT,
                        font=("Segoe UI", 10, "bold"),
                        width=3, padx=4, pady=2,
                    )
                    badge.pack(side="left", padx=(0, 10))
                    tk.Label(row, text=motivo,
                             bg=COR_WHITE, fg=COR_TEXT_DARK,
                             font=("Segoe UI", 10), anchor="w",
                             wraplength=900, justify="left").pack(side="left", fill="x", expand=True)
                tk.Frame(card_mot, bg=COR_WHITE, height=10).pack()
        except Exception as e:
            tk.Label(body, text=f"⚠ Erro lendo planilha: {e}",
                     bg=COR_BEIGE, fg=COR_DANGER).pack(pady=20)

    def _stat_barra(self, parent, label, qtd, pct, cor):
        """Linha com label/qtd e barra horizontal (trilha + preenchimento)."""
        wrap = tk.Frame(parent, bg=COR_WHITE)
        wrap.pack(fill="x", padx=18, pady=(0, 10))

        cab = tk.Frame(wrap, bg=COR_WHITE)
        cab.pack(fill="x", pady=(0, 4))
        tk.Label(cab, text=label, bg=COR_WHITE, fg=COR_TEXT_DARK,
                 font=("Segoe UI", 10)).pack(side="left")
        valor_txt = f"{qtd} · {pct:.1f}%".replace(".", ",")
        tk.Label(cab, text=valor_txt, bg=COR_WHITE, fg=COR_TEXT_MUTED,
                 font=("Segoe UI", 10)).pack(side="right")

        # Trilha + preenchimento via place (relwidth = proporção)
        trilha = tk.Frame(wrap, bg=self._COR_TRILHA, height=9)
        trilha.pack(fill="x")
        trilha.pack_propagate(False)
        if pct > 0:
            fill = tk.Frame(trilha, bg=cor)
            fill.place(relx=0, rely=0, relwidth=max(pct / 100, 0.02), relheight=1)

    # ---- Aba Regras (editor de regras.json) ---------------------

    def _build_regras(self):
        f = ttk.Frame(self.tab_regras)
        f.pack(fill="both", expand=True, padx=15, pady=15)

        ttk.Label(f, text=f"Edite regras.json diretamente. Chaves começadas com '_' são docs/exemplos (ignoradas pelo pipeline).",
                  foreground="#666", wraplength=900).pack(anchor="w", pady=(0, 8))

        self.regras_text = scrolledtext.ScrolledText(
            f, font=("Consolas", 9), wrap="none",
        )
        self.regras_text.pack(fill="both", expand=True)

        bar = ttk.Frame(f)
        bar.pack(fill="x", pady=(10, 0))
        ttk.Button(bar, text="💾 Salvar", command=self._salvar_regras).pack(side="left", padx=5)
        ttk.Button(bar, text="🔄 Recarregar do disco", command=self._carregar_regras_no_editor).pack(side="left", padx=5)
        ttk.Button(bar, text="✓ Validar JSON", command=self._validar_regras_json).pack(side="left", padx=5)

        self._carregar_regras_no_editor()

    def _carregar_regras_no_editor(self):
        try:
            if REGRAS_FILE.exists():
                conteudo = REGRAS_FILE.read_text(encoding="utf-8")
            else:
                conteudo = "{}"
            self.regras_text.delete("1.0", "end")
            self.regras_text.insert("1.0", conteudo)
        except Exception as e:
            messagebox.showerror("Erro carregando regras.json", str(e))

    def _validar_regras_json(self):
        try:
            json.loads(self.regras_text.get("1.0", "end"))
            messagebox.showinfo("JSON válido", "Sintaxe OK!")
        except json.JSONDecodeError as e:
            messagebox.showerror("JSON inválido", str(e))

    def _salvar_regras(self):
        try:
            conteudo = self.regras_text.get("1.0", "end").strip()
            json.loads(conteudo)  # valida antes de salvar
        except json.JSONDecodeError as e:
            messagebox.showerror("JSON inválido — não salvou", str(e))
            return
        try:
            REGRAS_FILE.write_text(conteudo + "\n", encoding="utf-8")
            self.gui_q.put(("log", f"💾 regras.json salvo ({len(conteudo)} chars)"))
            messagebox.showinfo("Salvo", "regras.json salvo com sucesso!\n\n"
                                "Reinicie o pipeline pra que mudanças entrem em vigor.")
        except Exception as e:
            messagebox.showerror("Erro salvando", str(e))

    # ---- App icon + Sobre dialog ---------------------------------

    def _aplicar_icone_janela(self):
        """Carrega admitir-logo.{ico,png} como ícone da janela.
        Gera .ico automaticamente do .png via Pillow se faltar.
        """
        # Auto-gera .ico do .png se faltar
        if _HAS_PIL and ADMITER_LOGO_PNG.exists() and not ADMITER_LOGO_ICO.exists():
            try:
                img = Image.open(ADMITER_LOGO_PNG)
                if img.mode != "RGBA":
                    img = img.convert("RGBA")
                img.save(
                    ADMITER_LOGO_ICO, format="ICO",
                    sizes=[(16, 16), (32, 32), (48, 48), (64, 64),
                           (128, 128), (256, 256)],
                )
            except Exception as e:
                print(f"[icone] falha gerando ICO: {e}")

        # Windows nativo: .ico funciona melhor (taskbar + title bar)
        if ADMITER_LOGO_ICO.exists():
            try:
                self.iconbitmap(default=str(ADMITER_LOGO_ICO))
                return
            except tk.TclError:
                pass

        # Fallback: iconphoto com PNG via Pillow
        if _HAS_PIL and ADMITER_LOGO_PNG.exists():
            try:
                img = Image.open(ADMITER_LOGO_PNG)
                self._icon_ref = ImageTk.PhotoImage(img)
                self.iconphoto(True, self._icon_ref)
            except Exception as e:
                print(f"[icone] falha aplicando iconphoto: {e}")

    def _abrir_sobre(self):
        """Modal 'Sobre' com logo, versão, descrição amigável e by CrosaraTech."""
        win = tk.Toplevel(self)
        win.title(f"Sobre o {APP_NAME}")
        win.geometry("540x720")
        win.resizable(False, False)
        win.transient(self)
        win.configure(bg=COR_WHITE)

        # Logo grande
        if _HAS_PIL and ADMITER_LOGO_PNG.exists():
            try:
                img = Image.open(ADMITER_LOGO_PNG)
                if img.mode != "RGBA":
                    img = img.convert("RGBA")
                img.thumbnail((130, 130), Image.LANCZOS)
                self._sobre_logo_ref = ImageTk.PhotoImage(img)
                tk.Label(win, image=self._sobre_logo_ref, bg=COR_WHITE).pack(pady=(25, 8))
            except Exception:
                pass

        # Título grande
        tk.Label(win, text=APP_NAME, font=("Segoe UI", 30, "bold"),
                 fg=COR_ORANGE, bg=COR_WHITE).pack()
        tk.Label(win, text=APP_TAGLINE, font=("Segoe UI", 11),
                 fg=COR_TEXT_MUTED, bg=COR_WHITE).pack()
        tk.Label(win, text=f"Versão {APP_VERSION}", font=("Segoe UI", 9),
                 fg=COR_TEXT_MUTED, bg=COR_WHITE).pack(pady=(2, 18))

        # Pitch curto
        pitch = (
            "Seu assistente automático de admissões.\n"
            "Lê os documentos do email, organiza tudo e cadastra\n"
            "no eContador sem você precisar digitar nada."
        )
        tk.Label(win, text=pitch, font=("Segoe UI", 10),
                 fg=COR_TEXT_DARK, bg=COR_WHITE, justify="center").pack(pady=(0, 18))

        # Caixa "O que ele faz por você"
        box = tk.Frame(win, bg=COR_BEIGE, padx=20, pady=15)
        box.pack(padx=30, fill="x")

        tk.Label(box, text="O que ele faz por você",
                 font=("Segoe UI", 10, "bold"), fg=COR_TEXT_DARK,
                 bg=COR_BEIGE).pack(anchor="w", pady=(0, 8))

        pontos = [
            ("📬", "Monitora o Gmail à procura de novas admissões"),
            ("📄", "Lê PDFs, fotos e documentos escaneados — até manuscritos"),
            ("🔎", "Confirma os dados do CPF em fontes oficiais"),
            ("✅", "Valida tudo e envia direto pro eContador"),
            ("⚠️", "Quando algo falta, abre uma pendência clara pra você resolver"),
            ("🛡️", "Bloqueia cadastros duplicados antes de enviar"),
            ("🧮", "Conta cada pendência uma vez — não mais o histórico inteiro"),
            ("🩹", "Distingue erro técnico nosso de informação que falta do cliente"),
            ("💸", "Mostra quanto custou cada admissão em tempo real"),
        ]
        for ico, txt in pontos:
            row = tk.Frame(box, bg=COR_BEIGE)
            row.pack(fill="x", pady=2, anchor="w")
            tk.Label(row, text=ico, bg=COR_BEIGE, font=("Segoe UI", 11),
                     width=2, anchor="w").pack(side="left")
            tk.Label(row, text=txt, bg=COR_BEIGE, fg=COR_TEXT_DARK,
                     font=("Segoe UI", 9), anchor="w",
                     wraplength=420, justify="left").pack(side="left", padx=(4, 0))

        # Assinatura
        tk.Label(win, text="", bg=COR_WHITE).pack(pady=6)
        tk.Label(win, text="feito por", font=("Segoe UI", 9),
                 fg=COR_TEXT_MUTED, bg=COR_WHITE).pack()
        tk.Label(win, text=APP_VENDOR, font=("Segoe UI", 16, "bold"),
                 fg=COR_ORANGE, bg=COR_WHITE).pack()

        ttk.Button(win, text="Fechar", command=win.destroy).pack(pady=20)

    def _atualizar_status_dot(self):
        """Cor do dot da pill: verde rodando, laranja aguardando, cinza parado."""
        if not hasattr(self, "_status_dot"):
            return
        s = (self.status_var.get() or "").lower()
        if "executando" in s or "rodando" in s:
            self._status_dot.configure(fg=COR_SUCCESS)
        elif "aguard" in s or "próxima" in s or "parando" in s:
            self._status_dot.configure(fg=COR_ORANGE)
        else:
            self._status_dot.configure(fg="#888")

    # ---- Toast notification + dark mode -------------------------

    def _notify_toast(self, titulo: str, msg: str):
        """Tenta toast Windows via plyer; fallback popup Tk topmost."""
        if _HAS_PLYER:
            try:
                _plyer_notif.notify(
                    title=titulo, message=msg, app_name="Crosara DP", timeout=6,
                )
                return
            except Exception:
                pass
        # Fallback: popup leve no canto
        popup = tk.Toplevel(self)
        popup.title(titulo)
        popup.attributes("-topmost", True)
        popup.geometry("400x100+100+100")
        ttk.Label(popup, text=titulo, font=("Segoe UI", 11, "bold")).pack(padx=15, pady=(10, 0))
        ttk.Label(popup, text=msg, wraplength=380).pack(padx=15, pady=(2, 10))
        popup.after(6000, popup.destroy)

    def _toggle_dark_mode(self):
        """Aplica paleta dark/light varrendo TODO o widget tree (tk) +
        ttk.Style. Mapeia bg/fg por correspondência slot-a-slot entre
        _PALETA_LIGHT e _PALETA_DARK.

        Limitação conhecida: widgets criados DEPOIS do toggle (ex: dialog
        Resolver Pendência aberto depois de ativar dark) começam em light.
        Pra contornar, abrir e fechar dialogs sob o tema escolhido.
        """
        is_dark = bool(self.dark_mode_var.get())
        pal_atual = _PALETA_DARK if is_dark else _PALETA_LIGHT
        pal_outro = _PALETA_LIGHT if is_dark else _PALETA_DARK

        # Constrói mapa: cor_atual_no_outro_palette → cor_no_palette_alvo
        # (mais cores adicionais que aparecem em ambos como "acentos" puros
        # — ex: COR_ORANGE — não entram no map, ficam intactos)
        bg_map: dict[str, str] = {}
        fg_map: dict[str, str] = {}
        for slot, novo in pal_atual.items():
            antigo = pal_outro[slot]
            if antigo == novo:
                continue
            # bg slots
            if slot in ("main_bg", "card_bg", "sidebar_bg", "log_bg", "sub_bg"):
                bg_map[antigo] = novo
            # fg slots
            elif slot in ("text_dark", "text_light", "text_muted"):
                fg_map[antigo] = novo

        # Caso especial: foreground do log_text que era COR_BEIGE
        # (cor "text_light" no contexto do log)
        if is_dark:
            fg_map[COR_BEIGE] = pal_atual["text_light"]
        else:
            fg_map[_PALETA_DARK["text_light"]] = COR_BEIGE

        # Caminha recursivamente, trocando bg/fg
        def remapear(w):
            for attr, mapa in (("bg", bg_map), ("background", bg_map),
                                ("fg", fg_map), ("foreground", fg_map)):
                try:
                    cur = str(w.cget(attr))
                except tk.TclError:
                    continue
                if cur in mapa:
                    try:
                        w.configure({attr: mapa[cur]})
                    except tk.TclError:
                        pass
            for child in w.winfo_children():
                remapear(child)

        remapear(self)

        # ttk styles — reconfigura cores baseadas na paleta atual
        style = ttk.Style()
        p = pal_atual
        style.configure("TFrame", background=p["main_bg"])
        style.configure("TLabel", background=p["main_bg"], foreground=p["text_dark"])
        style.configure("TLabelframe", background=p["main_bg"], foreground=p["text_dark"])
        style.configure("TLabelframe.Label", background=p["main_bg"], foreground=p["text_dark"])
        style.configure("Card.TFrame", background=p["card_bg"])
        style.configure("Card.TLabel", background=p["card_bg"], foreground=p["text_dark"])
        style.configure("TCheckbutton", background=p["main_bg"], foreground=p["text_dark"])
        style.map("TCheckbutton",
                  background=[("active", p["main_bg"])],
                  foreground=[("active", p["text_dark"])])
        style.configure("TButton", background=p["card_bg"], foreground=p["text_dark"])
        style.map("TButton",
                  background=[("active", p["sub_bg"])],
                  foreground=[("active", p["text_dark"])])
        style.configure("Treeview",
                        background=p["card_bg"], fieldbackground=p["card_bg"],
                        foreground=p["text_dark"])
        style.configure("Treeview.Heading",
                        background=p["sidebar_bg"], foreground=COR_WHITE)
        style.configure("TEntry",
                        fieldbackground=p["card_bg"], foreground=p["text_dark"])
        style.configure("TSpinbox",
                        fieldbackground=p["card_bg"], foreground=p["text_dark"])
        style.configure("TCombobox",
                        fieldbackground=p["card_bg"], foreground=p["text_dark"])

        # Atualiza nav buttons do sidebar (cores explícitas — não vêm via ttk)
        if hasattr(self, "_nav_buttons"):
            sidebar_bg = p["sidebar_bg"]
            for k, btn in self._nav_buttons.items():
                # Botão ativo (laranja) mantém. Inativos seguem a paleta.
                if str(btn.cget("bg")) != COR_ORANGE:
                    btn.configure(bg=sidebar_bg, fg=p["text_light"],
                                  activebackground=COR_NAVY_HOVER if not is_dark else "#2A2D34",
                                  activeforeground=COR_WHITE)

        self.gui_q.put(("log", f"Modo {'escuro' if is_dark else 'claro'} ativado"))

    # ---- Resolver pendência ------------------------------------

    def _iter_filhos_admissao(self, tree: ttk.Treeview):
        """Itera sobre TODOS os filhos (admissões) das árvores agrupadas por dia.
        Substitui `tree.get_children('')` que agora retorna apenas os nós pai
        (agrupadores). v2.8.0+."""
        for pai_iid in tree.get_children(""):
            for filho_iid in tree.get_children(pai_iid):
                yield filho_iid

    def _get_iid_admissao_selecionada(
        self, tree: ttk.Treeview, nome_acao: str = "ação"
    ) -> str | None:
        """Retorna o iid da admissão selecionada (filho), ou None se nada.
        Se o operador clicou no nó agrupador (dia), alterna expandido/colapsado
        em vez de errar. Helper compartilhado por todos os handlers que dependem
        da seleção de uma admissão específica (v2.8.0)."""
        sel = tree.selection()
        if not sel:
            messagebox.showinfo(
                "Selecione",
                f"Selecione uma admissão na tabela antes de {nome_acao}."
            )
            return None
        iid = sel[0]
        if not tree.parent(iid):
            # Clicou no agrupador (dia) — alterna expandido/colapsado e silencia
            tree.item(iid, open=not tree.item(iid, "open"))
            return None
        return iid

    def _reconstruir_ts_da_linha(
        self, tree: ttk.Treeview, iid: str
    ) -> str:
        """Recompõe 'YYYY-MM-DD HH:MM:SS' a partir do nó pai (dia) + linha
        filha (hora). Usado pra passar timestamp completo pro dialog de
        Resolver Pendência (header mostra 'Quando: ...')."""
        import re
        parent_iid = tree.parent(iid)
        if not parent_iid:
            return ""
        dia_label = str(tree.item(parent_iid, "text") or "")
        values = tree.item(iid)["values"] or []
        hora = str(values[0]) if values else ""
        m = re.search(r"(\d{2}/\d{2}/\d{4})", dia_label)
        if not m:
            return hora
        try:
            d = datetime.strptime(m.group(1), "%d/%m/%Y")
            iso = d.strftime("%Y-%m-%d")
            return f"{iso} {hora}" if hora else iso
        except ValueError:
            return hora

    def _resolver_selecionada(self):
        iid = self._get_iid_admissao_selecionada(self.tree_pend, "resolver")
        if not iid:
            return
        values_originais = self.tree_pend.item(iid)["values"] or []
        # Reconstrói o ts completo (dia+hora) — o dialog usa pra header
        ts_completo = self._reconstruir_ts_da_linha(self.tree_pend, iid)
        # Substitui col 0 (hora) pelo ts completo pra compat com dialog antigo
        if values_originais:
            values = (ts_completo,) + tuple(values_originais[1:])
        else:
            values = tuple(values_originais)
        # v2.13.2: passa o msg_id Gmail (vem das tags da linha) pro dialog —
        # antes o dialog precisava adivinhar olhando os valores das colunas, o
        # que falhava quando CNPJ era confundido com msg_id (caso ELIANE).
        msg_id_gmail = self._msg_id_from_selection(self.tree_pend)
        ResolverPendenciaDialog(
            parent=self,
            valores_linha=tuple(str(v) for v in values),
            config=self.config,
            gui_q=self.gui_q,
            on_resolved=self._refresh_tabelas,
            msg_id_gmail=msg_id_gmail,
        )

    def _achar_payload_por_nome(self, nome: str) -> Path | None:
        """Procura em payloads/<ts>_<msgid>.json o arquivo cujo conteúdo
        contém esse nome (mais recente primeiro).

        v2.11.3: rejeita nomes INVÁLIDOS (vazio, "?", curtos demais) — antes
        o "?" das pendências sem nome batia em qualquer payload e mostrava
        dados de OUTRO funcionário. Caso real: Rayssa (falha técnica 429)
        sem nome → busca por "?" → pegou payload do Yuri.
        """
        if not PAYLOADS_DIR.exists() or not nome:
            return None
        nome_limpo = nome.strip()
        # Filtro: nomes inválidos / placeholders
        if (
            not nome_limpo
            or nome_limpo in ("?", "-", "—", "(?)", "(sem nome)", "(?", "?")
            or len(nome_limpo) < 5
        ):
            return None
        nome_upper = nome_limpo.upper()
        for f in sorted(PAYLOADS_DIR.glob("*.json"), reverse=True):
            try:
                content = f.read_text(encoding="utf-8")
                if nome_upper in content.upper():
                    return f
            except Exception:
                pass
        return None

    def _enviar_mesmo_assim(self):
        """POSTa o payload de uma pendência mesmo com campos faltantes —
        exceto empresa e função, que são obrigatórios (definem PRA ONDE vai
        a admissão). DP completa o resto manualmente no Desktop."""
        iid = self._get_iid_admissao_selecionada(self.tree_pend, "enviar")
        if not iid:
            return
        sel = (iid,)  # mantém var pra compat com código antigo abaixo

        values = tuple(str(v) for v in self.tree_pend.item(sel[0])["values"])
        # v2.8.0: col 0 agora é "hora" (não ts completo). Layout: 7 cols.
        # Layouts legados ainda suportados (6 / 5 cols pra planilhas antigas).
        if len(values) == 7:
            _hora, nome, empresa, cnpj, _tipo, _cargo_ia, _proc = values
        elif len(values) == 6:
            _hora, nome, empresa, cnpj, _tipo, _proc = values
        else:
            nome, empresa, cnpj, _tipo, _proc = values

        payload_path = self._achar_payload_por_nome(nome)
        if not payload_path:
            # Pendências antigas (antes do fix de salvar_payload v2.4.x) não
            # têm payload no disco. Oferece reprocessar — vai usar a regra do
            # ASO + pré-enrichment + sanitizar, dando uma chance real de subir.
            msg_id = None
            try:
                tags = [str(t) for t in (self.tree_pend.item(sel[0])["tags"] or [])]
                # msg_id é a 2ª tag (1ª é o tipo: interno/cliente)
                for t in tags:
                    if t and t not in ("interno", "cliente") and not t.startswith("velha_"):
                        msg_id = t
                        break
            except Exception:
                pass
            opt = messagebox.askyesnocancel(
                "Sem payload salvo",
                f"Não encontrei payload salvo para '{nome}' em payloads/.\n\n"
                f"Provavelmente esta pendência foi criada antes do fix de v2.4.x "
                f"(que estava engolindo erros silenciosamente ao salvar) ou houve "
                f"falha de gravação.\n\n"
                f"SIM   → Reprocessar o email agora (custa ~US$ 0,15 em Claude; "
                f"vai gerar payload novo aplicando regra do ASO + enrichment)\n"
                f"NÃO   → Cancelar (use 'Marcar como resolvido manualmente' se "
                f"já cadastrou direto no eContador)"
            )
            if opt is True and msg_id:
                threading.Thread(
                    target=self._reprocessar_worker,
                    args=(msg_id,), daemon=True,
                ).start()
            elif opt is True and not msg_id:
                messagebox.showerror(
                    "Sem msg_id",
                    "Não consegui identificar o msg_id desta linha pra "
                    "reprocessar. Use 'Marcar como resolvido manualmente'."
                )
            return

        try:
            doc = json.loads(payload_path.read_text(encoding="utf-8"))
        except Exception as e:
            messagebox.showerror("Payload inválido", f"Falha lendo {payload_path.name}:\n{e}")
            return

        payload = doc.get("payload") or {}
        attrs_full = (payload.get("data") or {}).get("attributes") or {}
        rels = ((payload.get("data") or {}).get("relationships") or {})
        empresa_rel = (rels.get("empresa") or {}).get("data") or {}
        funcao_rel = (rels.get("funcao") or {}).get("data") or {}
        estado_rel = (rels.get("estado") or {}).get("data") or {}
        empresa_id = empresa_rel.get("id")
        funcao_id = funcao_rel.get("id")
        estado_id = estado_rel.get("id")

        faltam_obrigatorios = []
        if not empresa_id:
            faltam_obrigatorios.append("empresa (CNPJ não resolvido no eContador)")
        if not funcao_id:
            faltam_obrigatorios.append("função (cargo não cadastrado no eContador)")

        # Endereço obrigatório (cep, rua, bairro, cidade, estado). `numero`
        # continua opcional — 0 = "sem número" (DP marca checkbox no Desktop).
        endereco_faltando = []
        for campo in ("cep", "rua", "bairro", "cidade"):
            valor = attrs_full.get(campo)
            if not valor or (isinstance(valor, str) and not valor.strip()):
                endereco_faltando.append(campo)
        if not estado_id:
            endereco_faltando.append("estado")
        if endereco_faltando:
            faltam_obrigatorios.append(
                "endereço (" + ", ".join(endereco_faltando) + ")"
            )

        if faltam_obrigatorios:
            messagebox.showwarning(
                "Não dá pra enviar",
                "Empresa, função e endereço completo são obrigatórios mesmo "
                "no envio forçado.\n\n"
                "Faltam: " + ", ".join(faltam_obrigatorios) + "\n\n"
                "Use 'Resolver pendência' pra preencher manualmente."
            )
            return

        endereco_resumo = (
            f"{attrs_full.get('rua', '?')}, {attrs_full.get('numero', '?')} — "
            f"{attrs_full.get('bairro', '?')}, {attrs_full.get('cidade', '?')}/"
            f"UF#{estado_id} (CEP {attrs_full.get('cep', '?')})"
        )
        resumo = (
            f"Nome: {attrs_full.get('nome', '(?)')}\n"
            f"CPF: {attrs_full.get('cpf', '(?)')}\n"
            f"Empresa ID: {empresa_id}\n"
            f"Função ID: {funcao_id}\n"
            f"Endereço: {endereco_resumo}\n"
            f"Admissão: {attrs_full.get('admissao', '(?)')}\n"
            f"Salário: {attrs_full.get('salario', '(?)')}"
        )
        if not messagebox.askyesno(
            "Tem certeza?",
            "Vai POSTar a admissão MESMO com campos faltantes.\n"
            "O DP precisará completar o resto manualmente no Alterdata Desktop.\n\n"
            f"{resumo}\n\n"
            "Continuar?"
        ):
            return

        # v2.14.1 — UI consulta duplicata (pra mostrar o messagebox) e depois
        # delega POST/registro/log/label pro wrapper. permitir_duplicata=True
        # quando o usuário já confirmou aqui, pra que o wrapper não pule.
        cnpj_doc = str((doc.get("resolucao") or {}).get("cnpj_empresa") or cnpj or "")
        try:
            hits_preview = idempotencia.consultar_duplicata(attrs_full.get("cpf"), cnpj_doc)
        except Exception:
            hits_preview = []
        permitir_dup = False
        if hits_preview:
            if not messagebox.askyesno(
                "⚠ POSSÍVEL DUPLICATA",
                "Este CPF JÁ FOI POSTado no eContador:\n\n"
                + idempotencia.descricao_duplicatas(hits_preview)
                + "\n\nPOSTar de novo cria um candidato DUPLICADO que o DP "
                  "terá que excluir manualmente no eContador.\n\n"
                  "POSTar mesmo assim?",
                default=messagebox.NO,
            ):
                return
            permitir_dup = True

        msg_id_doc = str(doc.get("msg_id", ""))
        api = EContadorAPI(self.config.base_url, self.config.token)
        gmail_para_label = None
        try:
            # gmail pra aplicar label processado após sucesso (ITEM 7) —
            # opcional, falha silenciosa se OAuth não estiver disponível
            try:
                if msg_id_doc:
                    gmail_para_label = GmailClient()
            except Exception:
                gmail_para_label = None
            res = postar_candidato_registrado(
                api, payload,
                cpf=attrs_full.get("cpf"), cnpj=cnpj_doc, nome=nome,
                origem="ui_envio_forcado", msg_id=msg_id_doc,
                permitir_duplicata=permitir_dup,
                payload_path=payload_path,
                gmail=gmail_para_label,
                label_processado=self.config.label_processado,
                label_pendente_remover=[msg_id_doc] if msg_id_doc else None,
            )
            if res.ok:
                proc_extra = " — POST pulado (idempotência)" if res.pulou else ""
                registrar_admissao_planilha(
                    nome=nome, empresa=empresa, cnpj=cnpj,
                    procedencia=(
                        f"Cadastrado (envio forçado) — candidato "
                        f"{res.candidato_id}{proc_extra}"
                    ),
                    msg_id=msg_id_doc,
                )
                if res.pulou:
                    messagebox.showinfo(
                        "Já estava cadastrado",
                        f"O candidato {res.candidato_id} já existia no eContador "
                        f"para essa empresa.\n\nNão criei um novo. Marquei o email "
                        f"como processado pra fechar o ciclo."
                    )
                else:
                    messagebox.showinfo(
                        "Enviado",
                        f"Candidato criado: {res.candidato_id}\n\n"
                        "Lembre o DP de completar os campos pendentes no Desktop."
                    )
                self.gui_q.put((
                    "log",
                    f"🚀 Envio forçado de '{nome}' → "
                    f"candidato {res.candidato_id}{proc_extra}"
                ))
                self._refresh_tabelas()
            else:
                # ITEM 8: erro HTTP = falha técnica, NÃO pendência cliente
                messagebox.showerror(
                    "Falha técnica no POST",
                    f"O eContador rejeitou o cadastro: {res.erro_ref}\n\n"
                    f"{(res.body_err or '')[:800]}\n\n"
                    "(Esta é uma falha do nosso lado — não cobre o cliente.)"
                )
        finally:
            api.close()
            if gmail_para_label is not None:
                try:
                    gmail_para_label.close()
                except Exception:
                    pass

    # ---- Fechamento --------------------------------------------

    def _on_close(self):
        if self.polling:
            if not messagebox.askyesno("Polling ativo",
                                       "O polling está rodando. Tem certeza que quer fechar?\n"
                                       "(A passada atual termina antes de fechar)"):
                return
            self.polling = False
        self.destroy()


# ============================================================
# Diálogo: Resolver Pendência
# ============================================================

class ResolverPendenciaDialog(tk.Toplevel):
    """Dialog modal pra resolver uma pendência. 2 modos:

    1. **Form estruturado** (default quando procedência tem "faltam: A, B, C"):
       um Entry por campo faltante, com hint de formato. Submit injeta no
       payload e POSTa. Mais amigável que JSON cru.
    2. **Editor JSON** (fallback ou opt-in): edição livre do payload
       inteiro. Útil pra pendências internas ou casos complexos.

    Botão extra "Marcar como resolvido manualmente" pra quando DP
    cadastrou direto no eContador Desktop sem POST."""

    def __init__(self, parent, valores_linha, config, gui_q, on_resolved,
                 msg_id_gmail: str | None = None):
        super().__init__(parent)
        self.config = config
        self.gui_q = gui_q
        self.on_resolved = on_resolved
        self.valores = valores_linha
        # v2.13.2: msg_id Gmail é passado explicitamente pelo caller (vem das
        # tags da treeview, não dos valores). Antes a gente tentava adivinhar
        # olhando as colunas e o CNPJ batia no filtro de hex.
        self.msg_id_gmail = msg_id_gmail or None
        # Compat: 7 cols (v2.2.0+: +cargo_ia), 6 (com timestamp), 5 (legacy sem ts)
        if len(valores_linha) == 7:
            ts, nome, empresa, cnpj, tipo, _cargo_ia, procedencia = valores_linha
        elif len(valores_linha) == 6:
            ts, nome, empresa, cnpj, tipo, procedencia = valores_linha
        else:
            nome, empresa, cnpj, tipo, procedencia = valores_linha
            ts = ""

        self.title(f"Resolver Pendência — {nome}")
        self.geometry("950x720")
        self.transient(parent)

        # Header com info
        f_head = ttk.LabelFrame(self, text="Admissão", padding=12)
        f_head.pack(fill="x", padx=15, pady=(15, 8))
        ttk.Label(f_head, text=nome, font=("Segoe UI", 13, "bold")).pack(anchor="w")
        ttk.Label(f_head, text=f"Empresa: {empresa}").pack(anchor="w")
        ttk.Label(f_head, text=f"CNPJ: {cnpj}").pack(anchor="w")
        if ts:
            ttk.Label(f_head, text=f"Quando: {ts}").pack(anchor="w")
        cor_tipo = "#e65100" if tipo == "interno" else "#1565c0"
        ttk.Label(f_head, text=f"Tipo de pendência: {tipo.upper()}",
                  foreground=cor_tipo, font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(5, 0))
        ttk.Label(f_head, text=f"Motivo: {procedencia}",
                  wraplength=900, justify="left").pack(anchor="w", pady=(5, 0))

        # Procura o payload correspondente
        self.payload_path = self._achar_payload(nome)

        # Detecta campos faltando da procedência → habilita form estruturado
        # Padrões aceitos: "faltam: A, B, C", "faltam A, B e C", "falta X"
        self.campos_faltando = self._parsear_campos_faltando(procedencia)

        # Notebook interno: Form vs JSON
        nb_modo = ttk.Notebook(self)
        nb_modo.pack(fill="both", expand=True, padx=15, pady=(8, 0))

        # ---- Aba Form Estruturado ----
        f_form = ttk.Frame(nb_modo)
        nb_modo.add(f_form, text="📝 Form estruturado")
        self._build_form_estruturado(f_form)

        # ---- Aba JSON Editor ----
        f_json = ttk.Frame(nb_modo)
        nb_modo.add(f_json, text="{ }  Editor JSON")
        if self.payload_path:
            ttk.Label(f_json, text=f"Payload: {self.payload_path.name}",
                      foreground="#666").pack(anchor="w", padx=5, pady=(5, 3))
        else:
            ttk.Label(f_json, text="⚠ Sem payload em payloads/",
                      foreground="#c62828").pack(anchor="w", padx=5, pady=(5, 3))

        self.txt = scrolledtext.ScrolledText(f_json, font=("Consolas", 9), wrap="none")
        self.txt.pack(fill="both", expand=True, padx=5, pady=5)

        if self.payload_path:
            try:
                doc = json.loads(self.payload_path.read_text(encoding="utf-8"))
                payload = doc.get("payload") or {}
                self.txt.insert("1.0", json.dumps(payload, ensure_ascii=False, indent=2))
            except Exception as e:
                self.txt.insert("1.0", f"// Erro carregando payload:\n// {e}")
        else:
            self.txt.insert("1.0",
                "// Sem payload disponível.\n"
                "// Use 'Marcar como resolvido manualmente' se você cadastrou no eContador direto."
            )

        # Se há campos faltando, abre direto no form. Senão, no JSON.
        if self.campos_faltando:
            nb_modo.select(f_form)
        else:
            nb_modo.select(f_json)

        # Botões inferiores
        f_btn = ttk.Frame(self)
        f_btn.pack(fill="x", padx=15, pady=12)
        # "Aplicar form e POSTar" SEMPRE aparece quando há campos no form.
        # Se não houver payload salvo (pendência antiga, antes da refatoração
        # que salva payload pra toda pendência interna), o handler oferece
        # reprocessar pra gerar o payload primeiro.
        if self.campos_faltando:
            ttk.Button(f_btn, text="📤  Aplicar form e POSTar",
                       command=self._aplicar_form_e_postar).pack(side="left", padx=5)
        if self.payload_path:
            ttk.Button(f_btn, text="📤  POSTar JSON editado",
                       command=self._postar_json_cru).pack(side="left", padx=5)
        ttk.Button(f_btn, text="✅  Marcar como resolvido manualmente",
                   command=self._marcar_resolvido).pack(side="left", padx=5)
        ttk.Button(f_btn, text="Cancelar", command=self.destroy).pack(side="right", padx=5)

    @staticmethod
    def _parsear_data(valor: str) -> str:
        """Aceita dd/mm/aaaa, dd-mm-aaaa ou aaaa-mm-dd e retorna no formato
        ISO 8601 (YYYY-MM-DD) que o eContador espera. Levanta ValueError se
        não conseguir parsear ou se a data for inválida (ex: 31/02)."""
        import datetime as _dt
        v = (valor or "").strip()
        # Tenta formatos comuns
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
            try:
                d = _dt.datetime.strptime(v, fmt).date()
                return d.isoformat()
            except ValueError:
                continue
        raise ValueError(f"Data '{v}' não está em dd/mm/aaaa, dd-mm-aaaa nem aaaa-mm-dd")

    def _parsear_campos_faltando(self, procedencia: str) -> list[str]:
        """Extrai lista de labels faltantes de 'Pendente cliente — faltam: A, B'.
        Também detecta casos especiais (data no passado, etc.) e injeta o
        campo correspondente.
        Retorna [] se não conseguir parsear."""
        import re

        # Caso especial 1: "data de admissão ... ficou no passado"
        # ou "admissão precisa ser >= hoje" → pede nova data de admissão
        # Inclui também "data de admissão não informada" / "ausente"
        if re.search(r"data de admiss[ãa]o.*passado|"
                     r"admiss[ãa]o precisa ser|"
                     r"data de admiss[ãa]o.*n[ãa]o.*informad|"
                     r"data de admiss[ãa]o.*ausente|"
                     r"data de admiss[ãa]o.*n[ãa]o.*localizad",
                     procedencia, re.IGNORECASE):
            return ["Data de Admissão"]

        # Caso especial 2: "Não consegui identificar a data de admissão nem
        # a data do exame admissional (ASO)" → pede uma das 2
        if re.search(r"data de admiss[ãa]o.*exame admissional|"
                     r"identificar.*data.*admiss[ãa]o",
                     procedencia, re.IGNORECASE):
            return ["Data de Admissão", "Data ASO"]

        # Caso especial 3: função não bateu nos X-marcados → pede código
        if re.search(r"fun[çc][ãa]o n[ãa]o encontrada|"
                     r"fun[çc][ãa]o.*X-marcado|"
                     r"X-marcado.*planilha CBO|"
                     r"sem cargo.*marcado",
                     procedencia, re.IGNORECASE):
            return ["Função (código)"]

        # Caso especial 3.1 (v2.11.5): ESTÁGIO sem alias salvo → pede código
        # também (a função pra estágio tem código diferente da CLT, mas o campo
        # do form é o mesmo).
        if re.search(
            r"est[áa]gio.*sem alias|"
            r"est[áa]gio.*detectado.*sem.*alias|"
            r"estagi[áa]ri[oa].*sem alias|"
            r"sem alias.*est[áa]gio|"
            r"alias de est[áa]gio",
            procedencia, re.IGNORECASE,
        ):
            return ["Função (código)"]

        # Caso especial 4: endereço (CEP, comprovante, etc) — comum quando
        # Claude marca _pendente por falta de comprovante limpo.
        if re.search(
            r"CEP.*n[ãa]o.*localizado|"
            r"comprovante.*endere[çc]o|"
            r"endere[çc]o.*terceiro|"
            r"endere[çc]o.*n[ãa]o.*identificad",
            procedencia, re.IGNORECASE,
        ):
            return ["CEP", "Rua", "Bairro", "Cidade"]

        # Caso especial 5: CTPS faltando — NÃO mostra como campo!
        # CTPS é derivada do CPF automaticamente (regra do escritório).
        # Quando o Claude marca pendente por CTPS, ignoramos e o pipeline
        # gera de int(CPF[:7]) / CPF[7:11]. Se motivo é APENAS sobre CTPS,
        # retorna lista vazia → cai no fallback genérico.
        if re.search(r"CTPS.*n[ãa]o.*localiza|"
                     r"falta.*CTPS|sem CTPS",
                     procedencia, re.IGNORECASE):
            return []  # CTPS não precisa ser preenchida manualmente

        # Caso especial 6: RG faltando
        if re.search(r"RG.*n[ãa]o.*localizad|"
                     r"identidade.*ausente|"
                     r"RG.*n[ãa]o convencional",
                     procedencia, re.IGNORECASE):
            return ["RG", "Data RG", "Órgão Emissor"]

        # Caso especial 7: salário não informado / sem valor / ausente
        # (Caso real GABRIELY 10/06/2026: "Salário não informado — email diz
        # apenas 'SALARIO BASE' sem valor numérico").
        if re.search(r"sal[áa]rio.*n[ãa]o.*informad|"
                     r"sem.*sal[áa]rio|"
                     r"sal[áa]rio.*ausente|"
                     r"sal[áa]rio.*sem.*valor|"
                     r"sal[áa]rio.*contratual|"
                     r"sem.*valor.*num[ée]rico",
                     procedencia, re.IGNORECASE):
            return ["Salário Base"]

        # Padrão geral: "faltam: A, B" / "falta X"
        m = re.search(r"falta[m]?:?\s*(.+?)(?:\.|$)", procedencia, re.IGNORECASE)
        if not m:
            # Fallback: Claude marcou _pendente mas motivo não é parseável.
            # Mostra form genérico com campos críticos editáveis pra operador.
            if re.search(r"Pendente cliente|Claude marcou|_pendente",
                         procedencia, re.IGNORECASE):
                return ["Nome", "CPF", "Data de Nascimento", "Nome da Mãe",
                        "Data de Admissão", "Salário Base",
                        "CEP", "Rua", "Bairro", "Cidade"]
            return []
        bruto = m.group(1).strip()
        # Split por vírgula ou ' e '
        parts = re.split(r",\s*|\s+e\s+", bruto)
        return [p.strip() for p in parts if p.strip()]

    def _build_form_estruturado(self, parent: ttk.Frame):
        """Form com 1 Entry por campo faltante. Se não há campos parseáveis,
        mostra info amigável."""
        self.form_vars: dict[str, tk.StringVar] = {}
        self._salvar_alias_var = tk.BooleanVar(value=False)

        # ── BUG 2 fix (v2.5.0) ────────────────────────────────────────
        # Filtra campos_faltando removendo o que JÁ está preenchido no
        # payload salvo. O parser de motivo (_parsear_campos_faltando)
        # cai num fallback hardcoded de 10 campos quando não bate regex
        # específico — ignorando completamente os 32+ campos que o Claude
        # extraiu. Caso real: GABRIELY com motivo "Salário não informado"
        # mas form mostrando Nome/CPF/Data/Endereço/etc. Defesa em
        # profundidade: mesmo se o parser falhar, filtramos pelo payload.
        if self.payload_path and self.campos_faltando:
            try:
                doc = json.loads(self.payload_path.read_text(encoding="utf-8"))
                payload = doc.get("payload") or {}
                attrs = (payload.get("data") or {}).get("attributes") or {}
                rels = (payload.get("data") or {}).get("relationships") or {}

                def _ja_tem(lbl: str) -> bool:
                    attr = PSEUDO_LABELS.get(lbl) or LABEL_PRA_ATTR.get(lbl)
                    if not attr:
                        return False
                    # Pseudo-labels (_codigo_funcao): existe se relationship
                    # funcao tem id != placeholder "1"
                    if attr == "_codigo_funcao":
                        funcao_rel = (rels.get("funcao") or {}).get("data") or {}
                        return bool(funcao_rel.get("id")) and funcao_rel.get("id") != "1"
                    # Relationships diretos (Empresa, Departamento)
                    if attr in rels:
                        rel_id = (rels.get(attr) or {}).get("data") or {}
                        return bool(rel_id.get("id")) and rel_id.get("id") != "1"
                    # Attributes — vazio = string vazia, None ou 0 não-significativo
                    v = attrs.get(attr)
                    if v is None or v == "":
                        return False
                    return True

                campos_originais = list(self.campos_faltando)
                self.campos_faltando = [c for c in campos_originais if not _ja_tem(c)]
                if len(self.campos_faltando) < len(campos_originais):
                    removidos = [c for c in campos_originais if c not in self.campos_faltando]
                    logging.getLogger("admissao.ui").info(
                        f"   📋 Form filtrado: removidos {len(removidos)} campos já "
                        f"preenchidos no payload ({', '.join(removidos)})"
                    )
            except Exception as e:
                logging.getLogger("admissao.ui").warning(
                    f"   ⚠ Falha filtrando campos_faltando contra payload "
                    f"({self.payload_path.name if self.payload_path else '?'}): "
                    f"{type(e).__name__}: {e}"
                )

        canvas_frame = ttk.Frame(parent)
        canvas_frame.pack(fill="both", expand=True, padx=5, pady=5)

        if not self.campos_faltando:
            ttk.Label(
                canvas_frame,
                text=("Esta pendência não tem campos estruturados pra preencher "
                      "(provavelmente é pendência interna ou de Claude marcando "
                      "_pendente). Use a aba 'Editor JSON' pra editar o payload "
                      "manualmente, ou clique em 'Marcar como resolvido manualmente'."),
                wraplength=850,
                foreground="#666",
            ).pack(padx=10, pady=20)
            return

        ttk.Label(
            canvas_frame,
            text=f"Campos faltantes ({len(self.campos_faltando)}). "
                 f"Preencha e clique 'Aplicar form e POSTar':",
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w", pady=(0, 10))

        for label in self.campos_faltando:
            # Pseudo-labels (ex: "Função (código)") usam PSEUDO_LABELS;
            # demais mapeiam direto via LABELS_AMIGAVEIS reverso.
            attr = PSEUDO_LABELS.get(label) or LABEL_PRA_ATTR.get(label)
            row = ttk.Frame(canvas_frame)
            row.pack(fill="x", pady=4)
            ttk.Label(row, text=label, width=35, anchor="w").pack(side="left")

            v = tk.StringVar()
            self.form_vars[label] = v
            ent = ttk.Entry(row, textvariable=v, width=40)
            ent.pack(side="left", padx=5)

            hint = HINTS_FORM.get(attr or "", "")
            if hint:
                ttk.Label(row, text=hint, foreground="#888",
                          font=("Segoe UI", 8)).pack(side="left", padx=8)
            if not attr:
                ttk.Label(row, text="(label não mapeado — adicionar manualmente via JSON)",
                          foreground="#c62828", font=("Segoe UI", 8)).pack(side="left", padx=5)

        if "Função (código)" in self.campos_faltando:
            # v2.11.0: ajusta label do checkbox quando pendência é de estágio
            # — operador precisa saber que esse alias vai pra fatia separada
            eh_estagio_aviso = False
            try:
                if hasattr(self, "_payload_doc") and self._payload_doc:
                    resolucao_doc = self._payload_doc.get("resolucao") or {}
                    if "eh_estagio" in resolucao_doc:
                        eh_estagio_aviso = bool(resolucao_doc.get("eh_estagio"))
                    else:
                        erro_doc = (resolucao_doc.get("erro") or "").upper()
                        eh_estagio_aviso = "ESTÁGIO" in erro_doc or "ESTAGIO" in erro_doc
            except Exception:
                pass
            if eh_estagio_aviso:
                label_alias = (
                    "💾🎓 Salvar como alias permanente PARA ESTAGIÁRIOS deste cargo "
                    "(separado dos aliases CLT — próximos estágios do mesmo cargo sobem auto)"
                )
            else:
                label_alias = (
                    "💾 Salvar como alias permanente (aplica em todas as "
                    "admissões futuras deste cargo)"
                )
            ttk.Checkbutton(
                canvas_frame,
                text=label_alias,
                variable=self._salvar_alias_var,
            ).pack(anchor="w", padx=15, pady=(8, 0))

        # v2.12.0: checkbox pra salvar salário como padrão pra cliente + cargo.
        # Aparece SÓ quando "Salário Base" está nos campos faltantes.
        # Próxima admissão do mesmo CNPJ + cargo com "salário base" sem valor
        # → sistema pega automaticamente esse valor.
        self._salvar_salario_padrao_var = tk.BooleanVar(value=False)
        if "Salário Base" in self.campos_faltando:
            # Pega cargo e CNPJ pra mostrar contexto no label
            cargo_disp = "este cargo"
            cnpj_disp = "este cliente"
            try:
                if hasattr(self, "_payload_doc") and self._payload_doc:
                    resolucao_doc = self._payload_doc.get("resolucao") or {}
                    cargo_doc = resolucao_doc.get("cargo_extraido") or ""
                    cnpj_doc = resolucao_doc.get("cnpj_empresa") or ""
                    if cargo_doc:
                        cargo_disp = f"cargo '{cargo_doc}'"
                    if cnpj_doc:
                        cnpj_disp = f"CNPJ {cnpj_doc}"
            except Exception:
                pass
            label_sal = (
                f"💾💰 Salvar este valor como salário padrão "
                f"({cnpj_disp} + {cargo_disp}) — próximas admissões com "
                f"\"salário base\" sem valor sobem automáticas"
            )
            ttk.Checkbutton(
                canvas_frame,
                text=label_sal,
                variable=self._salvar_salario_padrao_var,
            ).pack(anchor="w", padx=15, pady=(4, 0))

    def _extrair_msg_id_dos_valores(self) -> str | None:
        """Retorna o msg_id Gmail pra reprocessar o email.

        v2.13.2: prioriza `self.msg_id_gmail` (passado pelo caller a partir
        das tags da treeview). Fallback: tenta encontrar nos valores das
        colunas (compat com chamadas antigas).

        Sobre o fallback: msg_id Gmail é hex com letras (a-f). Exigir
        `any(letra)` evita pegar CNPJ por engano (caso ELIANE 12/06/2026).
        """
        if self.msg_id_gmail:
            return self.msg_id_gmail
        for v in (self.valores or ()):
            s = str(v or "")
            if (
                len(s) >= 12
                and all(c in "0123456789abcdefABCDEF" for c in s)
                and any(c in "abcdefABCDEF" for c in s)  # ao menos 1 letra
            ):
                return s
        return None

    def _salvar_salario_padrao_se_marcado(self, salario: float | None) -> None:
        """v2.13.1: extraído do _aplicar_form_e_postar pra reusar em ambos os
        caminhos (POST direto E reprocessar). Garante que o salário marcado
        com o checkbox é salvo MESMO SE o reprocessamento for usado depois."""
        try:
            if (
                not hasattr(self, "_salvar_salario_padrao_var")
                or not self._salvar_salario_padrao_var.get()
                or not salario
            ):
                return
            from salarios_padrao import salvar as _salvar_sal_padrao
            cnpj_para_salvar = ""
            cargo_para_salvar = ""
            razao_para_salvar = ""
            # 1ª prioridade: resolução do payload salvo
            if hasattr(self, "_payload_doc") and self._payload_doc:
                resolucao_doc = self._payload_doc.get("resolucao") or {}
                cnpj_para_salvar = resolucao_doc.get("cnpj_empresa") or ""
                cargo_para_salvar = resolucao_doc.get("cargo_extraido") or ""
                razao_para_salvar = resolucao_doc.get("razao_social") or ""
            # 2ª prioridade: campos da linha da tabela (quando não tem payload)
            if not cnpj_para_salvar or not cargo_para_salvar:
                v = self.valores
                # v2.8.0+: (hora, nome, empresa, cnpj, tipo, cargo_ia, procedencia)
                if len(v) >= 7:
                    if not razao_para_salvar:
                        razao_para_salvar = str(v[2] or "")
                    if not cnpj_para_salvar:
                        cnpj_para_salvar = str(v[3] or "")
                    if not cargo_para_salvar:
                        cargo_para_salvar = str(v[5] or "")  # cargo_ia
            if cnpj_para_salvar and cargo_para_salvar:
                _salvar_sal_padrao(
                    cnpj=cnpj_para_salvar,
                    cargo=cargo_para_salvar,
                    salario=float(salario),
                    razao_social=razao_para_salvar,
                    fonte="manual",
                )
                self.gui_q.put((
                    "log",
                    f"💾💰 Salário padrão salvo: "
                    f"CNPJ {cnpj_para_salvar} + '{cargo_para_salvar}' "
                    f"→ R$ {float(salario):.2f}"
                ))
            else:
                self.gui_q.put((
                    "log",
                    f"⚠ Checkbox de salário padrão marcado mas faltou "
                    f"CNPJ ({cnpj_para_salvar!r}) ou cargo ({cargo_para_salvar!r}) "
                    f"— não salvei."
                ))
        except Exception as e:
            self.gui_q.put((
                "log",
                f"⚠ Falha salvando salário padrão: {type(e).__name__}: {e}"
            ))

    def _aplicar_form_e_postar(self):
        """Pega os valores do form, injeta nos attributes do payload e POSTa.

        Sem payload salvo (pendência antiga, antes da refatoração que persiste
        payload pra toda pendência interna): oferece reprocessar o email pra
        gerar o payload novo, ou marca como resolvido manualmente.
        """
        # v2.13.1: salva salário padrão (se marcado) ANTES de qualquer
        # caminho de saída — assim mesmo se o usuário cair no popup de
        # "Sem payload salvo" e for reprocessar, o cadastro do salário
        # já fica persistido. Aproveita o valor que o operador digitou
        # no campo Salário Base.
        try:
            salario_digitado = None
            sal_var = (self.form_vars or {}).get("Salário Base")
            if sal_var is not None:
                texto = (sal_var.get() or "").strip()
                if texto:
                    salario_digitado = float(texto.replace(",", "."))
            self._salvar_salario_padrao_se_marcado(salario_digitado)
        except Exception as e:
            self.gui_q.put((
                "log",
                f"⚠ Falha lendo salário antes do save: {type(e).__name__}: {e}"
            ))

        if not self.payload_path:
            msg_id = self._extrair_msg_id_dos_valores()
            opt = messagebox.askyesnocancel(
                "Sem payload salvo",
                "Esta pendência foi criada antes da versão atual e não tem "
                "payload salvo no disco.\n\n"
                "Sim  → Reprocessar o email (custa ~US$ 0,20 em Claude pra "
                "gerar o payload novo; depois você completa o form e POSTa)\n"
                "Não  → Cancelar (use 'Marcar como resolvido manualmente' se "
                "já cadastrou direto no eContador)"
            )
            if opt is True and msg_id:
                self.destroy()
                threading.Thread(
                    target=self.master._reprocessar_worker,  # type: ignore
                    args=(msg_id,), daemon=True,
                ).start()
            elif opt is True and not msg_id:
                messagebox.showerror(
                    "Sem msg_id",
                    "Não achei o msg_id desta linha pra reprocessar. "
                    "Use 'Marcar como resolvido manualmente'.\n\n"
                    "Dica: a pendência precisa ter um ID Gmail válido "
                    "(hexadecimal com letras) no campo de tags da linha — "
                    "pendências antigas/manuais nem sempre têm."
                )
            return

        try:
            doc = json.loads(self.payload_path.read_text(encoding="utf-8"))
        except Exception as e:
            messagebox.showerror("Erro lendo payload", str(e))
            return

        self._payload_doc = doc

        payload = doc.get("payload") or {}
        if "data" not in payload:
            messagebox.showerror("Payload inválido", "Estrutura sem 'data'.")
            return
        attrs = payload["data"].get("attributes") or {}

        # Aplica cada valor preenchido
        aplicados = []
        nao_mapeados = []
        admissao_mudou = False
        for label, var in self.form_vars.items():
            valor = (var.get() or "").strip()
            if not valor:
                continue  # vazio = não muda

            # Pseudo-label "Função (código)": busca pela coluna `codigo` da
            # planilha CBO em memória e set relationship.funcao do payload.
            if PSEUDO_LABELS.get(label) == "_codigo_funcao":
                planilha = getattr(self.master, "planilha", None) or []
                codigo_limpo = valor.strip()
                achadas = [
                    f for f in planilha
                    if str(f.get("codigo") or "").strip() == codigo_limpo
                    or str(f.get("codigo") or "").strip().lstrip("0") == codigo_limpo.lstrip("0")
                ]
                if not achadas:
                    messagebox.showerror(
                        "Código não encontrado",
                        f"Código '{codigo_limpo}' não encontrado na planilha CBO em memória.\n\n"
                        "Possibilidades:\n"
                        "  (a) Você digitou o código errado — confira a coluna 'codigo' na planilha\n"
                        "  (b) A função foi adicionada recentemente — clique '🔄 Recarregar planilha CBO'\n"
                        "  (c) A planilha está sem a coluna 'codigo' preenchida pra essa função"
                    )
                    return
                if len(achadas) > 1:
                    messagebox.showerror(
                        "Código ambíguo",
                        f"{len(achadas)} funções com código '{codigo_limpo}'. "
                        "Edite a planilha e deixe o código único."
                    )
                    return
                funcao = achadas[0]
                rels = payload["data"].setdefault("relationships", {})
                rels["funcao"] = {"data": {"type": "funcoes",
                                           "id": str(funcao["funcao_id"])}}
                attrs_payload = payload.setdefault("data", {}).setdefault("attributes", {})
                attrs_payload["nomecargo"] = str(funcao.get("nome_cargo") or "").strip()
                aplicados.append(
                    f"{label} = {funcao.get('nome_cargo')} (id {funcao['funcao_id']})"
                )
                try:
                    if hasattr(self, "_salvar_alias_var") and self._salvar_alias_var.get():
                        # Pega cargo extraído pelo Claude da resolucao salva no doc
                        cargo_orig = ""
                        eh_estagio_doc = False
                        if hasattr(self, "_payload_doc"):
                            resolucao_doc = self._payload_doc.get("resolucao") or {}
                            cargo_orig = resolucao_doc.get("cargo_extraido") or ""
                            # v2.11.0: flag persistida pelo pipeline (mais confiável
                            # que parse do texto de erro). Fallback no parse pra
                            # payloads antigos.
                            if "eh_estagio" in resolucao_doc:
                                eh_estagio_doc = bool(resolucao_doc.get("eh_estagio"))
                            else:
                                erro = (resolucao_doc.get("erro") or "").upper()
                                eh_estagio_doc = "ESTÁGIO" in erro or "ESTAGIO" in erro
                        if not cargo_orig:
                            cargo_orig = str(funcao.get("nome_cargo") or "")
                        salvar_funcao_alias(
                            cargo_orig,
                            str(funcao["funcao_id"]),
                            funcao.get("nome_cargo", ""),
                            eh_estagio=eh_estagio_doc,
                        )
                        # v2.11.2: o sanity check em salvar_funcao_alias pode
                        # forçar ESTÁGIO mesmo quando eh_estagio_doc=False
                        # (se nome_cargo contém ESTAGIO/ESTAGIARIO). Lê o
                        # estado REAL do alias salvo pra mostrar o tipo correto.
                        from funcao import consultar_funcao_alias as _consultar
                        tipo_alias = "CLT"
                        try:
                            alias_estagio_salvo = _consultar(cargo_orig, eh_estagio=True)
                            if alias_estagio_salvo and alias_estagio_salvo.get("funcao_id") == str(funcao["funcao_id"]):
                                tipo_alias = "ESTÁGIO 🎓"
                        except Exception:
                            tipo_alias = "ESTÁGIO 🎓" if eh_estagio_doc else "CLT"
                        self.gui_q.put(((
                            "log",
                            f"💾 Alias {tipo_alias} salvo: '{cargo_orig}' → "
                            f"{funcao.get('nome_cargo')} (id {funcao['funcao_id']})"
                        )))
                except Exception as e:
                    self.gui_q.put(("log", f"⚠ Falha salvando alias: {e}"))
                continue

            attr = LABEL_PRA_ATTR.get(label)
            if not attr:
                nao_mapeados.append(label)
                continue
            # Conversão por tipo
            try:
                if attr in DATE_ATTRS:
                    valor_conv = self._parsear_data(valor)
                elif attr in ("salario",):
                    valor_conv = float(valor.replace(",", "."))
                elif attr in ("cpf", "ctps", "diascontratoexperiencia"):
                    import re
                    digs = re.sub(r"\D", "", valor)
                    valor_conv = int(digs) if digs else None
                    if valor_conv is None:
                        raise ValueError(f"{label} precisa ter dígitos")
                elif attr == "primeiroemprego":
                    valor_conv = valor.lower() in ("true", "sim", "1", "y", "yes")
                else:
                    valor_conv = valor
            except Exception as e:
                messagebox.showerror(
                    "Valor inválido",
                    f"Campo '{label}' ({attr}): {e}\n\n"
                    f"Confira o formato e tente de novo."
                )
                return
            attrs[attr] = valor_conv
            aplicados.append(f"{label} = {valor_conv}")
            if attr == "admissao":
                admissao_mudou = True

        # Se admissão mudou, recalcula término de contrato = admissão + dias
        if admissao_mudou:
            try:
                import datetime as _dt
                adm = _dt.date.fromisoformat(str(attrs.get("admissao")))
                dias = int(attrs.get("diascontratoexperiencia") or 30)
                novo_termino = (adm + _dt.timedelta(days=dias)).isoformat()
                attrs["dataterminocontrato"] = novo_termino
                aplicados.append(f"Término de contrato recalculado = {novo_termino}")
            except Exception:
                pass  # não bloqueia se falhar — POST vai dizer se houver problema

        if nao_mapeados:
            messagebox.showwarning(
                "Campos não mapeados",
                f"Estes campos não pude aplicar automaticamente: {nao_mapeados}.\n"
                f"Use a aba 'Editor JSON' pra adicionar manualmente."
            )
            return

        if not aplicados:
            messagebox.showinfo("Nada pra aplicar", "Nenhum campo foi preenchido.")
            return

        payload["data"]["attributes"] = attrs

        if not messagebox.askyesno(
            "Confirmar POST",
            f"Vou aplicar {len(aplicados)} campo(s) e POSTar:\n\n"
            + "\n".join(f"  • {x}" for x in aplicados)
            + "\n\nConfirma?"
        ):
            return

        # v2.13.1: salário padrão já foi tentado no início do método (antes
        # do popup "Sem payload salvo"), então o checkbox já foi processado.

        self._postar(payload)

    def _postar_json_cru(self):
        """POSTa o JSON do editor (sem usar form)."""
        try:
            payload = json.loads(self.txt.get("1.0", "end"))
        except json.JSONDecodeError as e:
            messagebox.showerror("JSON inválido", f"{e}")
            return
        if not messagebox.askyesno("Confirmar POST", "POSTar este payload em /candidatos?"):
            return
        self._postar(payload)

    def _achar_payload(self, nome: str) -> Path | None:
        """Procura em payloads/<ts>_<msgid>.json o arquivo cujo conteúdo
        contém esse nome (mais recente primeiro).

        v2.11.4: rejeita nomes INVÁLIDOS (vazio, "?", curtos demais) — mesmo
        fix de `_achar_payload_por_nome`. Pendências sem nome (falha 429,
        Claude desistiu antes de extrair nome) NÃO devem aparecer com payload
        de outro funcionário.
        """
        if not PAYLOADS_DIR.exists() or not nome:
            return None
        nome_limpo = nome.strip()
        if (
            not nome_limpo
            or nome_limpo in ("?", "-", "—", "(?)", "(sem nome)", "(?", "?")
            or len(nome_limpo) < 5
        ):
            return None
        nome_upper = nome_limpo.upper()
        files = sorted(PAYLOADS_DIR.glob("*.json"), reverse=True)
        for f in files:
            try:
                content = f.read_text(encoding="utf-8")
                if nome_upper in content.upper():
                    return f
            except Exception:
                pass
        return None

    def _postar(self, payload: dict):
        """POSTa o payload (já validado JSON) em /candidatos. Reusado pelos
        2 fluxos: form estruturado e editor JSON cru.

        Re-aplica sanitizar_attributes antes de POST — o payload salvo no
        disco pode ter campos longos demais (ex: orgaoemissoridentidade com
        nome completo da SSP), e a UI por padrão não passa pelo sanitize.
        """
        from payload_builder import sanitizar_attributes
        try:
            attrs = (payload.get("data") or {}).get("attributes") or {}
            payload["data"]["attributes"] = sanitizar_attributes(attrs)
        except Exception:
            pass  # sanitize é defensivo — não bloqueia POST se falhar

        # v2.14.1 — UI consulta duplicata pra mostrar messagebox; wrapper único
        # cuida do resto (POST, registro, atualiza payload, label, log NDJSON).
        attrs_pre = (payload.get("data") or {}).get("attributes") or {}
        cnpj_pre = ""
        msg_id_pre = ""
        empresa_pre = ""
        if self.payload_path:
            try:
                _doc_pre = json.loads(self.payload_path.read_text(encoding="utf-8"))
                cnpj_pre = str((_doc_pre.get("resolucao") or {}).get("cnpj_empresa") or "")
                msg_id_pre = str(_doc_pre.get("msg_id", ""))
                empresa_pre = str((_doc_pre.get("resolucao") or {}).get("razao_social") or "")
            except Exception:
                pass
        try:
            hits_preview = idempotencia.consultar_duplicata(attrs_pre.get("cpf"), cnpj_pre)
        except Exception:
            hits_preview = []
        permitir_dup = False
        if hits_preview:
            if not messagebox.askyesno(
                "⚠ POSSÍVEL DUPLICATA",
                "Este CPF JÁ FOI POSTado no eContador:\n\n"
                + idempotencia.descricao_duplicatas(hits_preview)
                + "\n\nPOSTar de novo cria um candidato DUPLICADO que o DP "
                  "terá que excluir manualmente no eContador.\n\n"
                  "POSTar mesmo assim?",
                default=messagebox.NO,
            ):
                return
            permitir_dup = True

        nome_attrs = str(attrs_pre.get("nome") or "?").strip()
        # Fallback empresa/cnpj/nome a partir das colunas da tree se o payload
        # salvo não tiver (legacy/manual): v2.8.0+ tem 7 cols, 6 (sem cargo_ia),
        # 5 (legacy sem ts). Idx fixos pelo final (nome=v[idx_nome]).
        empresa = empresa_pre
        cnpj = cnpj_pre
        try:
            v = self.valores
            if (not empresa or not cnpj or nome_attrs == "?") and len(v) >= 4:
                idx_nome = 1 if len(v) >= 5 else 0
                if not nome_attrs or nome_attrs == "?":
                    nome_attrs = str(v[idx_nome] or nome_attrs)
                empresa = empresa or str(v[idx_nome + 1] or "")
                cnpj = cnpj or str(v[idx_nome + 2] or "")
        except Exception:
            pass

        api = EContadorAPI(self.config.base_url, self.config.token)
        gmail_para_label = None
        try:
            try:
                if msg_id_pre:
                    gmail_para_label = GmailClient()
            except Exception:
                gmail_para_label = None
            res = postar_candidato_registrado(
                api, payload,
                cpf=attrs_pre.get("cpf"), cnpj=cnpj or cnpj_pre, nome=nome_attrs,
                origem="ui_resolver", msg_id=msg_id_pre,
                permitir_duplicata=permitir_dup,
                payload_path=self.payload_path,
                gmail=gmail_para_label,
                label_processado=self.config.label_processado,
                label_pendente_remover=[msg_id_pre] if msg_id_pre else None,
            )
            if res.ok:
                proc_extra = " — POST pulado (idempotência)" if res.pulou else ""
                try:
                    registrar_admissao_planilha(
                        nome=nome_attrs or "?",
                        empresa=empresa or None,
                        cnpj=cnpj or None,
                        procedencia=(
                            f"Cadastrado — candidato {res.candidato_id} "
                            f"(resolvido via UI){proc_extra}"
                        ),
                        msg_id=msg_id_pre,
                    )
                except Exception as e:
                    self.gui_q.put((
                        "log",
                        f"⚠ POST OK (candidato {res.candidato_id}) mas falhou "
                        f"registrando na planilha: {type(e).__name__}: {e}"
                    ))
                if res.pulou:
                    messagebox.showinfo(
                        "Já estava cadastrado",
                        f"O candidato {res.candidato_id} já existia no eContador "
                        f"para essa empresa.\n\nNão criei outro. Email marcado "
                        f"como processado."
                    )
                else:
                    messagebox.showinfo("Sucesso", f"Candidato criado: {res.candidato_id}")
                self.gui_q.put((
                    "log",
                    f"✓ Pendência resolvida via UI → "
                    f"candidato {res.candidato_id}{proc_extra}"
                ))
                self.on_resolved()
                self.destroy()
            else:
                # ITEM 8: falha do eContador NÃO vira pendência cliente
                messagebox.showerror(
                    "Falha técnica no POST",
                    f"O eContador rejeitou o cadastro: {res.erro_ref}\n\n"
                    f"{(res.body_err or '')[:800]}\n\n"
                    "(Esta é uma falha do nosso lado — não cobre o cliente.)"
                )
        finally:
            api.close()
            if gmail_para_label is not None:
                try:
                    gmail_para_label.close()
                except Exception:
                    pass

    def _marcar_resolvido(self):
        if not messagebox.askyesno(
            "Confirmar",
            "Marcar essa pendência como resolvida manualmente?\n\n"
            "(NÃO faz POST — você marca quando já cadastrou no eContador direto.\n"
            "Aparece nova linha 'Cadastrado manualmente' na planilha.)"
        ):
            return
        # v2.11.2: extração robusta (5/6/7 cols) — antes só tratava 5/6 e
        # bug silencioso em 7 cols (v2.8.0+) sumia com o registro na planilha.
        v = self.valores
        if len(v) == 7:
            _hora, nome, empresa, cnpj, _tipo, _cargo_ia, _proc = v
        elif len(v) == 6:
            _ts, nome, empresa, cnpj, _tipo, _proc = v
        elif len(v) == 5:
            nome, empresa, cnpj, _tipo, _proc = v
        else:
            # Fallback ultra-defensivo: pega pelos índices comuns
            nome = str(v[1] if len(v) > 1 else "?")
            empresa = str(v[2] if len(v) > 2 else "")
            cnpj = str(v[3] if len(v) > 3 else "")
        registrar_admissao_planilha(
            nome=nome, empresa=empresa, cnpj=cnpj,
            procedencia=f"Cadastrado manualmente — via UI ({datetime.now().strftime('%d/%m/%Y %H:%M')})",
        )
        self.gui_q.put(("log", f"✓ {nome} marcado como resolvido manualmente"))
        self.on_resolved()
        self.destroy()


# ============================================================
# Importação manual de arquivos (v2.7.0)
# ============================================================

class ImportarArquivosDialog(tk.Toplevel):
    """Permite o operador subir admissão sem ter um email Gmail —
    seleciona PDFs/imagens locais e (opcional) adiciona contexto em texto.

    Caso de uso: cliente mandou docs por WhatsApp/Drive/pendrive, ou DP
    digitalizou contrato físico. Sem precisar criar um email fake e mandar
    pra si mesmo, importa direto e processa pelo mesmo pipeline."""

    EXTENSOES_ACEITAS = (
        ("Todos suportados", "*.pdf *.jpg *.jpeg *.png *.gif *.webp"),
        ("PDF", "*.pdf"),
        ("Imagens", "*.jpg *.jpeg *.png *.gif *.webp"),
        ("Todos arquivos", "*.*"),
    )

    def __init__(self, parent, config, claude, planilha_cbo, gui_q):
        super().__init__(parent)
        self.config = config
        self.claude = claude
        self.planilha_cbo = planilha_cbo
        self.gui_q = gui_q
        self.arquivos_selecionados: list[Path] = []

        self.title("📥 Importar arquivos manualmente")
        self.geometry("780x600")
        self.transient(parent)
        try:
            self.configure(bg=COR_BEIGE)
        except NameError:
            pass

        # Header explicativo
        f_head = ttk.LabelFrame(self, text="Importação manual", padding=12)
        f_head.pack(fill="x", padx=15, pady=(15, 8))
        ttk.Label(
            f_head,
            text=(
                "Use quando o cliente mandou docs por WhatsApp/Drive, ou quando "
                "você digitalizou contrato físico.\n"
                "Selecione os PDFs/imagens e (opcional) adicione contexto em texto. "
                "Será processado pelo mesmo pipeline de email — sobe sucesso ou vira pendência."
            ),
            wraplength=720,
            foreground="#555",
            justify="left",
        ).pack(anchor="w")

        # Seção 1: arquivos
        f_files = ttk.LabelFrame(self, text="1. Arquivos da admissão", padding=10)
        f_files.pack(fill="both", expand=False, padx=15, pady=8)

        bar_files = ttk.Frame(f_files)
        bar_files.pack(fill="x", pady=(0, 6))
        ttk.Button(bar_files, text="➕  Selecionar arquivos…",
                   command=self._selecionar_arquivos).pack(side="left", padx=3)
        ttk.Button(bar_files, text="🗑  Limpar lista",
                   command=self._limpar_lista).pack(side="left", padx=3)
        self.lbl_count = ttk.Label(bar_files, text="Nenhum arquivo selecionado",
                                    foreground="#888")
        self.lbl_count.pack(side="left", padx=12)

        # Lista de arquivos
        list_frame = ttk.Frame(f_files)
        list_frame.pack(fill="both", expand=True)
        self.lst_arquivos = tk.Listbox(list_frame, height=8, font=("Consolas", 9))
        sb = ttk.Scrollbar(list_frame, orient="vertical",
                           command=self.lst_arquivos.yview)
        self.lst_arquivos.configure(yscrollcommand=sb.set)
        self.lst_arquivos.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        # Seção 2: contexto opcional
        f_ctx = ttk.LabelFrame(self, text="2. Contexto extra (opcional)", padding=10)
        f_ctx.pack(fill="both", expand=False, padx=15, pady=8)
        ttk.Label(
            f_ctx,
            text=(
                "Texto adicional pra dar contexto ao Claude. Útil quando os documentos "
                "não trazem alguma info (ex: CNPJ da empresa, salário combinado, data de "
                "admissão diferente da ASO)."
            ),
            wraplength=720,
            foreground="#666",
            font=("Segoe UI", 8),
        ).pack(anchor="w", pady=(0, 4))
        self.txt_corpo = tk.Text(f_ctx, height=5, wrap="word",
                                  font=("Segoe UI", 9))
        self.txt_corpo.pack(fill="both", expand=True)

        # Botões finais
        f_btns = ttk.Frame(self)
        f_btns.pack(fill="x", padx=15, pady=12)
        ttk.Button(f_btns, text="🚀  Processar agora",
                   command=self._processar).pack(side="left", padx=3)
        ttk.Button(f_btns, text="Cancelar",
                   command=self.destroy).pack(side="right", padx=3)

    def _selecionar_arquivos(self):
        from tkinter import filedialog
        paths = filedialog.askopenfilenames(
            parent=self,
            title="Selecione PDFs ou imagens da admissão",
            filetypes=self.EXTENSOES_ACEITAS,
        )
        if not paths:
            return
        for p in paths:
            path = Path(p)
            if path not in self.arquivos_selecionados:
                self.arquivos_selecionados.append(path)
        self._refresh_lista()

    def _limpar_lista(self):
        self.arquivos_selecionados = []
        self._refresh_lista()

    def _refresh_lista(self):
        self.lst_arquivos.delete(0, "end")
        for p in self.arquivos_selecionados:
            try:
                tam_kb = p.stat().st_size / 1024
                tam_str = f"{tam_kb:.0f} KB" if tam_kb < 1024 else f"{tam_kb/1024:.1f} MB"
            except OSError:
                tam_str = "?"
            self.lst_arquivos.insert("end", f"  📄  {p.name}  ({tam_str})")
        n = len(self.arquivos_selecionados)
        if n == 0:
            self.lbl_count.configure(text="Nenhum arquivo selecionado",
                                      foreground="#888")
        else:
            self.lbl_count.configure(
                text=f"{n} arquivo{'s' if n != 1 else ''} selecionado{'s' if n != 1 else ''}",
                foreground="#2e7d32",
            )

    def _ler_arquivos_pra_anexos(self) -> list[dict]:
        """Carrega cada arquivo em bytes + infere MIME pela extensão."""
        EXT_TO_MIME = {
            ".pdf": "application/pdf",
            ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".gif": "image/gif",
            ".webp": "image/webp",
        }
        anexos: list[dict] = []
        for p in self.arquivos_selecionados:
            try:
                data = p.read_bytes()
            except OSError as e:
                messagebox.showerror(
                    "Erro lendo arquivo",
                    f"Não consegui ler '{p.name}':\n{type(e).__name__}: {e}"
                )
                return []
            mime = EXT_TO_MIME.get(p.suffix.lower(), "application/octet-stream")
            anexos.append({
                "filename": p.name,
                "mime": mime,
                "data": data,
            })
        return anexos

    def _processar(self):
        if not self.arquivos_selecionados:
            messagebox.showwarning(
                "Sem arquivos",
                "Selecione ao menos 1 arquivo antes de processar."
            )
            return

        anexos = self._ler_arquivos_pra_anexos()
        if not anexos:
            return  # erro já foi mostrado

        corpo = self.txt_corpo.get("1.0", "end").strip()

        if not messagebox.askyesno(
            "Confirmar importação",
            f"Processar {len(anexos)} arquivo(s) agora?\n\n"
            f"Vai consumir ~US$ 0,15 em Claude.\n"
            f"O resultado aparece em Pendentes ou Processadas como qualquer email.",
        ):
            return

        # Roda em background pra não travar UI
        self.destroy()
        threading.Thread(
            target=self._worker,
            args=(anexos, corpo),
            daemon=True,
        ).start()

    def _worker(self, anexos: list[dict], corpo: str):
        from main import processar_arquivos_avulsos, EContadorAPI
        self.gui_q.put((
            "log",
            f"📥 Importando {len(anexos)} arquivo(s) manualmente — processando..."
        ))
        api = EContadorAPI(self.config.base_url, self.config.token)
        try:
            processar_arquivos_avulsos(
                arquivos=anexos,
                corpo_texto=corpo,
                claude=self.claude,
                api=api,
                planilha_cbo=self.planilha_cbo,
                config=self.config,
            )
            self.gui_q.put((
                "log",
                "✅ Importação manual concluída — confira em Pendentes ou Processadas"
            ))
            self.gui_q.put(("refresh", None))
        except Exception as e:
            self.gui_q.put((
                "log",
                f"❌ Falha na importação manual: {type(e).__name__}: {e}"
            ))
        finally:
            try:
                api.close()
            except Exception:
                pass


# ============================================================
# Entry point
# ============================================================

def main():
    app = PipelineGUI()
    app.mainloop()


if __name__ == "__main__":
    main()
