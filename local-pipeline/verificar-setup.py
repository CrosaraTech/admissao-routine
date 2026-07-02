"""verificar-setup.py — pre-flight check pra deployment.

Roda uma série de verificações sequenciais e mostra resultado [OK]/[FAIL]
pra cada uma. Útil quando subindo o pipeline num servidor novo.

Uso:
    .\\.venv\\Scripts\\python.exe verificar-setup.py
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

ROOT = Path(__file__).parent
TOTAL_OK = 0
TOTAL_FAIL = 0


def _print(msg: str, color: str = ""):
    """Print com cores ANSI (funciona no Windows Terminal e PowerShell modernos)."""
    cores = {
        "ok":    "\x1b[32m",   # verde
        "fail":  "\x1b[31m",   # vermelho
        "warn":  "\x1b[33m",   # amarelo
        "info":  "\x1b[36m",   # ciano
        "reset": "\x1b[0m",
    }
    if color and color in cores:
        print(f"{cores[color]}{msg}{cores['reset']}")
    else:
        print(msg)


def check(nome: str, fn):
    """Roda fn(). Se passar, imprime OK. Se levantar ou retornar falso, FAIL."""
    global TOTAL_OK, TOTAL_FAIL
    _print(f"  • {nome}...", "info")
    try:
        resultado = fn()
        if resultado is False:
            _print(f"    [FAIL]", "fail")
            TOTAL_FAIL += 1
            return False
        if isinstance(resultado, str):
            _print(f"    [OK] {resultado}", "ok")
        else:
            _print(f"    [OK]", "ok")
        TOTAL_OK += 1
        return True
    except Exception as e:
        _print(f"    [FAIL] {e}", "fail")
        TOTAL_FAIL += 1
        return False


def secao(titulo: str):
    print()
    _print(f"━━━━ {titulo} ━━━━", "info")


# ============================================================
# Checks
# ============================================================

def check_python_version() -> str:
    v = sys.version_info
    if v.major != 3 or v.minor < 11:
        raise RuntimeError(f"Python {v.major}.{v.minor} — precisa ser 3.11+")
    return f"Python {v.major}.{v.minor}.{v.micro}"


def check_dep(modulo: str, instalavel: str | None = None) -> str:
    try:
        m = __import__(modulo)
        version = getattr(m, "__version__", "?")
        return f"{modulo} {version}"
    except ImportError as e:
        pacote = instalavel or modulo
        raise RuntimeError(f"módulo '{modulo}' não instalado. Rode: pip install {pacote}")


def check_arquivo(nome: str, descricao: str = ""):
    p = ROOT / nome
    if not p.exists():
        raise FileNotFoundError(f"{nome} não existe na pasta {ROOT}")
    tam = p.stat().st_size
    return f"{tam:,} bytes" + (f" — {descricao}" if descricao else "")


def check_env_var(nome: str, min_len: int = 10) -> str:
    valor = os.getenv(nome)
    if not valor:
        raise RuntimeError(f"{nome} não definida no .env (ou variável de ambiente)")
    if len(valor) < min_len:
        raise RuntimeError(f"{nome} parece truncada ({len(valor)} chars)")
    # Mascarar pra log
    if len(valor) > 30:
        mascara = valor[:8] + "..." + valor[-6:]
    else:
        mascara = "***"
    return f"presente ({len(valor)} chars: {mascara})"


def check_econtador_api() -> str:
    import httpx
    token = os.getenv("ECONTADOR_TOKEN")
    if not token:
        raise RuntimeError("ECONTADOR_TOKEN ausente — defina antes de testar")
    base = "https://dp.pack.alterdata.com.br/api/v1"
    r = httpx.get(
        f"{base}/empresas",
        params={"page[limit]": 1},
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/vnd.api+json",
        },
        timeout=15.0,
    )
    if r.status_code == 401:
        raise RuntimeError("HTTP 401 — token inválido ou expirado")
    if r.status_code != 200:
        raise RuntimeError(f"HTTP {r.status_code}: {r.text[:200]}")
    n = r.json().get("meta", {}).get("totalResourceCount", "?")
    return f"OK ({n} empresas no cadastro)"


def check_gmail_api() -> str:
    import json
    raw = os.getenv("GMAIL_TOKEN")
    if not raw:
        raise RuntimeError("GMAIL_TOKEN ausente")
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        raise RuntimeError(f"GMAIL_TOKEN não é JSON válido: {e}")
    obrig = ["token", "refresh_token", "token_uri", "client_id", "client_secret"]
    falt = [k for k in obrig if not data.get(k)]
    if falt:
        raise RuntimeError(f"GMAIL_TOKEN sem chaves obrigatórias: {falt}")
    # Testa conexão real
    from gmail_client import GmailClient
    gmail = GmailClient()
    labels = gmail.service.users().labels().list(userId="me").execute().get("labels", [])
    return f"OK ({len(labels)} labels na conta)"


def check_anthropic_api() -> str:
    import anthropic
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY ausente")
    client = anthropic.Anthropic(api_key=api_key)
    # Teste cheap — só pra validar chave
    msg = client.messages.create(
        model="claude-haiku-4-5",
        max_tokens=10,
        messages=[{"role": "user", "content": "Diga oi em 2 palavras"}],
    )
    return f"OK (modelo respondeu: '{msg.content[0].text[:30]}')"


def check_imports_pipeline() -> str:
    """Importa os módulos principais pra pegar erros de sintaxe/import cedo."""
    modulos = [
        "claude_client", "ecotador_client", "gmail_client",
        "departamento", "funcao", "payload_builder", "main", "interface",
    ]
    for m in modulos:
        __import__(m)
    return f"{len(modulos)} módulos importam sem erro"


def check_planilha_cbo() -> str:
    from openpyxl import load_workbook
    p = ROOT / "funcoes_cbo.xlsx"
    if not p.exists():
        raise FileNotFoundError("funcoes_cbo.xlsx ausente — rode gerar_planilha_funcoes.py")
    wb = load_workbook(p, read_only=True)
    ws = wb.active
    rows = sum(1 for _ in ws.iter_rows()) - 1  # menos header
    return f"{rows:,} cargos"


# ============================================================
# Run
# ============================================================

def main() -> int:
    # Carrega .env primeiro
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except ImportError:
        pass  # check_dep vai pegar isso depois

    print()
    _print("Crosara — Pipeline de Admissão — Verificação de Setup", "info")
    print()

    secao("Ambiente Python")
    check("Python 3.11+", check_python_version)
    check("dependência: httpx", lambda: check_dep("httpx"))
    check("dependência: dotenv", lambda: check_dep("dotenv", "python-dotenv"))
    check("dependência: anthropic", lambda: check_dep("anthropic"))
    check("dependência: openpyxl", lambda: check_dep("openpyxl"))
    check("dependência: google_auth_oauthlib", lambda: check_dep("google_auth_oauthlib"))
    check("dependência: googleapiclient", lambda: check_dep("googleapiclient", "google-api-python-client"))
    check("dependência: PIL (opcional, logo da UI)",
          lambda: check_dep("PIL", "Pillow"))
    check("dependência: plyer (opcional, toast Windows)",
          lambda: check_dep("plyer"))
    check("dependência: rarfile (opcional, descompactar .rar do cliente)",
          lambda: check_dep("rarfile"))
    check("unrar.exe (opcional, suporte a .rar)",
          lambda: check_arquivo("unrar.exe",
                                "baixe em rarlab.com/rar_add.htm e ponha na pasta"))

    secao("Arquivos essenciais")
    check("config.json", lambda: check_arquivo("config.json"))
    check("briefing.md", lambda: check_arquivo("briefing.md", "system prompt do Claude"))
    check("lookups.json", lambda: check_arquivo("lookups.json"))
    check("departamentos.json", lambda: check_arquivo("departamentos.json"))
    check("regras.json", lambda: check_arquivo("regras.json"))
    check("funcoes_cbo.xlsx", check_planilha_cbo)

    secao("Variáveis de ambiente (.env)")
    check("ECONTADOR_TOKEN", lambda: check_env_var("ECONTADOR_TOKEN", 100))
    check("GMAIL_TOKEN (JSON)", lambda: check_env_var("GMAIL_TOKEN", 200))
    check("ANTHROPIC_API_KEY", lambda: check_env_var("ANTHROPIC_API_KEY", 30))

    secao("Conectividade das APIs (chamadas reais)")
    check("eContador GET /empresas", check_econtador_api)
    check("Gmail API (listar labels)", check_gmail_api)
    check("Anthropic API (claude-haiku-4-5, teste mínimo)", check_anthropic_api)

    secao("Imports do pipeline")
    check("Módulos do pipeline importam", check_imports_pipeline)

    # Resumo
    print()
    _print("━" * 60, "info")
    if TOTAL_FAIL == 0:
        _print(f"✓ {TOTAL_OK} checks passaram, 0 falharam — setup OK!", "ok")
        _print("  Pode rodar run-gui.bat ou run-once.bat.", "ok")
        return 0
    else:
        _print(f"✗ {TOTAL_OK} OK, {TOTAL_FAIL} falharam.", "fail")
        _print("  Conserte os FAIL acima e rode de novo.", "warn")
        return 1


if __name__ == "__main__":
    sys.exit(main())
